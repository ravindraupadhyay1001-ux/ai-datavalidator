

# Data Validation AGENT - Enhanced FastAPI application.
# Supports: Compare, Data Quality, Column Mapping, Data Governance.
# Reference docs (data dictionary, business rules, mapping spec) can be
# uploaded to enrich each analysis module.


import asyncio
import concurrent.futures
import csv as _csv
import io
import json
import math
import os
import re
from datetime import datetime
from pathlib import Path


import time
import uuid
from collections import Counter
from difflib import SequenceMatcher
from itertools import combinations
from typing import Optional
import xml.etree.ElementTree as ET

# Thread-pool for CPU-bound analysis work -- keeps the async event loop free
_cpu_pool = concurrent.futures.ThreadPoolExecutor(max_workers=4, thread_name_prefix="dva_cpu")

import boto3
import chardet


import numpy as np
import pdfplumber
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.exceptions import RequestValidationError
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates

load_dotenv()

# -- License system


from license.license_manager import (
    load_license as _lic_load,
    get_state as _lic_state,
    is_feature_allowed as _lic_feature,
    activate_license as _lic_activate,
    heartbeat as _lic_heartbeat,
)

# -- Dataset Memory (feedback store shared with LangChain edition)
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from agent.feedback_store import (
    compute_fingerprint as _fp_compute,
    save_rule as _fp_save,
    get_rules as _fp_get_rules,
    get_rules_as_text as _fp_rules_text,
    delete_rule as _fp_delete,


    update_rule as _fp_update,
    get_dataset_label as _fp_get_label,
    resolve_fingerprint as _fp_resolve,
    list_all_datasets as _fp_list_datasets,
    _STORE_PATH as _feedback_store_path,
    copy_rules as _fp_copy_rules,
    delete_user_data as _fp_delete_user,
)

app = FastAPI(title="Data Validation AGENT")
templates = Jinja2Templates(directory="templates")

from starlette.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1024)

# Ensure tojson filter is available (Starlette includes it; add as fallback)
import json as _json
# NOTE (reconstruction): the source pages here duplicated this block twice (a
# page-repeat artifact seen elsewhere in this file) -- kept one occurrence.
if "tojson" not in templates.env.filters:
    templates.env.filters["tojson"] = lambda v: _json.dumps(v, ensure_ascii=False, default=str)

_chat_contexts: dict[str, dict] = {}
_results_store: dict[str, dict] = {}  # full results keyed by session_id for download/email
_session_quality_cache: dict = {}     # session_id -> {file_name -> quality_report}
_agent_chat_history: dict[str, list] = {}  # session_id -> [{"role","content"}, ...] for Agent mode


# -- Workspace initialisation

try:
    from workspace.db import init_db as _ws_init_db
    from workspace.auth import (
        WorkspaceAuthMiddleware,
        get_current_user as _ws_get_user,
        _resolve_username as _ws_resolve_username,
    )
    from workspace.sso import (
        sso_enabled as _sso_enabled,
        sso_mode as _sso_mode,


        saml_login_redirect_url as _saml_login_url,
        saml_process_response as _saml_process,
        oidc_login_redirect_url as _oidc_login_url,
        oidc_exchange_code as _oidc_exchange,
        create_sso_session_token as _sso_create_token,
    )

    from workspace.scheduler import (
        start_scheduler as _ws_start_scheduler,
        stop_scheduler as _ws_stop_scheduler,
        load_all_jobs as _ws_load_jobs,
        schedule_job as _ws_schedule_job,
        unregister_job as _ws_unregister_job,
        trigger_job_now as _ws_trigger_now,
        trigger_job_now_background as _ws_trigger_now_bg,
    )

    from workspace import connectors as _ws_connectors
    from workspace import db as _ws_db

    app.add_middleware(WorkspaceAuthMiddleware)


    @app.on_event("startup")
    async def _workspace_startup():
        _lic_load()  # validate license before anything else starts
        _ws_init_db()
        _ws_start_scheduler()
        _ws_load_jobs()
        # Daily license heartbeat -- runs every 24 hrs via APScheduler
        from workspace.scheduler import _scheduler as _sch
        if _sch and not _sch.get_job("_license_heartbeat"):
            _sch.add_job(
                _lic_heartbeat,
                trigger="interval",
                hours=24,
                id="_license_heartbeat",
                replace_existing=True,
                kwargs={"usage": {}},
            )

    @app.on_event("shutdown")


    async def _workspace_shutdown():
        _ws_stop_scheduler()

    _WS_ENABLED = True
except Exception as _ws_import_err:
    import logging as _logging
    _logging.getLogger(__name__).warning(
        "Workspace feature disabled -- import error: %s", _ws_import_err
    )
    _WS_ENABLED = False


@app.exception_handler(RequestValidationError)
async def validation_error_handler(request: Request, exc: RequestValidationError):
    detail = "; ".join(
        f"{'.'.join(str(l) for l in e['loc'])}: {e['msg']}" for e in exc.errors()
    )
    return JSONResponse(status_code=400, content={"detail": f"Validation error -- {detail}"})


# ---------------------------------------------------------------------
# License helper -- gates features at the API level
# ---------------------------------------------------------------------

def _require_feature(feature: str):
    # Raise HTTP 402 if the current license does not include this feature.
    if not _lic_feature(feature):
        state = _lic_state()
        if not state.get("valid"):
            raise HTTPException(
                status_code=402,
                detail=f"License invalid or expired: {state.get('error', 'Unknown')}. "
                "Please contact your provider.",
            )
        raise HTTPException(
            status_code=402,


            detail=f"Feature '{feature}' is not included in your '{state.get('tier','unknown')}' plan. "
            "Please upgrade.",
        )


# ---------------------------------------------------------------------
# License / Settings API
# ---------------------------------------------------------------------

@app.get("/api/license/status")
async def license_status():
    """Return current license state including tier comparison for upgrade modal."""
    s = _lic_state()
    from license.license_manager import FEATURE_TIERS, TIER_LIMITS, TIER_PRICING, FEATURE_LABELS
    return JSONResponse({
        "valid":       s.get("valid"),
        "client_id":   s.get("client_id"),
        "tier":        s.get("tier"),


        "features":     s.get("features", []),
        "expires_at":   s.get("expires_at"),
        "error":        s.get("error"),
        "last_check":   s.get("last_check"),
        "limits":       s.get("limits", {}),
        "all_tiers":    FEATURE_TIERS,
        "all_limits":   TIER_LIMITS,
        "tier_pricing": TIER_PRICING,
        "feature_labels": FEATURE_LABELS,
    })


@app.get("/api/usage")
async def get_usage(request: Request):
    """Return token usage summary for current user."""
    try:
        username = _ws_resolve_username(request) or "default"
    except Exception:
        username = "default"


    from workspace.db import get_token_usage_summary, get_token_usage_month_total
    from datetime import datetime as _dt
    month_key = _dt.utcnow().strftime("%Y-%m")
    return JSONResponse({
        "this_month": get_token_usage_month_total(username, month_key),
        "history":    get_token_usage_summary(username, months=3),
        "month":      month_key,
    })


def _require_settings_admin(request: Request) -> None:
    """Block Settings & License access to everyone except the admin role --
    but only once real login is actually enforced. With no login mechanism
    active there is only one implicit operator and no other user to protect
    this from, so access is left open (matches the fallback used throughout
    the rest of the app for the no-auth-configured case)."""
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED
        from workspace.sso import LDAP_ENABLED as _LDAP_ENABLED
    except Exception:
        LOCAL_AUTH_ENABLED = False
        _LDAP_ENABLED = False
    if not (LOCAL_AUTH_ENABLED or _LDAP_ENABLED):
        return
    try:
        username = _ws_resolve_username(request) or ""
    except Exception:
        username = ""
    role = "analyst"
    if username:
        try:
            role = getattr(request.state, "role", None) or _ws_db.get_user_role(username)
        except Exception:
            role = "analyst"
    if role != "admin":
        raise HTTPException(403, "Admin access required to view or change Settings.")


def _require_not_readonly(request: Request) -> None:
    """Run Only (readonly) users can run analysis modules and view results on
    screen, but cannot export data outside the app -- no file download, no
    emailed report."""
    try:
        username = _ws_resolve_username(request) or ""
    except Exception:
        username = ""
    role = "analyst"
    if username:
        try:
            role = getattr(request.state, "role", None) or _ws_db.get_user_role(username)
        except Exception:
            role = "analyst"
    if role == "readonly":
        raise HTTPException(403, "Run Only accounts cannot download or email reports.")


@app.post("/api/license/activate")
async def license_activate(request: Request):
    """Activate a new license key. Client pastes their key into Settings UI. Admin only."""
    _require_settings_admin(request)
    body  = await request.json()
    token = body.get("license_key", "").strip()
    if not token:
        return JSONResponse({"error": "No license key provided."}, status_code=400)
    result = _lic_activate(token)


    if result.get("valid"):
        return JSONResponse({"ok": True, "tier": result.get("tier"),
                    "expires_at": result.get("expires_at"),
                    "client_id": result.get("client_id")})
    return JSONResponse({"error": result.get("error", "Activation failed.")}, status_code=400)


@app.post("/api/license/heartbeat")
async def license_heartbeat_manual(request: Request):
    """Trigger an immediate heartbeat check (called from Settings UI 'Check Now' button). Admin only."""
    _require_settings_admin(request)
    ok = _lic_heartbeat()
    return JSONResponse({"ok": ok, "state": _lic_state()})


@app.get("/api/settings")
async def get_settings(request: Request):
    """Return non-secret runtime config that clients can adjust. Admin only."""
    _require_settings_admin(request)
    provider = os.getenv("LLM_PROVIDER", "bedrock").strip().lower()


    return JSONResponse({
        "llm": {
            "provider": provider,
            # Bedrock
            "model_id":   os.getenv("BEDROCK_MODEL_ID", ""),
            "region":     os.getenv("AWS_DEFAULT_REGION", "us-east-1"),
            "profile":    os.getenv("AWS_PROFILE", ""),
            # Groq
            "groq_model_id": os.getenv("GROQ_MODEL_ID", "llama-3.3-70b-versatile"),
            # Gemini
            "gemini_model_id": os.getenv("GEMINI_MODEL_ID", "gemini-1.5-flash"),
            # OpenAI
            "openai_model_id": os.getenv("OPENAI_MODEL_ID", "gpt-4o-mini"),
            # Anthropic direct
            "anthropic_model_id": os.getenv("ANTHROPIC_MODEL_ID", "claude-opus-4-8"),
        },
        "storage": {
            "backend":    os.getenv("WORKSPACE_DB", "sqlite"),


            "sqlite_path": os.getenv("WORKSPACE_SQLITE_PATH", "workspace.db"),
            "mssql_server": os.getenv("MSSQL_SERVER", ""),
            "mssql_database": os.getenv("MSSQL_DATABASE", ""),
        },
        "email": {
            "from_address": os.getenv("EMAIL_FROM", ""),
            "smtp_host":  os.getenv("SMTP_HOST", ""),
            "smtp_port":  os.getenv("SMTP_PORT", "25"),
        },
        "license_server": os.getenv("LICENSE_SERVER_URL", ""),
    })


@app.post("/api/settings")
async def save_settings(request: Request):

    # Persist config changes to .env file.
    # Client controls their own LLM keys, DB, storage -- none sent to provider.
    # Admin only -- these include API keys, SMTP password, and DB connection details.
    _require_settings_admin(request)


    body = await request.json()
    env_path = Path(".env")
    lines: list[str] = []
    if env_path.exists():
        lines = env_path.read_text(encoding="utf-8").splitlines()

    def _set(key: str, val: str):
        nonlocal lines
        val = val.strip()
        for i, line in enumerate(lines):
            if line.startswith(f"{key}=") or line.startswith(f"{key} ="):
                lines[i] = f"{key}={val}"
                return
        lines.append(f"{key}={val}")

    # LLM settings (BYOK -- client's own API account)
    if "llm" in body:
        llm = body["llm"]
        if llm.get("provider"):    _set("LLM_PROVIDER",    llm["provider"])


        # Bedrock
        if llm.get("model_id"):     _set("BEDROCK_MODEL_ID",  llm["model_id"])
        if llm.get("region"):       _set("AWS_DEFAULT_REGION", llm["region"])
        if llm.get("profile"):      _set("AWS_PROFILE",       llm["profile"])
        # Groq
        if llm.get("groq_api_key"):   _set("GROQ_API_KEY",    llm["groq_api_key"])
        if llm.get("groq_model_id"):  _set("GROQ_MODEL_ID",   llm["groq_model_id"])
        # Gemini
        if llm.get("gemini_api_key"):  _set("GEMINI_API_KEY",  llm["gemini_api_key"])
        if llm.get("gemini_model_id"): _set("GEMINI_MODEL_ID", llm["gemini_model_id"])
        # OpenAI
        if llm.get("openai_api_key"):   _set("OPENAI_API_KEY",  llm["openai_api_key"])
        if llm.get("openai_model_id"):  _set("OPENAI_MODEL_ID", llm["openai_model_id"])
        # Anthropic direct
        if llm.get("anthropic_api_key"): _set("ANTHROPIC_API_KEY", llm["anthropic_api_key"])
        if llm.get("anthropic_model_id"): _set("ANTHROPIC_MODEL_ID",
llm["anthropic_model_id"])


    # Storage settings (client's own DB)
    if "storage" in body:
        st = body["storage"]
        if st.get("backend"):        _set("WORKSPACE_DB",         st["backend"])
        if st.get("sqlite_path"):    _set("WORKSPACE_SQLITE_PATH", st["sqlite_path"])
        if st.get("mssql_server"):   _set("MSSQL_SERVER",         st["mssql_server"])
        if st.get("mssql_database"): _set("MSSQL_DATABASE",       st["mssql_database"])

    # Email settings
    if "email" in body:
        em = body["email"]
        if em.get("from_address"): _set("EMAIL_FROM",     em["from_address"])
        if em.get("smtp_host"):    _set("SMTP_HOST",      em["smtp_host"])
        if em.get("smtp_port"):    _set("SMTP_PORT",      em["smtp_port"])
        if em.get("smtp_user"):    _set("SMTP_USER",      em["smtp_user"])
        if em.get("smtp_pass"):    _set("SMTP_PASSWORD",  em["smtp_pass"])

    # License server URL


    if "license_server" in body and body["license_server"]:
        _set("LICENSE_SERVER_URL", body["license_server"])

    env_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return JSONResponse({"ok": True, "message": "Settings saved. Restart app to apply LLM/DB changes."})


# ---------------------------------------------------------------------
# AWS Bedrock
# ---------------------------------------------------------------------

MODEL_ID = os.getenv(
    "BEDROCK_MODEL_ID",
    "arn:aws:bedrock:us-east-1:331137948115:application-inference-profile/og6ymxr571ro",  # OCR-UNCERTAIN: long ARN id digits hard to verify precisely
)


def _get_bedrock_client():


    session = boto3.Session(
        profile_name=os.getenv("AWS_PROFILE", "claudecode"),
        region_name=os.getenv("AWS_DEFAULT_REGION", "us-east-1"),
    )
    kwargs = {"service_name": "bedrock-runtime"}
    if ca := os.getenv("AWS_CA_BUNDLE"):
        kwargs["verify"] = ca
    return session.client(**kwargs)


def _sanitize_json(obj):
    # Recursively sanitize a value so JSONResponse never raises.
    # Handles:
    # - NaN / Inf / -Inf floats (Python and numpy)
    # - numpy integer and float scalars
    # - numpy bool_ scalars
    # - pandas NA / NaT / Timestamp
    # - bytes (decoded as UTF-8, falling back to repr)

    # - sets (converted to sorted lists)
    # - any other non-serialisable type (converted to str)

    # -- dict / list
    if isinstance(obj, dict):
        return {k: _sanitize_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize_json(v) for v in obj]
    if isinstance(obj, set):
        return [_sanitize_json(v) for v in sorted(obj, key=str)]

    # -- None / bool (must come before int check -- bool is subclass of int) --
    if obj is None:
        return None
    if isinstance(obj, bool):
        return obj

    # -- numpy scalars
    try:
        import numpy as _np
        if isinstance(obj, _np.integer):
            return int(obj)
        if isinstance(obj, _np.floating):
            if _np.isnan(obj) or _np.isinf(obj):
                return None
            return float(obj)
        if isinstance(obj, _np.bool_):
            return bool(obj)
        if isinstance(obj, _np.ndarray):
            return [_sanitize_json(v) for v in obj.tolist()]
    except ImportError:
        pass

    # -- pandas NA / NaT / Timestamp
    try:
        import pandas as _pd
        if obj is _pd.NA or obj is _pd.NaT:

            return None
        if isinstance(obj, _pd.Timestamp):
            return obj.isoformat()
        if isinstance(obj, _pd.Series):
            return [_sanitize_json(v) for v in obj.tolist()]
    except ImportError:
        pass

    # -- Python float (NaN / Inf)
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    # -- Python int / str
    if isinstance(obj, (int, str)):
        return obj

    # -- bytes
    if isinstance(obj, (bytes, bytearray)):
        try:
            return obj.decode("utf-8")
        except Exception:
            return repr(obj)

    # -- fallback: anything else becomes a string
    return str(obj)


def _ask_llm_bedrock(messages: list[dict], system: str = "",
                      _module: str = "unknown", _call_type: str = "chat",
                      _username: str = "") -> str:
    client = _get_bedrock_client()
    kwargs = {"modelId": MODEL_ID, "messages": messages}
    if system:
        kwargs["system"] = [{"text": system}]

    response = client.converse(**kwargs)

    # -- Token usage tracking
    try:
        usage = response.get("usage", {})
        in_tok  = int(usage.get("inputTokens",  0))
        out_tok = int(usage.get("outputTokens", 0))
        if in_tok or out_tok:
            from workspace.db import log_token_usage as _log_usage
            _log_usage(
                username=_username or "",
                module=_module,
                call_type=_call_type,
                input_tokens=in_tok,
                output_tokens=out_tok,
                model=MODEL_ID,
            )
    except Exception:
        pass

    return response["output"]["message"]["content"][0]["text"]


def _ask_llm(messages: list[dict], system: str = "",
             _module: str = "unknown", _call_type: str = "chat",
             _username: str = "") -> str:
    # Multi-provider fallback: try LLM_PROVIDER first, then the rest of the
    # fallback order, so a misconfigured/unreachable Bedrock (no AWS creds,
    # wrong profile, etc.) doesn't take down every AI feature in the app.
    # Bedrock's native message format ({"content": [{"text": ...}]}) is used
    # by every caller of _ask_llm; the other providers want plain strings.
    from agent.llm import LLM_PROVIDER, _ask_groq, _ask_gemini, _ask_openai, _ask_anthropic

    plain_messages = [
        {
            "role": m["role"],
            "content": m["content"][0]["text"] if isinstance(m.get("content"), list) else m.get("content", ""),
        }
        for m in messages
    ]
    providers = [LLM_PROVIDER] + [p for p in ("groq", "gemini", "bedrock", "openai", "anthropic") if p != LLM_PROVIDER]

    errors: list[str] = []
    for provider in providers:
        try:
            if provider == "bedrock":
                return _ask_llm_bedrock(messages, system, _module, _call_type, _username)
            elif provider == "groq":
                return _ask_groq(plain_messages, system)
            elif provider == "gemini":
                return _ask_gemini(plain_messages, system)
            elif provider == "openai":
                return _ask_openai(plain_messages, system)
            elif provider == "anthropic":
                return _ask_anthropic(plain_messages, system)
        except Exception as exc:
            errors.append(f"{provider}: {exc}")
            continue
    raise RuntimeError("All LLM providers failed -- " + "; ".join(errors) if errors else "No LLM provider configured.")


def _llm_error_message(e: Exception) -> str:
    """Short, user-facing message for an _ask_llm() failure -- collapses the
    raw multi-provider error dump (which can include internal API details)
    into a clear "no LLM connection" signal for the UI, per-provider errors
    otherwise kept but trimmed."""
    msg = str(e)
    if "All LLM providers failed" in msg or "No LLM provider configured" in msg:
        return "No LLM connection configured."
    return f"AI summary unavailable ({msg[:160]})."


# ---------------------------------------------------------------------
# File loading
# ---------------------------------------------------------------------

def _detect_encoding(raw: bytes) -> str:
    return chardet.detect(raw).get("encoding") or "utf-8"


def _sniff_delimiter(text: str) -> str | None:
    # Robustly detect the delimiter from text.
    # Strategy:
    # 1. Count-and-confirm -- count every candidate delimiter in the header line.
    #    Winner must appear >=2 times in the header AND the same count (+-1) in
    #    each of the next 5 data rows. Handles pipe-delimited files where
    #    column names contain spaces (e.g. StockLoan "Anticip Recv - CNS").
    # 2. csv.Sniffer fallback for ambiguous cases.
    # 3. Single-char scan -- if a non-alphanumeric char appears at the same
    #    position across all rows, it is likely a fixed delimiter.
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return None

    header = lines[0]
    data_lines = lines[1:6]  # up to 5 data rows

    # All candidate delimiters including rare ones used in BFSI files
    candidates = ["\t", "|", ";", ",", "~", "^", "\x01", ":", "=", "\x1f"]

    # -- 1. Count-and-confirm
    counts = {d: header.count(d) for d in candidates}
    for delim in sorted(candidates, key=lambda d: counts[d], reverse=True):
        hdr_count = counts[delim]
        if hdr_count < 2:
            continue
        consistent = all(
            abs(ln.count(delim) - hdr_count) <= max(1, hdr_count // 10)
            for ln in data_lines if ln.strip()
        )
        if consistent:
            return delim

    # -- 2. csv.Sniffer
    sample = "\n".join(lines[:50])
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=",\t|;:~^=")
        if dialect.delimiter and header.count(dialect.delimiter) >= 1:
            return dialect.delimiter
    except _csv.Error:
        pass

    # -- 3. Position-scan -- find a char that appears at the same column index
    # across all lines (suggests fixed-field delimiter like space-padded)
    if len(lines) >= 3:
        min_len = min(len(ln) for ln in lines[:4])
        for pos in range(min_len):
            char = lines[0][pos]
            if char.isalnum() or char in (' ', '\t'):
                continue
            if all(len(ln) > pos and ln[pos] == char for ln in lines[1:4]):
                return char

    return None


def _decode_raw(raw: bytes) -> tuple[str, str]:
    # Robustly decode raw bytes to text.
    # Returns (text, encoding_used).
    # Strips BOM, normalises CRLF -> LF, removes null bytes.
    # Tries: UTF-8-BOM -> UTF-8 -> chardet -> latin-1 (never fails).

    # Strip common BOMs
    for bom, enc in [
        (b"\xef\xbb\xbf", "utf-8-sig"),
        (b"\xff\xfe",     "utf-16-le"),
        (b"\xfe\xff",     "utf-16-be"),
        (b"\x00\x00\xfe\xff", "utf-32-be"),
        (b"\xff\xfe\x00\x00", "utf-32-le"),
    ]:
        if raw.startswith(bom):
            try:
                text = raw[len(bom):].decode(enc, errors="replace")
                text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\x00", "")
                return text, enc
            except Exception:
                pass

    # Try UTF-8
    try:
        text = raw.decode("utf-8")
        text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\x00", "")
        return text, "utf-8"
    except UnicodeDecodeError:
        pass

    # Try chardet
    try:
        import chardet as _cd
        detected = _cd.detect(raw[:8192]).get("encoding") or "latin-1"
        text = raw.decode(detected, errors="replace")
        text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\x00", "")
        return text, detected
    except Exception:
        pass

    # Final fallback: latin-1 never fails
    text = raw.decode("latin-1", errors="replace")
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\x00", "")
    return text, "latin-1"


def _repair_and_parse(text: str, sep: str, expected_cols: int) -> pd.DataFrame | None:
    # Attempt to parse text with sep; if rows have column count mismatches,
    # apply auto-repair strategies before giving up:
    #
    # 1. Normal parse (on_bad_lines=skip for truly unparseable rows)
    # 2. If quoting errors, try different quote chars (" -> ' -> none)
    # 3. If extra/missing columns, pad/truncate each line to expected_cols
    # 4. Strip inline comments (lines starting with # that aren't the header)
    # 5. Skip metadata header rows (lines without the delimiter)

    if not sep:
        return None

    def _try_parse(t: str, quotechar: str = '"') -> pd.DataFrame | None:
        try:
            kw = dict(
                sep=sep, engine="python",
                on_bad_lines="skip",
                dtype=str,
            )
            if quotechar:
                kw["quotechar"] = quotechar
            df = pd.read_csv(io.StringIO(t), **kw)
            return df if len(df.columns) > 1 else None
        except Exception:
            return None

    # -- Pass 1: normal parse
    df = _try_parse(text)
    if df is not None and len(df) > 0:
        return df

    # -- Pass 2: try different quote chars
    for qc in ('"', "'", None):
        df = _try_parse(text, quotechar=qc or "")
        if df is not None and len(df) > 0:
            return df

    # -- Pass 3: strip comment/metadata lines, then retry
    lines = text.splitlines()
    # Find the first line that looks like a header (contains the delimiter)
    header_idx = 0
    for i, ln in enumerate(lines):
        if ln.count(sep) >= 1:
            header_idx = i
            break
    if header_idx > 0:
        cleaned = "\n".join(lines[header_idx:])
        df = _try_parse(cleaned)
        if df is not None and len(df) > 0:
            df.attrs["_skipped_header_rows"] = header_idx
            return df


    # -- Pass 4: pad/truncate rows to match header column count
    if expected_cols > 1:
        repaired_lines = []
        for i, ln in enumerate(lines):
            parts = ln.split(sep)
            if i == 0:
                repaired_lines.append(ln)  # header as-is
                expected_cols = len(parts)
                continue
            if len(parts) < expected_cols:
                # Pad with empty fields
                parts += [""] * (expected_cols - len(parts))
            elif len(parts) > expected_cols:
                # Merge excess fields into the last column
                parts = parts[:expected_cols - 1] + [sep.join(parts[expected_cols - 1:])]
            repaired_lines.append(sep.join(parts))
        repaired = "\n".join(repaired_lines)
        df = _try_parse(repaired)


        if df is not None and len(df) > 0:
            df.attrs["_auto_repaired"] = True
            return df

    return None


def _read_delimited(text: str, sep: str) -> pd.DataFrame | None:
    # Parse text as a delimited file with full auto-repair.
    # Returns None only if the delimiter is genuinely wrong.

    # Quick count check: separator must appear at least once in the first line
    first_line = text.split("\n")[0] if text else ""
    if first_line.count(sep) < 1:
        return None

    expected_cols = first_line.count(sep) + 1
    return _repair_and_parse(text, sep, expected_cols)


def _parse_txt(raw: bytes, encoding: str, delimiter: str | None = None) -> pd.DataFrame:
    # Parse any text/delimited file into a DataFrame.
    #
    # Handles all BFSI file variants:
    # - All delimiter types: tab, pipe, semicolon, comma, colon, tilde, caret, SOH,
    #   fixed-width, space-padded
    # - All encodings: UTF-8, UTF-8 BOM, UTF-16, Latin-1, Windows-1252
    # - Line ending variants: LF, CRLF, CR
    # - Quote chars: double-quote, single-quote, no quoting
    # - Bad rows: too many/few fields -> auto-padded or merged
    # - Comment/metadata header rows -> skipped automatically
    # - BOM stripping, null byte removal
    # - NDJSON (one JSON object per line)
    #
    # Detection order:
    # 1. NDJSON
    # 2. User-supplied delimiter
    # 3. Robust count-and-confirm sniffer (all delimiters)
    # 4. Trial of every known delimiter with auto-repair
    # 5. Whitespace / fixed-width
    # 6. Single-value fallback
    #
    # Sets df.attrs:
    #   _detected_delimiter  -- separator actually used
    #   _auto_repaired       -- True if row padding/truncation was applied
    #   _skipped_header_rows -- N if metadata rows were skipped
    #   _parse_warnings      -- list of warning strings
    warnings: list[str] = []

    # -- Robust decode (handles BOM, encoding detection, CRLF, null bytes) --
    text, used_enc = _decode_raw(raw)
    all_lines = text.splitlines()
    lines = [ln for ln in all_lines if ln.strip()]
    if not lines:
        raise ValueError("Empty file -- no content after decoding")

    # -- Pre-strip comment/metadata lines before delimiter detection
    # Skip leading lines that start with # or // or contain no potential delimiter
    # characters -- they are metadata and confuse the sniffer
    _COMMENT_PREFIXES = ("#", "//", "--", "/*", "!")
    skip_count = 0
    for ln in lines:
        if any(ln.lstrip().startswith(p) for p in _COMMENT_PREFIXES):
            skip_count += 1
        else:
            break
    if skip_count > 0:
        warnings.append(f"Skipped {skip_count} comment/metadata rows before the data header")
        lines = lines[skip_count:]
        text = "\n".join(lines)
    if not lines:
        raise ValueError("Empty file -- only comment lines found")


    def _mark(df: pd.DataFrame, sep: str) -> pd.DataFrame:
        df.attrs["_detected_delimiter"] = sep
        df.attrs["_encoding_used"]      = used_enc
        if df.attrs.get("_auto_repaired"):
            warnings.append("Some rows had mismatched column counts and were auto-repaired")
        if df.attrs.get("_skipped_header_rows"):
            warnings.append(f"Skipped {df.attrs['_skipped_header_rows']} metadata/comment rows before the header")
        df.attrs["_parse_warnings"] = warnings
        return df

    # -- 1. NDJSON
    try:
        records = [json.loads(ln) for ln in lines]
        if all(isinstance(r, dict) for r in records):
            return _mark(pd.DataFrame(records), "json")
    except Exception:
        pass

    # -- 2. User-supplied delimiter
    if delimiter:
        # Normalise escape sequences typed in a text box (e.g. "\t" -> real tab)
        sep = delimiter.encode("raw_unicode_escape").decode("unicode_escape") \
            if len(delimiter) > 1 else delimiter
        df = _read_delimited(text, sep)
        if df is not None:
            return _mark(df, sep)
        warnings.append(f"User-specified delimiter {repr(sep)} did not produce a valid table -- trying auto-detection")

    # -- 3. Robust sniffer
    detected = _sniff_delimiter(text)
    if detected and detected != (delimiter or ""):
        df = _read_delimited(text, detected)
        if df is not None:
            return _mark(df, detected)


    # -- 4. Trial of ALL known delimiters with auto-repair
    tried = {delimiter or "", detected or ""}
    # Ordered by BFSI frequency: tab, pipe, semicolon, comma, tilde, caret,
    # SOH (\x01 used in FIX), colon, equals, unit-separator (\x1f)
    for sep in ["\t", "|", ";", ",", "~", "^", "\x01", ":", "=", "\x1f", " "]:
        if sep in tried:
            continue
        tried.add(sep)
        df = _read_delimited(text, sep)
        if df is not None:
            return _mark(df, sep)

    # -- 5. Fixed-width / whitespace-delimited
    try:
        df = pd.read_fwf(io.StringIO(text))
        if len(df.columns) > 1 and len(df) > 0:
            return _mark(df, "fixed-width")
    except Exception:
        pass

    try:
        df = pd.read_csv(io.StringIO(text), sep=r"\s+", engine="python",
                          on_bad_lines="skip", dtype=str)
        if len(df.columns) > 1:
            return _mark(df, " ")
    except Exception:
        pass

    # -- 6. Last resort: one value per line
    warnings.append("Could not detect a delimiter -- loaded as single-column text")
    return _mark(pd.DataFrame({"value": lines}), "")


def _parse_xml(raw: bytes) -> pd.DataFrame:
    # Parse XML into a flat tabular DataFrame.
    # Strategy 1 -- Repeating same-tag children of root (most common tabular XML).
    # Uses full recursive dot-path flattening so 3+ level nesting works.
    # Strategy 2 -- Repeating same-tag elements found by BFS at any depth
    # (handles <root><group><row>...</row></group></root> patterns).
    # Strategy 3 -- Root has heterogeneous children -> one row, recursive flattening.
    # Strategy 4 -- Root is a leaf node.

    def _strip_ns(tag: str) -> str:
        return re.sub(r"\{[^}]+\}", "", tag)

    def _elem_to_dict(elem, prefix: str = "") -> dict:
        # Recursively flatten elem into {dot.path: value} pairs (unlimited depth).
        result: dict = {}
        tag = _strip_ns(elem.tag)
        path = f"{prefix}.{tag}" if prefix else tag
        for k, v in elem.attrib.items():
            result[f"{path}@{_strip_ns(k)}"] = v
        text = (elem.text or "").strip()
        children = list(elem)


        if children:
            for child in children:
                result.update(_elem_to_dict(child, path))
            # keep inline text if present alongside children
            if text:
                result[path] = text
        else:
            result[path] = text  # OCR-UNCERTAIN: both if/else branches read as identical "result[path] = text" -- verify against source
        return result

    def _row_from_elem(elem) -> dict:
        # Build a flat row from a row-level element using shallow + recursive fields.
        rec: dict = {f"@{_strip_ns(k)}": v for k, v in elem.attrib.items()}
        for child in elem:
            ctag = _strip_ns(child.tag)
            sub = list(child)
            for k, v in child.attrib.items():
                rec[f"{ctag}@{_strip_ns(k)}"] = v
            if sub:
                # Recurse: flatten nested children with dot-path
                for k, v in _elem_to_dict(child, "").items():
                    rec[k] = v
            else:
                rec[ctag] = (child.text or "").strip()
        return rec

    def _find_repeating(elem, depth: int = 0):
        # BFS to find the first group of 2+ sibling elements with the same tag.
        if depth > 8:
            return None
        children = list(elem)
        if children:
            tags = [_strip_ns(c.tag) for c in children]
            dominant, count = Counter(tags).most_common(1)[0]
            if count >= 2 or (count == 1 and len(list(children[0])) > 0):
                return [c for c in children if _strip_ns(c.tag) == dominant]
            # Search one level deeper (pick the child with most children)
            best = max(children, key=lambda c: len(list(c)), default=None)


            if best is not None:
                return _find_repeating(best, depth + 1)
        return None

    root = ET.fromstring(raw)
    root_tag = _strip_ns(root.tag)
    children = list(root)

    # Strategy 1: root -> repeating children
    if children:
        child_tags = [_strip_ns(c.tag) for c in children]
        dominant_tag = Counter(child_tags).most_common(1)[0][0]
        row_elems = [c for c in children if _strip_ns(c.tag) == dominant_tag]
        if len(row_elems) >= 2 or (len(row_elems) == 1 and len(list(row_elems[0])) > 0):
            records = [_row_from_elem(r) for r in row_elems]
            if records:
                return pd.DataFrame(records)

    # Strategy 2: BFS for repeating elements deeper in the tree

    # Skip root-level children -- Strategy 1 already tried those
    deeper_start = max(list(root), key=lambda c: len(list(c)), default=None)
    row_elems = _find_repeating(deeper_start) if deeper_start is not None else None
    if row_elems and len(row_elems) >= 2:
        records = [_row_from_elem(r) for r in row_elems]
        if records:
            return pd.DataFrame(records)

    # Strategy 3: root has children but no repeating pattern -> one row
    if children:
        return pd.DataFrame([_elem_to_dict(root)])

    # Strategy 4: leaf node
    row: dict = {f"@{_strip_ns(k)}": v for k, v in root.attrib.items()}
    if (root.text or "").strip():
        row["value"] = root.text.strip()
    row["tag"] = root_tag
    return pd.DataFrame([row])


# ---------------------------------------------------------------------
# SWIFT MT message parser
# ---------------------------------------------------------------------

# Field tag name lookup (most common MT tags)
_SWIFT_FIELD_NAMES: dict[str, str] = {
    "20":  "Transaction Reference",
    "21":  "Related Reference",
    "23B": "Bank Operation Code",
    "23E": "Instruction Code",
    "25":  "Account Identification",
    "25P": "Account / BIC",
    "26T": "Transaction Type Code",
    "28C": "Statement Number / Sequence",
    "32A": "Value Date / Currency / Amount",
    "32B": "Currency / Amount",
    "33B": "Currency / Instructed Amount",
    "36":  "Exchange Rate",
    "50A": "Ordering Customer (BIC)",
    "50K": "Ordering Customer",
    "51A": "Sending Institution",
    "52A": "Ordering Institution (BIC)",
    "52D": "Ordering Institution",
    "53A": "Sender's Correspondent (BIC)",
    "53B": "Sender's Correspondent",
    "54A": "Receiver's Correspondent (BIC)",
    "56A": "Intermediary (BIC)",
    "57A": "Account With Institution (BIC)",
    "57D": "Account With Institution",
    "59":  "Beneficiary Customer",
    "59A": "Beneficiary Customer (BIC)",
    "60F": "Opening Balance",
    "60M": "Intermediate Opening Balance",
    "61":  "Statement Line",
    "62F": "Closing Balance",
    "62M": "Intermediate Closing Balance",
    "70":  "Remittance Information",


    "71A": "Details of Charges",
    "71F": "Sender's Charges",
    "71G": "Receiver's Charges",
    "72":  "Sender to Receiver Information",
    "77B": "Regulatory Reporting",
    "77T": "Envelope Contents",
    "86":  "Information to Account Owner",
}

_MT_BLOCK_NAMES: dict[str, str] = {
    "1": "Basic Header",
    "2": "Application Header",
    "3": "User Header",
    "4": "Text Block",
    "5": "Trailer",
}


def _parse_swift_mt(raw: bytes, enc: str) -> pd.DataFrame:
    # Parse SWIFT MT messages into wide format: one row per message.
    #
    # Strategies (tried in order):
    # 1. Block-envelope {1:...}{2:...}{4:...:TAG:value...-} -- standard FIN file
    # 2. Bare :TAG: message blocks separated by bare '-' lines
    #    (files that contain multiple messages outside block envelopes)
    # 3. Single bare :TAG: pass over the whole text
    text = raw.decode(enc, errors="replace")

    def _col_name(tag: str) -> str:
        name = _SWIFT_FIELD_NAMES.get(tag, "")
        if name:
            return f"{tag}_{re.sub(r'[^A-Za-z0-9]+', '_', name).strip('_')}"
        return tag

    def _fields_from_text(content: str) -> dict:
        # Extract :TAG: value pairs from a block of text.
        rec: dict = {}
        field_re = re.compile(
            r"^:([0-9]{2}[A-Z]?):([ \t]*[^\n]*(?:\n(?![ \t]*:[0-9]|\-)[ \t]*[^\n]*)*)",  # OCR-UNCERTAIN: complex regex, exact escaping/brackets hard to verify from photo
            re.MULTILINE,
        )
        for m in field_re.finditer(content):
            tag = m.group(1).strip()
            val = m.group(2).strip()
            col = _col_name(tag)
            if col in rec:
                i = 2
                while f"{col}_{i}" in rec:
                    i += 1
                col = f"{col}_{i}"
            rec[col] = val
        return rec

    # -- Strategy 1: block-envelope {1:...}{4:...-}
    block_pat = re.compile(r"\{(\d):([^{}]*(?:\{[^{}]*\}[^{}]*)*)\}", re.DOTALL)  # OCR-UNCERTAIN: regex braces/escaping hard to verify from photo
    msg_splits = re.split(r"(?=\{1:)", text.strip())


    msg_splits = [m.strip() for m in msg_splits if m.strip()]

    records: list[dict] = []

    for msg_idx, msg_text in enumerate(msg_splits):
        blocks = {bid: bc for bid, bc in block_pat.findall(msg_text)}
        if not blocks:
            continue

        rec: dict = {"message_index": msg_idx + 1}

        if "1" in blocks:
            hdr = blocks["1"].strip()
            rec["basic_header"] = hdr
            bic_m = re.search(r"F\d{2}([A-Z]{8,11})", hdr)
            if bic_m:
                rec["sender_bic"] = bic_m.group(1)

        if "2" in blocks:
            hdr2 = blocks["2"].strip()
            rec["app_header"] = hdr2
            mt_m = re.search(r"^I?(\d{3})", hdr2)  # OCR-UNCERTAIN: leading caret/optional-I pattern hard to verify from photo
            if mt_m:
                rec["mt_type"] = mt_m.group(1)


        if "3" in blocks:
            for sm in re.finditer(r"\{([A-Z0-9]+):([^}]*)\}", blocks["3"]):  # OCR-UNCERTAIN: brace/escape sequence hard to verify from photo
                rec[f"blk3_{sm.group(1)}"] = sm.group(2).strip()

        if "4" in blocks:
            rec.update(_fields_from_text(blocks["4"]))

        # Only accept if block 4 yielded at least one real :TAG: field
        has_mt_fields = any(re.match(r"^\d{2}", k) for k in rec)
        if has_mt_fields:
            records.append(rec)

    if records:
        return pd.DataFrame(records)

    # -- Strategy 2: bare :TAG: blocks separated by lines containing only '-'
    # Strip comment lines (starting with #) and section markers first
    clean_lines = [
        ln for ln in text.splitlines()
        if not ln.strip().startswith("#") and not re.match(r"^=[\w_]+=?$", ln.strip())  # OCR-UNCERTAIN: exact regex characters hard to verify from photo
    ]
    clean_text = "\n".join(clean_lines)


    # Split on message terminators: a line that is just '-'
    msg_blocks = re.split(r"(?m)^\s*-\s*$", clean_text)
    msg_blocks = [b.strip() for b in msg_blocks if b.strip()]

    records = []
    for idx, block in enumerate(msg_blocks):
        fields = _fields_from_text(block)
        if fields:
            fields["message_index"] = idx + 1
            records.append(fields)

    if records:
        return pd.DataFrame(records)

    # -- Strategy 3: entire file as one message ----------------------------
    bare = _fields_from_text(text)
    if bare:
        bare["message_index"] = 1
        return pd.DataFrame([bare])

    return pd.DataFrame({"raw": [text]})


def _is_swift_mt(raw: bytes, enc: str) -> bool:
    # Heuristic: file looks like a SWIFT MT message.
    #
    # Deliberately strict to avoid false-positives on XML/CSV/config files:
    # - Block-envelope header {1:F01...} is definitive.
    # - {4:\n: block header is definitive.
    # - Bare :TAG: fields are only accepted when they appear at the start of
    #   a line AND at least 2 distinct occurrences exist in the sample -- a
    #   single colon-number-colon pattern is too common in non-SWIFT content.
    try:
        sample = raw[:2000].decode(enc, errors="replace")
        if re.search(r"\{1:F\d{2}", sample):
            return True
        if re.search(r"\{4:\s*\n?\s*:", sample):  # OCR-UNCERTAIN: exact escape sequence hard to verify from photo
            return True
        # Require SWIFT field tags anchored at the start of a line, >=2 matches


        field_tags = re.findall(r"(?m)^\s*:[0-9]{2}[A-Z]?:", sample)  # OCR-UNCERTAIN: exact regex chars hard to verify from photo
        if len(field_tags) >= 2:
            return True
        return False
    except Exception:
        return False


# ---------------------------------------------------------------------
# FIX protocol parser
# ---------------------------------------------------------------------

# Common FIX tag numbers -> names
_FIX_TAG_NAMES: dict[str, str] = {
    "8": "BeginString", "9": "BodyLength", "35": "MsgType",
    "49": "SenderCompID", "56": "TargetCompID", "34": "MsgSeqNum",
    "52": "SendingTime", "11": "ClOrdID", "37": "OrderID",
    "17": "ExecID", "150": "ExecType", "39": "OrdStatus",
    "55": "Symbol", "54": "Side", "38": "OrderQty",


    "44": "Price", "40": "OrdType", "58": "Text",
    "10": "CheckSum", "60": "TransactTime", "14": "CumQty",
    "151": "LeavesQty", "6": "AvgPx", "1": "Account",
    "48": "SecurityID", "22": "SecurityIDSource",
}


def _parse_fix(raw: bytes, enc: str) -> pd.DataFrame:
    # Parse FIX protocol messages.
    # Handles two layouts:
    # A. Standard FIX stream: SOH / pipe-delimited numeric tag=value pairs
    # (e.g. 8=FIX.4.2|35=D|49=SENDER...)
    # B. Named-key flat-file format: one KEY=VALUE per line, records separated
    # by ---RECORD N--- or blank lines (investment-bank key/value style)

    text = raw.decode(enc, errors="replace")

    # -- Layout A: standard numeric FIX (SOH or pipe delimited) -------------

    has_soh = "\x01" in text
    has_pipe = "|" in text and re.search(r"\b\d+=", text)

    if has_soh or has_pipe:
        delim = "\x01" if has_soh else "|"

        def _parse_fix_fields(msg: str) -> dict:
            rec: dict = {}
            for part in msg.split(delim):
                if "=" not in part:
                    continue
                tag, _, val = part.partition("=")
                tag = tag.strip()
                if not re.match(r"^\d+$", tag):
                    continue
                name = _FIX_TAG_NAMES.get(tag, "")
                col = f"{tag}_{name}" if name else tag
                rec[col] = val.strip()
            return rec

        raw_msgs = re.split(r"(?=8=FIX)", text)
        records = [r for m in raw_msgs if (r := _parse_fix_fields(m.strip()))]
        if records:
            return pd.DataFrame(records)

    # -- Layout B: named KEY=VALUE flat file --------------------------------
    # Records delimited by ---RECORD N--- headers or blank lines
    record_blocks: list[str] = []

    # Try splitting on ---RECORD--- markers first
    marker_blocks = re.split(r"---RECORD\s+\d+---", text)
    marker_blocks = [b.strip() for b in marker_blocks if b.strip()]

    # Discard the file header block (lines starting with # or = markers)
    content_blocks = []
    for b in marker_blocks:
        lines = [l for l in b.splitlines() if l.strip() and not l.strip().startswith("#") and not re.match(r"^=[\w_]+=?$", l.strip())]  # OCR-UNCERTAIN: exact regex chars hard to verify from photo
        if lines:
            content_blocks.append("\n".join(lines))

    if content_blocks:
        record_blocks = content_blocks
    else:
        # Fall back: blank-line-separated blocks
        record_blocks = [b.strip() for b in re.split(r"\n\s*\n", text) if b.strip()]

    records = []
    for block in record_blocks:
        rec: dict = {}
        for line in block.splitlines():
            line = line.strip()
            if not line or line.startswith("#") or re.match(r"^=[\w_]+=?$", line):
                continue
            if "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip()
            if key:
                rec[key] = val
        if rec:
            records.append(rec)

    if records:
        return pd.DataFrame(records)

    return pd.DataFrame({"raw": [text]})


def _is_fix(raw: bytes, enc: str) -> bool:
    # Heuristic: file looks like FIX protocol.
    try:

        sample = raw[:500].decode(enc, errors="replace")

        return bool(re.match(r"8=FIX\.\d", sample) or
                    (re.search(r"8=FIX\.\d", sample) and ("35=" in sample)))
    except Exception:

        return False


# ----------------------------------------------------------------------
# ISO 20022 / SWIFT MX XML parser
# ----------------------------------------------------------------------

def _is_iso20022(raw: bytes) -> bool:
    # Heuristic: contains urn:iso:std:iso:20022 namespace.
    try:

        sample = raw[:1000].decode("utf-8", errors="replace")

        return "urn:iso:std:iso:20022" in sample or "swift.com/xsd" in sample
    except Exception:

        return False


def _parse_iso20022(raw: bytes) -> pd.DataFrame:
    # Parse ISO 20022 / SWIFT MX XML messages.
    # Returns one row per leaf element with columns: path, element, value, attributes.

    def _strip_ns(tag: str) -> str:
        return re.sub(r"\{[^}]+\}", "", tag)

    def _collect_leaves(elem, path: str, rows: list):
        tag = _strip_ns(elem.tag)
        cur_path = f"{path}/{tag}" if path else tag
        attrs = {f"@{_strip_ns(k)}": v for k, v in elem.attrib.items()}
        children = list(elem)
        if not children:
            text = (elem.text or "").strip()
            row = {"path": cur_path, "element": tag, "value": text}
            row.update(attrs)
            rows.append(row)
        else:
            for child in children:
                _collect_leaves(child, cur_path, rows)

    root = ET.fromstring(raw)
    rows: list[dict] = []

    _collect_leaves(root, "", rows)
    return pd.DataFrame(rows) if rows else pd.DataFrame({"tag": [_strip_ns(root.tag)]})


    # ----------------------------------------------------------------------
    # JSON - enhanced nested flattening
    # ----------------------------------------------------------------------

def _parse_json_nested(raw: bytes) -> pd.DataFrame:
    # Parse JSON into a flat DataFrame.
    #
    # Handles: array of objects, nested wrapper dicts, single object,
    # array of primitives, multi-level nesting.
    #
    # Wrapper-dict unwrapping: recursively searches for the first list-of-dicts
    # value at any depth (BFS) so patterns like {"response":{"data":[{...}]}}
    # are correctly flattened.
    obj = json.loads(raw)


    def _find_records(o, depth: int = 0) -> list | None:
        # BFS: return the first list-of-dicts found within *o*.
        if depth > 5:
            return None

        if isinstance(o, list):
            if o and all(isinstance(i, dict) for i in o):
                return o
            # Heterogeneous list -- try each element
            for item in o:
                found = _find_records(item, depth + 1)
                if found:
                    return found

        if isinstance(o, dict):
            # Prefer keys that look like record arrays
            priority = [k for k in o if isinstance(o[k], list)
                        and o[k] and isinstance(o[k][0], dict)]
            others = [k for k in o if k not in priority]
            for key in priority + others:
                found = _find_records(o[key], depth + 1)
                if found:
                    return found

        return None

    # Root is a list
    if isinstance(obj, list):
        if not obj:
            return pd.DataFrame()

        if all(isinstance(i, dict) for i in obj):
            return pd.json_normalize(obj, sep=".")

        if all(isinstance(i, (str, int, float, bool)) for i in obj):
            return pd.DataFrame({"value": obj})

        # Mixed or nested -- search inside
        records = _find_records(obj)
        if records:
            return pd.json_normalize(records, sep=".")

        try:
            return pd.json_normalize(obj, sep=".")


        except Exception:
            return pd.DataFrame({"value": [json.dumps(i) for i in obj]})

    # Root is a dict -- unwrap wrapper patterns first
    if isinstance(obj, dict):
        records = _find_records(obj)
        if records:
            return pd.json_normalize(records, sep=".")

        # Flat or nested dict -> single row
        return pd.json_normalize([obj], sep=".")

    # Primitive
    return pd.DataFrame({"value": [obj]})


def _parse_fpml(raw: bytes) -> pd.DataFrame:
    # Parse FpML (Financial products Markup Language) XML into a flat DataFrame.

    # Each top-level trade / dataDocument child becomes one row.


    # Leaf elements are flattened to dot-path column names (namespace-stripped).

    # Handles FpML 4.x and 5.x envelopes (FpML, dataDocument, requestConfirmation,
    # etc.).


    def _strip_ns(tag: str) -> str:


        return re.sub(r"\{[^}]+\}", "", tag)


    def _flatten(elem, prefix: str = "") -> dict:


        tag = _strip_ns(elem.tag)

        path = f"{prefix}.{tag}" if prefix else tag

        result: dict = {}

        for k, v in elem.attrib.items():

            result[f"{path}@{_strip_ns(k)}"] = v

        children = list(elem)

        if not children:

            result[path] = (elem.text or "").strip()

        else:

            for child in children:

                result.update(_flatten(child, path))


        return result


    root = ET.fromstring(raw)

    root_tag = _strip_ns(root.tag)

    # Identify repeating trade/product containers
    TRADE_TAGS = {"trade", "dataDocument", "requestConfirmation",
                  "requestClearing", "executionNotification"}

    rows: list[dict] = []

    if root_tag in TRADE_TAGS or root_tag == "FpML":

        # Recurse into top-level children looking for trade elements
        for child in root:

            child_tag = _strip_ns(child.tag)

            if child_tag == "trade" or child_tag.endswith("trade"):

                rows.append(_flatten(child))

        if not rows:

            # Single document -- flatten entire root as one row
            rows.append(_flatten(root))


    else:

        rows.append(_flatten(root))


    return pd.DataFrame(rows)


def _parse_xbrl(raw: bytes) -> pd.DataFrame:
    # Parse XBRL instance documents (inline or standard) into a flat DataFrame.

    # Each context/fact pair becomes one row with columns:
    # context_id, entity, period_start, period_end, concept, value, unit, decimals.
    # Handles both XBRL 2.1 (.xbrl) and iXBRL embedded in HTML (.html/.htm).


    def _strip_ns(tag: str) -> str:


        return re.sub(r"\{[^}]+\}", "", tag)


    text = raw.decode("utf-8", errors="replace")

    # iXBRL: strip HTML wrapper and extract the xbrl portion


    if re.search(r"<html", text, re.IGNORECASE):

        # Pull out ix:nonFraction / ix:nonNumeric tags for inline XBRL
        rows: list[dict] = []

        for m in re.finditer(
            r'<ix:(?:nonFraction|nonNumeric)[^>]*name="([^"]+)"[^>]*'  # OCR-UNCERTAIN: complex regex, exact escaping hard to verify from photo
            r'(?:contextRef="([^"]*)")?[^>]*(?:unitRef="([^"]*)")?[^>]*'  # OCR-UNCERTAIN: complex regex, exact escaping hard to verify from photo
            r'(?:decimals="([^"]*)")?[^>]*>(.*?)</ix:(?:nonFraction|nonNumeric)>',  # OCR-UNCERTAIN: complex regex, exact escaping hard to verify from photo
            text, re.DOTALL | re.IGNORECASE,
        ):

            concept, ctx, unit, decimals, value = m.groups()

            rows.append({
                "concept": concept,
                "context_id": ctx or "",
                "value": re.sub(r"<[^>]+>", "", value).strip(),
                "unit": unit or "",
                "decimals": decimals or "",
            })

        if rows:

            return pd.DataFrame(rows)


    # Standard XBRL: parse contexts then facts
    root = ET.fromstring(raw)

    ns_map: dict[str, str] = {}

    for k, v in root.attrib.items():

        if k.startswith("xmlns:"):

            ns_map[k[6:]] = v


    # Collect contexts: id -> {entity, period_start, period_end}
    contexts: dict[str, dict] = {}

    for ctx in root.iter():

        if _strip_ns(ctx.tag) == "context":

            cid = ctx.attrib.get("id", "")

            entity, p_start, p_end = "", "", ""

            for child in ctx.iter():

                t = _strip_ns(child.tag)

                if t == "identifier":

                    entity = (child.text or "").strip()

                elif t == "startDate":


                    p_start = (child.text or "").strip()

                elif t == "endDate":

                    p_end = (child.text or "").strip()

                elif t == "instant":

                    p_start = p_end = (child.text or "").strip()

            contexts[cid] = {"entity": entity, "period_start": p_start, "period_end": p_end}


    # Collect facts (leaf elements that reference a contextRef)
    rows_xbrl: list[dict] = []

    for elem in root:

        tag = _strip_ns(elem.tag)

        if tag in ("context", "unit", "schemaRef"):

            continue

        ctx_ref = elem.attrib.get("contextRef", "")

        if ctx_ref:

            ctx_info = contexts.get(ctx_ref, {})

            rows_xbrl.append({
                "concept": tag,
                "context_id": ctx_ref,


                "entity": ctx_info.get("entity", ""),
                "period_start": ctx_info.get("period_start", ""),
                "period_end": ctx_info.get("period_end", ""),
                "value": (elem.text or "").strip(),
                "unit": elem.attrib.get("unitRef", ""),
                "decimals": elem.attrib.get("decimals", ""),
            })


    return pd.DataFrame(rows_xbrl) if rows_xbrl else pd.DataFrame()


def _parse_nacha(raw: bytes, enc: str) -> pd.DataFrame:
    # Parse NACHA / ACH fixed-width flat files into a DataFrame.
    # Each record type is identified by its leading digit (1=File Header,
    # 5=Batch Header, 6=Entry Detail, 7=Addenda, 8=Batch Control, 9=File Control).
    # Returns one row per Entry Detail (record type 6) enriched with its
    # parent Batch Header (record type 5) fields for easy reconciliation.

    RECORD_SPECS: dict[str, dict] = {
        "1": {  # File Header
            "record_type": (0, 1), "priority_code": (1, 3),
            "immediate_destination": (3, 13), "immediate_origin": (13, 23),
            "file_creation_date": (23, 29), "file_creation_time": (29, 33),
            "file_id_modifier": (33, 34), "record_size": (34, 37),
            "blocking_factor": (37, 39), "format_code": (39, 40),
            "destination_name": (40, 63), "origin_name": (63, 86),
            "reference_code": (86, 94),
        },
        "5": {  # Batch Header
            "record_type": (0, 1), "service_class_code": (1, 4),
            "company_name": (4, 20), "company_discretionary_data": (20, 40),
            "company_id": (40, 50), "standard_entry_class": (50, 53),
            "company_entry_description": (53, 63), "company_descriptive_date": (63, 69),
            "effective_entry_date": (69, 75), "settlement_date": (75, 78),
            "originator_status_code": (78, 79), "odfi_routing": (79, 87),
            "batch_number": (87, 94),
        },

        "6": {  # Entry Detail
            "record_type": (0, 1), "transaction_code": (1, 3),
            "rdfi_routing": (3, 11), "check_digit": (11, 12),
            "dfi_account_number": (12, 29), "amount": (29, 39),
            "individual_id": (39, 54), "individual_name": (54, 76),
            "discretionary_data": (76, 78), "addenda_indicator": (78, 79),
            "trace_number": (79, 94),
        },
        "7": {  # Addenda
            "record_type": (0, 1), "addenda_type_code": (1, 3),
            "payment_related_info": (3, 83), "sequence_number": (83, 87),
            "entry_detail_sequence": (87, 94),
        },
        "8": {  # Batch Control
            "record_type": (0, 1), "service_class_code": (1, 4),
            "entry_addenda_count": (4, 10), "entry_hash": (10, 20),
            "total_debit": (20, 32), "total_credit": (32, 44),
            "company_id": (44, 54), "odfi_routing": (84, 92),  # OCR-UNCERTAIN: offset overlaps with batch_number below, exact numbers hard to verify from photo
            "batch_number": (87, 94),

        },
    }

    text = raw.decode(enc, errors="replace")
    lines = [ln for ln in text.splitlines() if len(ln.strip()) == 94]

    rows: list[dict] = []
    current_batch: dict = {}

    for line in lines:
        rtype = line[0]
        spec = RECORD_SPECS.get(rtype)
        if not spec:
            continue
        record = {field: line[s:e].strip() for field, (s, e) in spec.items()}
        if rtype == "5":
            current_batch = {f"batch_{k}": v for k, v in record.items()}
        elif rtype == "6":
            entry = {**current_batch, **record}

            # Convert amount to decimal (NACHA stores cents without decimal point)
            try:
                entry["amount"] = int(entry.get("amount", "0")) / 100
            except ValueError:
                pass
            rows.append(entry)
        elif rtype == "7":
            # Attach addenda info to last entry row
            if rows:
                rows[-1]["addenda_payment_info"] = record.get("payment_related_info", "")

    # If no entry detail records found, return all records generically
    if not rows:
        for line in lines:
            rtype = line[0]
            spec = RECORD_SPECS.get(rtype)
            if spec:
                rows.append({f: line[s:e].strip() for f, (s, e) in spec.items()})

    return pd.DataFrame(rows)


def _parse_pdf(raw: bytes) -> tuple[pd.DataFrame, str]:
    # Extract a DataFrame from a PDF.
    # Strategy:
    # 1. Use pdfplumber to extract tables directly (works well for PDFs with
    # embedded table structures -- reports, statements, confirmations).
    # 2. If no tables found, extract all text and pass to parse_unstructured()
    # so Claude LLM can infer the structure.

    # Returns (DataFrame, extracted_text_for_llm_fallback).
    # The caller should use the DataFrame when it has >1 column/row, otherwise
    # fall through to the LLM path using the returned text.

    # Group tables by column count so we return the richest consistent schema
    # Tables with only 1 column are almost always layout artefacts (logos,
    # section headers) and are excluded from the structured output.

    tables_by_ncols: dict[int, list[pd.DataFrame]] = {}
    all_text_lines: list[str] = []

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or len(table[0]) < 2:
                    continue  # skip 1-column layout artefacts
                headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(table[0])]
                # De-duplicate repeated header names
                seen: dict[str, int] = {}
                deduped: list[str] = []
                for h in headers:
                    if h in seen:
                        seen[h] += 1
                        deduped.append(f"{h}_{seen[h]}")
                    else:
                        seen[h] = 0
                        deduped.append(h)

                rows = [[str(c).strip() if c is not None else "" for c in row] for row in table[1:]]
                if rows:
                    df_tbl = pd.DataFrame(rows, columns=deduped)
                    tables_by_ncols.setdefault(len(deduped), []).append(df_tbl)
            # Always collect text for LLM fallback
            text = page.extract_text()
            if text:
                all_text_lines.append(text)

    full_text = "\n".join(all_text_lines)

    if tables_by_ncols:
        # Pick the column-count group with the most total rows (richest data)
        best_ncols = max(tables_by_ncols, key=lambda n: sum(len(t) for t in tables_by_ncols[n]))
        df = pd.concat(tables_by_ncols[best_ncols], ignore_index=True)
        return df, full_text

    # No multi-column tables -- return empty DataFrame so caller uses LLM path

    return pd.DataFrame(), full_text


def _parse_bloomberg_dlx(raw: bytes, enc: str) -> pd.DataFrame:
    # Parse Bloomberg Data License eXchange (BDL/DLX) pipe-delimited files.
    # Bloomberg DL files have a metadata header section (START-OF-FILE ...
    # END-OF-FIELDS) followed by a data section (START-OF-DATA ... END-OF-DATA).
    # Each data row is pipe-delimited and aligns to the field list declared in
    # the header.  Files that lack Bloomberg envelope markers are treated as
    # plain pipe-delimited CSV.

    text = raw.decode(enc, errors="replace")
    lines = text.splitlines()

    fields: list[str] = []
    data_rows: list[list[str]] = []
    in_fields = False
    in_data = False

    for line in lines:
        stripped = line.strip()
        if stripped == "START-OF-FIELDS":
            in_fields = True
            continue
        if stripped == "END-OF-FIELDS":
            in_fields = False
            continue
        if stripped == "START-OF-DATA":
            in_data = True
            continue
        if stripped == "END-OF-DATA":
            in_data = False
            continue
        if in_fields and stripped:
            fields.append(stripped)
        elif in_data and stripped:
            data_rows.append(stripped.split("|"))

    if not fields and not data_rows:
        # Fallback: plain pipe-delimited
        return pd.read_csv(io.BytesIO(raw), sep="|", encoding=enc,
                            engine="python", on_bad_lines="skip")

    if fields and data_rows:
        # Pad/trim rows to match field count
        n = len(fields)
        normalised = [row[:n] + [""] * max(0, n - len(row)) for row in data_rows]
        return pd.DataFrame(normalised, columns=fields)

    # Header-only or data-only edge cases
    return pd.read_csv(io.BytesIO(raw), sep="|", encoding=enc,
                        engine="python", on_bad_lines="skip")


def _parse_murex_mxml(raw: bytes) -> pd.DataFrame:
    # Parse Murex MXML trade export files into a flat DataFrame.


    # MXML wraps trades in <TRADES><TRADE>...</TRADE></TRADES> or
    # <MXMLDocument><Trade>...</Trade></MXMLDocument> envelopes.
    # Each trade element is dot-path flattened (namespace-stripped) into one row.
    # Falls back to generic XML flattening when no known Murex envelope is found.

    def _strip_ns(tag: str) -> str:


        return re.sub(r"\{[^}]+\}", "", tag)


    def _flatten(elem, prefix: str = "") -> dict:


        tag = _strip_ns(elem.tag)

        path = f"{prefix}.{tag}" if prefix else tag

        result: dict = {}

        for k, v in elem.attrib.items():

            result[f"{path}@{_strip_ns(k)}"] = v

        children = list(elem)

        if not children:

            result[path] = (elem.text or "").strip()

        else:


            for child in children:

                result.update(_flatten(child, path))

        return result


    root = ET.fromstring(raw)

    root_tag = _strip_ns(root.tag).upper()


    TRADE_TAGS = {"TRADE", "DEAL", "TRANSACTION"}

    rows: list[dict] = []


    # Look for repeating trade children at root or one level deep
    for candidate in [root] + list(root):

        for child in candidate:

            if _strip_ns(child.tag).upper() in TRADE_TAGS:

                rows.append(_flatten(child))

        if rows:

            break


    if not rows:


        rows.append(_flatten(root))


    return pd.DataFrame(rows)


def _parse_markitwire(raw: bytes) -> pd.DataFrame:
    # Parse MarkitWire / TradeWeb XML confirmation files.

    # MarkitWire confirmations are standard XML with a <tradeConfirmed> or
    # <FpML> root containing one or more <trade> elements.  Each trade is
    # dot-path flattened into one row.  Delegates to _parse_fpml for FpML
    # envelopes; falls back to generic XML for non-standard layouts.

    # MarkitWire often embeds FpML -- reuse that parser
    text = raw.decode("utf-8", errors="replace")
    if re.search(r"<FpML|fpml\.org|tradeConfirmed", text, re.IGNORECASE):
        return _parse_fpml(raw)

    return _parse_xml(raw)


def _is_bloomberg_dlx(raw: bytes, enc: str) -> bool:
    # Return True if the file looks like a Bloomberg DL/BDL envelope.
    try:

        head = raw[:2048].decode(enc, errors="replace")

        return "START-OF-FILE" in head or "START-OF-FIELDS" in head
    except Exception:

        return False


def _parse_dtcc_gtr(raw: bytes, enc: str) -> pd.DataFrame:
    # Parse DTCC GTR / TRACE regulatory reporting files.

    # DTCC GTR submissions are pipe-delimited flat files with a one-line header.
    # TRACE files follow the same convention (tab or pipe separated, fixed header).
    # The parser auto-detects the delimiter and returns one row per trade report.

    text = raw.decode(enc, errors="replace")
    first_line = text.split("\n")[0] if text else ""

    # Detect delimiter from header line
    pipe_count = first_line.count("|")
    tab_count = first_line.count("\t")
    delim = "|" if pipe_count >= tab_count else "\t"

    return pd.read_csv(io.BytesIO(raw), sep=delim, encoding=enc,
                        engine="python", on_bad_lines="skip")


def _parse_reuters_ric(raw: bytes, enc: str) -> pd.DataFrame:
    # Parse Reuters/Refinitiv RIC market-data feed files.
    # RIC feed files are CSV/TSV exports from Refinitiv Eikon or DataScope with
    # a standard header row.  Common column names include RIC, Date, Open, High,
    # Low, Close, Volume, Bid, Ask.  Returns one row per record.

    text = raw.decode(enc, errors="replace")
    first_line = text.split("\n")[0] if text else ""

    # Detect delimiter
    pipe_count = first_line.count("|")
    tab_count = first_line.count("\t")
    comma_count = first_line.count(",")
    if pipe_count >= tab_count and pipe_count >= comma_count:
        delim = "|"
    elif tab_count >= comma_count:
        delim = "\t"
    else:
        delim = ","

    return pd.read_csv(io.BytesIO(raw), sep=delim, encoding=enc,
                        engine="python", on_bad_lines="skip")


def _is_junk_archive_member(filename: str) -> bool:
    base = os.path.basename(filename)
    return (not base or filename.startswith("__MACOSX/")
            or base.startswith(".") or base.lower() in ("thumbs.db", "desktop.ini"))


def _extract_zip_members(raw: bytes, archive_name: str) -> list[tuple[bytes, str]]:
    import zipfile
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        raise HTTPException(400, f"'{archive_name}' is not a valid zip file.")
    members: list[tuple[bytes, str]] = []
    for info in zf.infolist():
        if info.is_dir() or _is_junk_archive_member(info.filename):
            continue
        try:
            members.append((zf.read(info), os.path.basename(info.filename)))
        except Exception:
            continue  # skip a corrupt member rather than failing the whole archive
    if not members:
        raise HTTPException(400, f"Zip file '{archive_name}' contains no usable files.")
    return members


def _extract_tar_members(raw: bytes, archive_name: str) -> list[tuple[bytes, str]]:
    import tarfile
    try:
        tf = tarfile.open(fileobj=io.BytesIO(raw), mode="r:*")
    except tarfile.TarError as exc:
        raise HTTPException(400, f"'{archive_name}' is not a valid tar archive: {exc}")
    members: list[tuple[bytes, str]] = []
    try:
        for info in tf.getmembers():
            if not info.isfile() or _is_junk_archive_member(info.name):
                continue
            try:
                fh = tf.extractfile(info)
                if fh is None:
                    continue
                members.append((fh.read(), os.path.basename(info.name)))
            except Exception:
                continue
    finally:
        tf.close()
    if not members:
        raise HTTPException(400, f"Tar archive '{archive_name}' contains no usable files.")
    return members


def _extract_gz_member(raw: bytes, archive_name: str) -> list[tuple[bytes, str]]:
    import gzip
    try:
        data = gzip.decompress(raw)
    except Exception as exc:
        raise HTTPException(400, f"'{archive_name}' is not a valid gzip file: {exc}")
    inner_name = re.sub(r"\.gz$", "", archive_name, flags=re.IGNORECASE) or "data"
    return [(data, os.path.basename(inner_name))]


def _extract_archive_members(raw: bytes, name: str) -> list[tuple[bytes, str]] | None:
    """Every usable file inside a recognised archive (zip / tar / tar.gz / tgz /
    tar.bz2 / gz), skipping junk (macOS resource forks, hidden/system files) and
    any individual corrupt member without failing the whole upload. Returns None
    if `name` isn't a recognised archive extension (i.e. nothing to extract)."""
    lower = name.lower()
    try:
        if lower.endswith((".tar.gz", ".tgz", ".tar.bz2", ".tbz2", ".tar")):
            return _extract_tar_members(raw, name)
        if lower.endswith(".zip"):
            return _extract_zip_members(raw, name)
        if lower.endswith(".gz"):
            return _extract_gz_member(raw, name)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Could not open '{name}' as an archive: {exc}") from exc
    return None


def _raise_if_looks_like_broken_xml(raw: bytes, enc: str, xml_exc: Exception) -> None:
    # Auto-detection falls through to a generic delimited-text parser when
    # XML parsing fails -- that parser never truly "fails", it just mangles
    # tag fragments into meaningless columns and returns them silently. If
    # the content clearly looks like XML (an <?xml declaration, a DOCTYPE,
    # or a bare '<' as the first non-whitespace byte) but didn't parse, that
    # silent garbage is worse than a clear error, so raise one instead of
    # letting the cascade continue.
    try:
        sample = raw.lstrip()[:200].decode(enc, errors="replace").lstrip()
    except Exception:
        return
    if sample.startswith(("<?xml", "<!DOCTYPE", "<")):
        raise ValueError(f"This file looks like XML but is not well-formed: {xml_exc}")


def _load_file(upload, delimiter=None) -> pd.DataFrame:
    # Load any supported file format into a DataFrame.
    #
    # Sets df.attrs['_format'] and df.attrs['_delimiter'] on the returned frame.
    #
    # Parameters
    # ----------
    # upload : UploadFile from the multipart form
    # delimiter : Optional explicit delimiter character (user-supplied).
    #     Applies to CSV/TXT/unknown files.  Pass None to auto-detect.
    #     Accepts escape sequences typed as strings (e.g. "\t" -> tab).
    #
    # Supported formats
    # -----------------
    # CSV - TSV - PSV - XLSX/XLS - Parquet - JSON (nested) - JSON Lines
    # (.jsonl/.ndjson) - XML - ISO 20022 MX - SWIFT MT (.mt/.fin/.swift)
    # FIX Protocol (.fix) - FpML (.fpml) - XBRL (.xbrl) - NACHA/ACH
    # (.ach/.nacha) - Avro (.avro) - ORC (.orc) - PDF (.pdf)
    # Bloomberg DLX (.bbg/.dlx/.bdl) - Murex MXML (.mxml)
    # MarkitWire (.mwire/.markitwire) - DTCC GTR/TRACE (.gtr/.trace)
    # Reuters RIC (.ric) - plain TXT
    # Text files without a recognised extension are auto-sniffed for
    # SWIFT MT -> FIX -> delimiter-separated -> single-value fallback.
    #
    # Archives -- .zip / .tar / .tar.gz / .tgz / .tar.bz2 / .gz are
    # transparently unpacked first; multi-file loaders (_load_file_multi_
    # from_upload, used by Reconciliation/Cross Reference) expand every
    # usable file inside into a separate dataset, not just the largest.
    raw = upload.file.read()
    name = upload.filename or ""
    return _load_file_from_bytes(raw, name, delimiter)


def _load_file_from_bytes(raw: bytes, name: str, delimiter=None) -> pd.DataFrame:
    # Core of _load_file(), operating on already-read bytes so archive members
    # (extracted separately, see _load_file_multi_from_upload) can be parsed
    # through the same format-detection logic without re-reading an upload.
    ext = os.path.splitext(name)[1].lower()

    _resolved_name = name
    members = _extract_archive_members(raw, name)
    if members:
        # Single-file callers get the largest member (the actual data file,
        # not a bundled README/manifest) -- use _load_file_multi_from_upload
        # to get every member back as a separate dataset.
        raw, _resolved_name = max(members, key=lambda m: len(m[0]))
        ext = os.path.splitext(_resolved_name)[1].lower()
        name = _resolved_name

    enc = _detect_encoding(raw)

    # Normalise escape sequences that users type in a text box (e.g. "\t" -> tab)
    if delimiter and len(delimiter) > 1:
        try:
            delimiter = delimiter.encode("raw_unicode_escape").decode("unicode_escape")
        except Exception:
            pass  # keep as-is if normalisation fails

    def _labelled(df: pd.DataFrame, fmt: str, delim: str = "") -> pd.DataFrame:
        df.attrs["_format"] = fmt
        df.attrs["_delimiter"] = delim
        # Surface any parse warnings -- _load_file() has no logger of its own, so
        # warnings are left on df.attrs for the caller's own _log() to pick up.
        df.attrs["_parse_warnings"] = df.attrs.pop("_parse_warnings", [])
        # Resolved filename (differs from upload.filename when the upload was a
        # .zip -- callers should prefer this over upload.filename for display).
        df.attrs["_source_filename"] = _resolved_name
        return df

    def _try_read(sep: str) -> pd.DataFrame | None:
        # Attempt pd.read_csv with a given separator; return None on failure or 1-column result.
        try:
            df = pd.read_csv(io.BytesIO(raw), sep=sep, encoding=enc,
                              engine="python", on_bad_lines="skip", low_memory=False)
            return df if len(df.columns) > 1 else None
        except Exception:
            return None

    def _csv_with_delim(fmt_label: str) -> pd.DataFrame:
        # Parse a CSV / delimited file with reliable auto-detection.
        # Uses the robust _sniff_delimiter (count-and-confirm + sniffer fallback)
        # so files like SmartLoan.csv that are actually pipe-delimited parse correctly
        # even when the extension says .csv.
        #
        # Detection order:
        # 1. User-supplied delimiter (trust it)
        # 2. Robust count-and-confirm sniffer (_sniff_delimiter)
        # 3. Trial of all BFSI delimiters with auto-repair (_read_delimited)
        # 4. Single-column fallback
        if delimiter:
            df = pd.read_csv(io.BytesIO(raw), sep=delimiter, encoding=enc,
                              engine="python", on_bad_lines="skip", dtype=str)
            return _labelled(df, fmt_label, delimiter)

        # Decode once for sniffing
        text, used_enc = _decode_raw(raw)

        # -- Robust sniffer (handles pipe-delimited .csv, spaces in col names, etc.)
        sniffed = _sniff_delimiter(text)
        if sniffed:
            df = _read_delimited(text, sniffed)
            if df is not None:
                return _labelled(df, fmt_label, sniffed)

        # -- Trial of all BFSI delimiters with auto-repair
        tried = {sniffed or ""}
        for sep in [",", "\t", "|", ";", "~", "^", ":", "=", "\x01"]:
            if sep in tried:
                continue
            tried.add(sep)
            df = _read_delimited(text, sep)
            if df is not None:
                return _labelled(df, fmt_label, sep)

        # -- Single-column fallback
        try:
            df = pd.read_csv(io.StringIO(text), sep=",", on_bad_lines="skip", dtype=str)
            return _labelled(df, fmt_label, ",")
        except Exception:
            return _labelled(pd.DataFrame({"value": text.splitlines()}), fmt_label, "")

    try:
        if ext == ".csv":
            return _csv_with_delim("CSV")

        if ext == ".tsv":
            # TSV is explicitly tab-separated; honour user override if supplied
            sep = delimiter or "\t"
            df = pd.read_csv(io.BytesIO(raw), sep=sep, encoding=enc,
                              engine="python", on_bad_lines="skip")
            return _labelled(df, "TSV", sep)

        if ext == ".psv":
            # Explicitly pipe-separated
            sep = delimiter or "|"
            df = pd.read_csv(io.BytesIO(raw), sep=sep, encoding=enc,
                              engine="python", on_bad_lines="skip")
            return _labelled(df, "PSV", sep)

        if ext in (".jsonl", ".ndjson"):
            # JSON Lines: one JSON object per line (common for event/log exports)
            df = pd.read_json(io.BytesIO(raw), lines=True, encoding=enc)
            return _labelled(df, "JSON Lines")

        if ext in (".xlsx", ".xls", ".xlsm"):
            # .xlsm (macro-enabled Excel) is the same OOXML zip container as
            # .xlsx -- openpyxl reads the cell data fine, it just never
            # touches the embedded VBA macros.
            return _labelled(pd.read_excel(io.BytesIO(raw)), "Excel")

        if ext == ".parquet":
            return _labelled(pd.read_parquet(io.BytesIO(raw)), "Parquet")

        # -- JSON (including nested / deeply-nested)
        if ext == ".json":
            return _labelled(_parse_json_nested(raw), "JSON")

        # -- XML family (detect ISO 20022 from namespace)
        if ext == ".xml":
            if _is_iso20022(raw):
                return _labelled(_parse_iso20022(raw), "ISO 20022 XML")
            return _labelled(_parse_xml(raw), "XML")

        # -- ISO 20022 / MX files explicitly named
        if ext in (".mx", ".pacs", ".pain", ".camt", ".sepa"):
            return _labelled(_parse_iso20022(raw), "ISO 20022 MX")

        # -- SWIFT MT / FIN files
        if ext in (".mt", ".fin", ".swift"):
            return _labelled(_parse_swift_mt(raw, enc), "SWIFT MT")

        # -- FIX protocol
        if ext == ".fix":
            return _labelled(_parse_fix(raw, enc), "FIX Protocol")

        # -- PDF: table extraction -> LLM fallback
        if ext == ".pdf":
            df_pdf, pdf_text = _parse_pdf(raw)
            if not df_pdf.empty and len(df_pdf.columns) > 1:
                return _labelled(df_pdf, "PDF")
            # No embedded tables -- use the LLM to interpret the text. `name`
            # is this function's own filename parameter -- there is no
            # `upload` object in scope here (that's the caller's parameter),
            # referencing it raised a NameError on every text-only PDF.
            if pdf_text.strip():
                result = parse_unstructured(pdf_text.encode("utf-8"), name or "file.pdf")
                # parse_unstructured returns "rows" as a list of dicts already
                # (record-oriented, keyed by column name) -- zipping them
                # against `cols` treated each dict as an iterable of its own
                # keys, silently producing {"col": "col"} garbage instead of
                # the actual parsed values.
                if not result.get("error") and result.get("rows"):
                    cols = result["columns"]
                    return _labelled(pd.DataFrame(result["rows"], columns=cols), "PDF (AI-parsed)")
            raise ValueError("Could not extract any structured data from this PDF.")

        # -- FpML (Financial products Markup Language)
        if ext == ".fpml":
            return _labelled(_parse_fpml(raw), "FpML")

        # -- XBRL instance documents (standard or inline)
        if ext in (".xbrl",):
            df_xbrl = _parse_xbrl(raw)
            if df_xbrl.empty:
                raise ValueError("No XBRL facts found in this file.")
            return _labelled(df_xbrl, "XBRL")

        # -- NACHA / ACH fixed-width payment files
        if ext in (".ach", ".nacha"):
            return _labelled(_parse_nacha(raw, enc), "NACHA/ACH")

        # -- Apache Avro
        if ext == ".avro":
            try:
                import fastavro
                reader = fastavro.reader(io.BytesIO(raw))
                records = list(reader)
                return _labelled(pd.DataFrame(records), "Avro")
            except ImportError:
                raise ImportError("fastavro is required for Avro files. Install it with: pip install fastavro")

        # -- Apache ORC
        if ext == ".orc":
            try:
                import pyarrow.orc as orc
                table = orc.read_table(io.BytesIO(raw))
                return _labelled(table.to_pandas(), "ORC")
            except ImportError:
                raise ImportError("pyarrow is required for ORC files (already in requirements).")

        # -- Bloomberg Data License (BDL/DLX)
        if ext in (".bbg", ".dlx", ".bdl"):
            return _labelled(_parse_bloomberg_dlx(raw, enc), "Bloomberg DLX")

        # -- Murex MXML trade exports
        if ext == ".mxml":
            return _labelled(_parse_murex_mxml(raw), "Murex MXML")

        # -- MarkitWire / TradeWeb XML confirmations
        if ext in (".mwire", ".markitwire"):
            return _labelled(_parse_markitwire(raw), "MarkitWire")

        # -- DTCC GTR / TRACE regulatory reporting
        if ext in (".gtr", ".trace"):
            return _labelled(_parse_dtcc_gtr(raw, enc), "DTCC GTR/TRACE")

        # -- Reuters / Refinitiv RIC market-data feeds
        if ext == ".ric":
            return _labelled(_parse_reuters_ric(raw, enc), "Reuters RIC")

        # -- Text files: sniff format (JSON -> XML -> SWIFT MT -> FIX -> delimited)
        if ext in (".txt", ""):
            if not delimiter:
                # Try JSON first -- catches .txt files containing JSON
                try:
                    df = _parse_json_nested(raw)
                    if len(df.columns) > 1 or len(df) > 1:
                        return _labelled(df, "JSON (auto-detected)")
                except Exception:
                    pass

                # Try XML next
                try:
                    df = _parse_xml(raw)
                    if len(df.columns) > 1 or len(df) > 1:
                        return _labelled(df, "XML (auto-detected)")
                except Exception as _xml_exc:
                    _raise_if_looks_like_broken_xml(raw, enc, _xml_exc)

                if _is_bloomberg_dlx(raw, enc):
                    return _labelled(_parse_bloomberg_dlx(raw, enc), "Bloomberg DLX (auto-detected)")

                if _is_swift_mt(raw, enc):
                    return _labelled(_parse_swift_mt(raw, enc), "SWIFT MT (auto-detected)")

                if _is_fix(raw, enc):
                    return _labelled(_parse_fix(raw, enc), "FIX Protocol (auto-detected)")

            df = _parse_txt(raw, enc, delimiter)
            used_delim = df.attrs.pop("_detected_delimiter", delimiter or "")
            used_enc_attr = df.attrs.pop("_encoding_used", enc)
            parse_warns = df.attrs.pop("_parse_warnings", [])
            label = f"Text/CSV ({used_enc_attr})" if used_enc_attr != "utf-8" else "Text/CSV"
            if parse_warns:
                df.attrs["_parse_warnings"] = parse_warns
            return _labelled(df, label, used_delim)

        # -- Unknown extension: CSV -> JSON -> XML -> SWIFT MT -> FIX -> plain text
        try:
            df = _csv_with_delim(ext.lstrip(".").upper() or "Data")
            if len(df.columns) > 1:
                return df  # already labelled by _csv_with_delim
        except Exception:
            pass

        if not delimiter:
            # Try JSON/XML before SWIFT MT to avoid false-positive detection
            try:
                df = _parse_json_nested(raw)
                if len(df.columns) > 1 or len(df) > 1:
                    return _labelled(df, "JSON (auto-detected)")
            except Exception:
                pass

            try:
                df = _parse_xml(raw)
                if len(df.columns) > 1 or len(df) > 1:
                    return _labelled(df, "XML (auto-detected)")
            except Exception as _xml_exc:
                _raise_if_looks_like_broken_xml(raw, enc, _xml_exc)

            if _is_swift_mt(raw, enc):
                return _labelled(_parse_swift_mt(raw, enc), "SWIFT MT (auto-detected)")

            if _is_fix(raw, enc):
                return _labelled(_parse_fix(raw, enc), "FIX Protocol (auto-detected)")

        df = _parse_txt(raw, enc, delimiter)
        used_delim = df.attrs.pop("_detected_delimiter", delimiter or "")
        return _labelled(df, "Text", used_delim)

    # OCR-UNCERTAIN: exact enclosing try/function for this except block not visible on this page
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Cannot parse '{name}': {exc}") from exc


def _load_file_multi_from_upload(upload, delimiter=None) -> list[tuple[str, pd.DataFrame]]:
    """Like _load_file(), but when the upload is an archive containing multiple
    usable files, returns ALL of them as separate (filename, DataFrame) entries
    instead of just the largest -- so one zip/tar/tgz upload can expand into
    several datasets (e.g. multiple monthly extracts to combine, or several
    sources for cross reference). A member that fails to parse is skipped
    rather than failing the whole batch; only raises if none could be parsed."""
    raw = upload.file.read()
    name = upload.filename or ""
    members = _extract_archive_members(raw, name)
    if not members:
        return [(name, _load_file_from_bytes(raw, name, delimiter))]

    results: list[tuple[str, pd.DataFrame]] = []
    skipped: list[str] = []
    for member_raw, member_name in members:
        try:
            results.append((member_name, _load_file_from_bytes(member_raw, member_name, delimiter)))
        except Exception:
            skipped.append(member_name)
            continue
    if not results:
        raise HTTPException(400, f"None of the {len(members)} file(s) inside '{name}' could be parsed.")
    return results


# ------------------------------------------------------------------------
# Reference document parsing - auto-classification + multi-file support
# ------------------------------------------------------------------------


def _df_from_connection(conn_id: str, username: str) -> tuple[str, pd.DataFrame]:
    # Fetch a DataFrame from a saved workspace connection.
    # Returns (connection_name, DataFrame).
    # Raises HTTPException on failure so the user gets a clear error.

    if not _WS_ENABLED:
        raise HTTPException(503, "Workspace feature not available -- cannot load from saved connection.")

    rec = _ws_db.get_connection(conn_id, username)
    if not rec:
        raise HTTPException(404, f"Workspace connection '{conn_id}' not found for this user.")

    # Retry with backoff on transient failures (network blips, momentary DB
    # unavailability, rate limits) -- enterprise connectors expect this;
    # without it a single dropped packet fails an entire scheduled job.
    # Config errors (bad table name, missing field) fail the same way on
    # every attempt, so they still surface promptly after the retries.
    _max_attempts = 3
    _last_exc: Exception | None = None
    for _attempt in range(1, _max_attempts + 1):
        try:
            connector = _ws_connectors.BaseConnector.from_type(rec["source_type"], rec["config"])
            df = connector.fetch()
            df.attrs["_format"] = rec["source_type"]
            df.attrs["_delimiter"] = ""
            return rec["name"], df
        except HTTPException:
            raise
        except Exception as exc:
            _last_exc = exc
            if _attempt < _max_attempts:
                # _log() is a per-request closure defined inside each route
                # handler -- this is a module-level function with no access
                # to it, so calling it here raised NameError on every retry
                # (masking whatever the real connection error was).
                import logging as _logging
                _logging.getLogger(__name__).warning(
                    "Connection '%s' fetch attempt %d/%d failed (%s) -- retrying...",
                    rec["name"], _attempt, _max_attempts, exc,
                )
                time.sleep(0.5 * (2 ** (_attempt - 1)))

    raise HTTPException(
        502, f"Failed to fetch data from '{rec['name']}' after {_max_attempts} attempts: {_last_exc}"
    ) from _last_exc


# Column-name keyword sets used for auto-detection
_DD_KEYS = {"description", "owner", "sensitivity", "data_type", "nullable",
            "is_pk", "format_pattern", "allowed_values", "business_term",
            "classification", "steward", "pii", "retention"}

_RULES_KEYS = {"rule_type", "rule_name", "threshold", "operator",
               "condition", "expectation", "check_type", "dq_rule",
               "min_value", "max_value", "allowed_values", "pattern",
               "not_null", "unique"}

_MAP_KEYS = {"source_column", "source_field", "source_attribute",
             "target_column", "target_field", "target_attribute",
             "transformation", "mapping", "src", "tgt"}


def _classify_ref_doc(df: pd.DataFrame) -> str:
    # Auto-detect reference document type by scoring its column names.
    # Returns one of: 'data_dict' | 'rules' | 'mapping' | 'general'


    if df.empty:
        return "general"

    cols = {c.strip().lower().replace(" ", "_").replace("-", "_") for c in df.columns}

    # Mapping: needs BOTH a source-like and a target-like column
    has_src = bool(cols & {"source_column", "source_field", "source_attribute", "src", "source"})
    has_tgt = bool(cols & {"target_column", "target_field", "target_attribute", "tgt", "target"})

    if has_src and has_tgt:
        return "mapping"

    # Score each type by overlap with keyword sets
    dd_score = len(cols & _DD_KEYS)
    rules_score = len(cols & _RULES_KEYS)

    # Rules: must have a rule_type or similar plus a column reference
    has_col_ref = bool(cols & {"column_name", "column", "field_name", "field", "attribute"})
    if rules_score >= 1 and has_col_ref:
        return "rules"


    if rules_score >= 2:  # even without column ref (dataset-level rules)
        return "rules"

    # Data dictionary: column reference + at least one metadata keyword
    if has_col_ref and dd_score >= 1:
        return "data_dict"

    # Fallback: if many dd keywords present even without explicit column ref
    if dd_score >= 3:
        return "data_dict"

    return "general"


def _extract_kb_as_ref_doc(raw_bytes: bytes, filename: str) -> dict:
    # Knowledge Base extraction -- handles ANY file format as a reference document.
    # For structured/tabular files (CSV, Excel, JSON, XML, SWIFT, FIX, Parquet):
    # -> _load_file() -> DataFrame -> _classify_ref_doc() -> existing pipeline
    #
    # For unstructured files (PDF free-text, plain-text prose, Notepad, Word/HTML):
    # -> Extract text -> LLM reads it -> returns data_dict + rules in standard format
    #
    # Returns same structure as _load_and_classify_ref_docs() per-file output:
    # {"doc_type", "data_dict", "rules", "mapping_spec", "raw_text"}
    import io as _io

    fname_lower = filename.lower()

    # -- Try structured parse first --

    class _FakeUpload:
        def __init__(self, fname, data):
            self.filename = fname
            self.file = _io.BytesIO(data)

    try:
        df = _load_file(_FakeUpload(filename, raw_bytes))
        doc_type = _classify_ref_doc(df)
        if doc_type != "general" or len(df.columns) > 1:
            # Successfully parsed as tabular -- use existing pipeline
            result = {"doc_type": doc_type, "data_dict": {}, "rules": [], "mapping_spec": []}
            if doc_type == "data_dict":
                result["data_dict"] = _parse_data_dictionary(df)
            elif doc_type == "rules":
                result["rules"] = _parse_business_rules(df)
            elif doc_type == "mapping":
                result["mapping_spec"] = _parse_mapping_spec(df)
            else:
                # Tabular but unclassified -- try all parsers and take best result
                dd = _parse_data_dictionary(df)
                rls = _parse_business_rules(df)
                if dd:
                    result["doc_type"] = "data_dict"
                    result["data_dict"] = dd
                elif rls:
                    result["doc_type"] = "rules"
                    result["rules"] = rls
            return result
    except Exception:
        pass  # Fall through to unstructured LLM extraction

    # -- Unstructured extraction via LLM --------------------------------------
    # Handles: PDF free-text, Notepad prose, Word/HTML/Confluence exports,
    #     any file where tabular parsing failed or produced 1 column
    try:
        import chardet as _cd
        enc = _cd.detect(raw_bytes[:4096]).get("encoding") or "utf-8"
        raw_text = raw_bytes.decode(enc, errors="replace")
    except Exception:
        raw_text = raw_bytes.decode("utf-8", errors="replace")

    # Trim to 8000 chars -- enough for a full data dictionary
    text_sample = raw_text[:8000]

    # NOTE (reconstruction): this prompt string had lost its opening/closing
    # markers embedded mid-string during OCR assembly. Restored as an f-string
    # with the markers removed (they are pipeline artifacts, not original
    # content) -- verify against source in code review.
    prompt = f"""You are a data dictionary and business rules extractor for a BFSI data platform.

The following document is a reference document (data dictionary, business rules, glossary, or mapping spec).

Extract ALL column definitions, business rules, and glossary terms from it.

Document: "{filename}"
Content:
---
{text_sample}
---

Return ONLY valid JSON in this exact format:
{{
  "doc_type": "data_dict" | "rules" | "mapping" | "glossary",
  "data_dict": [
    {{
      "column_name": "trade_id",
      "description": "Unique trade identifier",
      "data_type": "string",
      "nullable": false,
      "sensitivity": "Internal Use Only",
      "owner": "Trade Operations",
      "business_term": "Trade Reference",
      "allowed_values": "",
      "format_pattern": "",
      "min_value": "",
      "max_value": ""
    }}
  ],
  "rules": [
    {{
      "column_name": "notional",
      "rule_type": "positive",
      "description": "Notional must be positive",
      "severity": "major"
    }}
  ],
  "glossary": [
    {{
      "term": "CUSIP",
      "definition": "9-character security identifier",
      "domain": "Reference Data"
    }}
  ]
}}

Extract everything you can find. Return empty arrays [] for sections not present in the
document.

Return ONLY the JSON -- no explanation."""

    try:
        raw_llm = _ask_llm(
            [{"role": "user", "content": [{"text": prompt}]}],
            system="You are a data dictionary extraction engine. Output ONLY valid JSON."
        )
        import re as _re
        m = _re.search(r'\{.*\}', raw_llm, _re.DOTALL)
        parsed = json.loads(m.group(0)) if m else {}
    except Exception:
        parsed = {}

    doc_type = parsed.get("doc_type", "general")

    # Convert LLM data_dict list → our dict[col_name, dict] format
    dd_out: dict = {}
    for entry in parsed.get("data_dict", []):
        col = entry.get("column_name", "").strip()
        if col:
            dd_out[col] = {
                "description":   entry.get("description", ""),
                "data_type":     entry.get("data_type", ""),
                "nullable":      entry.get("nullable", True),
                "sensitivity":   entry.get("sensitivity", ""),
                "owner":         entry.get("owner", ""),
                "business_term": entry.get("business_term", ""),
                "allowed_values": entry.get("allowed_values", ""),
                "format_pattern": entry.get("format_pattern", ""),
                "min_value":      entry.get("min_value", ""),
                "max_value":      entry.get("max_value", ""),
            }

    # Convert LLM rules list → our rules format
    rules_out = []
    for r in parsed.get("rules", []):
        col = r.get("column_name", "")
        rt = r.get("rule_type", "")
        if col and rt:
            rules_out.append({
                "column_name": col,
                "rule_type":   rt,
                "description": r.get("description", ""),
                "severity":    r.get("severity", "major"),
                "value":       r.get("value", ""),
            })


    # Store glossary terms in data_dict with a special marker so they
    # can be used for sensitivity/description enrichment
    for g in parsed.get("glossary", []):
        term = g.get("term", "").strip()
        if term and term not in dd_out:
            dd_out[term] = {
                "description":    g.get("definition", ""),
                "business_term":  term,
                "domain":         g.get("domain", ""),
                "_from_glossary": True,
            }

    return {
        "doc_type":     doc_type,
        "data_dict":    dd_out,
        "rules":        rules_out,
        "mapping_spec": [],
        "raw_text":     text_sample[:500],  # for logging
    }


def _load_and_classify_ref_docs(uploads: list, kb_raw: list | None = None) -> dict:
    # Knowledge Base document loader -- accepts ANY format from ANY source.
    #
    # Inputs:
    #   uploads: list of FastAPI UploadFile objects (local file upload)
    #   kb_raw:  list of (filename, bytes) tuples (fetched from connectors --
    #            SharePoint, OneDrive, Confluence, S3, SFTP, etc.)
    #
    # ALL formats handled:
    #   Structured (CSV, Excel, JSON, XML, SWIFT, FIX, Parquet, TXT-delimited):
    #     → existing tabular pipeline (_classify_ref_doc)
    #   Unstructured (PDF free-text, Notepad prose, Word/HTML, Confluence):
    #     → LLM extraction (_extract_kb_as_ref_doc)
    #
    # Returns:
    # {
    #   'data_dict':       dict[str, dict],  # merged column-level metadata
    #   'rules':           list[dict],       # business rules
    #   'mapping_spec':    list[dict],       # column mappings
    #   'general':         list[str],        # unclassified filenames
    #   'classifications': list[dict]        # per-file log
    # }
    merged_dict: dict[str, dict] = {}
    merged_rules: list[dict]     = []
    merged_mapping: list[dict]   = []
    general_docs: list[str]      = []
    log: list[dict]              = []

    # Build unified list of (filename, raw_bytes) from both upload sources
    sources: list[tuple[str, bytes]] = []

    # 1. Local file uploads (UploadFile objects)
    for upload in (uploads or []):
        if not upload or not getattr(upload, "filename", None):
            continue
        try:
            upload.file.seek(0)
            raw = upload.file.read()
            sources.append((upload.filename, raw))
        except Exception as exc:
            log.append({"file": getattr(upload, "filename", "?"), "type": "error", "detail": str(exc)})

    # 2. Connector-fetched bytes (SharePoint, OneDrive, S3, SFTP, Confluence etc.)
    for fname, raw in (kb_raw or []):
        if fname and raw:
            sources.append((fname, raw))

    # Process each source through the unified KB extractor
    for fname, raw_bytes in sources:
        try:
            result = _extract_kb_as_ref_doc(raw_bytes, fname)


            doc_type = result.get("doc_type", "general")

            log_entry = {"file": fname, "type": doc_type}
            if result.get("data_dict"):
                log_entry["columns_extracted"] = len(result["data_dict"])
            if result.get("rules"):
                log_entry["rules_extracted"] = len(result["rules"])
            if result.get("raw_text"):
                log_entry["detail"] = f"LLM-extracted from unstructured text"
            log.append(log_entry)

            if result.get("data_dict"):
                merged_dict.update(result["data_dict"])
            if result.get("rules"):
                merged_rules.extend(result["rules"])
            if result.get("mapping_spec"):
                merged_mapping.extend(result["mapping_spec"])
            if doc_type == "general" and not result.get("data_dict") and not result.get("rules"):
                general_docs.append(fname)

        except Exception as exc:
            log.append({"file": fname, "type": "error", "detail": str(exc)})

    return {
        "data_dict":       merged_dict,
        "rules":           merged_rules,
        "mapping_spec":    merged_mapping,
        "general":         general_docs,
        "classifications": log,
    }


def _parse_data_dictionary(df: pd.DataFrame) -> dict[str, dict]:
    # Expects columns (case-insensitive):
    # column_name | description | owner | sensitivity | data_type |
    # nullable | is_pk | format_pattern | allowed_values | min_value | max_value | business_term
    #
    # Returns dict keyed by column_name.
    if df.empty:
        return {}

    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    col_field = next((c for c in df.columns if "column" in c or "field" in c or "name" in c), df.columns[0])  # OCR-UNCERTAIN: condition text ran off right edge of photo; fallback default reconstructed from context

    result = {}
    for _, row in df.iterrows():
        key = str(row.get(col_field, "")).strip()
        if not key:
            continue
        result[key] = {
            "description": str(row.get("description", "") or ""),
            "owner":       str(row.get("owner", "") or ""),
            "sensitivity": str(row.get("sensitivity", "") or ""),
            "data_type":   str(row.get("data_type", "") or ""),
            "nullable":    str(row.get("nullable", "true")).lower() not in ("false", "no", "0"),
            "is_pk":          str(row.get("is_pk", "false")).lower() in ("true", "yes", "1"),
            "format_pattern": str(row.get("format_pattern", "") or ""),
            "allowed_values": [v.strip() for v in str(row.get("allowed_values", "") or "").split("|") if v.strip()],  # OCR-UNCERTAIN: line ran off right edge of photo; split/filter pattern reconstructed from matching code seen on page 0134
            "min_value":      row.get("min_value"),
            "max_value":      row.get("max_value"),
            "business_term":  str(row.get("business_term", "") or ""),
        }
    return result


def _parse_business_rules(df: pd.DataFrame) -> list[dict]:
    # Expects columns: column_name | rule_name | rule_type | value | description
    # rule_types: not_null | unique | min | max | min_length | max_length |
    #   pattern | allowed_values | freshness_days | row_count_min | row_count_max  # OCR-UNCERTAIN: "row_count_max" inferred, text ran off right edge showing only "row_count_m"
    if df.empty:
        return []

    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    rules = []
    for _, row in df.iterrows():
        col = str(row.get("column_name", "") or "").strip()
        rules.append({
            "column_name": col,
            "rule_name":   str(row.get("rule_name", "") or col),
            "rule_type":   str(row.get("rule_type", "") or "").strip().lower(),
            "value":       row.get("value"),
            "description": str(row.get("description", "") or ""),
        })
    return [r for r in rules if r["rule_type"]]


def _parse_mapping_spec(df: pd.DataFrame) -> list[dict]:
    # Reads extended mapping spec with full validation rule support.
    #
    # Required columns: source_column, target_column
    # Optional: data_type, mandatory, not_null, unique, value_in_list,
    # min_value, max_value, regex_pattern, condition,
    # transformation, business_rule, severity, description
    #
    # value_in_list is split on '|' into a Python list.
    # mandatory / not_null / unique are coerced to bool.
    # min_value / max_value are coerced to float when present.
    if df.empty:
        return []

    # Normalise column headers
    df = df.copy()
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    # Locate required columns (flexible name matching)
    src = next((c for c in df.columns if "source" in c), df.columns[0] if len(df.columns) > 0 else None)
    tgt = next((c for c in df.columns if "target" in c), df.columns[1] if len(df.columns) > 1 else None)


    if not src or not tgt:
        return []

    def _bool(val, default: bool = True) -> bool:
        # Coerce spreadsheet boolean-like text to Python bool.
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return default
        return str(val).strip().lower() not in ("false", "no", "0", "")

    def _float_or_none(val):
        # Parse numeric threshold; return None when absent or non-numeric.
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    def _str(val) -> str:
        # Return stripped string; treat NaN / None as empty string.
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return ""
        return str(val).strip()

    specs = []
    for _, row in df.iterrows():
        source_col = _str(row.get(src, ""))
        target_col = _str(row.get(tgt, ""))
        if not source_col and not target_col:
            continue

        # Parse value_in_list: "JEG|JIL|JLL" -> ["JEG", "JIL", "JLL"]
        vil_raw = _str(row.get("value_in_list", ""))
        value_in_list = [v.strip() for v in vil_raw.split("|") if v.strip()] if vil_raw else []

        specs.append({
            "source_column": source_col,
            "target_column": target_col,
            "data_type":      _str(row.get("data_type", "string")) or "string",
            "mandatory":      _bool(row.get("mandatory", True)),
            "not_null":       _bool(row.get("not_null", False)),
            "unique":         _bool(row.get("unique", False), default=False),
            "value_in_list":  value_in_list,
            "min_value":      _float_or_none(row.get("min_value")),
            "max_value":      _float_or_none(row.get("max_value")),
            "regex_pattern":  _str(row.get("regex_pattern", "")),
            "condition":      _str(row.get("condition", "")),
            "transformation": _str(row.get("transformation", "")),
            "business_rule":  _str(row.get("business_rule", "")),
            "severity":       _str(row.get("severity", "ERROR")).upper() or "ERROR",
            "description":    _str(row.get("description", "")),
        })
    return specs


# ------------------------------------------------------------------
# Smart key inference -- multi-phase, domain-aware
# ------------------------------------------------------------------
# Columns that look like system-generated surrogate keys or audit timestamps.
# These are COMPLETELY excluded from key inference because:
# - Auto-increment IDs (id, pk, row_id) have no business meaning and differ
#   between files even when the underlying data is the same record.
# - Audit timestamps (created_at, updated_at) are metadata, not identifiers.
# A column named "trade_id" is NOT excluded -- only bare single-token names a  # OCR-UNCERTAIN: comment line ran off right edge of photo
_SYSTEM_GEN_EXACT = frozenset({
    # bare auto-increment / surrogate key names
    "id", "pk", "row_id", "rowid", "rownum", "row_num", "row_number",
    "record_id", "recordid", "auto_id", "autoid", "sys_id", "sysid",
    "seq_id", "seqid", "_id", "oid", "surrogate_key",
    # parser-generated sequence/index columns
    "message_index", "msg_index", "msg_idx", "message_idx",
    "record_index", "record_idx", "line_index", "line_number", "line_num",
    "row_index", "index",
    # audit / ETL timestamps
    "created_at", "updated_at", "created_date", "modified_date",
    "insert_time", "update_time", "load_date", "etl_date",
    "created_ts", "updated_ts", "modified_ts", "last_updated",
    "inserted_at", "deleted_at",
})

# ------------------------------------------------------------------
# Metadata column auto-detection
# ------------------------------------------------------------------
# Primary detection: _metadata_score() -- purely data-driven, no column names
# needed.
# Fallback (all-null / single-file): _is_metadata_col() -- universal name patterns only.
# No domain-specific column names appear anywhere in this section.


def _normalize_colname(col: str) -> str:
    # Lower-case + camelCase-to-snake, compress all non-alphanumeric runs to '_'.
    # e.g. 'UpdatedBy' -> 'updated_by', 'last-update' -> 'last_update'.
    s = col.strip()

    # camelCase / PascalCase -> snake
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)
    s = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1_\2", s)
    s = s.lower()
    # replace any non-alphanumeric char with _
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s


_METADATA_PATTERNS: list = [
    # --- Audit timestamps ---
    re.compile(r"^(created|modified|updated|deleted|inserted|load|etl|changed)"
               r"(_(at|on|date|time|ts|timestamp|dt|dtime))?$"),
    re.compile(r"^(create|modify|update|delete|insert)_(date|time|ts|timestamp|dt)$"),
    re.compile(r"^(date|time|ts|timestamp)_(created|modified|updated|deleted|inserted)$"),
    # --- "last ..." audit patterns ---
    re.compile(r"^last_(updated?|modified?|changed?|refresh)(_(at|on|date|time|ts|by))?$"),
    re.compile(r"^(last|prev|previous)_(update|modify|change|refresh)_?(date|time|ts)?$"),
    # --- "modified by / created by / updated by" ---
    re.compile(r"^(created|modified|updated|deleted|inserted|changed)_by$"),
    re.compile(r"^(create|modify|update|insert|delete)_(by|user|usr|userid|username)$"),
    re.compile(r"^(modified|updated|created)_(user|username|userid|usr)$"),
    # --- ETL / load plumbing ---
    re.compile(r"^etl_(id|date|time|ts|batch|run|run_id|job|jobid|source|flag)$"),
    re.compile(r"^load_(date|time|ts|by|id|batch|run|flag|indicator)$"),
    re.compile(r"^(batch|job)_(id|run|date|ts|time)$"),
    re.compile(r"^(dw|dwh|ods|stg|staging)_(insert|update|load|etl)_(date|ts|time|flag)$"),
    # --- Row versioning / hash / soft-delete ---
    re.compile(r"^(row|record|rec)_(version|ver|hash|checksum|crc|fingerprint)$"),
    re.compile(r"^(is|flag)_(deleted|active|current|valid|latest|live)$"),
    re.compile(r"^(deleted|active|current|valid|latest|is_latest)$"),
    re.compile(r"^(effective|valid)_(from|to|start|end)$"),
    # --- Bare system names already in _SYSTEM_GEN_EXACT that also appear in comparisons ---
    re.compile(r"^(row_?id|record_?id|auto_?id|sys_?id|seq_?id)$"),
]


def _is_metadata_col(col: str, dtype=None) -> bool:

    # Return True if *col* looks like a metadata / audit column that should be
    # excluded from comparison.

    # Checks (in order):
    # 1. Normalised name is in _SYSTEM_GEN_EXACT.
    # 2. Normalised name matches any _METADATA_PATTERNS regex.
    # 3. Column has a datetime dtype AND its name contains an audit keyword.

    norm = _normalize_colname(col)


    # Check 1 -- already a known system-gen name
    if norm in _SYSTEM_GEN_EXACT:
        return True

    # Check 2 -- pattern match
    for pat in _METADATA_PATTERNS:
        if pat.match(norm):
            return True

    # Check 3 -- datetime dtype + unambiguous audit keyword in name.
    # Deliberately excludes "date", "time", "expiry", "valid" -- these appear in
    # business columns (TradeDate, ExpiryDate, MaturityDate) and must not be
    # classified as metadata solely because they have a datetime dtype.
    if dtype is not None and pd.api.types.is_datetime64_any_dtype(dtype):
        audit_words = {
            "created", "modified", "updated", "deleted", "inserted",
            "load", "etl", "changed", "last",
        }


        name_tokens = set(norm.split("_"))
        if name_tokens & audit_words:
            return True

    return False


def _metadata_score(col: str, df1: "pd.DataFrame", df2: "pd.DataFrame") -> tuple[bool, str]:
    # Decide whether *col* is a metadata column using data-behaviour signals first,
    # then name patterns as a secondary signal. No predefined column list required.
    #
    # Returns (is_metadata: bool, reason: str).
    #
    # Data signals (checked regardless of column name):
    # S1  Disjoint integers, 100% unique in each file → surrogate/auto-increment key
    # S2  Datetime dtype with >80% divergence between files → audit timestamp
    # S3  String column parseable as datetime with >80% divergence → audit timestamp
    # String column with <15% value overlap and low cardinality → audit user/process
    # S4  Nearly identical value sets (>=90% overlap) → business dimension, NOT metadata
    #     (hard override -- protects any column whose values are stable across files)
    #
    # Name signal (weak fallback, only when data is inconclusive):
    # N1  Name matches universal audit/ETL patterns, confirmed only when S4 did not fire
    _SAMPLE = 500
    if col not in df1.columns or col not in df2.columns:
        return _is_metadata_col(col), "name-only (column missing from one file)"

    s1 = df1[col].dropna()
    s2 = df2[col].dropna()
    if len(s1) == 0 or len(s2) == 0:
        return _is_metadata_col(col), "name-only (column all-null)"

    s1s = s1.head(_SAMPLE)
    s2s = s2.head(_SAMPLE)

    set1 = set(s1s.astype(str))
    set2 = set(s2s.astype(str))
    union = set1 | set2
    overlap_ratio = len(set1 & set2) / len(union) if union else 1.0

    # S4 -- nearly identical value sets → business dimension (fires first, hard override)
    if overlap_ratio >= 0.90:
        return False, f"business (value overlap {overlap_ratio*100:.0f}%)"

    # S1 -- disjoint unique integers → surrogate / auto-increment key
    if pd.api.types.is_integer_dtype(s1) and pd.api.types.is_integer_dtype(s2):
        u1 = s1s.nunique() / len(s1s)
        u2 = s2s.nunique() / len(s2s)
        if overlap_ratio == 0.0 and u1 > 0.80 and u2 > 0.80:
            return True, "surrogate key (disjoint high-unique integers)"

    # S2 -- datetime dtype + high divergence + audit name → audit timestamp.
    # Name guard is required: business date columns (TradeDate, ExpiryDate,
    # MaturityDate) from different report periods will have high divergence but
    # are not metadata. Only flag when the name independently looks like an
    # audit/ETL timestamp (created_at, load_date, updated_on etc.).
    if pd.api.types.is_datetime64_any_dtype(s1):
        divergence = 1.0 - overlap_ratio
        if divergence > 0.80 and _is_metadata_col(col):
            return True, f"audit timestamp (datetime + name, {divergence*100:.0f}% values differ)"

    # S3 -- string column: looks like a timestamp string or high-unique with low overlap
    if s1.dtype == object or pd.api.types.is_string_dtype(s1):
        # Quick sniff: only attempt datetime parse if first non-null value looks date-like
        _DATE_SNIFF = re.compile(r"^\d{4}[-/]\d{2}[-/]\d{2}")
        first_val = str(s1s.iloc[0]) if len(s1s) > 0 else ""
        if _DATE_SNIFF.match(first_val):
            try:
                pd.to_datetime(s1s.head(20), errors="raise")
                divergence = 1.0 - overlap_ratio
                # Same name guard as S2 -- string date columns like TradeDate / ExpiryDate
                # stored as strings will also have high divergence across report periods.
                if divergence > 0.80 and _is_metadata_col(col):
                    return True, f"audit timestamp (string datetime + name, {divergence*100:.0f}% values differ)"
            except Exception:
                pass

    # High-unique string with very low cross-file overlap and low total cardinality
    # → likely an audit-user or process column, BUT only when the column name also
    # matches a known audit/ETL pattern. Without the name guard, business dimension
    # columns (Market, Exchange, Status, Side) are wrongly classified as metadata
    # simply because they carry different valid values across two different systems.
    u1 = s1s.nunique() / len(s1s) if len(s1s) > 0 else 0
    u2 = s2s.nunique() / len(s2s) if len(s2s) > 0 else 0
    total_distinct = len(union)
    if (overlap_ratio < 0.15 and total_distinct <= 50 and u1 < 0.5 and u2 < 0.5
            and _is_metadata_col(col)):
        return True, f"audit user/process ({total_distinct} distinct values, {overlap_ratio*100:.0f}% overlap)"

    # N1 -- name pattern only fires when values are highly divergent AND the column
    # passes the name check without relying on dtype (avoids false positives on
    # business date columns like TradeDate, ExpiryDate whose values naturally differ
    # across two different trade populations but are not audit/ETL metadata).
    if _is_metadata_col(col):
        divergence = 1.0 - overlap_ratio
        if divergence > 0.60:
            return True, f"name pattern + values highly divergent ({divergence*100:.0f}%)"

    return False, f"business (no metadata signal, {overlap_ratio*100:.0f}% value overlap)"


def _split_meta_cols(
    cols: list[str],
    df1: "pd.DataFrame | None" = None,
    df2: "pd.DataFrame | None" = None,
) -> tuple[list[str], list[str]]:

    # Partition *cols* into (business_cols, meta_cols).


    # When both DataFrames are available, uses _metadata_score which detects
    # metadata purely from data behaviour -- works on any column name.
    # Falls back to universal name-pattern detection when only one file is available.

    business: list[str] = []
    meta: list[str] = []
    for c in cols:
        if df1 is not None and df2 is not None:
            is_meta, _ = _metadata_score(c, df1, df2)
        else:
            dtype = None
            if df1 is not None and c in df1.columns:
                dtype = df1[c].dtype
            elif df2 is not None and c in df2.columns:
                dtype = df2[c].dtype
            is_meta = _is_metadata_col(c, dtype)

        if is_meta:


            meta.append(c)
        else:
            business.append(c)
    return business, meta


# Extended keyword sets for key-like column names.
# NOTE: bare "id" and "pk" removed -- they are in _SYSTEM_GEN_EXACT instead.
# Compound names like "trade_id" still match via _KEY_SUFFIX_RE suffix rules.
_KEY_HINTS_EXACT = {
    # Generic meaningful identifiers
    "code", "uuid", "uid", "num", "number",
    "ref", "reference", "identifier", "sequence", "seq",
    # Financial / securities
    "isin", "cusip", "ticker", "symbol", "ric", "sedol", "figi", "lei",
    "account", "acct", "trade", "tradeid", "order", "orderid", "deal",
    "dealid", "position", "portfolio", "client", "clientid", "counterparty",
    "cpty", "entity", "entityid", "instrument", "security", "contract",
    # SWIFT


    "txnref", "msgref", "swiftref", "transactionreference",
    # FIX
    "clordid", "execid",
    # General business
    "invoice", "invoiceno", "lineitem", "transid",
    "transaction", "batch", "batchid", "messageid", "correlationid",
}

_KEY_SUFFIX_RE = re.compile(
    r"(?:_id|_key|_no|_num|_ref|_code|_cd|_uuid|_uid|_seq|_identifier)"
    r"|(?:id_|ref_|key_|no_)|(?:_id\d+)$",
    re.IGNORECASE,
)

# Transaction/event IDs that are unique within one system but carry no cross-file
# join meaning. Excluded from key inference only -- they remain in the column map
# and data comparison as normal business columns.
# Rule: these are IDs that identify a system-internal event/transaction, not a


# real-world business entity (instrument, account, client) that exists in multiple
# systems.


def _col_uniqueness(series: pd.Series) -> float:
    # Uniqueness ratio over non-null values.
    #
    # Returns 0.0 if >40% of values are null (column is too sparse to be a key).
    n = len(series)
    if n == 0:
        return 0.0

    null_ratio = series.isna().sum() / n
    if null_ratio > 0.40:
        return 0.0

    non_null = series.dropna()
    if len(non_null) == 0:
        return 0.0

    return non_null.nunique() / len(non_null)


_UNIQUENESS_SAMPLE = 5_000  # sample size for composite-key uniqueness checks


def _combined_uniqueness(df: pd.DataFrame, cols: list[str]) -> float:
    # Uniqueness ratio for a combination of columns (nulls treated as a distinct value).

    # Uses a random sample on large frames -- accurate enough for key candidacy
    # scoring.
    n = len(df)
    if n == 0:
        return 0.0

    sample = df if n <= _UNIQUENESS_SAMPLE else df.sample(_UNIQUENESS_SAMPLE, random_state=42)

    ns = len(sample)

    return sample[cols].fillna("__NULL__").astype(str).drop_duplicates().shape[0] / ns


def _name_score(col: str) -> float:
    # Score a column name on how key-like it looks.
    # Returns a value in [0.0, 0.5] -- combined with uniqueness for final ranking.
    #
    # Uses _normalize_colname for PascalCase/camelCase splitting so compound names
    # like 'AccountIdRef' or 'PrincipalLeId' correctly match the _id suffix pattern.
    norm = _normalize_colname(col)   # handles PascalCase / camelCase
    tokens = set(norm.split("_")) | {norm}

    # Exact token match with known key words -- strongest signal
    if tokens & _KEY_HINTS_EXACT:
        return 0.50

    # Suffix / prefix pattern match on normalised form
    if _KEY_SUFFIX_RE.search(norm):
        return 0.45

    # Underscore-aware substring hints (only meaningful after normalisation)
    if any(hint in norm for hint in ("_id", "id_", "uuid", "code", "ref", "key", "no_", "_no")):
        return 0.30

    return 0.0


def _dtype_bonus(series: pd.Series) -> float:
    # Small bonus for data types commonly used as keys.
    # Integers > strings > float (unless integer-valued) > other.
    dtype = series.dtype
    if pd.api.types.is_integer_dtype(dtype):
        return 0.05

    if dtype == object or pd.api.types.is_string_dtype(dtype):
        return 0.03

    if pd.api.types.is_float_dtype(dtype):
        non_null = series.dropna()
        # Integer-valued floats (1.0, 2.0 ...) are acceptable keys
        if len(non_null) > 0:
            try:
                if (non_null == non_null.astype("int64").astype("float64")).all():
                    return 0.02
            except Exception:
                pass
        return 0.0

    return 0.0


def _is_constant(series: pd.Series) -> bool:
    # True if the column has only one distinct value -- useless as a key.
    # Single-row series are never considered constant: any column in a 1-row
    # file is 100% unique and can serve as a meaningful key.
    if len(series) <= 1:
        return False

    return series.nunique(dropna=True) <= 1


def _col_score(col: str, s1: pd.Series, s2: pd.Series) -> float:
    # Composite score for a single candidate key column.
    # Components (all in [0,1] range):
    # - uniqueness:  min(uniqueness_f1, uniqueness_f2)  -- most important
    # - name_score:  keyword / pattern match            -- secondary signal
    # - dtype_bonus: integer > string > float            -- tie-breaker
    u = min(_col_uniqueness(s1), _col_uniqueness(s2))
    return u + _name_score(col) + _dtype_bonus(s1)


_OVERLAP_SAMPLE = 5_000  # max rows to sample when computing cross-file value overlap


def _cross_overlap(s1: pd.Series, s2: pd.Series) -> float:
    # Fraction of values from s1 that appear anywhere in s2 (Jaccard-style lookup).
    # Uses a sampled set intersection so it runs in O(sample) not O(n*m).
    #
    # Returns 0.0 when the two series share no values at all -- a strong signal that
    # the column is a system-generated surrogate key (e.g. TradeID, AlertID) whose
    # values have no cross-file join relationship even if the column name matches.
    # Returns 1.0 if every value in s1 appears in s2.
    v1 = s1.dropna()
    v2 = s2.dropna()
    if len(v1) == 0 or len(v2) == 0:
        return 0.0

    # Sample both sides to keep this O(sample) on large frames
    if len(v1) > _OVERLAP_SAMPLE:
        v1 = v1.iloc[:_OVERLAP_SAMPLE]
    if len(v2) > _OVERLAP_SAMPLE:
        v2 = v2.iloc[:_OVERLAP_SAMPLE]

    set1 = set(v1.astype(str))
    set2 = set(v2.astype(str))
    union = set1 | set2
    if not union:
        return 0.0

    return len(set1 & set2) / len(union)


def infer_keys(
    df1: pd.DataFrame,
    df2: pd.DataFrame,
    common_cols: list[str],
    matched_cols: list[str] | None = None,
) -> tuple[list[str], str]:

    # Multi-phase unique key inference.

    # Phase 1 -- Perfect single-column key (100% unique in BOTH files).
    # Among all perfect candidates, prefer the one with the highest
    # name score, then dtype bonus.

    # Phase 2 -- 2-column composite: try all pairs from top-ranked candidates
    # until a combination achieves >= 99.5% combined uniqueness.
    # Near-perfect single columns (< 100%) are skipped in favour of
    # a composite that is fully unique, avoiding duplicate-index issues.


    # Phase 3 -- 3-column composite: same logic up to triples.

    # Phase 4 -- Best-effort single column (most unique available, even if
    # imperfect). Returned with a "may have duplicates" label so
    # the caller knows to handle duplicate index entries.

    # Phase 5 -- Content-based fallback (no usable key found at all).

    # Columns are disqualified if:
    # • Name (normalised, lower-case) is in _SYSTEM_GEN_EXACT (auto-increment IDs,
    # audit timestamps -- these are system-generated and have no cross-file meaning)
    # • > 40% null values
    # • Only 1 distinct value (constant column)

    # matched_cols: when provided (from analyze_mapping exact+fuzzy results), key
    # candidates are restricted to this set -- columns confirmed semantically
    # equivalent across both files. System-specific IDs that happen to share a
    # name but carry no cross-file join meaning are excluded as a result.


    if not common_cols or len(df1) == 0 or len(df2) == 0:
        return [], "content-based (empty dataframe)"

    # Restrict candidate pool to semantically matched columns when available.
    # This prevents system-internal IDs (e.g. Trade ID, Global Alert ID) that
    # are unique per file but have no cross-file join relationship from being
    # selected as keys.
    if matched_cols:
        common_cols = [c for c in common_cols if c in matched_cols]
        if not common_cols:
            return [], "content-based (no matched columns available for key inference)"

    # ---------- pre-filter: drop columns that cannot be meaningful keys ----------
    # Two candidate pools:
    #  solo_candidates -- used for Phase 1 (perfect single key); must pass null filter
    #  all_candidates  -- used for composite search; high-null dimension columns
    #                     included
    #                     because _combined_uniqueness treats NULLs as a distinct value


    #
    # Zero-overlap exclusion: a column whose values share nothing between the two
    # files
    # cannot produce any matches when used as a join key -- it is a system-generated
    # surrogate (e.g. TradeID, AlertID) even if the name looks like a business key.
    # Pre-compute overlap once per column so the same result is reused across pools.
    _overlap_cache: dict[str, float] = {
        c: _cross_overlap(df1[c], df2[c]) for c in common_cols
        if c in df1.columns and c in df2.columns
    }


    def _not_meta_not_const(c):
        return (not _is_metadata_col(c, df1[c].dtype if c in df1.columns else None)
                and not _is_constant(df1[c]) and not _is_constant(df2[c]))


    def _has_cross_overlap(c):
        # Exclude columns with zero value overlap across the two files -- they are
        # system-generated surrogates and would produce zero join matches.
        return _overlap_cache.get(c, 0.0) > 0.0


    solo_candidates = [
        c for c in common_cols
        if _not_meta_not_const(c)
        and _has_cross_overlap(c)
        and _col_uniqueness(df1[c]) > 0
        and _col_uniqueness(df2[c]) > 0
    ]

    all_candidates = [c for c in common_cols if _not_meta_not_const(c) and _has_cross_overlap(c)]

    if not all_candidates:
        # Relax metadata filter but keep the overlap guard -- no point joining on
        # a column whose values never appear on both sides.
        all_candidates = [
            c for c in common_cols
            if not _is_constant(df1[c]) and not _is_constant(df2[c])
            and _has_cross_overlap(c)
        ]


    if not all_candidates:
        return [], "content-based (no candidate columns -- all share zero cross-file value overlap)"

    if not solo_candidates:
        solo_candidates = all_candidates

    # ---------- rank candidates by composite score ----------
    # Solo candidates ranked by standard score (respects null penalty).
    # All candidates also ranked -- composite search uses this wider pool.
    ranked = sorted(
        [(c, _col_score(c, df1[c], df2[c])) for c in solo_candidates],
        key=lambda x: x[1],
        reverse=True,
    )

    # Wider ranked list for composite search -- includes high-null cols, ranked by name
    # score + dtype
    ranked_all = sorted(
        [(c, _name_score(c) + _dtype_bonus(df1[c])) for c in all_candidates],
        key=lambda x: x[1],
        reverse=True,
    )

    # -- Phase 1: perfect single-column key --
    # Reuse uniqueness already computed during _col_score to avoid redundant scans
    u_cache = {c: (_col_uniqueness(df1[c]), _col_uniqueness(df2[c])) for c in solo_candidates}

    perfect = [
        (c, sc) for c, sc in ranked
        if u_cache[c][0] == 1.0 and u_cache[c][1] == 1.0
    ]

    if perfect:
        best = perfect[0][0]
        return [best], f"single-column key '{best}' (100% unique, auto)"

    # -- Phases 2-4: composite key search (2- through 4-column combos) --
    # Pre-compute a per-column string array (sampled) once, then combine columns
    # with numpy string concatenation -- avoids repeated fillna+astype+drop_duplicates
    # per combo (16ms → <1ms per combo).
    top_cols = [c for c, _ in ranked_all[:12]]


    _PERFECT = 0.995
    _GOOD    = 0.95

    # Pre-materialise sampled string arrays for each candidate column
    def _col_str_arr(df: pd.DataFrame, col: str) -> np.ndarray:
        s = df[col] if len(df) <= _UNIQUENESS_SAMPLE else df[col].iloc[:_UNIQUENESS_SAMPLE]
        return s.fillna("__NULL__").astype(str).to_numpy()


    col_arrs1 = {c: _col_str_arr(df1, c) for c in top_cols}
    col_arrs2 = {c: _col_str_arr(df2, c) for c in top_cols}
    n1 = len(next(iter(col_arrs1.values())))
    n2 = len(next(iter(col_arrs2.values())))


    def _combo_uniqueness_fast(arrs: dict, cols: list, n: int) -> float:
        # Uniqueness of a column combo using pre-built string arrays + numpy.
        if len(cols) == 1:
            return len(np.unique(arrs[cols[0]])) / n
        # Join col values with a separator char -- unique joined strings = unique combos


        sep = np.full(n, "\x00", dtype=object)
        combined = arrs[cols[0]].astype(object)
        for c in cols[1:]:
            combined = combined + sep + arrs[c].astype(object)
        return len(np.unique(combined)) / n


    best_combo: list[str] = []
    best_combo_u: float = 0.0

    for width, combos in [
        (2, combinations(top_cols, 2)),
        (3, combinations(top_cols, 3)),
        (4, combinations(top_cols, 4)),
    ]:
        if best_combo_u >= _GOOD:
            break
        for combo in combos:
            combo = list(combo)
            u = min(_combo_uniqueness_fast(col_arrs1, combo, n1),


                    _combo_uniqueness_fast(col_arrs2, combo, n2))
            if u > best_combo_u:
                best_combo_u, best_combo = u, combo
            if best_combo_u >= _PERFECT:
                cols_str = ", ".join(best_combo)
                return best_combo, f"composite key [{cols_str}] ({width}-col, auto)"
            if best_combo_u >= _GOOD:
                break

        if best_combo_u >= 0.70:
            pct = round(best_combo_u * 100, 1)
            cols_str = ", ".join(best_combo)
            return best_combo, (
                f"best-effort composite [{cols_str}] ({pct}% unique -- may have duplicates)"
            )

    # -- Phase 5: best-effort single column --
    best_col, best_score = ranked[0]
    u1 = round(_col_uniqueness(df1[best_col]) * 100, 1)


    u2 = round(_col_uniqueness(df2[best_col]) * 100, 1)
    if min(u1, u2) >= 50:      # at least half-unique -- still useful
        return [best_col], (
            f"best-effort key '{best_col}' ({u1}%/{u2}% unique -- may have duplicates)"  # OCR-UNCERTAIN: closing word "duplicates)" ran off right edge of photo, reconstructed from matching phrase used elsewhere in this function
        )

    # -- Phase 6: content-based fallback --
    return [], "content-based (no sufficiently unique key found)"


# ------------------------------------------------------------------
# Comparison engine
# ------------------------------------------------------------------


def _normalise(df: pd.DataFrame) -> pd.DataFrame:
    clean_cols = [str(c).strip() for c in df.columns]
    needs_copy = (list(df.columns) != clean_cols) or (df.index != pd.RangeIndex(len(df))).any()
    if not needs_copy:


        return df

    out = df.copy()
    out.columns = clean_cols
    return out.reset_index(drop=True)


_MAX_DIFF_ROWS = 500  # UI row cap -- keeps browser responsive on large datasets
_MAX_DATA_COLS = 25   # max data columns to carry per row (keeps payload manageable)


_NULL_SENTINELS = {"null", "none", "nan", "n/a", "na", "#n/a", ""}


def _safe_str(v) -> str:
    # Convert a cell value to a canonical string, treating NaN/None/NULL sentinels as
    # empty.
    # Normalising sentinel values prevents false exceptions when the same logical
    # null is represented differently across file formats (e.g. 'NULL' in TXT vs
    # NaN in JSON/CSV).

    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    s = str(v).strip()
    if s.lower() in _NULL_SENTINELS:
        return ""
    return s


def _detect_null_column_exceptions(df1, df2, data_cols: list) -> list[dict]:

    # Detect columns that have data on one side but are entirely null on the other.
    # These are structural/population exceptions -- separate from row-level value breaks.
    # Returns a list of dicts: {column, has_data_in, all_null_in, non_null_count,
    # sample_values}

    exceptions = []


    for col in data_cols:
        in1 = col in df1.columns
        in2 = col in df2.columns
        if not in1 or not in2:
            continue
        null1 = df1[col].isna().all() or (df1[col].astype(str).str.strip().eq("").all())  # OCR-UNCERTAIN: character inside eq(...) unclear in photo, reconstructed as empty-string check to match null2 line below
        null2 = df2[col].isna().all() or (df2[col].astype(str).str.strip().eq("").all())
        if null1 == null2:
            continue  # both have data or both are null -- not an exception
        if null1:
            # File 1 is all null, File 2 has data
            non_null = df2[col].dropna()
            non_null = non_null[non_null.astype(str).str.strip() != ""]
            exceptions.append({
                "column":     col,
                "has_data_in": "File 2",
                "all_null_in": "File 1",
                "non_null_count": int(len(non_null)),
                "sample_values": non_null.astype(str).head(5).tolist(),


            })
        else:
            # File 2 is all null, File 1 has data
            non_null = df1[col].dropna()
            non_null = non_null[non_null.astype(str).str.strip() != ""]
            exceptions.append({
                "column":     col,
                "has_data_in": "File 1",
                "all_null_in": "File 2",
                "non_null_count": int(len(non_null)),
                "sample_values": non_null.astype(str).head(5).tolist(),
            })
    return exceptions


def _build_modified_rows(changed_df, keys, use_cols, make_kv, out_list):
    # Vectorised extraction of changed cell pairs -- avoids iterrows on the diff result.
    if changed_df.empty:
        return

    def _norm_col(series: pd.Series) -> pd.Series:
        # Normalise a data column for comparison: NaN + all null sentinels -> "".
        return (
            series.fillna("")
            .astype(str)
            .str.strip()
            .apply(lambda v: "" if v.lower() in _NULL_SENTINELS else v)
        )

    for c in use_cols:
        c1, c2 = f"{c}__f1", f"{c}__f2"
        if c1 not in changed_df.columns or c2 not in changed_df.columns:
            continue
        # Compare normalised values -- treats NaN, "", "NULL", "null", "N/A", "nan"
        # etc. as identical so cross-format nulls don't produce false exceptions.
        v1 = _norm_col(changed_df[c1])
        v2 = _norm_col(changed_df[c2])
        mask = v1 != v2


        changed_df[f"_diff_{c}"] = mask
        # Store normalised values so _safe_str display is already clean
        changed_df[f"_n1_{c}"] = v1
        changed_df[f"_n2_{c}"] = v2


    records = changed_df.to_dict("records")
    for row in records:
        key_val = (tuple(row[k] for k in keys) if len(keys) > 1 else row[keys[0]])
        changes = {}
        for c in use_cols:
            if row.get(f"_diff_{c}"):
                v1 = row.get(f"_n1_{c}", "")
                v2 = row.get(f"_n2_{c}", "")
                if v1 == "" and v2 == "":
                    continue
                changes[c] = {"file1": v1, "file2": v2}
        if changes:
            out_list.append({
                "key_values":    make_kv(key_val),


                "changes":       changes,
                "changed_col_count": len(changes),
                "break_type":    _classify_break(changes),
            })


def _series_differs(v1: pd.Series, v2: pd.Series) -> pd.Series:
    # String-level inequality first -- cheap, and already correct for the
    # (common) case where both sides match exactly.
    raw_diff = v1 != v2
    if not raw_diff.any():
        return raw_diff
    # For rows that differ as strings, check whether both sides are actually
    # the same number just formatted differently (e.g. "100.50" vs "100.5",
    # "300" vs "300.00", "200.25" vs "200.250") -- extremely common across
    # BFSI source systems and not a real reconciliation break.
    n1 = pd.to_numeric(v1, errors="coerce")
    n2 = pd.to_numeric(v2, errors="coerce")
    both_numeric = n1.notna() & n2.notna()
    numerically_equal = both_numeric & (n1.sub(n2).abs() < 1e-9)
    still_diff = raw_diff & ~numerically_equal

    # For remaining differences where at least one side isn't a plain number,
    # check whether both sides parse to the same calendar date -- e.g.
    # "2026-01-15" vs "01/15/2026" vs "15-Jan-2026" are the same date just
    # formatted differently, also a very common cross-system artefact.
    candidates = still_diff & ~both_numeric
    if candidates.any():
        try:
            d1 = pd.to_datetime(v1[candidates], errors="coerce", format="mixed")
            d2 = pd.to_datetime(v2[candidates], errors="coerce", format="mixed")
            dates_equal = d1.notna() & d2.notna() & (d1 == d2)
            still_diff.loc[dates_equal.index[dates_equal]] = False
        except Exception:
            pass
    return still_diff


def _key_based_diff(df1, df2, keys, common_cols, force_data_cols: list[str] | None = None):
    # Key-based comparison.
    #
    # Returns three sections -- no ADDED/REMOVED labels, just:
    #   file1_only – rows present in File 1 but not File 2 (actual column data)
    #   file2_only – rows present in File 2 but not File 1 (actual column data)
    #   modified   – rows in both files whose non-key values differ
    #
    # Each row carries:
    #   key_values  {key_col: value, ...}
    #   row_data    {col: value, ...} (file1_only / file2_only)
    #   changes     {col: {file1, file2}, ...} (modified only)
    #
    # force_data_cols: when provided, skip _split_meta_cols entirely and use
    # these columns as the data columns to compare. Used when the caller knows
    # the exact value column(s) and they would otherwise be misclassified as
    # metadata (e.g. aggregated numeric columns that look like surrogate keys).

    # Normalise key columns to str so cross-format type differences (e.g. int in
    # JSON vs str in XML/TXT for the same numeric ID) never cause false mismatches.
    # NaN and null-sentinel strings ("NULL", "null", "nan", "none", "n/a") are all
    # collapsed to "" so that the same logical null matches across formats -- e.g.
    # ProductRegion=NULL in a TXT file (parsed as NaN) matches ProductRegion="NULL"
    # in a JSON file (stored as a literal string).
    def _norm_key_col(s: pd.Series) -> pd.Series:
        s = s.fillna("").astype(str).str.strip()
        return s.where(~s.str.lower().isin(_NULL_SENTINELS), "")

    for k in keys:
        df1[k] = _norm_key_col(df1[k])
        df2[k] = _norm_key_col(df2[k])

    d1 = df1.set_index(keys).sort_index()
    d2 = df2.set_index(keys).sort_index()
    idx1 = set(d1.index.tolist())
    idx2 = set(d2.index.tolist())
    only1_keys  = idx1 - idx2     # in File 1, not in File 2
    only2_keys  = idx2 - idx1     # in File 2, not in File 1
    common_keys = idx1 & idx2

    def _safe_sort_key(k):
        # Convert any key (including tuples with NaN) to a sortable string.
        if isinstance(k, tuple):
            return tuple("" if (isinstance(v, float) and pd.isna(v)) else str(v) for v in k)
        return "" if (isinstance(k, float) and pd.isna(k)) else str(k)

    all_data_cols = [c for c in common_cols if c not in keys]
    if force_data_cols is not None:
        # Caller provided explicit value columns -- skip metadata classification entirely.
        # This is used when aggregated numeric columns would be wrongly classified as
        # surrogate keys by _metadata_score.
        data_cols = [c for c in force_data_cols if c in all_data_cols]
        excluded_meta_cols = [c for c in all_data_cols if c not in data_cols]
    else:
        data_cols, excluded_meta_cols = _split_meta_cols(all_data_cols, df1, df2)

    def extract_row(df_idx, key) -> dict:
        # Return up to _MAX_DATA_COLS non-key, non-metadata columns as a clean
        # string dict.
        try:
            r = df_idx.loc[key]
            if isinstance(r, pd.DataFrame):
                r = r.iloc[0]
            return {c: _safe_str(r.get(c, "")) for c in data_cols[:_MAX_DATA_COLS]}
        except Exception:
            return {}


    def make_kv(key):
        return dict(zip(keys, key)) if isinstance(key, tuple) else {keys[0]: key}

    # -- File 1 only rows
    file1_only = []
    for k in sorted(only1_keys, key=_safe_sort_key)[:_MAX_DIFF_ROWS]:
        file1_only.append({"key_values": make_kv(k), "row_data": extract_row(d1, k)})

    # -- File 2 only rows
    file2_only = []
    for k in sorted(only2_keys, key=_safe_sort_key)[:_MAX_DIFF_ROWS]:
        file2_only.append({"key_values": make_kv(k), "row_data": extract_row(d2, k)})

    # -- Modified rows
    # ------------------------------------------------------------------
    # For unique keys: vectorised merge (1-to-1, safe).
    # For non-unique / best-effort keys: positional alignment within each group
    # (avoids Cartesian explosion when one key value maps to 1000s of rows).
    modified_rows = []
    _true_modified_count = 0  # real total, independent of the _MAX_DIFF_ROWS display cap

    if common_keys and data_cols:
        use_cols = data_cols[:_MAX_DATA_COLS]
        # Determine if key is truly unique in both files
        _key_unique = (
            d1.index.is_unique and d2.index.is_unique
        )

        if _key_unique:
            # Fast path: 1-to-1 merge, no explosion risk
            d1r = d1[use_cols].reset_index()
            d2r = d2[use_cols].reset_index()
            merged = d1r.merge(d2r, on=keys, suffixes=("__f1", "__f2"), how="inner")
            diff_mask = pd.Series(False, index=merged.index)
            for c in use_cols:
                c1, c2 = f"{c}__f1", f"{c}__f2"
                if c1 in merged.columns and c2 in merged.columns:
                    v1 = merged[c1].fillna("").astype(str).str.strip().apply(
                        lambda v: "" if v.lower() in _NULL_SENTINELS else v)
                    v2 = merged[c2].fillna("").astype(str).str.strip().apply(
                        lambda v: "" if v.lower() in _NULL_SENTINELS else v)
                    diff_mask |= _series_differs(v1, v2)

            _true_modified_count = int(diff_mask.sum())
            changed_df = merged[diff_mask].head(_MAX_DIFF_ROWS)
            _build_modified_rows(changed_df, keys, use_cols, make_kv, modified_rows)
        else:
            # Vectorised positional alignment for non-unique keys.

            def _add_pos(df_idx: pd.DataFrame) -> pd.DataFrame:
                r = df_idx.reset_index()
                r["_row_pos"] = r.groupby(keys).cumcount()
                return r

            d1r = _add_pos(d1[use_cols])
            d2r = _add_pos(d2[use_cols])
            join_cols = keys + ["_row_pos"]
            merged = d1r.merge(d2r, on=join_cols, suffixes=("__f1", "__f2"), how="inner")

            diff_mask = pd.Series(False, index=merged.index)
            for c in use_cols:
                c1, c2 = f"{c}__f1", f"{c}__f2"
                if c1 in merged.columns and c2 in merged.columns:
                    v1 = merged[c1].fillna("").astype(str).str.strip().apply(
                        lambda v: "" if v.lower() in _NULL_SENTINELS else v)
                    v2 = merged[c2].fillna("").astype(str).str.strip().apply(
                        lambda v: "" if v.lower() in _NULL_SENTINELS else v)
                    diff_mask |= _series_differs(v1, v2)

            _true_modified_count = int(diff_mask.sum())
            changed_df = merged[diff_mask].head(_MAX_DIFF_ROWS)
            _build_modified_rows(changed_df, keys, use_cols, make_kv, modified_rows)

    return {
        "comparison_type": "key-based",
        "key_columns":     keys,
        "data_columns":    data_cols[:_MAX_DATA_COLS],
        "excluded_meta_cols": excluded_meta_cols,
        "file1_only":      file1_only,
        "file2_only":      file2_only,
        "modified_rows":   modified_rows[:_MAX_DIFF_ROWS],
        "file1_only_count": len(only1_keys),
        "file2_only_count": len(only2_keys),
        "modified_count":   _true_modified_count,
        # aliases kept for Excel/email export backward-compat
        "added_count":      len(only2_keys),
        "removed_count":    len(only1_keys),
        "added_rows":       file2_only,
        "removed_rows":     file1_only,
        # Columns with data on one side but entirely null on the other
        "null_column_exceptions": _detect_null_column_exceptions(df1, df2, data_cols),
    }


def _content_based_diff(df1, df2, common_cols):
    # Content-based comparison (no unique key available).
    # Detects rows that are present in one file but absent in the other using
    # a Counter-based approach on the full row content.
    biz_cols, excluded_meta_cols = _split_meta_cols(common_cols, df1, df2)

    show_cols = biz_cols[:_MAX_DATA_COLS]

    # Cap rows before hashing to avoid OOM on very large DataFrames
    _ROW_HASH_CAP = 10_000
    truncated = len(df1) > _ROW_HASH_CAP or len(df2) > _ROW_HASH_CAP
    d1 = df1[show_cols].iloc[:_ROW_HASH_CAP].astype(str).fillna("")
    d2 = df2[show_cols].iloc[:_ROW_HASH_CAP].astype(str).fillna("")

    c1 = Counter(tuple(r) for r in d1.itertuples(index=False))
    c2 = Counter(tuple(r) for r in d2.itertuples(index=False))

    f1_tuples = [r for r, n in c1.items() for _ in range(max(0, n - c2.get(r, 0)))]
    f2_tuples = [r for r, n in c2.items() for _ in range(max(0, n - c1.get(r, 0)))]

    def to_entry(t):
        row_data = dict(zip(show_cols, t))
        return {"key_values": {}, "row_data": row_data}

    file1_only = [to_entry(t) for t in f1_tuples[:_MAX_DIFF_ROWS]]

    file2_only = [to_entry(t) for t in f2_tuples[:_MAX_DIFF_ROWS]]

    result = {
        "comparison_type": "content-based",
        "key_columns":      [],   # no key -- content-based uses full-row hashing
        "data_columns":     show_cols,
        "excluded_meta_cols": excluded_meta_cols,
        "file1_only":       file1_only,
        "file2_only":       file2_only,
        "modified_rows":    [],
        "file1_only_count": len(f1_tuples),
        "file2_only_count": len(f2_tuples),
        "modified_count":   0,
        "added_count":      len(f2_tuples),
        "removed_count":    len(f1_tuples),
        "added_rows":       file2_only,
        "removed_rows":     file1_only,
    }
    if truncated:
        result["truncation_note"] = (
            f"Row hashing capped at {_ROW_HASH_CAP:,} rows per file to avoid OOM; "
            "results reflect sampled data only."
        )
    return result


def _col_stats(df1: pd.DataFrame, df2: pd.DataFrame, common_cols: list[str]) -> list[dict]:
    # Per-column statistics for both files side-by-side. Capped at 60 cols for
    # performance.
    _MAX_STAT_COLS = 60
    stats = []
    total = max(len(df1), len(df2), 1)
    for c in common_cols[:_MAX_STAT_COLS]:
        s1, s2 = df1[c], df2[c]
        f1_null = int(s1.isna().sum())
        f2_null = int(s2.isna().sum())
        # Column-level match rate -- reset index so element-wise compare never raises
        _n1 = s1.fillna("").astype(str).str.strip().reset_index(drop=True)


        _n2 = s2.fillna("").astype(str).str.strip().reset_index(drop=True)
        _denom  = min(len(_n1), len(_n2))
        _matched = int((_n1.iloc[:_denom] == _n2.iloc[:_denom]).sum())
        match_rate_pct = round(_matched / _denom * 100, 1) if _denom > 0 else 100.0
        mismatch_count = _denom - _matched
        row: dict = {
            "col": c,
            "column": c,  # alias used by _reconCard JS
            "f1_dtype": str(s1.dtype), "f2_dtype": str(s2.dtype),
            "f1_rows": len(s1), "f2_rows": len(s2),
            "f1_null": f1_null, "f2_null": f2_null,
            "f1_unique": int(s1.nunique(dropna=True)),
            "f2_unique": int(s2.nunique(dropna=True)),
            "dtype_changed": str(s1.dtype) != str(s2.dtype),
            "match_rate_pct": match_rate_pct,
            "mismatch_count": mismatch_count,
        }
        if pd.api.types.is_numeric_dtype(s1) and pd.api.types.is_numeric_dtype(s2):
            f1_sum = float(s1.sum())


            f2_sum = float(s2.sum())
            row["f1_sum"]   = round(f1_sum, 4)
            row["f2_sum"]   = round(f2_sum, 4)
            row["sum_diff"] = round(f2_sum - f1_sum, 4)
            n1 = len(s1) - f1_null
            n2 = len(s2) - f2_null
            row["f1_mean"] = round(f1_sum / n1, 4) if n1 > 0 else None
            row["f2_mean"] = round(f2_sum / n2, 4) if n2 > 0 else None
            # Gross break = sum of |differences| on matched rows; net break = algebraic sum
            try:
                _s1r = s1.reset_index(drop=True)
                _s2r = s2.reset_index(drop=True)
                _len = min(len(_s1r), len(_s2r))
                diffs = _s2r.iloc[:_len] - _s1r.iloc[:_len]
                diffs = diffs.dropna()
                row["gross_break"] = round(float(diffs.abs().sum()), 4)
                row["net_break"]   = round(float(diffs.sum()), 4)
            except Exception:


                pass
        stats.append(row)
    return stats


# Break classification -- tag each modified row as TIMING / AMOUNT / VALUE / MISSING
_DATE_COL_RE = re.compile(
    r"(date|dt|time|ts|timestamp|day|month|year|period|settle|value_dt|trade_dt)",
    re.IGNORECASE,
)
_AMT_COL_RE = re.compile(
    r"(amount|amt|price|rate|qty|quantity|notional|value|bal|balance|vol|volume|nav|pnl|market)",
    re.IGNORECASE,
)


def _classify_break(changes: dict) -> str:
    # Return break category: TIMING | AMOUNT | MISSING | VALUE.


    if not changes:
        return "MISSING"
    cols = list(changes.keys())
    only_date = all(_DATE_COL_RE.search(c) for c in cols)
    has_amount = any(_AMT_COL_RE.search(c) for c in cols)
    if only_date:
        return "TIMING"
    if has_amount:
        return "AMOUNT"
    return "VALUE"


def _build_waterfall(file1_only_count: int, file2_only_count: int,
                      modified_rows: list[dict], data_cols: list[str]) -> dict:
    # Waterfall break summary:
    # - Missing source (in target only)
    # - Missing target (in source only)
    # - Value breaks by type (TIMING / AMOUNT / VALUE)
    # - Per-column break counts for the top broken columns
    type_counts: dict[str, int] = {"TIMING": 0, "AMOUNT": 0, "VALUE": 0}
    col_break_counts: dict[str, int] = {}
    for row in modified_rows:
        changes = row.get("changes", {})
        cat = _classify_break(changes)
        type_counts[cat] = type_counts.get(cat, 0) + 1
        for col in changes:
            col_break_counts[col] = col_break_counts.get(col, 0) + 1

    top_cols = sorted(col_break_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    return {
        "missing_source": file2_only_count,  # in target, not in source
        "missing_target": file1_only_count,  # in source, not in target
        "value_breaks":  len(modified_rows),
        "timing_breaks":  type_counts.get("TIMING", 0),
        "amount_breaks":  type_counts.get("AMOUNT", 0),
        "other_breaks":  type_counts.get("VALUE", 0),
        "top_broken_cols": [{"col": c, "count": n} for c, n in top_cols],
    }


def compare_dataframes(
    df1,
    df2,
    manual_keys,
    auto_keys,
    exclude_cols: list[str] | None = None,
    matched_cols: list[str] | None = None,
    user_hints: dict | None = None,
    force_data_cols: list[str] | None = None,
):
    # Compare two DataFrames.
    #
    # Key selection priority:
    # 1. Manual keys supplied by user (comma-separated in UI)
    # 2. key_hints from user_hints (if no manual keys given)
    # 3. Auto-inferred keys via infer_keys() -- always attempted when no
    # manual keys are given, regardless of the auto_keys flag.
    # infer_keys() tries: 100% unique single col → 2-col composite
    # → 3-col composite → best-effort single col → content-based fallback.
    #
    # exclude_cols: columns to drop from comparison entirely (user-specified).
    # force_data_cols: explicit list of value columns to compare; bypasses
    # _split_meta_cols so aggregated numeric columns are never misclassified.
    # matched_cols: semantically matched columns from analyze_mapping() (exact +
    # fuzzy). When provided, key inference is restricted to these columns so
    # that system-specific IDs are never selected as join keys.
    # user_hints: optional dict from UI criteria panel; recognised keys:
    #   key_hints   -- comma-separated expected key column name(s)
    #   exclude_hints -- comma-separated columns to skip in comparison
    hints = user_hints or {}
    df1, df2 = _normalise(df1), _normalise(df2)

    # Merge hint-based exclusions with explicit exclude_cols
    # ==== GAP: the source page(s) between here and "page 195" were never
    # captured -- the physical page re-photographed at this point was a
    # duplicate of an earlier page, not new content, so the original code
    # computing `common_cols` could not be recovered from the scan.
    # RECONSTRUCTED (not verified against source) from how common_cols is
    # used throughout the rest of this function (shared columns minus
    # exclusions) -- verify against the original file if it becomes available.
    exclude_hints = [c.strip() for c in str(hints.get("exclude_hints", "")).split(",") if c.strip()]
    all_excludes = set(exclude_cols or []) | set(exclude_hints)
    common_cols = [c for c in df1.columns if c in df2.columns and c not in all_excludes]
    # ==== END GAP ====

    # just JSON serialising every value as a string -- not a meaningful schema diff.
    _fmt1 = df1.attrs.get("_format", "").upper()
    _fmt2 = df2.attrs.get("_format", "").upper()
    _cross_format = _fmt1 != _fmt2 and bool(_fmt1) and bool(_fmt2)
    _string_formats = {"JSON", "XML"}

    def _is_format_artefact(t1: str, t2: str) -> bool:
        if not _cross_format:
            return False
        one_string = (t1 == "object" or t2 == "object")
        other_numeric = any(k in (t1 + t2) for k in ("int", "float", "bool"))
        either_string_fmt = _fmt1 in _string_formats or _fmt2 in _string_formats
        return one_string and other_numeric and either_string_fmt

    type_mismatch = {}
    for c in common_cols:
        t1, t2 = str(df1[c].dtype), str(df2[c].dtype)
        if t1 != t2:
            type_mismatch[c] = {
                "file1": t1,
                "file2": t2,
                "format_note": _is_format_artefact(t1, t2),
            }

    # key_hints from UI criteria panel -- treated as user-supplied manual keys
    # when no explicit manual_keys were provided via the key column picker
    if not manual_keys and hints.get("key_hints"):
        _hint_keys = [k.strip() for k in hints["key_hints"].split(",") if k.strip()]
        _valid_hint_keys = [k for k in _hint_keys if k in common_cols]
        if _valid_hint_keys:
            manual_keys = _valid_hint_keys

    if manual_keys:
        miss = [k for k in manual_keys if k not in common_cols]
        if miss:
            raise HTTPException(400, f"Key column(s) not found in both files: {miss}")
        keys = manual_keys
        key_method = f"manual ({', '.join(manual_keys)})"
    else:
        # Always try to find a key automatically -- infer_keys() handles all phases.
        # Pass matched_cols so inference is restricted to semantically equivalent
        # columns.
        keys, key_method = infer_keys(df1, df2, common_cols,
                                       matched_cols=matched_cols)

    diff = _key_based_diff(df1, df2, keys, common_cols, force_data_cols) if keys else \
        _content_based_diff(df1, df2, common_cols)


    # -- Duplicate row detection within each file ----------------------------------------

    _DUP_MAX = 200


    def _collect_dups(df: pd.DataFrame) -> tuple[list, int]:
        # Return duplicate rows and count. Uses key cols only for the dup check on large files
        # to avoid O(n*cols) hashing cost; falls back to full-row check for small files.
        check_cols = keys if keys and all(k in df.columns for k in keys) else list(df.columns)
        mask = df.duplicated(subset=check_cols, keep=False)
        count = int(mask.sum())


        if count == 0:
            return [], 0

        show_cols = list(df.columns)[:_MAX_DATA_COLS]

        rows = [
            {c: _safe_str(r[c]) for c in show_cols}
            for _, r in df[mask].head(_DUP_MAX).iterrows()
        ]

        return rows, count


    file1_dups, file1_dup_count = _collect_dups(df1)
    file2_dups, file2_dup_count = _collect_dups(df2)


    _cs = _col_stats(df1, df2, common_cols)
    # ==== GAP: the added/removed schema column lists were not recoverable from the
    # scan -- RECONSTRUCTED (unverified) below from df1/df2.columns, matching how
    # schema_added_columns ("Extra columns from File 2") and schema_removed_columns
    # ("Extra columns from File 1") are labelled by the Excel export elsewhere in
    # this file; verify against source if available.
    schema_added = [c for c in df2.columns if c not in df1.columns]
    schema_removed = [c for c in df1.columns if c not in df2.columns]
    diff.update({
        "key_method": key_method,
        "col_stats": _cs,
        "schema_added_columns": schema_added,
        "schema_removed_columns": schema_removed,
        "type_mismatches": type_mismatch,


        "file1_rows": len(df1), "file2_rows": len(df2),
        "common_columns": common_cols,
        "file1_columns": list(df1.columns),
        "file2_columns": list(df2.columns),
        # Duplicate rows within each individual file
        "file1_duplicate_count": file1_dup_count,
        "file2_duplicate_count": file2_dup_count,
        "file1_duplicate_rows": file1_dups,
        "file2_duplicate_rows": file2_dups,
        "file1_dup_columns": list(df1.columns)[:_MAX_DATA_COLS],
        "file2_dup_columns": list(df2.columns)[:_MAX_DATA_COLS],
        # Waterfall + net/gross break summary
        "waterfall": _build_waterfall(
            diff.get("file1_only_count", 0),
            diff.get("file2_only_count", 0),
            diff.get("modified_rows", []),
            diff.get("data_columns", []),
        ),
        "gross_break_total": round(sum(


                abs(c.get("net_break", 0) or 0) for c in _cs
                if c.get("net_break") is not None
            ), 4),
            "net_break_total": round(sum(
                c.get("net_break", 0) or 0 for c in _cs
                if c.get("net_break") is not None
            ), 4),
        })

    return diff


def _values_differ_scalar(a: str, b: str, try_date: bool = True) -> bool:
    # Scalar counterpart to _series_differs -- numeric- and date-aware equality
    # for two already-stringified cell values, used by the N-way engine's
    # per-key Python loop (compare_dataframes' vectorised path doesn't apply
    # once there are more than two sources to line up per key).
    # try_date=False skips the pd.to_datetime fallback entirely -- callers
    # that already know a column isn't date-like (e.g. currency codes,
    # statuses) use this to avoid dateutil's expensive parser running on
    # every mismatching pair of plain text values.
    if a == b:
        return False
    try:
        if abs(float(a) - float(b)) < 1e-9:
            return False
    except (ValueError, TypeError):
        pass
    if try_date:
        try:
            da, db = pd.to_datetime(a, errors="raise"), pd.to_datetime(b, errors="raise")
            if da == db:
                return False
        except Exception:
            pass
    return True


def compare_dataframes_nway(
    dataframes: list[tuple[str, pd.DataFrame]],
    manual_keys: list[str] | None = None,
    exclude_cols: list[str] | None = None,
) -> dict:
    # N-way reconciliation across 2+ (meaningfully, 3+) sources. Reuses the
    # same key normalisation and numeric/date-aware value comparison as the
    # 2-file engine (_key_based_diff / _series_differs), but instead of a
    # single file1-vs-file2 diff, builds:
    #   - a presence matrix: which sources contain each key (fully matched /
    #     partial / present in only one source)
    #   - a value-break table: for keys present in 2+ sources, every column
    #     where the sources disagree, showing ALL sources' values side by
    #     side (not just a pair)
    #
    # Key is inferred from the first two sources (infer_keys is inherently a
    # pairwise algorithm) and then applied across every source -- a
    # deliberate simplification rather than a new N-way key-inference
    # algorithm, which would be a much larger and riskier undertaking for
    # comparatively little benefit (the same business key almost always
    # applies uniformly across every source in real reconciliation work).
    if len(dataframes) < 2:
        raise HTTPException(400, "N-way reconciliation requires at least 2 files.")

    names = [n for n, _ in dataframes]
    dfs = [_normalise(df) for _, df in dataframes]

    all_excludes = set(exclude_cols or [])
    common_cols = set(dfs[0].columns)
    for df in dfs[1:]:
        common_cols &= set(df.columns)
    common_cols = [c for c in dfs[0].columns if c in common_cols and c not in all_excludes]
    if not common_cols:
        raise HTTPException(400, f"The {len(dfs)} uploaded files have no common columns to reconcile on.")

    if manual_keys:
        keys = [k for k in manual_keys if k in common_cols]
        method = f"manual ({', '.join(keys)})" if keys else "content-based (manual keys not found in common columns)"
    else:
        keys, method = infer_keys(dataframes[0][1], dataframes[1][1], common_cols)
    if not keys:
        raise HTTPException(400, f"No usable join key found across these {len(dfs)} files ({method}). "
                                  f"Specify key column(s) manually and rerun.")

    excluded_meta_cols = [c for c in common_cols if c not in keys
                           and _is_metadata_col(c, dfs[0][c].dtype if c in dfs[0].columns else None)]
    value_cols = [c for c in common_cols if c not in keys and c not in excluded_meta_cols][:_MAX_DATA_COLS]

    def _norm_key_col(s: pd.Series) -> pd.Series:
        s = s.fillna("").astype(str).str.strip()
        return s.where(~s.str.lower().isin(_NULL_SENTINELS), "")

    indexed: list[pd.DataFrame | None] = []
    duplicate_counts: list[int] = []
    for df in dfs:
        missing_keys = [k for k in keys if k not in df.columns]
        if missing_keys:
            indexed.append(None)
            duplicate_counts.append(0)
            continue
        df = df.copy()
        for k in keys:
            df[k] = _norm_key_col(df[k])
        duplicate_counts.append(int(df.duplicated(subset=keys).sum()))
        df_idx = df.set_index(keys)
        indexed.append(df_idx[~df_idx.index.duplicated(keep="first")])

    key_sets = [set(d.index.tolist()) if d is not None else set() for d in indexed]
    all_keys: set = set()
    for ks in key_sets:
        all_keys |= ks

    def _safe_sort_key(k):
        if isinstance(k, tuple):
            return tuple("" if (isinstance(v, float) and pd.isna(v)) else str(v) for v in k)
        return "" if (isinstance(k, float) and pd.isna(k)) else str(k)

    def make_kv(key):
        return dict(zip(keys, key)) if isinstance(key, tuple) else {keys[0]: key}

    all_keys_sorted = sorted(all_keys, key=_safe_sort_key)
    key_pos = {k: i for i, k in enumerate(all_keys_sorted)}

    # Reindex every source to the shared, sorted key list once and pull each
    # value column out as a plain numpy array. The per-key loop below then
    # does O(1) positional array indexing instead of repeated
    # DataFrame.loc[key, col] scalar lookups, which do not scale -- profiling
    # showed 20k rows x 3 sources spending ~12s in .loc alone (of ~21s total).
    _empty_idx = pd.Index([], dtype=object)
    aligned_cols: list[dict[str, "np.ndarray | None"]] = []
    for df_idx in indexed:
        src = (df_idx if df_idx is not None else pd.DataFrame(index=_empty_idx)).reindex(all_keys_sorted)
        aligned_cols.append({
            c: (src[c].map(_safe_str).to_numpy() if c in src.columns else None)
            for c in value_cols
        })

    # Decide once per column whether it's worth attempting date-aware
    # comparison at all -- _values_differ_scalar's pd.to_datetime fallback is
    # expensive (dateutil parsing), and running it for every mismatching pair
    # in plainly non-date columns (currency codes, statuses, free text) was
    # the other dominant cost alongside the .loc lookups above.
    date_like_cols: set[str] = set()
    for c in value_cols:
        sample: list[str] = []
        for arr in aligned_cols:
            vals = arr.get(c)
            if vals is not None:
                sample.extend(v for v in vals[:50] if v)
            if len(sample) >= 50:
                break
        if sample:
            parsed = pd.to_datetime(pd.Series(sample[:50]), errors="coerce", format="mixed")
            if parsed.notna().mean() > 0.5:
                date_like_cols.add(c)

    fully_matched = 0
    partial_rows: list[dict] = []
    singleton_counts = {n: 0 for n in names}
    value_break_rows: list[dict] = []
    col_break_counts: dict[str, int] = {}

    for key in all_keys_sorted:
        presence = [i for i, ks in enumerate(key_sets) if key in ks]
        if len(presence) == len(names):
            fully_matched += 1
        elif len(presence) == 1:
            singleton_counts[names[presence[0]]] += 1
        elif len(partial_rows) < _MAX_DIFF_ROWS:
            partial_rows.append({
                "key_values": make_kv(key),
                "present_in": [names[i] for i in presence],
                "missing_from": [names[i] for i in range(len(names)) if i not in presence],
            })

        if len(presence) < 2:
            continue  # nothing to value-compare with only one (or zero) source holding this key

        pos = key_pos[key]
        row_changes: dict[str, dict[str, str]] = {}
        for col in value_cols:
            values: dict[str, str] = {}
            for i in presence:
                arr = aligned_cols[i].get(col)
                values[names[i]] = arr[pos] if arr is not None else ""
            distinct_vals = list(values.values())
            ref = distinct_vals[0]
            try_date = col in date_like_cols
            if any(_values_differ_scalar(ref, v, try_date=try_date) for v in distinct_vals[1:]):
                row_changes[col] = values
                col_break_counts[col] = col_break_counts.get(col, 0) + 1
        if row_changes and len(value_break_rows) < _MAX_DIFF_ROWS:
            value_break_rows.append({"key_values": make_kv(key), "changes": row_changes})

    top_broken_columns = [{"col": c, "count": n} for c, n in
                           sorted(col_break_counts.items(), key=lambda x: x[1], reverse=True)[:10]]

    return {
        "sources": [{"name": n, "rows": len(df), "columns": len(df.columns)} for n, df in zip(names, dfs)],
        "keys": keys,
        "method": method,
        "common_columns": common_cols,
        "excluded_meta_cols": excluded_meta_cols,
        "duplicate_counts": dict(zip(names, duplicate_counts)),
        "counts": {
            "fully_matched": fully_matched,
            "partial": len(partial_rows),
            "value_breaks": len(value_break_rows),
            "total_keys": len(all_keys),
        },
        "singleton_counts": singleton_counts,
        "partial_rows": partial_rows,
        "value_breaks": value_break_rows,
        "top_broken_columns": top_broken_columns,
    }


def _fuzzy_field_score(a: str, b: str, kind: str) -> float:
    # 0..1 similarity for one field. `kind` is "text" | "numeric" | "date",
    # decided once per field by the caller (from dtype), not re-sniffed per row.
    a, b = (a or "").strip(), (b or "").strip()
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    if kind == "numeric":
        try:
            na, nb = float(a), float(b)
            denom = max(abs(na), abs(nb), 1.0)
            return max(0.0, 1.0 - min(1.0, abs(na - nb) / denom))
        except (ValueError, TypeError):
            pass  # fall through to text similarity if not actually numeric
    elif kind == "date":
        try:
            da, db = pd.to_datetime(a, errors="raise"), pd.to_datetime(b, errors="raise")
            days = abs((da - db).days)
            return max(0.0, 1.0 - min(1.0, days / 30.0))
        except Exception:
            pass  # fall through to text similarity if not actually a date
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


_FUZZY_MAX_ROWS = 3000  # O(n*m) candidate generation -- keep the pairwise scan bounded


def _fuzzy_score_and_match(
    df1: pd.DataFrame, df2: pd.DataFrame,
    fuzzy_fields: list[str],
    weights: dict[str, float] | None,
    threshold: float,
) -> list[tuple[int, int, float]]:
    # Core scoring + greedy assignment, shared by the 2-file and N-way fuzzy
    # entry points. Every row in df1 is scored against every candidate row in
    # df2 across the chosen fields (SequenceMatcher for text, relative
    # distance for numeric, day-delta for dates), combined into a weighted
    # average, then greedily assigned best-match-first so each row matches at
    # most once. Returns (df1_idx, df2_idx, score) triples for matches
    # scoring at or above threshold. Callers are responsible for the
    # _FUZZY_MAX_ROWS check (O(n*m) candidate generation).
    weights = weights or {f: 1.0 for f in fuzzy_fields}
    total_weight = sum(weights.get(f, 1.0) for f in fuzzy_fields) or 1.0

    field_kind: dict[str, str] = {}
    for f in fuzzy_fields:
        dt = str(df1[f].dtype)
        if "int" in dt or "float" in dt:
            field_kind[f] = "numeric"
        elif "datetime" in dt or re.search(r"date|_dt$|^dt\b", f, re.IGNORECASE):
            field_kind[f] = "date"
        else:
            field_kind[f] = "text"

    rows1 = df1[fuzzy_fields].astype(str).to_dict(orient="records")
    rows2 = df2[fuzzy_fields].astype(str).to_dict(orient="records")

    candidates: list[tuple[float, int, int]] = []
    for i, r1 in enumerate(rows1):
        for j, r2 in enumerate(rows2):
            score = sum(
                weights.get(f, 1.0) * _fuzzy_field_score(r1[f], r2[f], field_kind[f])
                for f in fuzzy_fields
            ) / total_weight
            if score >= threshold:
                candidates.append((score, i, j))
    candidates.sort(key=lambda t: t[0], reverse=True)

    used1: set[int] = set()
    used2: set[int] = set()
    matches: list[tuple[int, int, float]] = []
    for score, i, j in candidates:
        if i in used1 or j in used2:
            continue
        used1.add(i)
        used2.add(j)
        matches.append((i, j, score))
    return matches


def fuzzy_match_dataframes(
    df1: pd.DataFrame, df2: pd.DataFrame,
    fuzzy_fields: list[str],
    weights: dict[str, float] | None = None,
    threshold: float = 0.75,
    exclude_cols: list[str] | None = None,
) -> dict:
    # Probabilistic row matching for when there's no clean shared key -- e.g.
    # matching a bank statement narrative against an internal ledger by
    # name/amount/date proximity, tolerating typos and formatting drift.
    # Not an extension of the exact key-based engine -- see
    # _fuzzy_score_and_match for the actual scoring/assignment.
    df1, df2 = _normalise(df1), _normalise(df2)
    all_excludes = set(exclude_cols or [])
    common_cols = [c for c in df1.columns if c in df2.columns and c not in all_excludes]
    fuzzy_fields = [f for f in fuzzy_fields if f in common_cols]
    if not fuzzy_fields:
        raise HTTPException(400, "None of the requested fuzzy-match fields exist in both files.")
    if len(df1) > _FUZZY_MAX_ROWS or len(df2) > _FUZZY_MAX_ROWS:
        raise HTTPException(400, f"Fuzzy matching is capped at {_FUZZY_MAX_ROWS:,} rows per file "
                                  f"(got {len(df1):,} and {len(df2):,}) -- the pairwise comparison "
                                  f"doesn't scale past that. Use an exact key for larger files.")

    matches = _fuzzy_score_and_match(df1, df2, fuzzy_fields, weights, threshold)
    used1 = {i for i, _, _ in matches}
    used2 = {j for _, j, _ in matches}
    rows1 = df1[fuzzy_fields].astype(str).to_dict(orient="records")
    rows2 = df2[fuzzy_fields].astype(str).to_dict(orient="records")
    data_cols = [c for c in common_cols if c not in fuzzy_fields][:_MAX_DATA_COLS]

    def _row_dict(df: pd.DataFrame, idx: int, cols: list[str]) -> dict:
        r = df.iloc[idx]
        return {c: _safe_str(r.get(c, "")) for c in cols}

    matched_rows = []
    modified_rows = []
    for i, j, score in matches:
        changes = {}
        for c in data_cols:
            v1, v2 = _safe_str(df1.iloc[i].get(c, "")), _safe_str(df2.iloc[j].get(c, ""))
            if _values_differ_scalar(v1, v2):
                changes[c] = {"file1": v1, "file2": v2}
        entry = {
            "match_score": round(score, 3),
            "file1_row": _row_dict(df1, i, fuzzy_fields + data_cols),
            "file2_row": _row_dict(df2, j, fuzzy_fields + data_cols),
        }
        if changes:
            entry["changes"] = changes
            modified_rows.append(entry)
        else:
            matched_rows.append(entry)

    file1_only = [{"row_data": _row_dict(df1, i, fuzzy_fields + data_cols)}
                  for i in range(len(rows1)) if i not in used1][:_MAX_DIFF_ROWS]
    file2_only = [{"row_data": _row_dict(df2, j, fuzzy_fields + data_cols)}
                  for j in range(len(rows2)) if j not in used2][:_MAX_DIFF_ROWS]

    return {
        "fuzzy": True,
        "fuzzy_fields": fuzzy_fields,
        "threshold": threshold,
        "counts": {
            "matched": len(matched_rows),
            "modified": len(modified_rows),
            "file1_only": len(file1_only),
            "file2_only": len(file2_only),
        },
        "matched_rows": matched_rows[:_MAX_DIFF_ROWS],
        "modified_rows": modified_rows[:_MAX_DIFF_ROWS],
        "file1_only": file1_only,
        "file2_only": file2_only,
    }


def fuzzy_match_dataframes_nway(
    dataframes: list[tuple[str, pd.DataFrame]],
    fuzzy_fields: list[str],
    weights: dict[str, float] | None = None,
    threshold: float = 0.75,
    exclude_cols: list[str] | None = None,
) -> dict:
    # Fuzzy matching across 3+ sources. Like compare_dataframes_nway's exact-key
    # engine, this reuses the pairwise algorithm against a reference source
    # (the first uploaded file) rather than a new N-way-native fuzzy algorithm
    # -- deliberately, for the same reason: a much larger and riskier
    # undertaking for comparatively little benefit, since one source usually
    # serves as the natural anchor in real reconciliation work anyway.
    # Every other source is fuzzy-matched against the reference independently;
    # a reference row that matched in ALL other sources is "fully matched", one
    # that matched in some but not all is "partial", one that matched in none
    # is a reference-only singleton. Each non-reference source's own unmatched
    # rows are its own singletons.
    if len(dataframes) < 2:
        raise HTTPException(400, "N-way fuzzy matching requires at least 2 files.")

    names = [n for n, _ in dataframes]
    dfs = [_normalise(df) for _, df in dataframes]
    ref_df = dfs[0]

    all_excludes = set(exclude_cols or [])
    common_cols = set(dfs[0].columns)
    for df in dfs[1:]:
        common_cols &= set(df.columns)
    common_cols = [c for c in dfs[0].columns if c in common_cols and c not in all_excludes]
    fuzzy_fields = [f for f in fuzzy_fields if f in common_cols]
    if not fuzzy_fields:
        raise HTTPException(400, "None of the requested fuzzy-match fields exist in every file.")
    if any(len(df) > _FUZZY_MAX_ROWS for df in dfs):
        raise HTTPException(400, f"Fuzzy matching is capped at {_FUZZY_MAX_ROWS:,} rows per file.")

    data_cols = [c for c in common_cols if c not in fuzzy_fields][:_MAX_DATA_COLS]

    def _row_dict(df: pd.DataFrame, idx: int) -> dict:
        r = df.iloc[idx]
        return {c: _safe_str(r.get(c, "")) for c in fuzzy_fields + data_cols}

    # ref_idx -> {other_source_name: (other_idx, score)}
    ref_matches: dict[int, dict[str, tuple[int, float]]] = {i: {} for i in range(len(ref_df))}
    other_unmatched: dict[str, list[int]] = {}

    for name, df in zip(names[1:], dfs[1:]):
        matches = _fuzzy_score_and_match(ref_df, df, fuzzy_fields, weights, threshold)
        matched_other_idx: set[int] = set()
        for ref_i, other_j, score in matches:
            ref_matches[ref_i][name] = (other_j, score)
            matched_other_idx.add(other_j)
        other_unmatched[name] = [j for j in range(len(df)) if j not in matched_other_idx]

    fully_matched = 0
    partial_rows: list[dict] = []
    ref_singletons = 0
    value_break_rows: list[dict] = []
    col_break_counts: dict[str, int] = {}

    for ref_i in range(len(ref_df)):
        present = ref_matches[ref_i]
        if not present:
            ref_singletons += 1
            continue
        if len(present) == len(names) - 1:
            fully_matched += 1
        elif len(partial_rows) < _MAX_DIFF_ROWS:
            partial_rows.append({
                "reference_row": _row_dict(ref_df, ref_i),
                "present_in": [names[0]] + list(present.keys()),
                "missing_from": [n for n in names[1:] if n not in present],
            })

        if len(present) < 1:
            continue
        row_changes: dict[str, dict[str, str]] = {}
        for col in data_cols:
            values = {names[0]: _safe_str(ref_df.iloc[ref_i].get(col, ""))}
            for name, (other_j, _score) in present.items():
                df = dfs[names.index(name)]
                values[name] = _safe_str(df.iloc[other_j].get(col, ""))
            distinct_vals = list(values.values())
            ref_val = distinct_vals[0]
            if any(_values_differ_scalar(ref_val, v) for v in distinct_vals[1:]):
                row_changes[col] = values
                col_break_counts[col] = col_break_counts.get(col, 0) + 1
        if row_changes and len(value_break_rows) < _MAX_DIFF_ROWS:
            value_break_rows.append({"reference_row": _row_dict(ref_df, ref_i), "changes": row_changes})

    top_broken_columns = [{"col": c, "count": n} for c, n in
                           sorted(col_break_counts.items(), key=lambda x: x[1], reverse=True)[:10]]

    return {
        "fuzzy": True,
        "nway": True,
        "fuzzy_fields": fuzzy_fields,
        "threshold": threshold,
        "sources": [{"name": n, "rows": len(df), "columns": len(df.columns)} for n, df in zip(names, dfs)],
        "counts": {
            "fully_matched": fully_matched,
            "partial": len(partial_rows),
            "value_breaks": len(value_break_rows),
            "reference_only": ref_singletons,
        },
        "singleton_counts": {names[0]: ref_singletons,
                             **{n: len(other_unmatched.get(n, [])) for n in names[1:]}},
        "partial_rows": partial_rows,
        "value_breaks": value_break_rows,
        "top_broken_columns": top_broken_columns,
    }


# ----------------------------------------------------------------------
# Data Quality - enhanced with business rules & data dictionary
# ----------------------------------------------------------------------

# -- BFSI format validation patterns
# -- Built-in reference domain lists for Accuracy dimension
# No external dependency - baked in, covers 95%+ of BFSI use cases.
_KNOWN_DOMAINS: dict[str, frozenset] = {


    "currency_code": frozenset([
        "AED","AFN","ALL","AMD","ANG","AOA","ARS","AUD","AWG","AZN","BAM","BBD","BDT",
        "BGN","BHD","BIF","BMD","BND","BOB","BRL","BSD","BTN","BWP","BYN","BZD","CAD",
        "CDF","CHF","CLP","CNY","COP","CRC","CUP","CVE","CZK","DJF","DKK","DOP","DZD",
        "EGP","ERN","ETB","EUR","FJD","FKP","GBP","GEL","GHS","GIP","GMD","GNF","GTQ",
        "GYD","HKD","HNL","HRK","HTG","HUF","IDR","ILS","INR","IQD","IRR","ISK","JMD",
        "JOD","JPY","KES","KGS","KHR","KMF","KPW","KRW","KWD","KYD","KZT","LAK","LBP",

        "LKR","LRD","LSL","LYD","MAD","MDL","MGA","MKD","MMK","MNT","MOP","MRU","MUR",

        "MVR","MWK","MXN","MYR","MZN","NAD","NGN","NIO","NOK","NPR","NZD","OMR","PAB",
        "PEN","PGK","PHP","PKR","PLN","PYG","QAR","RON","RSD","RUB","RWF","SAR","SBD",
        "SCR","SDG","SEK","SGD","SHP","SLL","SOS","SRD","STN","SVC","SYP","SZL","THB",
        "TJS","TMT","TND","TOP","TRY","TTD","TWD","TZS","UAH","UGX","USD","UYU","UZS",
        "VES","VND","VUV","WST","XAF","XCD","XOF","XPF","YER","ZAR","ZMW","ZWL",
    ]),
    "country_code": frozenset([
        "AD","AE","AF","AG","AI","AL","AM","AO","AQ","AR","AS","AT","AU","AW","AX","AZ",
        "BA","BB","BD","BE","BF","BG","BH","BI","BJ","BL","BM","BN","BO","BQ","BR","BS",
        "BT","BV","BW","BY","BZ","CA","CC","CD","CF","CG","CH","CI","CK","CL","CM","CN",


    "CO","CR","CU","CV","CW","CX","CY","CZ","DE","DJ","DK","DM","DO","DZ","EC","EE",
    "EG","EH","ER","ES","ET","FI","FJ","FK","FM","FO","FR","GA","GB","GD","GE","GF",
    "GG","GH","GI","GL","GM","GN","GP","GQ","GR","GS","GT","GU","GW","GY","HK","HM",
    "HN","HR","HT","HU","ID","IE","IL","IM","IN","IO","IQ","IR","IS","IT","JE","JM",
    "JO","JP","KE","KG","KH","KI","KM","KN","KP","KR","KW","KY","KZ","LA","LB","LC",
    "LI","LK","LR","LS","LT","LU","LV","LY","MA","MC","MD","ME","MF","MG","MH","MK",

    "ML","MM","MN","MO","MP","MQ","MR","MS","MT","MU","MV","MW","MX","MY","MZ","NA",
    "NC","NE","NF","NG","NI","NL","NO","NP","NR","NU","NZ","OM","PA","PE","PF","PG",
    "PH","PK","PL","PM","PN","PR","PS","PT","PW","PY","QA","RE","RO","RS","RU","RW",
    "SA","SB","SC","SD","SE","SG","SH","SI","SJ","SK","SL","SM","SN","SO","SR","SS",
    "ST","SV","SX","SY","SZ","TC","TD","TF","TG","TH","TJ","TK","TL","TM","TN","TO",
    "TR","TT","TV","TW","TZ","UA","UG","UM","US","UY","UZ","VA","VC","VE","VG","VI",
    "VN","VU","WF","WS","YE","YT","ZA","ZM","ZW",
]),
"asset_class": frozenset([

    "EQUITY","BOND","FX","RATES","CREDIT","COMMODITY","REAL_ESTATE","CASH",
    "DERIVATIVE","SWAP","FUTURE","OPTION","FORWARD","ETF","FUND","REPO",
    "SECURITISATION","STRUCTURED","FIXED_INCOME","MONEY_MARKET",


    "EQ","FI","CMDTY","CRNCY","CRED","ALTINV",
  ]),
  "side":
frozenset(["BUY","SELL","B","S","BUY/SELL","SHORT","LONG","COVER","OPEN"]),
  "trade_type":
frozenset(["OUTRIGHT","SWAP","OPTION","FORWARD","SPOT","REPO","REVERSE_REPO"]),
  "settlement_type": frozenset(["DVP","FOP","DFP","RVP","FREE"]),
  "day_count":
frozenset(["ACT/360","ACT/365","ACT/ACT","30/360","30E/360","BUS/252"]),
}


# --- Address completeness validator ---------------------------------------

_ADDRESS_PARTS = re.compile(
    r'(?P<number>\d+[A-Za-z]?)\s+'         # house/building number
    r'(?P<street>[A-Za-z][A-Za-z\s\.\-]{2,})'  # street name
)


_DQ_FORMAT_PATTERNS: dict[str, re.Pattern] = {


    "isin":   re.compile(r'^[A-Z]{2}[A-Z0-9]{9}[0-9]$'),
    "cusip":  re.compile(r'^[0-9A-Z]{9}$'),
    "sedol":  re.compile(r'^[0-9BCDFGHJKLMNPQRSTVWXYZ]{7}$'),
    "lei":    re.compile(r'^[A-Z0-9]{18}[0-9]{2}$'),
    "bic":    re.compile(r'^[A-Z]{4}[A-Z]{2}[A-Z0-9]{2}([A-Z0-9]{3})?$'),
    "iban":   re.compile(r'^[A-Z]{2}[0-9]{2}[A-Z0-9]{4}[0-9]{7}([A-Z0-9]?){0,16}$'),  # OCR-UNCERTAIN: pattern tail cut off at photo edge
    "email":  re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$'),
    "phone":  re.compile(r'^(?!-|\d+\.\d)(\+|[(]|\d)[\d\s\-()]{6,18}\d$'),  # OCR-UNCERTAIN: bracket/paren grouping hard to read clearly
    "date_iso":re.compile(r'^\d{4}-\d{2}-\d{2}$'),
    "url":    re.compile(r'^https?://[^\s]+$'),
    "currency_code": re.compile(r'^[A-Z]{3}$'),
    "mic":    re.compile(r'^[A-Z]{4}$'),

    # --- Address / Postcode formats -----------------------------------

    # UK postcode: SW1A 2AA / EC1A 1BB / W1A 0AX
    "postcode_uk": re.compile(r'^[A-Z]{1,2}[0-9][0-9A-Z]?\s?[0-9][A-Z]{2}$'),
    # US ZIP:    12345 or 12345-6789
    "postcode_us": re.compile(r'^\d{5}(-\d{4})?$'),
    # German PLZ: 5 digits


    "postcode_de": re.compile(r'^\d{5}$'),
    # French CP: 5 digits starting 01-95 or 97x
    "postcode_fr": re.compile(r'^(0[1-9]|[1-8]\d|9[0-5]|97[1-6])\d{3}$'),
    # Indian PIN: 6 digits starting 1-9
    "postcode_in": re.compile(r'^[1-9]\d{5}$'),
    # Generic postcode -- alphanumeric 3-10 chars
    "postcode":   re.compile(r'^[A-Z0-9][A-Z0-9\s\-]{2,9}$'),
    # Street address -- must contain a number and letters (very permissive)
    "street_address": re.compile(r'^\d+\s+.{3,}$'),
}


def _apply_rule(series: pd.Series, rule: dict) -> dict:
    # Apply a single business rule to a column. Returns pass/fail stats.
    rt       = rule.get("rule_type", "")
    val      = rule.get("value", rule.get("min_value", ""))
    severity = rule.get("severity", "major")
    total    = len(series)
    passed   = failed = 0


    failing_examples = []

    try:
        if rt == "not_null":
            mask = series.isna() | (series.astype(str).str.strip() == "")
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()


        elif rt == "unique":
            dups = series.duplicated(keep=False)
            failed = int(dups.sum())
            passed = total - failed
            failing_examples = series[dups].head(5).astype(str).tolist()


        elif rt in ("min", "min_value"):
            numeric = pd.to_numeric(series, errors="coerce")
            mask = numeric.notna() & (numeric < float(val))
            failed = int(mask.sum())


            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt in ("max", "max_value"):
            numeric = pd.to_numeric(series, errors="coerce")
            mask = numeric.notna() & (numeric > float(val))
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "range":
            # Fix: range was parsed but never validated -- now implemented
            lo = float(rule.get("min_value", val))
            hi = float(rule.get("max_value", val))
            numeric = pd.to_numeric(series, errors="coerce")
            mask = numeric.notna() & ((numeric < lo) | (numeric > hi))
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()


        elif rt == "min_length":
            mask = series.astype(str).str.len() < int(val)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "max_length":
            mask = series.astype(str).str.len() > int(val)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "exact_length":
            mask = series.astype(str).str.len() != int(val)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "pattern":
            mask = ~series.astype(str).str.fullmatch(str(val), na=False)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "allowed_values":
            # Accept both comma and pipe as separators (UI uses comma, legacy uses pipe)
            sep = "," if "," in str(val) else "|"
            allowed = {v.strip() for v in str(val).split(sep) if v.strip()}
            mask = series.notna() & ~series.astype(str).str.strip().isin(allowed)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "not_allowed_values":
            not_allowed = {v.strip() for v in str(val).split("|")}
            mask = series.astype(str).isin(not_allowed)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "freshness_days":
            dates = pd.to_datetime(series, errors="coerce")
            cutoff = pd.Timestamp.now() - pd.Timedelta(days=int(val))
            mask = dates < cutoff
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "date_format":
            # Validate dates parse correctly -- use pd.to_datetime with the format
            parsed = pd.to_datetime(series, format=str(val), errors="coerce")
            mask = series.notna() & parsed.isna()
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "date_range":
            # val = "YYYY-MM-DD|YYYY-MM-DD"
            parts = str(val).split("|")
            date_lo = pd.to_datetime(parts[0].strip(), errors="coerce")
            date_hi = pd.to_datetime(parts[1].strip(), errors="coerce") if len(parts) > 1 else pd.Timestamp.now()
            dates = pd.to_datetime(series, errors="coerce")
            mask = dates.notna() & ((dates < date_lo) | (dates > date_hi))
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "not_future_date":
            dates = pd.to_datetime(series, errors="coerce")
            mask = dates > pd.Timestamp.now()
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "not_past_date":
            dates = pd.to_datetime(series, errors="coerce")
            mask = dates < pd.Timestamp.now()
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "positive":
            numeric = pd.to_numeric(series, errors="coerce")
            mask = numeric.notna() & (numeric <= 0)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "non_negative":
            numeric = pd.to_numeric(series, errors="coerce")
            mask = numeric.notna() & (numeric < 0)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "negative":
            numeric = pd.to_numeric(series, errors="coerce")
            mask = numeric.notna() & (numeric >= 0)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "integer_only":
            numeric = pd.to_numeric(series, errors="coerce")
            mask = numeric.notna() & (numeric != numeric.round())
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "no_whitespace":
            mask = series.astype(str).str.contains(r'\s', regex=True)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "no_special_chars":
            mask = series.astype(str).str.contains(r'[^A-Za-z0-9\s\-_.]', regex=True)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "uppercase":
            mask = series.astype(str).str.upper() != series.astype(str)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "lowercase":
            mask = series.astype(str).str.lower() != series.astype(str)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "numeric_only":
            mask = pd.to_numeric(series, errors="coerce").isna() & series.notna()
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "no_duplicates_in_set":
            # Like unique but treats pipe-delimited sets within a value
            def _has_dup_in_set(v):
                parts = [x.strip() for x in str(v).split("|")]
                return len(parts) != len(set(parts))
            mask = series.apply(_has_dup_in_set)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        # NOTE (reconstruction): a duplicate "sum_equals" elif branch (simpler,
        # no-tolerance version) appeared here in the scanned pages -- dropped as
        # a page-repeat artifact in favor of the more complete tolerance-based
        # "sum_equals" kept below under Feature 3 (Control Totals). Since elif
        # chains match top-to-bottom, only one "sum_equals" branch can ever run;
        # verify against source if the original file becomes available.
        elif rt == "mean_range":
            # val = "lo|hi"
            parts = str(val).split("|")
            lo_m, hi_m = float(parts[0]), float(parts[1])
            numeric = pd.to_numeric(series, errors="coerce")
            mean_val = float(numeric.mean())
            if lo_m <= mean_val <= hi_m:
                passed, failed = total, 0
            else:
                passed, failed = 0, total
            failing_examples = [f"Mean={mean_val:.4f}, expected {lo_m}-{hi_m}"]  # OCR-UNCERTAIN

        elif rt == "std_max":
            numeric = pd.to_numeric(series, errors="coerce")
            std_val = float(numeric.std())
            if std_val <= float(val):
                passed, failed = total, 0
            else:
                passed, failed = 0, total
            failing_examples = [f"Std={std_val:.4f}, max allowed={float(val):.4f}"]

        elif rt == "completeness_pct":
            # % non-null must be >= val
            non_null = int(series.notna().sum())
            pct = non_null / total * 100 if total else 0
            if pct >= float(val):
                passed, failed = total, 0
            else:
                passed, failed = 0, total
            failing_examples = [f"Completeness={pct:.1f}%, required>={float(val):.1f}%"]

        elif rt == "uniqueness_pct":
            uniq = int(series.nunique(dropna=True))
            pct = uniq / total * 100 if total else 0
            if pct >= float(val):
                passed, failed = total, 0
            else:
                passed, failed = 0, total
            failing_examples = [f"Uniqueness={pct:.1f}%, required>={float(val):.1f}%"]

        # -- Custom regex rule (from Rule Builder UI) -----------------------------
        elif rt == "regex_format":
            # val holds the user-supplied regex pattern
            if not val:
                return {"skipped": True, "reason": "regex_format requires a pattern value"}
            try:
                _custom_pat = re.compile(str(val))
            except re.error as _re_err:
                return {"skipped": True, "reason": f"Invalid regex pattern: {_re_err}"}
            mask = ~series.astype(str).str.fullmatch(_custom_pat.pattern)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        # -- BFSI format checks -- all use vectorised str.fullmatch ----------------
        elif rt.endswith("_format"):
            fmt_key = rt[:-7]  # strip "_format" suffix
            pat = _DQ_FORMAT_PATTERNS.get(fmt_key)
            if pat is None:
                return {"skipped": True, "reason": f"Unknown format: {fmt_key}"}
            upper = series.astype(str).str.upper()
            mask = ~upper.str.fullmatch(pat.pattern)
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "decimal_places":
            # Vectorised: extract decimal part, strip trailing zeros, measure length
            max_dp = int(val)
            clean = series.dropna()
            dp = (clean.astype(str)
                  .str.extract(r'\.(\d+)$')[0]
                  .str.rstrip("0")
                  .str.len()
                  .fillna(0)
                  .astype(int))
            mask = dp > max_dp
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = clean[mask.values].head(5).astype(str).tolist()

        elif rt == "significant_figures":
            # Vectorised: format to 10 sig figs, strip non-digits, measure length
            max_sf = int(val)
            clean = series.dropna()
            sf = (clean.apply(lambda v: len(f"{float(v):.10g}".replace("-", "").replace(".", "").lstrip("0")) if str(v) not in ("nan", "") else 0))  # OCR-UNCERTAIN
            mask = sf > max_sf
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = clean[mask.values].head(5).astype(str).tolist()

        elif rt == "sla_hours":
            # Timeliness: values (timestamps) must be within N hours of now
            dates = pd.to_datetime(series, errors="coerce")
            cutoff = pd.Timestamp.now() - pd.Timedelta(hours=int(val))
            mask = dates < cutoff
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "update_frequency_days":
            # Timeliness: the most recent value must be within N days of today
            dates = pd.to_datetime(series, errors="coerce")
            most_recent = dates.max()
            if pd.isna(most_recent):
                passed, failed = 0, total
                failing_examples = ["No valid dates found"]
            elif (pd.Timestamp.now() - most_recent).days <= int(val):
                passed, failed = total, 0
            else:
                passed, failed = 0, total
                failing_examples = [f"Most recent: {most_recent.date()}, threshold: {val} days"]

        elif rt == "row_count_min":
            if total >= int(val):
                passed, failed = total, 0
            else:
                passed, failed = 0, total

        elif rt == "row_count_max":
            if total <= int(val):
                passed, failed = total, 0
            else:
                passed, failed = 0, total

        # -- Feature 3: Control Totals ---------------------------------------------
        elif rt == "sum_equals":
            # Column sum must equal a specific control total value
            numeric = pd.to_numeric(series, errors="coerce")
            actual = float(numeric.sum())
            target = float(val)
            tol   = abs(target) * float(rule.get("tolerance_pct", 0.001)) / 100
            if abs(actual - target) <= max(tol, 0.01):
                passed, failed = total, 0
            else:
                passed, failed = 0, total
            failing_examples = [f"Sum={actual:.4f}, expected={target:.4f}, diff={actual-target:.4f}"]

        elif rt == "sum_range":
            # Column sum must be within [min_value, max_value]
            numeric = pd.to_numeric(series, errors="coerce")
            actual = float(numeric.sum())
            lo = float(rule.get("min_value", val))
            hi = float(rule.get("max_value", val))
            if lo <= actual <= hi:
                passed, failed = total, 0
            else:
                passed, failed = 0, total
            failing_examples = [f"Sum={actual:.4f}, expected range [{lo}, {hi}]"]

        elif rt == "sum_netzero":
            # Column must net to zero (e.g. cashflows, P&L attribution)
            numeric = pd.to_numeric(series, errors="coerce")
            net   = float(numeric.sum())
            tol   = float(rule.get("tolerance", 0.01))
            if abs(net) <= tol:
                passed, failed = total, 0
            else:
                passed, failed = 0, total
            failing_examples = [f"Net={net:.6f} (tolerance={tol})"]

        # -- Feature 5: Business Day / Settlement Date Validation -------------------
        elif rt == "business_day":
            # All dates must fall on a business day (Mon-Fri, non-holiday)
            dates = pd.to_datetime(series, errors="coerce")
            valid = dates.dropna()
            # Weekends
            mask_weekend = valid.dt.dayofweek >= 5
            failed = int(mask_weekend.sum())
            passed = len(valid) - failed
            failing_examples = valid[mask_weekend].dt.date.head(5).astype(str).tolist()

        elif rt == "settlement_date_t2":
            # Settlement date must be T+2 business days from a paired trade_date column
            # rule.get("trade_date_col") specifies the trade date column name
            trade_col = rule.get("trade_date_col", "")
            if not trade_col:
                return {"skipped": True, "reason": "settlement_date_t2 requires trade_date_col"}
            # This rule is applied at the DataFrame level -- series is the settlement date
            # The check runs in _apply_cross_col_rules, not here
            return {"skipped": True, "reason": "Use cross_column rule type for settlement_date_t2"}

        elif rt == "no_future_date":
            dates = pd.to_datetime(series, errors="coerce")
            today = pd.Timestamp.now().normalize()
            mask = dates > today
            failed = int(mask.sum())
            passed = int(dates.notna().sum()) - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        elif rt == "no_weekend_date":
            dates = pd.to_datetime(series, errors="coerce")
            valid = dates.dropna()
            mask = valid.dt.dayofweek >= 5
            failed = int(mask.sum())
            passed = len(valid) - failed
            failing_examples = valid[mask].dt.date.head(5).astype(str).tolist()

        # -- Feature 1: Conditional / Cross-column Rules -----------------------------
        elif rt == "conditional":
            # if col_a == val_a then col_b must satisfy condition_b
            # rule: {when_col, when_val, then_col, then_condition, then_val}
            # Applied at DataFrame level -- skip here, handled by _apply_cross_col_rules
            return {"skipped": True, "reason": "conditional rules require DataFrame context -- use cross_column_rules hint"}

        # -- Feature 4: Referential Integrity -----------------------------------------
        elif rt == "ref_exists":
            # All values must exist in a reference set
            # rule: {ref_values: ["val1","val2",...]} OR {ref_col: "colname"} (cross-file)
            ref_values = rule.get("ref_values", [])
            if not ref_values:
                return {"skipped": True, "reason": "ref_exists requires ref_values list"}
            ref_set = {str(v).strip().lower() for v in ref_values}
            clean = series.dropna().astype(str).str.strip().str.lower()
            mask = ~clean.isin(ref_set)
            failed = int(mask.sum())
            passed = len(clean) - failed
            failing_examples = series[series.index.isin(clean[mask].index)].head(5).astype(str).tolist()

        # -- Address structural completeness -----------------------------------------
        elif rt == "address_complete":
            # Validates that a compound address field contains at least a
            # house number + street name. No geocoding -- pure structural.
            def _is_valid_address(v: str) -> bool:
                v = str(v).strip()
                if len(v) < 5:
                    return False
                return bool(_ADDRESS_PARTS.search(v))
            mask = series.dropna().astype(str).apply(lambda v: not _is_valid_address(v))
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask[mask].index].head(5).astype(str).tolist()

        elif rt == "address_parts_complete":
            # Validates that all required address sub-fields are non-empty.
            # val = comma-separated sub-field column names, e.g. "street,city,postcode,country"
            required_cols = [c.strip() for c in str(val).split(",") if c.strip()]
            # For a single compound column, check it contains comma or newline separating parts
            if required_cols and required_cols[0] in (series.name if hasattr(series, 'name') else ""):  # OCR-UNCERTAIN
                # Single column -- check non-empty parts count >= required
                mask = series.astype(str).str.count(r'[,\n]') < (len(required_cols) - 1)
            else:
                mask = series.isna() | (series.astype(str).str.strip() == "")
            failed = int(mask.sum())
            passed = total - failed
            failing_examples = series[mask].head(5).astype(str).tolist()

        # -- Known domain accuracy
        # ---------------------------------------------------------------------------
        elif rt == "domain_accuracy":
            # val = domain name from _KNOWN_DOMAINS (e.g. "currency_code", "country_code")
            domain = str(val).strip().lower()
            ref_set = _KNOWN_DOMAINS.get(domain)
            if ref_set is None:
                return {"skipped": True, "reason": f"Unknown domain: {domain}. Valid: {list(_KNOWN_DOMAINS)}"}
            clean = series.dropna().astype(str).str.strip().str.upper()
            mask = ~clean.isin(ref_set)
            failed = int(mask.sum())
            passed = len(clean) - failed
            failing_examples = series[series.index.isin(clean[mask].index)].head(5).astype(str).tolist()

        else:
            return {"skipped": True, "reason": f"Unknown rule type: {rt}"}

    except Exception as exc:
        return {"skipped": True, "reason": str(exc)}

    return {
        "skipped": False,
        "total": total,
        "passed": passed,
        "failed": failed,
        "pass_pct": round(passed / total * 100, 1) if total else 0,
        "failing_examples": failing_examples,
        "status": "PASS" if failed == 0 else ("WARN" if failed / total < 0.05 else "FAIL"),
        "severity": severity,
    }


# ---- Cross Reference Engine ----------------------------------------------

# BFSI identifier cross-mapping: columns that mean the same thing across systems


_XREF_ID_SYNONYMS: dict[str, list[str]] = {
    "isin":     ["isin", "isin_code", "security_id", "sec_id", "bond_isin"],
    "cusip":    ["cusip", "cusip_code", "cusip9"],
    "sedol":    ["sedol", "sedol_code", "sedol7"],
    "ticker":   ["ticker", "symbol", "bbg_ticker", "ric", "reuters_ticker"],
    "lei":      ["lei", "legal_entity_id", "lei_code", "reporting_entity"],
    "bbg_id":   ["bbg_id", "bloomberg_id", "bbg_global_id", "figi"],
    "trade_id": ["trade_id", "trade_ref", "tradeid", "deal_id", "order_id"],
    "account":  ["account", "acct", "account_id", "acct_id", "account_number"],
    "currency": ["currency", "ccy", "base_ccy", "currency_code", "iso_ccy"],
    "counterparty": ["counterparty", "cpty", "counterparty_id", "cp_id"],
}


def _xref_normalise_key(series: "pd.Series") -> "pd.Series":
    # Normalise an identifier column for cross-source matching:
    # strip whitespace, uppercase, remove dashes/spaces (ISIN/CUSIP format variants).
    return (


        series.fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .str.replace(r"[\s\-]", "", regex=True)
    )


def _xref_find_key_col(df: "pd.DataFrame", user_key: str) -> str | None:
    # Find the best key column in a DataFrame.
    # Tries: user-specified -> BFSI synonym match -> high-cardinality column.
    cols_lc = {c.lower(): c for c in df.columns}

    if user_key:
        # Exact or case-insensitive match
        if user_key in df.columns:
            return user_key
        if user_key.lower() in cols_lc:
            return cols_lc[user_key.lower()]


        # Partial match
        for col_lc, col in cols_lc.items():
            if user_key.lower() in col_lc or col_lc in user_key.lower():
                return col

    # BFSI synonym lookup
    for canonical, synonyms in _XREF_ID_SYNONYMS.items():
        for s in synonyms:
            if s in cols_lc:
                return cols_lc[s]

    # Fallback: highest-cardinality non-numeric column (likely an identifier)
    best_col, best_ratio = None, 0.0
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            continue
        ratio = df[col].nunique(dropna=True) / max(len(df), 1)
        if ratio > best_ratio:
            best_ratio, best_col = ratio, col


    return best_col if best_ratio > 0.5 else None


def analyze_cross_reference(
    sources: list[tuple[str, "pd.DataFrame"]],
    key_col:     str | None = None,
    compare_fields: list[str] | None = None,
    golden_source:  str | None = None,
    conflicts_only: bool     = False,
    show_coverage: bool     = True,
) -> dict:

    # Batch 2 & 3 -- N-way cross reference engine.

    # Args:
    # sources:    [(source_name, DataFrame), ...] -- 2 to 5 sources
    # key_col:    common identifier column name (auto-detected if None)
    # compare_fields: columns to compare across sources (all common if None)
    # golden_source:  name of the authoritative source (majority-wins if None)


    # conflicts_only: only return rows with conflicts
    # show_coverage:  include coverage matrix in output

    # Returns:
    # {
    # "summary":    {total_records, sources, matched, conflict_count, coverage_pct},
    # "key_col_used":  str,
    # "key_col_per_source": {source_name: col},
    # "compare_fields": [str],
    # "coverage_matrix": {field: {source_name: count_non_null}},
    # "conflicts":    [{key, field, values: {src: val}, conflict_type, golden_value}],
    # "only_in":    {source_name: [key_values]},
    # "golden_source": str,
    # "source_stats": {source_name: {rows, matched, conflicts, coverage_pct}},
    # }


    import pandas as _pd

    if len(sources) < 2:


        raise ValueError("Cross Reference requires at least 2 sources.")

    source_names = [s[0] for s in sources]
    source_dfs   = [s[1] for s in sources]

    # ---- Step 1: Find key column in each source -------------------------
    key_per_source: dict[str, str] = {}
    for name, df in sources:
        k = _xref_find_key_col(df, key_col or "")
        if not k:
            raise ValueError(f"Could not find a key column in source '{name}'. "
                f"Please specify the Common Identifier Key.")
        key_per_source[name] = k

    # ---- Step 2: Build normalised key sets per source --------------------
    keyed: dict[str, _pd.DataFrame] = {}
    for name, df in sources:
        k = key_per_source[name]
        d = df.copy()


        d["__xref_key__"] = _xref_normalise_key(d[k])
        d = d.drop_duplicates(subset=["__xref_key__"], keep="first")
        keyed[name] = d.set_index("__xref_key__")

    # ---- Step 3: Find common fields to compare ----------------------------
    # Only compare columns that appear in at least 2 sources
    all_cols: dict[str, int] = {}
    for name, df in sources:
        for col in df.columns:
            if col == key_per_source[name]:
                continue
            norm = col.lower().strip()
            all_cols[norm] = all_cols.get(norm, 0) + 1

    # Map normalised col name -> actual col name per source
    col_map: dict[str, dict[str, str]] = {}  # norm_col -> {source: actual_col}
    for name, df in sources:
        for col in df.columns:
            if col == key_per_source[name]:


                continue
            norm = col.lower().strip()
            col_map.setdefault(norm, {})[name] = col

    # Fields in >=2 sources
    common_norm_fields = [f for f, cnt in all_cols.items() if cnt >= 2]
    if compare_fields:
        cf_lower = {c.lower().strip() for c in compare_fields}
        common_norm_fields = [f for f in common_norm_fields if f in cf_lower]

    # ---- Step 4: Build universal key set (union of all source keys) ------------
    all_keys: set[str] = set()
    for df in keyed.values():
        all_keys.update(df.index.tolist())
    all_keys.discard("")

    # ---- Step 5: Coverage matrix -------------------------------------
    coverage_matrix: dict[str, dict[str, int]] = {}
    if show_coverage:


        for norm_col in common_norm_fields:
            coverage_matrix[norm_col] = {}
            for name in source_names:
                actual = col_map.get(norm_col, {}).get(name)
                if actual and actual in keyed.get(name, _pd.DataFrame()).columns:
                    non_null = int(keyed[name][actual].notna().sum())
                    coverage_matrix[norm_col][name] = non_null
                else:
                    coverage_matrix[norm_col][name] = 0

    # ---- Step 6: Conflict detection -------------------------------------
    conflicts: list[dict] = []
    only_in:  dict[str, list[str]] = {n: [] for n in source_names}

    # Keys present in all sources vs only some
    keys_per_source: dict[str, set] = {n: set(keyed[n].index) for n in source_names}
    keys_in_all = set.intersection(*keys_per_source.values()) if keys_per_source else set()

    for name in source_names:


        _other_key_sets = [keys_per_source[n] for n in source_names if n != name]
        exclusive = keys_per_source[name] - (set.union(*_other_key_sets) if _other_key_sets else set())
        only_in[name] = sorted(list(exclusive))[:500]

    # Value conflicts across sources for common keys
    for key in sorted(keys_in_all):
        for norm_col in common_norm_fields:
            values: dict[str, str] = {}
            for name in source_names:
                actual = col_map.get(norm_col, {}).get(name)
                if actual and actual in keyed[name].columns and key in keyed[name].index:
                    val = keyed[name].loc[key, actual]
                    values[name] = "" if (_pd.isna(val) if not isinstance(val, str) else False) else str(val).strip()
                else:
                    values[name] = None  # source doesn't have this field

            # Only flag if at least 2 sources have a non-null value
            present_vals = {v for v in values.values() if v not in (None, "")}
            if len(present_vals) < 2:


                continue

            # Normalise for comparison (case, whitespace)
            norm_vals = {k: v.upper().strip() if v else v for k, v in values.items()}
            unique_norm = {v for v in norm_vals.values() if v not in (None, "")}

            if len(unique_norm) <= 1:
                continue  # all sources agree

            # Classify conflict type
            try:
                num_vals = {k: float(v.replace(",", "")) for k, v in values.items()
                    if v not in (None, "")}
                if len(num_vals) == len(present_vals):
                    spread = max(num_vals.values()) - min(num_vals.values())
                    avg   = sum(num_vals.values()) / len(num_vals)
                    conflict_type = "NUMERIC_CONFLICT"
                    if avg != 0 and spread / abs(avg) < 0.001:
                        conflict_type = "ROUNDING_DIFFERENCE"


                else:
                    conflict_type = "VALUE_CONFLICT"
            except (ValueError, ZeroDivisionError):
                conflict_type = "VALUE_CONFLICT"

            # Determine golden value
            if golden_source and golden_source in values and values[golden_source] not in (None, ""):
                golden_value = values[golden_source]
            else:
                # Majority vote
                from collections import Counter as _Counter
                vote = _Counter(v for v in values.values() if v not in (None, ""))
                golden_value = vote.most_common(1)[0][0] if vote else ""

            conflicts.append({
                "key":     key,
                "field":   norm_col,
                "values":  values,


                "conflict_type": conflict_type,
                "golden_value": golden_value,
                "sources_agree": [n for n in source_names
                    if norm_vals.get(n, "").upper() == golden_value.upper()],
                "sources_differ": [n for n in source_names
                    if norm_vals.get(n, "") not in ("", None)
                    and norm_vals.get(n, "").upper() != golden_value.upper()],
            })

    if conflicts_only:
        # Include only keys with at least one conflict
        conflict_keys = {c["key"] for c in conflicts}
        for name in list(only_in):
            only_in[name] = [k for k in only_in[name] if k in conflict_keys]

    # ---- Step 7: Source-level statistics -------------------------------------
    source_stats: dict[str, dict] = {}
    total_conflicts = len(set((c["key"], c["field"]) for c in conflicts))
    for name in source_names:


        n_keys  = len(keys_per_source[name])
        n_match = len(keys_in_all & keys_per_source[name])
        n_conf  = sum(1 for c in conflicts if name in c.get("sources_differ", []))
        cov_pct = round(n_match / max(len(all_keys), 1) * 100, 1)
        source_stats[name] = {
            "rows":   n_keys,
            "matched": n_match,
            "exclusive": len(only_in.get(name, [])),
            "conflicts": n_conf,
            "coverage_pct": cov_pct,
        }

    total_matched = len(keys_in_all)
    overall_conflict_pct = round(len(conflicts) / max(total_matched *
        len(common_norm_fields), 1) * 100, 1)

    return {
        "summary": {
            "total_keys":   len(all_keys),


            "matched_in_all":  total_matched,
            "sources":     source_names,
            "source_count":   len(sources),
            "conflict_count":  len(conflicts),
            "conflict_field_pairs": total_conflicts,
            "conflict_pct":   overall_conflict_pct,
            "compare_fields":  common_norm_fields,
        },
        "key_col_used":  key_col or "(auto-detected)",
        "key_col_per_source": key_per_source,
        "compare_fields": common_norm_fields,
        "coverage_matrix": coverage_matrix,
        "conflicts":   conflicts[:1000],  # cap for display
        "only_in":    only_in,
        "golden_source": golden_source or "(majority-wins)",
        "source_stats": source_stats,
    }


def _apply_cross_col_rules(df: pd.DataFrame, rules: list[dict]) -> list[dict]:
    # Feature 1 -- Conditional / Cross-column Rules.
    # Feature 5 -- Settlement date T+2 validation.
    #
    # Evaluates rules that span multiple columns:
    #   conditional: if col_a == val_a then col_b must satisfy condition
    #   settlement_date_t2: settle_col must be >= trade_col + 2 business days
    #   referential_integrity: col_a values must all appear in col_b
    #
    # Returns list of result dicts compatible with rule_results format.
    import numpy as _np
    results = []
    total = len(df)
    if total == 0:
        return results

    for rule in rules:
        rt     = rule.get("rule_type", "")
        severity = rule.get("severity", "major")
        try:
            if rt == "conditional":
                # if when_col == when_val then then_col must satisfy then_condition
                when_col  = rule.get("when_col", "")
                when_val  = str(rule.get("when_val", ""))
                then_col  = rule.get("then_col", "")
                then_cond = rule.get("then_condition", "not_null")
                then_val  = rule.get("then_val", "")

                if when_col not in df.columns or then_col not in df.columns:
                    results.append({"skipped": True, "reason": f"Column not found: {when_col} or {then_col}",
                                     "rule_name": rule.get("description", rt), "rule_type": rt})
                    continue

                # Rows where condition is met
                when_mask = df[when_col].astype(str).str.strip().str.lower() == when_val.lower()

                subset   = df.loc[when_mask, then_col]
                sub_total = int(when_mask.sum())

                if then_cond == "not_null":
                    fail_mask = subset.isna() | (subset.astype(str).str.strip() == "")
                elif then_cond == "positive":
                    fail_mask = pd.to_numeric(subset, errors="coerce").fillna(-1) <= 0
                elif then_cond == "equals":
                    fail_mask = subset.astype(str).str.strip() != str(then_val)
                elif then_cond == "not_equals":
                    fail_mask = subset.astype(str).str.strip() == str(then_val)
                elif then_cond in ("gt", "greater_than"):
                    fail_mask = pd.to_numeric(subset, errors="coerce") <= float(then_val)
                elif then_cond in ("lt", "less_than"):
                    fail_mask = pd.to_numeric(subset, errors="coerce") >= float(then_val)
                else:
                    fail_mask = pd.Series(False, index=subset.index)

                failed = int(fail_mask.sum())

                passed = sub_total - failed
                examples = subset[fail_mask].head(5).astype(str).tolist()
                results.append({
                    "skipped": False, "rule_type": rt, "severity": severity,
                    "rule_name": rule.get("description", f"if {when_col}={when_val} then {then_col}{then_cond}"),
                    "column_name": then_col,
                    "total": sub_total, "passed": passed, "failed": failed,
                    "pass_pct": round(passed / sub_total * 100, 1) if sub_total else 100,
                    "failing_examples": examples,
                    "status": "PASS" if failed == 0 else ("WARN" if sub_total > 0 and failed / sub_total < 0.05 else "FAIL"),
                })

            elif rt == "settlement_date_t2":
                # settlement_col must be >= trade_col + 2 business days
                trade_col  = rule.get("trade_col",   "trade_date")
                settle_col = rule.get("settlement_col", "settlement_date")
                min_days  = int(rule.get("min_business_days", 2))

                if trade_col not in df.columns or settle_col not in df.columns:
                    results.append({"skipped": True, "reason": f"Columns not found: {trade_col}, {settle_col}",
                                     "rule_name": f"T+{min_days} settlement", "rule_type": rt})
                    continue

                trade_dt = pd.to_datetime(df[trade_col], errors="coerce")
                settle_dt = pd.to_datetime(df[settle_col], errors="coerce")
                both_valid = trade_dt.notna() & settle_dt.notna()

                # Calendar days proxy for business days (approx: 2 biz days ~= 2-4 cal days)
                # Use numpy busday_count for accurate business day count

                def _bus_days(t, s):
                    try:
                        return int(_np.busday_count(t.date(), s.date()))
                    except Exception:
                        return -1

                fail_indices = []

                for idx in df.index[both_valid]:
                    bd = _bus_days(trade_dt[idx], settle_dt[idx])
                    if bd < min_days:
                        fail_indices.append(idx)

                failed = len(fail_indices)
                sub_total = int(both_valid.sum())
                passed = sub_total - failed
                examples = [f"Trade:{trade_dt[i].date()} Settle:{settle_dt[i].date()}" for i in
                            fail_indices[:5]]
                results.append({
                    "skipped": False, "rule_type": rt, "severity": severity,
                    "rule_name": rule.get("description", f"Settlement must be T+{min_days} business days"),
                    "column_name": settle_col,
                    "total": sub_total, "passed": passed, "failed": failed,
                    "pass_pct": round(passed / sub_total * 100, 1) if sub_total else 100,
                    "failing_examples": examples,
                    "status": "PASS" if failed == 0 else ("WARN" if sub_total > 0 and failed /
                                                            sub_total < 0.05 else "FAIL"),
                })

            elif rt == "referential_integrity":
                # Feature 4 -- all values in col_a must appear in col_b (same DataFrame)
                col_a = rule.get("col_a", "")
                col_b = rule.get("col_b", "")
                if col_a not in df.columns or col_b not in df.columns:
                    results.append({"skipped": True, "reason": f"Columns not found: {col_a}, {col_b}",
                                     "rule_name": f"{col_a} ref {col_b}", "rule_type": rt})
                    continue
                ref_set = set(df[col_b].dropna().astype(str).str.strip().str.lower())
                col_vals = df[col_a].dropna().astype(str).str.strip().str.lower()
                mask   = ~col_vals.isin(ref_set)
                failed  = int(mask.sum())
                passed  = len(col_vals) - failed
                examples = df.loc[col_vals.index[mask], col_a].head(5).astype(str).tolist()
                results.append({
                    "skipped": False, "rule_type": rt, "severity": severity,
                    "rule_name": rule.get("description", f"{col_a} values must exist in {col_b}"),
                    "column_name": col_a,
                    "total": len(col_vals), "passed": passed, "failed": failed,
                    "pass_pct": round(passed / len(col_vals) * 100, 1) if col_vals.size else 100,
                    "failing_examples": examples,
                    "status": "PASS" if failed == 0 else ("WARN" if len(col_vals) > 0 and failed /
                                                           len(col_vals) < 0.05 else "FAIL"),
                })

        except Exception as exc:
            results.append({"skipped": True, "reason": str(exc),
                             "rule_name": rule.get("description", rt), "rule_type": rt})
    return results


def _dq_score(total_rows: int, dup_rows: int, cols: list[dict],
        rule_results: list[dict], consistency_issues: list[dict] | None = None,
        accuracy: float | None = None) -> dict:
    # Compute a 0-100 DQ score across 8 dimensions with dynamic weights.

    # Timeliness, Precision, and Accuracy are opt-in:
    # - Timeliness only activates when at least one column has a user-configured
    # freshness threshold (col_config timeliness_days or timeliness_hints).
    # Marker: col["freshness_user_configured"] = True set by analyze_quality.
    # - Precision only activates when at least one numeric column shows actual
    # decimal-place inconsistency (dp_range > 0), meaning there is something
    # real to measure.
    # - Accuracy activates when accuracy_ref_values are set in col_config for at
    # least one column, enabling spot-checking against known-good reference values.

    # When a dimension is inactive its weight is redistributed to Completeness,
    # which is always the most reliable automatic signal.

    # Base weights (active):
    # Completeness 20% - Uniqueness 12% - Validity 18% - Consistency 12%
    # Conformity 8% - Precision 8% - Timeliness 5% - Accuracy 17%

    # ---- Completeness (base 25%) ----------------------------------------
    scorable = [c for c in cols if not c.get("hint_nullable")]
    completeness = (
        sum(100 - c["null_pct"] for c in scorable) / len(scorable) if scorable else 100.0
    )

    # ---- Uniqueness (15%) ------------------------------------------------
    uniqueness = 100.0 - (dup_rows / total_rows * 100 if total_rows else 0)

    # ---- Validity (20%) -- severity-weighted ------------------------------
    non_skipped = [r for r in rule_results if not r.get("skipped")]
    _SEV_WEIGHT = {"critical": 3.0, "major": 2.0, "minor": 1.0}
    if non_skipped:
        weighted_total = sum(r.get("total", 0) * _SEV_WEIGHT.get(r.get("severity", "major"), 2.0) for r in non_skipped)
        weighted_passed = sum(r.get("passed", 0) * _SEV_WEIGHT.get(r.get("severity", "major"), 2.0) for r in non_skipped)
        validity = weighted_passed / weighted_total * 100 if weighted_total else 100.0

    else:
        validity = 100.0

    # ---- Severity breakdown ------------------------------------------------
    severity_breakdown = {
        "critical_fails": sum(1 for r in rule_results if r.get("status") == "FAIL" and r.get("severity") == "critical"),
        "major_fails":  sum(1 for r in rule_results if r.get("status") == "FAIL" and r.get("severity") == "major"),
        "minor_fails":  sum(1 for r in rule_results if r.get("status") == "FAIL" and r.get("severity") == "minor"),
    }

    # ---- Consistency (15%) ------------------------------------------------
    if consistency_issues:
        total_cons  = len(consistency_issues)
        failed_cons = sum(1 for c in consistency_issues if c.get("status") == "FAIL")
        consistency = max(0.0, 100.0 - (failed_cons / total_cons * 100))
    else:
        consistency = 100.0

    # ---- Conformity (10%) ------------------------------------------------
    # Outlier penalty is suppressed for high-skew columns (|skewness| > 2).
    # Prices, volumes, notionals are naturally right-skewed -- IQR will always
    # flag extreme values as "outliers" even when they are legitimate.
    # Empty-string penalty applies to all string columns regardless of skew.
    if cols:
        conf_scores = []
        for c in cols:
            skewness = abs(c.get("skewness", 0) or 0)
            outlier_pct = 0.0 if skewness > 2 else (c.get("outlier_pct", 0) or 0)
            empty_pct  = (c.get("empty_str_count", 0) or 0) / total_rows * 100 if total_rows else 0
            conf_scores.append(max(0.0, 100 - outlier_pct - empty_pct))
        conformity = sum(conf_scores) / len(conf_scores)
    else:
        conformity = 100.0

    # ---- Precision (10%) -- opt-in -----------------------------------------
    # Only active when at least one column has actual decimal inconsistency
    # (dp_range > 0) or a user-declared decimal_places constraint exists.
    # When inactive, its 10% weight moves to Completeness.
    prec_active = any(
        (c.get("dp_range") or 0) > 0 or c.get("precision_user_configured")
        for c in cols
    )
    if prec_active and cols:
        prec_scores = []
        for c in cols:
            if c.get("dp_range") is not None:
                prec_scores.append(max(50.0, 100.0 - c["dp_range"] * 5))
            elif c.get("avg_length") is not None:
                mn = c.get("min_length", 0) or 0
                mx = c.get("max_length", 0) or 0
                avg = c.get("avg_length", mx) or mx
                if avg <= 20 and c.get("cardinality") in ("low (enum-like)", "medium"):
                    prec_scores.append(max(0.0, 100.0 - min((mx - mn) * 5, 100)))

                else:
                    prec_scores.append(100.0)
            else:
                prec_scores.append(100.0)
        precision = sum(prec_scores) / len(prec_scores)
    else:
        precision = None  # inactive -- weight redistributed

    # ---- Timeliness (5%) -- opt-in -----------------------------------------
    # Only active when the user has explicitly set freshness_days on at least
    # one column via the config table or timeliness_hints. Auto-detected date
    # columns do NOT activate this -- that would silently penalise historical data.
    user_configured_date_cols = [c for c in cols if c.get("freshness_user_configured")]
    if user_configured_date_cols:
        tim_scores = []
        for c in user_configured_date_cols:
            fd = c.get("freshness_days") or 0
            threshold = c.get("freshness_threshold_days", 1)
            if fd <= threshold:

                tim_scores.append(100.0)
            elif fd <= threshold * 30:
                tim_scores.append(max(0.0, 100.0 - (fd - threshold) / (threshold * 29) * 20))
            else:
                tim_scores.append(max(0.0, 80.0 - (fd - threshold * 30) / (threshold * 335) * 80))
        timeliness = sum(tim_scores) / len(tim_scores)
    else:
        timeliness = None  # inactive -- weight redistributed

    # ---- Dynamic weight allocation ------------------------------------------
    # Inactive dimensions (None) redistribute their weight to Completeness.
    w_completeness = 0.20
    w_uniqueness  = 0.12
    w_validity   = 0.18
    w_consistency = 0.12
    w_conformity  = 0.08
    w_precision  = 0.08
    w_timeliness  = 0.05

    w_accuracy   = 0.17

    if precision is None:
        w_completeness += w_precision
        w_precision = 0.0
    if timeliness is None:
        w_completeness += w_timeliness
        w_timeliness = 0.0
    if accuracy is None:
        w_completeness += w_accuracy
        w_accuracy = 0.0

    score = (
        completeness   * w_completeness +
        uniqueness    * w_uniqueness +
        validity     * w_validity +
        consistency   * w_consistency +
        conformity    * w_conformity +
        (precision or 100.0) * w_precision +

        (timeliness or 100.0) * w_timeliness +
        (accuracy or 100.0) * w_accuracy
    )

    score = round(min(100.0, max(0.0, score)), 1)
    grade = "A" if score >= 90 else "B" if score >= 75 else "C" if score >= 60 else "D" if score >= 40 else "F"

    return {
        "score":     score,
        "grade":     grade,
        "completeness":  round(completeness, 1),
        "uniqueness":   round(uniqueness, 1),
        "validity":    round(validity, 1),
        "consistency":  round(consistency, 1),
        "conformity":   round(conformity, 1),
        "precision":   round(precision, 1) if precision is not None else None,
        "timeliness":   round(timeliness, 1) if timeliness is not None else None,
        "accuracy":    round(accuracy, 1) if accuracy is not None else None,
        "precision_active": precision is not None,


        "timeliness_active": timeliness is not None,
        "accuracy_active":  accuracy is not None,
        "severity_breakdown": severity_breakdown,
        "weights": {
            "completeness": round(w_completeness, 2),
            "uniqueness":  round(w_uniqueness, 2),
            "validity":   round(w_validity, 2),
            "consistency": round(w_consistency, 2),
            "conformity":  round(w_conformity, 2),
            "precision":  round(w_precision, 2),
            "timeliness":  round(w_timeliness, 2),
            "accuracy":   round(w_accuracy, 2),
        },
    }


def _detect_drift(cols: list[dict], baseline: dict | None) -> list[dict]:
    # Compare current column stats against a saved baseline snapshot.
    # Returns list of drift alerts: {column, metric, current, baseline, delta, severity, detail}

    if not baseline:
        return []
    alerts = []
    baseline_cols = {c["name"]: c for c in baseline.get("columns", [])}
    for c in cols:
        name = c["name"]
        b = baseline_cols.get(name)
        if not b:
            continue
        checks = [
            ("null_pct",    "Null %",    5.0, "WARN", 10.0, "FAIL"),
            ("uniqueness_pct", "Uniqueness %", 5.0, "WARN", 15.0, "FAIL"),
            ("outlier_pct",   "Outlier %",   3.0, "WARN", 8.0, "FAIL"),
        ]
        for metric, label, warn_thresh, warn_sev, fail_thresh, fail_sev in checks:
            cur = c.get(metric)
            bas = b.get(metric)
            if cur is None or bas is None:

                continue
            delta = abs(cur - bas)
            if delta >= fail_thresh:
                alerts.append({
                    "column": name, "metric": label,
                    "current": cur, "baseline": bas,
                    "delta": round(delta, 1),
                    "severity": fail_sev,
                    "detail": f"{label}: {bas}% -> {cur}% (delta {delta:.1f}pp)",
                })
            elif delta >= warn_thresh:
                alerts.append({
                    "column": name, "metric": label,
                    "current": cur, "baseline": bas,
                    "delta": round(delta, 1),
                    "severity": warn_sev,
                    "detail": f"{label}: {bas}% -> {cur}% (delta {delta:.1f}pp)",
                })
        # Numeric mean drift

        if c.get("mean") is not None and b.get("mean") is not None:
            cur_m, bas_m = c["mean"], b["mean"]
            if bas_m != 0:
                delta_pct = abs(cur_m - bas_m) / abs(bas_m) * 100
                if delta_pct >= 20:
                    sev = "FAIL" if delta_pct >= 50 else "WARN"
                    alerts.append({
                        "column": name, "metric": "Mean",
                        "current": cur_m, "baseline": bas_m,
                        "delta": round(delta_pct, 1),
                        "severity": sev,
                        "detail": f"Mean: {bas_m} -> {cur_m} ({delta_pct:.1f}% change)",
                    })
    return alerts


def _dq_schema_fingerprint(df: pd.DataFrame) -> str:
    # Stable fingerprint from sorted column names + dtypes for DQ rule persistence.
    import hashlib


    sig = "|".join(f"{c}:{str(df[c].dtype)}" for c in sorted(df.columns))
    return hashlib.md5(sig.encode()).hexdigest()[:12]


def _detect_schema_drift(df: pd.DataFrame, cols: list[dict], baseline_snapshot: dict | None) -> list[dict]:
    # Compare current schema against saved baseline snapshot.
    if not baseline_snapshot:
        return []
    changes = []
    baseline_cols = {c["name"]: c for c in baseline_snapshot.get("columns", [])}
    current_cols = {c["name"]: c for c in cols}
    for col in set(current_cols) - set(baseline_cols):
        changes.append({"type": "added", "column": col, "severity": "warn", "detail": "New column added"})
    for col in set(baseline_cols) - set(current_cols):
        changes.append({"type": "removed", "column": col, "severity": "fail", "detail": "Column removed from baseline"})
    for col in set(current_cols) & set(baseline_cols):
        bas = baseline_cols[col]; cur = current_cols[col]


        if bas.get("cardinality") and cur.get("cardinality") and bas["cardinality"] != cur["cardinality"]:
            changes.append({"type": "cardinality_change", "column": col, "severity": "warn",
                "detail": f"Cardinality: {bas['cardinality']} -> {cur['cardinality']}"})
        # Row count change > 50%
        bas_rows = baseline_snapshot.get("total_rows", 0)
        cur_rows = len(df)
        if bas_rows and abs(cur_rows - bas_rows) / bas_rows > 0.5:
            if col == cols[0]["name"]:  # Only add once
                changes.append({"type": "row_count_change", "column": "(dataset)", "severity": "warn",
                    "detail": f"Row count: {bas_rows} -> {cur_rows} ({abs(cur_rows - bas_rows) / bas_rows * 100:.0f}% change)"})
    return changes


def _detect_duplicates(df: pd.DataFrame, cols: list[dict], hints: dict) -> dict:
    # Enhanced duplicate detection beyond exact full-row duplicates.
    # Returns dict with keys: exact_count, exact_pct, subset_key_groups,
    # near_duplicate_count, near_duplicate_examples.

    total = len(df)
    result = {
        "exact_count": 0,
        "exact_pct": 0.0,
        "subset_key_groups": [],
        "near_duplicate_count": 0,
        "near_duplicate_examples": [],
    }
    if total == 0:
        return result

    # 1. Exact full-row duplicates -- cap at 50k rows to keep hashing fast
    _DUP_CAP = 50_000
    _df_sample = df if total <= _DUP_CAP else df.iloc[:_DUP_CAP]
    exact_sample = int(_df_sample.duplicated().sum())
    # Extrapolate to full dataset if sampled
    exact = int(exact_sample * total / len(_df_sample)) if total > _DUP_CAP else exact_sample

    result["exact_count"] = exact
    result["exact_pct"] = round(exact / total * 100, 1)
    if total > _DUP_CAP:
        result["exact_note"] = f"Estimated from {_DUP_CAP:,}-row sample"

    # 2. Subset-key duplicates: same identifier-ish column(s), different value in another col
    id_cols = [c["name"] for c in cols if c.get("cardinality") == "identifier-like" and c["name"] in df.columns][:3]
    non_id_cols = [c["name"] for c in cols if c.get("cardinality") != "identifier-like" and c["name"] in df.columns][:5]

    for key_col in id_cols:
        # Use value_counts -- much faster than duplicated() for large frames
        vc = df[key_col].value_counts()
        dup_keys = vc[vc > 1]
        if len(dup_keys) > 0:
            dup_key_count = int(dup_keys.sum())
            distinct_dup_keys = len(dup_keys)
            key_dups = df[df[key_col].isin(dup_keys.index.tolist()[:50])]


            result["subset_key_groups"].append({
                "key_column": key_col,
                "duplicate_row_count": dup_key_count,
                "distinct_key_values_with_dups": distinct_dup_keys,
                "examples": key_dups[[key_col] + non_id_cols[:3]].head(5).to_dict("records"),
            })

    # 3. Near-duplicate detection: same key + timestamp within N minutes (trade-level)
    date_cols = [c["name"] for c in cols if c.get("freshness_days") is not None and c["name"] in df.columns][:2]
    near_dup_threshold_minutes = int(hints.get("near_dup_threshold_minutes", 5))

    if id_cols and date_cols:
        key_col = id_cols[0]
        ts_col = date_cols[0]
        try:
            tmp = df[[key_col, ts_col]].copy()
            tmp[ts_col] = pd.to_datetime(tmp[ts_col], errors="coerce")
            tmp = tmp.dropna(subset=[key_col, ts_col])

            if len(tmp) > 1:
                tmp_sorted = tmp.sort_values([key_col, ts_col])
                tmp_sorted["_ts_diff"] = tmp_sorted.groupby(key_col)[ts_col].diff().dt.total_seconds() / 60
                near_dups = tmp_sorted[
                    tmp_sorted["_ts_diff"].notna() &
                    (tmp_sorted["_ts_diff"] >= 0) &
                    (tmp_sorted["_ts_diff"] <= near_dup_threshold_minutes)
                ]
                result["near_duplicate_count"] = len(near_dups)
                result["near_duplicate_examples"] = near_dups.drop(columns=["_ts_diff"]).head(5).astype(str).to_dict("records")
                result["near_dup_key_col"] = key_col
                result["near_dup_ts_col"] = ts_col
                result["near_dup_threshold_minutes"] = near_dup_threshold_minutes
        except Exception:
            pass

    return result


def analyze_quality(df: pd.DataFrame, name: str,
                     data_dict: dict | None = None,
                     rules: list[dict] | None = None,
                     user_hints: dict | None = None) -> dict:
    hints = user_hints or {}
    total = len(df)
    data_dict = data_dict or {}
    rules    = list(rules or [])

    # ---- Per-column config from the interactive config table ----------------
    # Each dict: {name, mandatory, exclude, null_threshold_pct, min_val, max_val,
    #     decimal_places, timeliness_days, allow_dups, force_unique, allowed_values,
    #     pattern, rule_type}
    col_config_list: list[dict] = list(hints.get("col_config", []))

    # ---- Inject BFSI Rule Pack validators into col_config --------------------
    # bfsi_validators is a list of "rule_type:column_name" strings e.g.
    # ["isin_format:isin", "positive:notional", "email_format:contact_email"]


    # For each entry, inject a col_config rule for that column if it exists in df.
    _bfsi_v_list = hints.get("bfsi_validators", [])
    if _bfsi_v_list:
        _existing_rules = {(c.get("name", ""), c.get("rule_type", ""), c.get("value", "")) for c in col_config_list}
        for _bv in _bfsi_v_list:
            if ":" not in _bv:
                continue
            _bv_rule, _bv_col = _bv.split(":", 1)
            _bv_col_lower = _bv_col.strip().lower()
            _bv_rule = _bv_rule.strip()

            # domain_accuracy_{domain_name} -> rule_type=domain_accuracy, value=domain_name
            _rule_type = _bv_rule
            _rule_value = ""
            if _bv_rule.startswith("domain_accuracy_"):
                _rule_type = "domain_accuracy"
                _rule_value = _bv_rule[len("domain_accuracy_"):]


            # Match case-insensitively against actual df columns
            _matched = next((c for c in df.columns if c.lower() == _bv_col_lower), None)
            if _matched and (_matched, _rule_type, _rule_value) not in _existing_rules:
                _entry = {"name": _matched, "rule_type": _rule_type, "severity": "major"}
                if _rule_value:
                    _entry["value"] = _rule_value
                col_config_list.append(_entry)
                _existing_rules.add((_matched, _rule_type, _rule_value))

    col_config_map: dict[str, dict] = {c["name"]: c for c in col_config_list if c.get("name")}

    # Columns the user wants to exclude from DQ entirely
    _config_excluded = {c["name"] for c in col_config_list if c.get("exclude")}

    # Columns the user marked as non-mandatory (nullable) -- suppresses completeness penalty
    _config_nullable = {c["name"] for c in col_config_list if not c.get("mandatory", True)}

    # Nullable columns named in hints are exempted from NOT-NULL checks and


    # have their null_pct excluded from completeness scoring.
    # Must be defined before the col_config injection loop which references it.
    _hint_nullable = (
        {c.strip().lower() for c in hints.get("nullable_hints", "").split(",") if c.strip()}
        | {c.lower() for c in _config_nullable}
    )

    # Columns where the user explicitly allows duplicates -- suppresses auto unique rule
    # force_unique=True (Unique checkbox in UI) overrides allow_dups
    _config_allow_dups = {c["name"] for c in col_config_list if c.get("allow_dups") and not c.get("force_unique")}
    _config_force_unique = {c["name"] for c in col_config_list if c.get("force_unique")}

    # Accuracy scores dict -- populated during col_config processing below
    accuracy_scores: dict[str, float] = {}

    # Inject rules from config table into the rules list
    for cc in col_config_list:
        cname = cc.get("name", "")


        if not cname or cc.get("exclude") or cname not in df.columns:
            continue
        rt = cc.get("rule_type", "auto")

        # Severity for this column config entry (user can override via col_config)
        _cc_severity = cc.get("severity", "major")

        # Mandatory checkbox -> not_null rule (user explicitly declared)
        if cc.get("mandatory") and cname.lower() not in _hint_nullable:
            rules.append({
                "rule_type":  "not_null",
                "column_name":  cname,
                # OCR-UNCERTAIN: source shows a checkbox glyph (checked box icon) inline here, rendered below as [x]
                "description":  "Not-null (Mandatory [x] in config table)",
                "severity":  _cc_severity,
                "_hint_injected": True,
            })

        # Unique checkbox -> unique rule (user explicitly declared)
        if cc.get("force_unique"):


            rules.append({
                "rule_type":  "unique",
                "column_name":  cname,
                # OCR-UNCERTAIN: source shows a checkbox glyph (checked box icon) inline here, rendered below as [x]
                "description":  "Unique (Unique [x] in config table)",
                "severity":  _cc_severity,
                "_hint_injected": True,
            })

        # Null threshold -> completeness_pct rule
        thresh = cc.get("null_threshold_pct", "")
        if thresh != "" and str(thresh).strip() != "":
            try:
                max_null = float(thresh)
                required_pct = 100 - max_null
                rules.append({
                    "rule_type": "completeness_pct",
                    "column_name": cname,
                    "value": required_pct,
                    "description": f"Completeness >= {required_pct:.0f}% (config table)",


                    "severity":  _cc_severity,
                    "_hint_injected": True,
                })
            except ValueError:
                pass

        # Numeric range from min/max
        min_v = cc.get("min_val", "")
        max_v = cc.get("max_val", "")
        if min_v != "" and max_v != "" and str(min_v).strip() and str(max_v).strip():
            try:
                rules.append({
                    "rule_type": "range",
                    "column_name": cname,
                    "min_value": float(min_v),
                    "max_value": float(max_v),
                    "description": f"Range {min_v}-{max_v} (config table)",
                    "severity":  _cc_severity,
                    "_hint_injected": True,


                })
            except ValueError:
                pass

        # Decimal places
        dp = cc.get("decimal_places", "")
        if dp != "" and str(dp).strip():
            try:
                rules.append({
                    "rule_type": "decimal_places",
                    "column_name": cname,
                    "value": int(dp),
                    "description": f"Max {dp} decimal places (config table)",
                    "severity":  _cc_severity,
                    "_hint_injected": True,
                })
            except ValueError:
                pass


        # Timeliness / freshness
        td = cc.get("timeliness_days", "")
        if td != "" and str(td).strip():
            try:
                rules.append({
                    "rule_type": "freshness_days",
                    "column_name": cname,
                    "value": int(td),
                    "description": f"Freshness <= {td} days (config table)",
                    "severity":  _cc_severity,
                    "_hint_injected": True,
                })
            except ValueError:
                pass

        # Allowed values -- user-declared enum list (e.g. "BUY,SELL,SHORT")
        av = str(cc.get("allowed_values", "") or "").strip()
        if av:
            rules.append({


                "rule_type": "allowed_values",
                "column_name": cname,
                "value": av,
                "description": f"Allowed values: {av} (config table)",
                "severity":  _cc_severity,
                "_hint_injected": True,
            })

        # Pattern -- user-declared regex every value must match
        pat = str(cc.get("pattern", "") or "").strip()
        if pat:
            rules.append({
                "rule_type": "pattern",
                "column_name": cname,
                "value": pat,
                "description": f"Pattern: {pat} (config table)",
                "severity":  _cc_severity,
                "_hint_injected": True,
            })


        # Accuracy reference values -- "expected" lookup for spot-checking
        # Format: comma-separated known-good values. % matching = accuracy score
        # for that column.
        av_ref = str(cc.get("accuracy_ref_values", "") or "").strip()
        if av_ref:
            ref_set = {v.strip() for v in av_ref.split(",") if v.strip()}
            col_series = df[cname]
            non_null = col_series.dropna()
            if len(non_null):
                match_n = int(non_null.astype(str).str.strip().isin(ref_set).sum())
                accuracy_scores[cname] = match_n / len(non_null) * 100
                rules.append({
                    "rule_type": "allowed_values",
                    "column_name": cname,
                    "value": av_ref,
                    "description": f"Accuracy reference check: {av_ref[:60]} (config table)",
                    "severity":  _cc_severity,
                    "_hint_injected": True,


                })

        # Explicit rule type from the Rule dropdown (unique, not_null, positive, etc.)
        # Also handles domain_accuracy, address_complete and other new rule types.
        # Skip types already handled above to avoid double-counting.
        _SKIP_RT = {"auto", "range", "freshness_days", "completeness_pct", "decimal_places"}
        if rt and rt not in _SKIP_RT:
            rules.append({
                "rule_type":  rt,
                "column_name":  cname,
                "value":     cc.get("value", ""),  # pass domain name for domain_accuracy
                "description": f"{rt} (config table)",
                "severity":  _cc_severity,
                "_hint_injected": True,
            })

    # Excluded columns from config table -- drop them from the dataframe before analy-
    # sis
    if _config_excluded:


        df = df.drop(columns=[c for c in _config_excluded if c in df.columns])

    # Range hints: parse "col 0-100, price 0-9999" into synthetic range rules.
    for part in hints.get("range_hints", "").split(","):
        part = part.strip()
        m = re.match(r'^(\S+)\s+([\d.]+)\s*[-]\s*([\d.]+)$', part)
        if m:
            col_h, lo, hi = m.group(1), float(m.group(2)), float(m.group(3))
            if col_h in df.columns:
                rules.append({
                    "rule_type": "range",
                    "column_name": col_h,
                    "min_value": lo,
                    "max_value": hi,
                    "description": f"Range {lo}-{hi} (user-supplied hint)",
                    "_hint_injected": True,
                })

    # Timeliness hints: parse "col_name N" -> freshness_days rule (max N days old).


    # Format: "trade_date 1, settlement_date 3"
    for part in hints.get("timeliness_hints", "").split(","):
        part = part.strip()
        m = re.match(r'^(\S+)\s+(\d+)$', part)
        if m:
            col_h, days = m.group(1), int(m.group(2))
            if col_h in df.columns:
                rules.append({
                    "rule_type": "freshness_days",
                    "column_name": col_h,
                    "value": days,
                    "description": f"Must be within {days} day(s) (timeliness hint)",
                    "_hint_injected": True,
                })

    # Precision hints: "col N" -> values must have at most N decimal places.
    # Format: "price 4, rate 6"
    for part in hints.get("precision_hints", "").split(","):
        part = part.strip()


        m = re.match(r'^(\S+)\s+(\d+)$', part)
        if m:
            col_h, decimals = m.group(1), int(m.group(2))
            if col_h in df.columns:
                rules.append({
                    "rule_type": "decimal_places",
                    "column_name": col_h,
                    "value": decimals,
                    "description": f"Max {decimals} decimal place(s) (precision hint)",
                    "_hint_injected": True,
                })

    # ---- Track which columns have user-configured freshness / precision ----------
    # Used by _dq_score to decide whether Timeliness and Precision are active.
    # (accuracy_scores already initialized before the col_config loop above)
    _user_freshness_cols = set()
    _user_precision_cols = set()

    # From config table


    for cc in col_config_list:
        cname = cc.get("name", "")
        if not cname or cc.get("exclude"):
            continue
        if str(cc.get("timeliness_days", "") or "").strip():
            _user_freshness_cols.add(cname)
        if str(cc.get("decimal_places", "") or "").strip():
            _user_precision_cols.add(cname)

    # From text hints
    for part in hints.get("timeliness_hints", "").split(","):
        m = re.match(r'^(\S+)\s+(\d+)$', part.strip())
        if m and m.group(1) in df.columns:
            _user_freshness_cols.add(m.group(1))
    for part in hints.get("precision_hints", "").split(","):
        m = re.match(r'^(\S+)\s+(\d+)$', part.strip())
        if m and m.group(1) in df.columns:
            _user_precision_cols.add(m.group(1))


    # ---- Column-level profiling ----
    cols = []
    for col in df.columns:
        s = df[col]
        null_n = int(s.isna().sum())
        # Cache str conversion once -- reused throughout this column's profiling block
        _s_str = s.astype(str) if s.dtype == object else None
        _s_str_stripped = _s_str.str.strip() if _s_str is not None else None
        empty_n = int((_s_str_stripped == "").sum()) - null_n if s.dtype == object else 0
        # Treat empty strings as nulls for completeness scoring (BFSI standard).
        # A column of 100 "" values should NOT score 100% complete.
        effective_null_n = null_n + max(0, empty_n)
        uniq_n = int(s.nunique(dropna=True))
        dd = data_dict.get(col, {})

        # Top value frequencies
        top_vals = (
            s.value_counts(dropna=True)
            .head(5)


            .reset_index()
            .rename(columns={"count": "count", col: "value", "index": "value"})
            .to_dict("records")
        )

        _is_hint_nullable = col.lower() in _hint_nullable
        info = {
            "name": col,
            "dtype": str(s.dtype),
            "null_count":    effective_null_n,  # includes empty strings
            "null_pct":      round(effective_null_n / total * 100, 1) if total else 0,
            "raw_null_count": null_n,          # original NaN count only
            "empty_str_count": max(0, empty_n),
            "unique_count": uniq_n,
            "uniqueness_pct": round(uniq_n / total * 100, 1) if total else 0,
            "top_values": top_vals,
            # From data dictionary
            "dd_description":   dd.get("description", ""),
            "dd_owner":      dd.get("owner", ""),


            "dd_sensitivity":   dd.get("sensitivity", ""),
            "dd_nullable":      dd.get("nullable", True),
            "dd_is_pk":     dd.get("is_pk", False),
            "dd_business_term": dd.get("business_term", ""),
            # Flag columns not documented in the data dictionary
            "missing_from_dict": col not in data_dict and bool(data_dict),
            # Hint-declared nullable -- suppresses completeness penalty
            "hint_nullable":    _is_hint_nullable,
        }

        _numeric_s = s
        _is_numeric_col = pd.api.types.is_numeric_dtype(s)
        if not _is_numeric_col and not pd.api.types.is_datetime64_any_dtype(s):
            # _load_file() loads every column as string (often pandas' modern
            # StringDtype, which is NOT == object) -- coerce here so
            # numeric-looking columns (the common case for CSV/TXT/Excel
            # uploads) still get real min/max/outlier stats.
            _coerced = pd.to_numeric(s, errors="coerce")
            _non_null = s.notna().sum()
            if _non_null and _coerced.notna().sum() / _non_null >= 0.9:
                _is_numeric_col = True
                _numeric_s = _coerced

        if _is_numeric_col:
            clean = _numeric_s.dropna()
            if len(clean):
                q1, q3 = float(clean.quantile(0.25)), float(clean.quantile(0.75))
                iqr = q3 - q1
                outliers = int(((clean < q1 - 1.5 * iqr) | (clean > q3 + 1.5 * iqr)).sum())
                info.update({
                    "min":      round(float(clean.min()), 4),
                    "max":      round(float(clean.max()), 4),


                    "mean":     round(float(clean.mean()), 4),
                    "median":   round(float(clean.median()), 4),
                    "std":      round(float(clean.std()), 4),
                    "variance": round(float(clean.var()), 4),
                    "skewness": round(float(clean.skew()), 4),
                    "kurtosis": round(float(clean.kurt()), 4),
                    "p25":      round(float(q1), 4),
                    "p75":      round(float(q3), 4),
                    "p95":      round(float(clean.quantile(0.95)), 4),
                    "p99":      round(float(clean.quantile(0.99)), 4),
                    "outlier_count": outliers,
                    "outlier_pct": round(outliers / total * 100, 1) if total else 0,
                    "zero_count":  int((clean == 0).sum()),
                    "negative_count": int((clean < 0).sum()),
                })
                # Decimal place range -- vectorised via string ops, capped at 200 rows
                _samp = clean.head(200).astype(str)
                _dp = _samp.str.extract(r'\.(\d+)$')[0].str.rstrip("0").str.len().fillna(0).astype(int)
                info["dp_range"] = int(_dp.max()) - int(_dp.min())


        elif pd.api.types.is_datetime64_any_dtype(s):
            clean = s.dropna()
            if len(clean):
                info.update({
                    "min":      str(clean.min()),
                    "max":      str(clean.max()),
                    "freshness_days": (pd.Timestamp.now() - clean.max()).days,
                    "date_range_days": (clean.max() - clean.min()).days,
                    "future_count":  int((clean > pd.Timestamp.now()).sum()),
                    "weekend_count": int(clean.dt.dayofweek.isin([5, 6]).sum()),
                })
        elif s.dtype == object:
            # Cap all per-row string ops at 500 rows to keep profiling fast
            str_s    = s.dropna().astype(str)
            str_s_samp = str_s.head(500)
            if len(str_s_samp):
                lens = str_s_samp.str.len()
                # Mixed-case: vectorised -- no Python apply
                _upper = str_s_samp.str.upper()


                _lower = str_s_samp.str.lower()
                # OCR-UNCERTAIN: this intermediate formula looks like a stray/duplicated
                # draft calculation (superseded by the "Simpler and correct" version below);
                # transcribed as photographed.
                mixed_case_n = int((str_s_samp != _upper).sum() - (str_s_samp == _lower).sum() +
                    (str_s_samp == _lower).sum() - (str_s_samp == _lower).sum())
                # Simpler and correct: has both upper and lower chars
                mixed_case_n = int(((str_s_samp != _upper) & (str_s_samp != _lower)).sum())
                info.update({
                    "min_length":    int(lens.min()),
                    "max_length":    int(lens.max()),
                    "avg_length":    round(float(lens.mean()), 1),
                    "leading_space_count": int(str_s_samp.str.startswith(" ").sum()),
                    "trailing_space_count": int(str_s_samp.str.endswith(" ").sum()),
                    "mixed_case_count":  mixed_case_n,
                })
                # Non-ASCII character detection -- vectorised regex, no Python apply loop
                non_ascii_n = int(str_s_samp.str.contains(r'[^\x00-\x7F]', regex=True, na=False).sum())
                if non_ascii_n > 0:
                    info["non_ascii_count"] = non_ascii_n
                    info["non_ascii_pct"]  = round(non_ascii_n / len(str_s_samp) * 100, 1)


                # BOM marker detection (first row only)
                if len(str_s_samp) > 0:
                    first_val = str(str_s_samp.iloc[0])
                    if first_val.startswith("﻿") or first_val.startswith('\xef\xbb\xbf'):
                        info["has_bom"] = True

                # Control characters (non-printable, non-whitespace)
                ctrl_n = int(str_s_samp.str.contains(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]',
                    regex=True).sum())
                if ctrl_n > 0:
                    info["control_char_count"] = ctrl_n
                # BFSI format detection -- sample 100 rows, use vectorised str.match.
                # Skip columns whose values are numeric/financial quantities: if >=80%
                # of the non-null sample parses as a number, no format rule applies.
                _samp100 = str_s_samp.head(100)
                _numeric_rate = pd.to_numeric(_samp100, errors="coerce").notna().mean()
                if _numeric_rate < 0.8:
                    sample_upper = _samp100.str.upper()


                    for fmt_name, pat in _DQ_FORMAT_PATTERNS.items():
                        match_rate = sample_upper.str.fullmatch(pat.pattern).mean()
                        if match_rate >= 0.8:
                            info["detected_format"] = fmt_name
                            # Count violations on full column (vectorised)
                            non_conform = int((~str_s.str.upper().str.fullmatch(pat.pattern)).sum())
                            info["format_violations"] = non_conform
                            info["format_violation_pct"] = round(non_conform / total * 100, 1) if total else 0
                            break

                # Auto-parse string date columns for Timeliness -- sample 50, then full only
                # if needed
                if "freshness_days" not in info:
                    parsed50 = pd.to_datetime(str_s_samp.head(50), errors="coerce")
                    if parsed50.notna().mean() >= 0.8:
                        # Parse a capped sample (1000 rows) for freshness -- avoids full-column
                        # parse
                        parsed_full = pd.to_datetime(str_s.head(1000), errors="coerce")
                        valid = parsed_full.dropna()


                        if len(valid):
                            info["freshness_days"] = (pd.Timestamp.now() - valid.max()).days
                            info["date_range_days"] = (valid.max() - valid.min()).days
                            info["future_count"]   = int((valid > pd.Timestamp.now()).sum())

        # Cardinality classification
        if total > 0:
            ratio = uniq_n / total
            info["cardinality"] = (
                "identifier-like" if ratio > 0.95
                else "high" if ratio > 0.5
                else "medium" if ratio > 0.05
                else "low (enum-like)"
            )

        # Mark opt-in dimension participation
        if col in _user_freshness_cols:
            info["freshness_user_configured"] = True
            # Store the user's threshold so _dq_score can use it


            _td = next(
                (cc.get("timeliness_days") for cc in col_config_list if cc.get("name") == col),
                None
            )
            if _td:
                info["freshness_threshold_days"] = int(_td)
        if col in _user_precision_cols:
            info["precision_user_configured"] = True

        # Mark accuracy opt-in participation
        if col in accuracy_scores:
            info["accuracy_user_configured"] = True
            info["accuracy_score"] = round(accuracy_scores[col], 1)

        cols.append(info)

    # -- Compute aggregate accuracy value
    accuracy_value = (sum(accuracy_scores.values()) / len(accuracy_scores)) if accuracy_scores else None


    # -- Auto-inject baseline validity rules where no explicit rule exists ----------
    # Rules are injected only when the user has not already declared one for the
    # column. Multiple rules can apply to the same column (e.g. not_null + unique).
    _explicit_rule_cols = {r["column_name"] for r in rules if r.get("column_name")}

    for col_info in cols:
        cname   = col_info["name"]
        s       = df[cname]
        null_pct = col_info.get("null_pct", 0)

        # Skip entirely if user already declared any rule for this column
        if cname in _explicit_rule_cols:
            continue

        injected_any = False

        # 1. not_null -- only when column has nulls AND is not declared nullable.
        #   Skip columns with 0% nulls: completeness already penalises nulls and


        #   a passing not_null rule on clean data inflates validity without
        #   testing anything real.
        # Severity: critical for identifier-like columns, major for others.
        _is_id_like = col_info.get("cardinality") == "identifier-like"
        if cname.lower() not in _hint_nullable and 0 < null_pct <= 50:
            rules.append({
                "rule_type":  "not_null",
                "column_name": cname,
                "description": f"Not-null (auto: {null_pct:.1f}% nulls found)",
                "severity":   "critical" if _is_id_like else "major",
                "_auto":      True,
            })
            injected_any = True

        # 2. unique -- identifier-like columns, unless user ticked Allow Dups.
        #   Independent of not_null so both can fire on the same column.
        if (col_info.get("cardinality") == "identifier-like"
            and cname not in _config_allow_dups):
            rules.append({


                "rule_type":  "unique",
                "column_name": cname,
                "description": "Unique (auto: identifier-like cardinality)",
                "severity":   "critical",
                "_auto":      True,
            })
            injected_any = True

        # 3. allowed_values -- low-cardinality string columns (enum-like, <=20 distinct
        #    values). The observed distinct values become the allowed set.
        #    Skipped if uniqueness already covers the column.
        elif (s.dtype == object
            and col_info.get("cardinality") == "low (enum-like)"
            and col_info.get("unique_count", 0) <= 20
            and col_info.get("unique_count", 0) >= 2):
            distinct_vals = sorted(s.dropna().astype(str).str.strip().unique().tolist())
            allowed_str = ",".join(distinct_vals)
            rules.append({
                "rule_type":  "allowed_values",


                "column_name": cname,
                "value":    allowed_str,
                "description": f"Allowed values: {allowed_str} (auto: enum-like column)",
                "severity":   "major",
                "_auto":    True,
            })
            injected_any = True

        # 4. BFSI format -- when a format was auto-detected during profiling
        elif col_info.get("detected_format") and not injected_any:
            fmt_rule = f"{col_info['detected_format']}_format"
            rules.append({
                "rule_type":  fmt_rule,
                "column_name": cname,
                "description": f"{fmt_rule} (auto-detected)",
                "severity":   "major",
                "_auto":    True,
            })
            injected_any = True


        # 5. numeric range -- only inject when the column has actual variance AND
        #    is not highly skewed (skewed columns like price/volume have misleading
        #    sigma-based bounds). Use observed min/max instead for skewed columns.
        if (not injected_any
            and pd.api.types.is_numeric_dtype(s)
            and col_info.get("std") is not None):
            skewness = abs(col_info.get("skewness", 0) or 0)
            std_v   = col_info.get("std", 0) or 0
            if std_v > 0:
                if skewness <= 2:
                    # Symmetric-ish: use mu±4sigma
                    mean_v = col_info.get("mean", 0) or 0
                    lo = round(mean_v - 4 * std_v, 4)
                    hi = round(mean_v + 4 * std_v, 4)
                    desc = f"Range mu±4sigma [{lo}, {hi}] (auto)"
                else:
                    # Skewed: use observed min/max with 10% headroom
                    obs_min = col_info.get("min", 0) or 0


                    obs_max = col_info.get("max", 0) or 0
                    headroom = abs(obs_max - obs_min) * 0.10
                    lo = round(obs_min - headroom, 4)
                    hi = round(obs_max + headroom, 4)
                    desc = f"Range [min-10%, max+10%] [{lo}, {hi}] (auto, skewed col)"
                rules.append({
                    "rule_type":  "range",
                    "column_name": cname,
                    "min_value":  lo,
                    "max_value":  hi,
                    "description": desc,
                    "severity":   "minor",
                    "_auto":    True,
                })

    # -- Business rule validation --
    rule_results = []
    # Separate cross-column rules (need full DataFrame) from single-column rules
    _CROSS_COL_TYPES = {"conditional", "settlement_date_t2", "referential_integrity"}


    col_rules    = [r for r in rules if r.get("column_name") and r.get("rule_type") not in _CROSS_COL_TYPES]
    dataset_rules = [r for r in rules if not r.get("column_name") and r.get("rule_type") not in _CROSS_COL_TYPES]
    cross_col_rules = [r for r in rules if r.get("rule_type") in _CROSS_COL_TYPES]

    for rule in col_rules:
        col_name = rule["column_name"]
        if col_name not in df.columns:
            rule_results.append({**rule, "skipped": True, "reason": f"Column '{col_name}' not found"})
            continue
        result = _apply_rule(df[col_name], rule)
        rule_results.append({**rule, **result})

    for rule in dataset_rules:
        result = _apply_rule(pd.Series([total]), rule)
        rule_results.append({**rule, **result})

    # Feature 1 & 4 & 5 -- cross-column rules (conditional, referential, settlement T+2)


    if cross_col_rules:
        cc_results = _apply_cross_col_rules(df, cross_col_rules)
        rule_results.extend(cc_results)

    # -- Remediation Fix List ----------
    # Structured list of failing rules with actionable fix suggestions.
    _FIX_SUGGESTIONS = {
        "not_null":    "Populate missing values or mark column as nullable if intentional.",
        "unique":     "Remove or merge duplicate records. Check upstream deduplication logic.",
        "range":     "Clamp or investigate out-of-range values. Check data pipeline filters.",
        "allowed_values": "Map values to the allowed set or extend the allowed list if legitimate.",
        "pattern":    "Reformat values to match the required pattern.",
        "freshness_days": "Refresh the data feed. Check ETL schedule and source system timestamps.",
        "isin_format":  "Validate ISIN codes against a reference dataset (e.g. OpenFIGI).",
        "cusip_format": "Validate CUSIP codes against a reference dataset.",
        "lei_format":  "Validate LEI codes against the GLEIF database.",


        "bic_format":  "Validate BIC/SWIFT codes against the SWIFT BIC directory.",
        "iban_format": "Validate IBAN using checksum algorithm (ISO 13616).",
        "currency_code_format": "Use ISO 4217 3-letter currency codes.",
        "decimal_places": "Round or truncate values to the specified decimal precision.",
        "positive":   "Investigate negative values -- may indicate sign errors in the source.",
        "non_negative": "Investigate negative values -- may indicate data entry errors.",
        "integer_only": "Strip decimal parts or fix upstream type conversion.",
        "uppercase":  "Apply UPPER() transformation in the ETL pipeline.",
        "lowercase":  "Apply LOWER() transformation in the ETL pipeline.",
        "no_whitespace": "Apply TRIM() or strip() to remove surrounding whitespace.",
        "regex_format": "Reformat values to match the required custom pattern.",
    }

    remediation_items = []
    for r in rule_results:
        if r.get("status") == "FAIL" and not r.get("skipped"):
            rt = r.get("rule_type", "")
            remediation_items.append({
                "column":  r.get("column_name", "(dataset)"),


                "rule_type": rt,
                "severity": r.get("severity", "major"),
                "failed_rows": r.get("failed", 0),
                "examples": r.get("failing_examples", [])[:3],
                "suggestion": _FIX_SUGGESTIONS.get(rt, f"Review {rt} rule failures and fix upstream data."),
                "description": r.get("description", ""),
            })

    # Sort by severity: critical first
    _SEV_ORDER = {"critical": 0, "major": 1, "minor": 2}
    remediation_items.sort(key=lambda x: _SEV_ORDER.get(x["severity"], 1))

    # -- Per-column DQ sub-score
    # Based on rules applicable to that column, plus a null penalty.
    _SEV_W = {"critical": 3.0, "major": 2.0, "minor": 1.0}
    for col_info in cols:
        cname = col_info["name"]
        col_rules = [r for r in rule_results if r.get("column_name") == cname and not r.get("skipped")]


        if not col_rules:
            col_score = round(100 - col_info.get("null_pct", 0), 1)
        else:
            wt = sum(r.get("total", 0) * _SEV_W.get(r.get("severity", "major"), 2.0) for r in col_rules)
            wp = sum(r.get("passed", 0) * _SEV_W.get(r.get("severity", "major"), 2.0) for r in col_rules)
            col_score = round(wp / wt * 100, 1) if wt else 100.0
            col_score = round(max(0, col_score - col_info.get("null_pct", 0) * 0.5), 1)
        col_grade = "A" if col_score >= 90 else "B" if col_score >= 75 else "C" if col_score >= 60 else "D" if col_score >= 40 else "F"
        col_info["dq_score"] = col_score
        col_info["dq_grade"] = col_grade

    # -- Consistency checks
    consistency_issues = []

    # 0. Auto-detected consistency issues (no user input required)
    # All checks use already-computed col_info stats or a capped sample -- no


    # full-column Python apply loops here.
    _DATE_PAT_RE = [
        re.compile(r'\d{4}-\d{2}-\d{2}'),   # ISO
        re.compile(r'\d{2}/\d{2}/\d{4}'),   # MM/DD/YYYY
        re.compile(r'\d{2}-\d{2}-\d{4}'),   # DD-MM-YYYY
        re.compile(r'\d{2}\.\d{2}\.\d{4}'), # DD.MM.YYYY
    ]

    _NULL_PLACEHOLDERS = {"NULL", "NA", "N/A", "NONE", "NIL", "-", "?", ""}

    for col_info in cols:
        cname = col_info["name"]
        s = df[cname]

        if s.dtype != object:
            continue  # checks below only apply to string columns

        # Use a capped sample for all string checks
        samp = s.dropna().astype(str).head(500)
        samp_n = len(samp)


        if samp_n == 0:
            continue
        actual_null_n = int(s.isna().sum())

        # 0a. Mixed null representation -- vectorised isin on sample
        null_like_n = int(samp.str.strip().str.upper().isin(_NULL_PLACEHOLDERS).sum())
        if null_like_n > actual_null_n:
            extra = null_like_n - actual_null_n
            consistency_issues.append({
                "check": "Mixed null representation",
                "column": cname,
                "failed": extra,
                "status": "WARN",
                "detail": f"{extra} rows use placeholder strings instead of true nulls",
            })

        # 0b. Date format inconsistency -- only for columns whose avg length looks date-like
        avg_len = col_info.get("avg_length", 0) or 0


        if 6 <= avg_len <= 11:
            pat_hits = [int(samp.str.fullmatch(p.pattern).sum()) for p in _DATE_PAT_RE]
            formats_present = sum(1 for h in pat_hits if h > 0)
            if formats_present > 1:
                consistency_issues.append({
                    "check": "Inconsistent date format",
                    "column": cname,
                    "failed": samp_n - max(pat_hits),
                    "status": "FAIL",
                    "detail": f"{formats_present} different date formats detected in same column",
                })

        # 0c. Mixed numeric/text -- use already-profiled stats where possible
        coerced = pd.to_numeric(samp, errors="coerce")
        numeric_n = int(coerced.notna().sum())
        non_null_n = int(s.notna().sum())
        if non_null_n > 0 and 0 < numeric_n < samp_n:
            mixed_n = samp_n - numeric_n


            if mixed_n / samp_n >= 0.01:
                consistency_issues.append({
                    "check": "Mixed numeric/text values",
                    "column": cname,
                    "failed": mixed_n,
                    "status": "FAIL" if mixed_n / samp_n >= 0.05 else "WARN",
                    "detail": f"{mixed_n} non-numeric values in an otherwise numeric column",
                })

        # 0d. Whitespace contamination -- reuse stats already computed during profiling
        ws_n = col_info.get("leading_space_count", 0) + col_info.get("trailing_space_count", 0)
        if non_null_n > 0 and ws_n / non_null_n >= 0.05:
            consistency_issues.append({
                "check": "Whitespace contamination",
                "column": cname,
                "failed": ws_n,
                "status": "WARN",
                "detail": f"{ws_n} values have leading/trailing spaces",


            })

    # 1. NOT NULL violations from data dictionary
    for col_name, dd in data_dict.items():
        if col_name in df.columns and not dd.get("nullable", True):
            if col_name.lower() in _hint_nullable:
                continue
            null_n = int(df[col_name].isna().sum())
            if null_n:
                consistency_issues.append({
                    "check": "NOT NULL (data dictionary)",
                    "column": col_name,
                    "failed": null_n,
                    "status": "FAIL",
                })

    # 2. Primary key enforcement -- PK columns must be unique and non-null
    pk_cols = [c for c, d in data_dict.items() if d.get("is_pk") and c in df.columns]
    if pk_cols:


        pk_nulls = int(df[pk_cols].isna().any(axis=1).sum())
        pk_dups  = int(df.duplicated(subset=pk_cols, keep=False).sum())
        if pk_nulls:
            consistency_issues.append({
                "check": "PK NOT NULL",
                "column": ", ".join(pk_cols),
                "failed": pk_nulls,
                "status": "FAIL",
            })
        if pk_dups:
            consistency_issues.append({
                "check": "PK UNIQUE",
                "column": ", ".join(pk_cols),
                "failed": pk_dups,
                "status": "FAIL",
            })

    # 3. Cross-column consistency: referential checks from hints
    # Format: "col_a=val_when:col_b>0" -- e.g. "side=BUY:quantity>0"


    for expr in hints.get("cross_column_rules", "").split(";"):
        expr = expr.strip()
        if not expr:
            continue
        try:
            # Simple form: "col_a op val => col_b op val"
            # Parse "colA=X:colB>0" -> when colA==X, colB must be >0
            m = re.match(r'^(\w+)\s*=\s*(.+?)\s*:\s*(\w+)\s*([><=!]+)\s*(.+)$', expr)
            if m and m.group(1) in df.columns and m.group(3) in df.columns:
                cond_col, cond_val, check_col, op, check_val = m.groups()
                cond_mask = df[cond_col].astype(str).str.strip() == cond_val.strip()
                num = pd.to_numeric(df.loc[cond_mask, check_col], errors="coerce")
                cv  = float(check_val)
                if op == ">":   fail_mask = num <= cv
                elif op == ">=": fail_mask = num < cv
                elif op == "<":  fail_mask = num >= cv
                elif op == "<=": fail_mask = num > cv
                elif op == "==" or op == "=": fail_mask = num != cv
                elif op == "!=": fail_mask = num == cv


                else: continue
                failed_n = int(fail_mask.sum())
                if failed_n:
                    consistency_issues.append({
                        "check": f"Cross-column: when {cond_col}={cond_val}, {check_col}{op}{check_val}",
                        "column": check_col,
                        "failed": failed_n,
                        "status": "FAIL" if failed_n / max(int(cond_mask.sum()), 1) >= 0.05 else "WARN",
                    })
        except Exception:
            pass

    # 4. Duplicate row detection (full row and subset-key)
    dup_rows = int(df.duplicated().sum())

    # 5. Referential integrity -- allowed_values cross-check between columns
    for expr in hints.get("referential_rules", "").split(";"):
        expr = expr.strip()


        # Format: "col_a->col_b"  means every value in col_a must appear in col_b
        m = re.match(r'^(\w+)\s*->\s*(\w+)$', expr)
        if m and m.group(1) in df.columns and m.group(2) in df.columns:
            src_col, ref_col = m.group(1), m.group(2)
            valid_vals = set(df[ref_col].dropna().astype(str))
            mask = ~df[src_col].astype(str).isin(valid_vals)
            failed_n = int(mask.sum())
            if failed_n:
                consistency_issues.append({
                    "check": f"Referential integrity: {src_col} -> {ref_col}",
                    "column": src_col,
                    "failed": failed_n,
                    "status": "FAIL" if failed_n / total >= 0.05 else "WARN",
                })

    # 6. Conflicting records -- same identifier key, different value in another column.
    # Auto-discovered: identifier-like columns (cardinality > 95%) are used as the
    # key; every other non-identifier column is checked for value consistency within
    # each key group. Capped at 5 key columns and 10 value columns for performance.


    id_cols = [
        c["name"] for c in cols
        if c.get("cardinality") == "identifier-like" and c["name"] in df.columns
    ][:5]
    if id_cols:
        val_cols = [
            c["name"] for c in cols
            if c.get("cardinality") not in ("identifier-like",)
            and c["name"] in df.columns
            and c["name"] not in id_cols
        ][:10]
        # Cap rows for conflict detection -- groupby on 300k rows is slow
        _CONFLICT_CAP = 50_000
        _df_conflict = df if total <= _CONFLICT_CAP else df.iloc[:_CONFLICT_CAP]
        for key_col in id_cols:
            for val_col in val_cols:
                try:
                    # For each key value, count distinct non-null values in val_col
                    grp = (


                        _df_conflict[[key_col, val_col]]
                        .dropna(subset=[key_col])
                        .groupby(key_col, sort=False)[val_col]
                        .nunique()
                    )
                    conflicting_keys = int((grp > 1).sum())
                    if conflicting_keys > 0:
                        conflict_pct = conflicting_keys / len(grp) * 100
                        consistency_issues.append({
                            "check": f"Conflicting records: {key_col} -> {val_col}",
                            "column": val_col,
                            "failed": conflicting_keys,
                            "status": "FAIL" if conflict_pct >= 5 else "WARN",
                            "detail": (
                            f"{conflicting_keys} value(s) of '{key_col}' map to "
                            f"multiple distinct '{val_col}' values"
                            ),
                        })
                except Exception:


                    pass

    # 7. Cross-file referential integrity
    # Format: "local_col->ref_filename:ref_col" semicolon-separated
    cross_file_ref_data: dict[str, pd.DataFrame] = hints.get("cross_file_ref_data", {})
    for expr in hints.get("cross_file_rules", "").split(";"):
        expr = expr.strip()
        m = re.match(r'^(\w+)\s*->\s*(.+?)\s*:\s*(\w+)$', expr)
        if not m:
            continue
        local_col, ref_fname, ref_col = m.group(1), m.group(2).strip(), m.group(3)
        if local_col not in df.columns:
            continue
        ref_df = cross_file_ref_data.get(ref_fname)
        if ref_df is None or ref_col not in ref_df.columns:
            consistency_issues.append({
                "check": f"Cross-file ref: {local_col} -> {ref_fname}:{ref_col}",
                "column": local_col,
                "failed": 0,


                "status": "WARN",
                "detail": f"Reference file '{ref_fname}' or column '{ref_col}' not available",
            })
            continue
        valid_vals = set(ref_df[ref_col].dropna().astype(str))
        mask = df[local_col].notna() & ~df[local_col].astype(str).isin(valid_vals)
        failed_n = int(mask.sum())
        if failed_n or True:  # always report
            consistency_issues.append({
                "check": f"Cross-file ref: {local_col} -> {ref_fname}:{ref_col}",
                "column": local_col,
                "failed": failed_n,
                "status": "FAIL" if failed_n > 0 else "PASS",
                "detail": (
                    f"{failed_n} value(s) in '{local_col}' not found in '{ref_fname}.{ref_col}'"
                    if failed_n else
                    f"All values in '{local_col}' match '{ref_fname}.{ref_col}' ✓"
                ),
            })


    # 8. Conditional completeness rules
    # Format: "target_col:condition_col=condition_val" semicolon-separated
    for expr in hints.get("conditional_completeness_rules", "").split(";"):
        expr = expr.strip()
        m = re.match(r'^(\w+)\s*:\s*(\w+)\s*=\s*(.+)$', expr)
        if not m:
            continue
        target_col, cond_col, cond_val = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
        if target_col not in df.columns or cond_col not in df.columns:
            continue
        cond_mask = df[cond_col].astype(str).str.strip() == cond_val
        target_null = df.loc[cond_mask, target_col].isna() | (df.loc[cond_mask, target_col].astype(str).str.strip() == "")
        failed_n = int(target_null.sum())
        cond_count = int(cond_mask.sum())
        if cond_count > 0:
            consistency_issues.append({


                "check": f"Conditional completeness: {target_col} when {cond_col}={cond_val}",
                "column": target_col,
                "failed": failed_n,
                "status": "FAIL" if failed_n > 0 else "PASS",
                "detail": (
                    f"{failed_n}/{cond_count} rows where {cond_col}={cond_val} have null '{target_col}'"
                    if failed_n else
                    f"All {cond_count} rows where {cond_col}={cond_val} have '{target_col}' populated ✓"
                ),
            })

    # 9. Encoding anomalies -- columns with non-ASCII or control characters
    for col_info in cols:
        if col_info.get("non_ascii_pct", 0) > 5:
            consistency_issues.append({
                "check": "Non-ASCII characters",
                "column": col_info["name"],


                "failed": col_info.get("non_ascii_count", 0),
                "status": "WARN",
                "detail": f"{col_info['non_ascii_pct']}% of values contain non-ASCII characters -- check encoding",
            })
        if col_info.get("has_bom"):
            consistency_issues.append({
                "check": "BOM marker detected",
                "column": col_info["name"],
                "failed": 1,
                "status": "WARN",
                "detail": "BOM (byte-order mark) found in column values -- strip before processing",
            })
        if col_info.get("control_char_count", 0) > 0:
            consistency_issues.append({
                "check": "Control characters",
                "column": col_info["name"],
                "status": "WARN",


                "detail": f"{col_info['control_char_count']} values contain non-printable control characters",
            })

    # -- Undocumented columns
    undocumented = [c for c in df.columns if c not in data_dict] if data_dict else []
    columns_in_dict_not_in_data = [c for c in data_dict if c not in df.columns] if data_dict else []

    # -- Format violation summary (from auto-detected BFSI formats) ----------
    format_violations = [
        {"column": c["name"], "format": c["detected_format"],
        "violations": c["format_violations"], "violation_pct": c["format_violation_pct"]}
        for c in cols if c.get("detected_format") and c.get("format_violations", 0) > 0
    ]

    # -- Enhanced duplicate analysis
    dup_analysis = _detect_duplicates(df, cols, hints)


    # -- Drift detection against saved baseline
    _baseline = hints.get("dq_baseline")
    drift_alerts = _detect_drift(cols, _baseline)

    # -- Schema drift detection

    schema_drift = _detect_schema_drift(df, cols, _baseline)

    # -- Schema fingerprint (for DQ rule persistence)
    _dq_fp = _dq_schema_fingerprint(df)

    dq = _dq_score(total, dup_rows, cols, rule_results, consistency_issues, accuracy=accuracy_value)

    return {
  "file_name":       name,
  "total_rows":      total,
  "total_cols":      len(df.columns),
  "duplicate_rows":    dup_rows,
  "has_data_dict":     bool(data_dict),


  "has_rules":       bool(rules),
  "pk_columns":      pk_cols,
  "undocumented_columns":  undocumented,
  "dict_columns_missing_in_data": columns_in_dict_not_in_data,
  "consistency_issues":   consistency_issues,
  "format_violations":   format_violations,
  "dq_score":        dq,
  "rule_results":      rule_results,
  "columns":        cols,
  "dup_analysis":      dup_analysis,
  "drift_alerts":      drift_alerts,
  "schema_drift":      schema_drift,
  "schema_fingerprint":   _dq_fp,
  "remediation_items":   remediation_items,
  "baseline_snapshot": {
    "file_name": name,
    "total_rows": total,
    "columns": [


      {k: c.get(k) for k in ["name", "null_pct", "uniqueness_pct", "outlier_pct", "mean",
  "std", "cardinality"]}
      for c in cols
    ],
  },
}

# ------------------------------------------------------------------------
# Shared column stats -- computed once, reused by quality, profile, governance
# ------------------------------------------------------------------------

def _compute_col_stats(df: pd.DataFrame) -> dict[str, dict]:

    # Single-pass column statistics shared across analyze_quality, analyze_profile
    # and analyze_governance.  Eliminates duplicate isna()/nunique()/dtype scans.

    # Returns a dict keyed by column name:
    # null_count, null_pct, unique_count, unique_pct, dtype,


    # is_numeric, is_datetime, is_object,
    # min, max, mean, std  (numeric only -- None otherwise)
    # top_values      (list of {value, count} -- up to 8)

    total = len(df)
    stats: dict[str, dict] = {}
    for col in df.columns:
        s = df[col]
        null_n  = int(s.isna().sum())
        null_pct = round(null_n / total * 100, 1) if total else 0.0
        uniq_n  = int(s.nunique(dropna=True))
        uniq_pct = round(uniq_n / total * 100, 1) if total else 0.0

        is_num = pd.api.types.is_numeric_dtype(s)
        is_dt  = pd.api.types.is_datetime64_any_dtype(s)
        is_obj = s.dtype == object

        mn = mx = mu = sd = None
        numeric_series = s
        if not is_num and not is_dt:
            # _load_file() loads every column as string (often pandas' modern
            # StringDtype, which is NOT == object) -- coerce here so
            # numeric-looking columns (the common case for CSV/TXT/Excel
            # uploads) still get real stats.
            coerced = pd.to_numeric(s, errors="coerce")
            non_null = s.notna().sum()
            if non_null and coerced.notna().sum() / non_null >= 0.9:
                is_num = True
                numeric_series = coerced

        if is_num:


            clean = numeric_series.dropna()
            if len(clean):
                mn = float(clean.min())
                mx = float(clean.max())
                mu = float(clean.mean())
                sd = float(clean.std())

        # value_counts via groupby + nlargest -- avoids full sort on large frames
        try:
            vc = s.groupby(s).size().nlargest(8)
            top_vals = [
                {"value": str(v), "count": int(c)}
                for v, c in vc.items()
            ]
        except Exception:
            top_vals = []

        stats[col] = {
            "null_count": null_n,


            "null_pct": null_pct,
            "unique_count": uniq_n,
            "unique_pct": uniq_pct,
            "dtype":    str(s.dtype),
            "is_numeric": is_num,
            "is_datetime": is_dt,
            "is_object": is_obj,
            "min":     mn,
            "max":     mx,
            "mean":    mu,
            "std":     sd,
            "top_values": top_vals,
        }
    return stats

    # ------------------------------------------------------------------------
    # Merged full DQ -- quality + profile + governance in one pass
    # ------------------------------------------------------------------------

def analyze_quality_full(
  df: pd.DataFrame,
  name: str,
  data_dict: dict | None = None,
  rules: list[dict] | None = None,
  user_hints: dict | None = None,
  df2: pd.DataFrame | None = None,
  name2: str | None = None,
) -> dict:

  # Single-call merged analysis combining:
  # 1. Data Quality  -- DQ score, completeness, validity, rule results
  # 2. Data Profile  -- semantic types, cardinality, top values, key candidates,
  # BFSI domain annotation, cross-column correlations
  # 3. Data Governance -- PII detection, BFSI identifiers, sensitivity tier,
  # regulatory frameworks, mandatory breach checks,
  # conditional violations, stewardship routing
  # 4. Mapping recon  -- if df2 is provided: type mismatches + recon FAILs


  # feed a governance penalty into the DQ score

  # Uses _compute_col_stats() so null/unique/dtype/min/max/mean/std are
  # computed once and shared across all three modules -- no redundant scans.

  # Returns the standard analyze_quality dict enriched with:
  # profile   -- per-column semantic/cardinality/top-values/BFSI domain
  # governance  -- per-column PII/BFSI/sensitivity/regulatory/stewardship
  # mapping_recon -- (optional) type mismatches + recon FAILs from df2
  # dq_score   -- governance_penalty applied on top of base score

  # -- Shared single-pass column stats
  col_stats = _compute_col_stats(df)

  # -- 1. Core DQ (passes col_stats hints so it skips redundant scans) ----------
  _hints = dict(user_hints or {})
  _hints["_col_stats"] = col_stats    # consumed by analyze_quality internally
  quality = analyze_quality(df, name, data_dict=data_dict, rules=rules,
              user_hints=_hints)


  # -- 2. Profile enrichment (reuses col_stats -- no re-scan) ----------
  total  = len(df)
  mem_mb = round(df.memory_usage(deep=False).sum() / 1_048_576, 2)

  _DATE_HINTS = {"date", "dt", "time", "ts", "timestamp", "created", "updated",
"modified"}
  _ID_HINTS  = {"id", "key", "pk", "code", "ref", "num", "no", "number"}
  _AMT_HINTS = {"amount", "amt", "price", "rate", "qty", "quantity", "notional",
      "value", "bal", "balance", "vol", "volume"}
  _FLAG_HINTS = {"flag", "ind", "indicator", "is_", "has_", "active", "status",
      "type", "category", "class", "group"}

  def _semantic(col: str, cs: dict) -> str:
    lower = col.lower()
    if cs["is_datetime"]:
      return "datetime"
    if any(h in lower for h in _DATE_HINTS) and cs["is_object"]:
      return "date-like string"


    if any(h == lower or lower.endswith(f"_{h}") or lower.startswith(f"{h}_")
        for h in _ID_HINTS):
      return "identifier"
    if any(h in lower for h in _AMT_HINTS) and cs["is_numeric"]:
      return "amount/quantity"
    if any(lower == h or lower.startswith(h) for h in _FLAG_HINTS):
      return "flag/category"
    if cs["is_numeric"]:
      return "numeric"
    if cs["is_object"]:
      sample = df[col].dropna().astype(str).head(200)
      if sample.str.match(r'^\d{4}[-/]\d{2}[-/]\d{2}').mean() > 0.7:
        return "date string"
      if sample.str.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$').mean() > 0.5:
        return "email"
      if sample.str.match(r'^\+?\d[\d\s\-().]{7,}$').mean() > 0.5:
        return "phone"
      avg_len = sample.str.len().mean() if len(sample) else 0


      return "free text" if avg_len > 60 else "text"
    return cs["dtype"]

  profile_cols = []
  numeric_col_names = []
  for col in df.columns:
    cs = col_stats[col]
    sem = _semantic(col, cs)
    uq = cs["unique_pct"]
    cardinality = (
      "identifier-like" if uq > 95 else
      "high"      if uq > 50 else
      "medium"     if uq > 5  else
      "low (enum)"
    )
    entry = {
      "name":    col,
      "dtype":   cs["dtype"],
      "semantic": sem,


      "bfsi_domain": _infer_operational_usage(col),
      "criticality": _infer_criticality(col, cs["null_pct"]),
      "null_count": cs["null_count"],
      "null_pct":  cs["null_pct"],
      "unique_count": cs["unique_count"],
      "unique_pct": uq,
      "cardinality": cardinality,
      "top_values": cs["top_values"],
    }
    if cs["is_numeric"] and cs["min"] is not None:
      clean = df[col].dropna()
      if not pd.api.types.is_numeric_dtype(clean):
        clean = pd.to_numeric(clean, errors="coerce").dropna()
      q1, q3 = float(clean.quantile(0.25)), float(clean.quantile(0.75))
      iqr  = q3 - q1
      out_n = int(((clean < q1 - 1.5 * iqr) | (clean > q3 + 1.5 * iqr)).sum())
      entry.update({
        "min":    round(cs["min"], 4),
        "max":    round(cs["max"], 4),
        "mean":    round(cs["mean"], 4),
        "std":    round(cs["std"], 4),


        "q1":    round(q1, 4),
        "q3":    round(q3, 4),
        "outlier_count": out_n,
        "outlier_pct":  round(out_n / total * 100, 1) if total else 0,
      })
      numeric_col_names.append(col)
    elif cs["is_datetime"]:
      clean = df[col].dropna()
      if len(clean):
        entry["min"] = str(clean.min())
        entry["max"] = str(clean.max())
        entry["freshness_days"] = (pd.Timestamp.now() - clean.max()).days
    elif cs["is_object"]:
      clean_str = df[col].dropna().astype(str)
      if len(clean_str):
        lens = clean_str.str.len()
        entry.update({
          "avg_length": round(float(lens.mean()), 1),
          "min_length": int(lens.min()),


          "max_length": int(lens.max()),
          "sample_values": clean_str.head(3).tolist(),
        })
    profile_cols.append(entry)

  key_candidates = [
    c["name"] for c in profile_cols
    if c["null_count"] == 0 and c["unique_count"] == total and total > 0
  ]
  near_key = [
    c["name"] for c in profile_cols
    if c["null_count"] == 0 and c["unique_pct"] >= 90
    and c["name"] not in key_candidates
  ]

  # Cross-column correlations (top 10 numeric pairs)
  correlations: list[dict] = []
  if len(numeric_col_names) >= 2:
    try:


      corr_matrix = df[numeric_col_names].corr()
      seen: set = set()
      for i, c1 in enumerate(numeric_col_names):
        for c2 in numeric_col_names[i + 1:]:
          if (c1, c2) not in seen:
            seen.add((c1, c2))
            v = corr_matrix.loc[c1, c2]
            if not math.isnan(v):
              correlations.append({
                "col1": c1, "col2": c2,
                "corr": round(float(v), 3),
                "strength": (
                  "strong"  if abs(v) >= 0.7 else
                  "moderate" if abs(v) >= 0.4 else
                  "weak"
                ),
              })
      correlations.sort(key=lambda x: abs(x["corr"]), reverse=True)
      correlations = correlations[:10]


    except Exception:
      pass

  type_counts: dict[str, int] = {}
  for c in profile_cols:
    sem = c["semantic"]
    type_counts[sem] = type_counts.get(sem, 0) + 1

  profile = {
    "memory_mb":    mem_mb,
    "key_candidates": key_candidates,
    "near_key_cols":  near_key,
    "type_breakdown": type_counts,
    "correlations":  correlations,
    "columns":     profile_cols,
  }

# -- 3. Governance -- load saved Dataset Controls overrides first ----------
# Rules are kept SEPARATE per module -- dc_governance_override rules only.


  # Search ALL fingerprints in feedback_rules.json for dc_governance_override
  # rules whose text mentions a column present in the CURRENT file.
  # This is column-name based, not fingerprint-based, so it works regardless
  # of which dataset the rule was originally saved from.
  _gov_hints = dict(user_hints or {})
  try:
    import json as _json_gov
    _fb_path = _feedback_store_path
    _all_saved: dict = {}
    if os.path.exists(_fb_path):
      with open(_fb_path, encoding="utf-8") as _fb:
        _all_saved = _json_gov.load(_fb)

    _col_overrides: dict = {}
    _col_map_lc = {c.lower(): c for c in df.columns}

    # Scan every fingerprint for dc_governance_* rules (override AND exclude)
    for _fp_entry in _all_saved.values():


      for _r in _fp_entry.get("rules", []):
        _cat = _r.get("category", "")
        if not _cat.startswith("dc_governance_"):
          continue
        _text = _r.get("rule", "").lower()
        _is_override = _cat.startswith("dc_governance_override")
        _is_exclude = _cat.startswith("dc_governance_exclude")
        # Parse all possible override actions from the rule text
        _new_sens = _parse_sensitivity_from_text(_text)
        _new_regs = _parse_regulatory_override(_text)
        _not_pii  = any(k in _text for k in ("not pii","no pii","false positive","not a","not personal"))
        _do_excl  = _is_exclude or any(k in _text for k in
("exclude","skip","ignore","omit"))
        # Match against columns in THIS file -- column-name based
        for _col_lc, _col_actual in _col_map_lc.items():
          if len(_col_lc) >= 3 and _col_lc in _text:
            _col_overrides.setdefault(_col_actual, {})
            if _do_excl:
              _col_overrides[_col_actual]["exclude"] = True


            if _new_sens:
              _col_overrides[_col_actual]["sensitivity"] = _new_sens
            if _not_pii:
              _col_overrides[_col_actual]["clear_pii"] = True
            if _new_regs is not None:
              _col_overrides[_col_actual]["regulatory"] = _new_regs

    if _col_overrides:
      _gov_hints["_dc_col_overrides"] = _col_overrides
  except Exception:
    pass

  governance = analyze_governance(df, name, data_dict=data_dict,
                   user_hints=_gov_hints)

  # -- 4. Mapping recon penalty (optional -- only when df2 provided) ----------
  mapping_recon: dict = {}
  governance_penalty = 0.0

  if df2 is not None:
      mapping = analyze_mapping(df2, df, name2 or "File 2", name,
                  mapping_spec=None, user_hints=user_hints)

      recon_fails  = [r for r in mapping.get("reconciliation", [])
              if r.get("status") == "FAIL"]

      type_mismatches = [e for e in mapping.get("exact", [])
               if not e.get("type_ok")]

      # Each recon FAIL on a critical BFSI column -> heavier penalty
      for r in recon_fails:
          crit = _infer_criticality(r["column"], 0)
          governance_penalty += 3.0 if crit == "High" else 1.5

      # Type mismatches on business columns -> small penalty
      for e in type_mismatches:
          if _infer_criticality(e["f1_col"], 0) == "High":
              governance_penalty += 1.0

      mapping_recon = {
          "mapping_completeness_pct": mapping.get("mapping_completeness_pct"),
          "type_mismatches": [
              {"column": e["f1_col"], "f1_type": e["f1_type"], "f2_type": e["f2_type"]}
              for e in type_mismatches
          ][:10],
          "recon_fails": [
              {"column": r["column"], "f1_sum": r["f1_sum"],
               "f2_sum": r["f2_sum"], "diff_pct": r["diff_pct"]}
              for r in recon_fails
          ][:10],
          "recon_warnings": [
              {"column": r["column"], "diff_pct": r["diff_pct"]}
              for r in mapping.get("reconciliation", [])
              if r.get("status") == "WARN"
          ][:10],
      }

      # Governance penalty: mandatory breaches + conditional warnings
      gov_breach_count = (
          len(governance.get("mandatory_breaches", [])) +
          len(governance.get("conditional_warnings", []))
      )

      governance_penalty += gov_breach_count * 2.0

  # Apply governance penalty to DQ score (floor 0)
  base_score = quality["dq_score"]["score"]
  penalised_score = round(max(0.0, base_score - governance_penalty), 1)
  penalised_grade = (
      "A" if penalised_score >= 90 else
      "B" if penalised_score >= 75 else
      "C" if penalised_score >= 60 else
      "D" if penalised_score >= 40 else "F"
  )

  quality["dq_score"]["score"]              = penalised_score
  quality["dq_score"]["grade"]              = penalised_grade
  quality["dq_score"]["governance_penalty"] = round(governance_penalty, 1)
  quality["dq_score"]["base_score"]         = base_score

  # -- Merge and return
  # -----------------------------------------------------------------------

  return {
      **quality,
      "profile":       profile,
      "governance":    governance,
      "mapping_recon": mapping_recon,
  }


# -----------------------------------------------------------------------
# Data Profiling - deep structural portrait of a dataset
# -----------------------------------------------------------------------


def analyze_profile(df: pd.DataFrame, name: str) -> dict:
    # Produce a deep structural profile of a DataFrame.
    #
    # Returns per-column stats (type, nulls, cardinality, min/max/mean/std,
    # top values, inferred semantic type, pattern samples) plus dataset-level
    # summary (shape, memory, duplicate rows, potential key candidates,
    # cross-column correlations for numeric pairs).
    total = len(df)

    # deep=False is 50x faster -- uses dtype sizes instead of traversing object values
    mem_mb = round(df.memory_usage(deep=False).sum() / 1_048_576, 2)

    # Cap duplicate check at 50k rows -- full hash of 300k rows is very slow
    _DUP_SAMPLE = min(total, 50_000)
    dup_rows = int(df.iloc[:_DUP_SAMPLE].duplicated().sum())
    if total > _DUP_SAMPLE:
        dup_rows = int(dup_rows * total / _DUP_SAMPLE)  # extrapolate

    # -- Infer semantic type from column name + content --
    _DATE_HINTS  = {"date", "dt", "time", "ts", "timestamp", "created", "updated", "modified"}
    _ID_HINTS    = {"id", "key", "pk", "code", "ref", "num", "no", "number"}
    _AMT_HINTS   = {"amount", "amt", "price", "rate", "qty", "quantity", "notional",
                    "value", "bal", "balance", "vol", "volume"}
    _FLAG_HINTS  = {"flag", "ind", "indicator", "is_", "has_", "active", "status",
                    "type", "category", "class", "group"}

    def _semantic(col: str, s: "pd.Series") -> str:
        lower = col.lower()

        if pd.api.types.is_datetime64_any_dtype(s):
            return "datetime"

        if any(h in lower for h in _DATE_HINTS) and s.dtype == object:
            return "date-like string"

        if any(h == lower or lower.endswith(f"_{h}") or lower.startswith(f"{h}_")
               for h in _ID_HINTS):
            return "identifier"

        if any(h in lower for h in _AMT_HINTS) and pd.api.types.is_numeric_dtype(s):
            return "amount/quantity"

        if any(lower == h or lower.startswith(h) for h in _FLAG_HINTS):
            return "flag/category"

        if pd.api.types.is_numeric_dtype(s):
            return "numeric"

        if s.dtype == object:
            # Sniff patterns in a sample of non-null values
            sample = s.dropna().astype(str).head(200)

            if sample.str.match(r'^\d{4}[-/]\d{2}[-/]\d{2}').mean() > 0.7:  # OCR-UNCERTAIN: regex pattern approximate
                return "date string"

            if sample.str.match(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$').mean() > 0.5:  # OCR-UNCERTAIN: regex pattern approximate
                return "email"

            if sample.str.match(r'^\+?\d[\d\s\-().]{7,}$').mean() > 0.5:  # OCR-UNCERTAIN: regex pattern approximate
                return "phone"

            if sample.str.match(r'^[A-Z]{2}\d{2}[A-Z0-9]{4}\d{7,}$').mean() > 0.4:  # OCR-UNCERTAIN: regex pattern approximate
                return "IBAN"

            avg_len = sample.str.len().mean() if len(sample) else 0
            if avg_len > 60:
                return "free text"
            return "text"

        return str(s.dtype)

    # -- Column-level profiles --
    col_profiles = []

    numeric_cols = []

    for col in df.columns:
        s = df[col]
        null_n   = int(s.isna().sum())
        null_pct = round(null_n / total * 100, 1) if total else 0
        uniq_n   = int(s.nunique(dropna=True))
        uniq_pct = round(uniq_n / total * 100, 1) if total else 0
        sem      = _semantic(col, s)

        top_vals = (
            s.value_counts(dropna=True)
            .head(8)
            .reset_index()
            .rename(columns={col: "value", "index": "value", "count": "count"})
            .to_dict("records")
        )

        entry: dict = {
            "name":  col,
            "dtype":     str(s.dtype),
            "semantic":  sem,
            "null_count": null_n,
            "null_pct":  null_pct,
            "unique_count": uniq_n,
            "unique_pct":  uniq_pct,
            "top_values": top_vals,
            "cardinality": (
                "identifier-like" if uniq_pct > 95
                else "high"     if uniq_pct > 50
                else "medium"   if uniq_pct > 5
                else "low (enum)"
            ),
        }

        if pd.api.types.is_numeric_dtype(s):
            clean = s.dropna()
            if len(clean):
                q1, q3 = float(clean.quantile(0.25)), float(clean.quantile(0.75))
                iqr  = q3 - q1
                out_n = int(((clean < q1 - 1.5 * iqr) | (clean > q3 + 1.5 * iqr)).sum())
                entry.update({
                    "min":    round(float(clean.min()), 4),
                    "max":    round(float(clean.max()), 4),
                    "mean":   round(float(clean.mean()), 4),
                    "median": round(float(clean.median()), 4),
                    "std":    round(float(clean.std()), 4),
                    "q1":     round(q1, 4),
                    "q3":     round(q3, 4),
                    "outlier_count": out_n,
                    "outlier_pct":  round(out_n / total * 100, 1) if total else 0,
                    "zeros": int((clean == 0).sum()),
                    "negatives": int((clean < 0).sum()),
                })

            numeric_cols.append(col)

        elif pd.api.types.is_datetime64_any_dtype(s):
            clean = s.dropna()
            if len(clean):
                entry.update({
                    "min": str(clean.min()),
                    "max": str(clean.max()),
                    "freshness_days": (pd.Timestamp.now() - clean.max()).days,
                })

        elif s.dtype == object:
            clean_str = s.dropna().astype(str)
            if len(clean_str):
                lens = clean_str.str.len()
                entry.update({
                    "avg_length": round(float(lens.mean()), 1),
                    "min_length": int(lens.min()),
                    "max_length": int(lens.max()),
                    "sample_values": clean_str.head(3).tolist(),
                })

        # -- Distribution data for visualisation --------------------------------
        if pd.api.types.is_numeric_dtype(s):
            clean_num = s.dropna()
            if len(clean_num) >= 2:
                try:
                    n_bins = min(10, max(4, int(len(clean_num) ** 0.5 // 2)))
                    counts, edges = np.histogram(clean_num, bins=n_bins)
                    entry["histogram"] = {
                        "bins":   [round(float(e), 4) for e in edges],
                        "counts": [int(c) for c in counts],
                    }
                except Exception:
                    pass

        elif pd.api.types.is_datetime64_any_dtype(s):
            # Monthly time-series count -- trend of rows per month
            clean_dt = s.dropna()
            if len(clean_dt) >= 2:
                try:
                    monthly = (
                        clean_dt.dt.to_period("M")
                        .value_counts()
                        .sort_index()
                    )
                    if len(monthly) >= 2:
                        entry["ts_trend"] = {
                            "labels": [str(p) for p in monthly.index],
                            "counts": [int(v) for v in monthly.values],
                        }
                except Exception:
                    pass

        else:
            # Top-N frequency chart for categorical / text columns
            vc = s.value_counts(dropna=True).head(8)
            if len(vc):
                entry["freq_chart"] = {
                    "labels": [str(k) for k in vc.index],
                    "counts": [int(v) for v in vc.values],
                }

        col_profiles.append(entry)

    # -- Potential key candidates (high cardinality, zero nulls) --
    key_candidates = [
        c["name"] for c in col_profiles
        if c["null_count"] == 0 and c["unique_count"] == total and total > 0
    ]

    near_key = [
        c["name"] for c in col_profiles
        if c["null_count"] == 0 and c["unique_pct"] >= 90 and c["name"] not in
        key_candidates
    ]

    # -- Pairwise correlation for numeric columns (top 10 pairs by |corr|) --
    correlations = []
    if len(numeric_cols) >= 2:
        try:
            corr_matrix = df[numeric_cols].corr()
            seen: set = set()
            for i, c1 in enumerate(numeric_cols):
                for c2 in numeric_cols[i + 1:]:
                    pair = (c1, c2)
                    if pair not in seen:
                        seen.add(pair)
                        v = corr_matrix.loc[c1, c2]
                        if not math.isnan(v):
                            correlations.append({
                                "col1": c1, "col2": c2,
                                "corr": round(float(v), 3),
                                "strength": (
                                    "strong"   if abs(v) >= 0.7
                                    else "moderate" if abs(v) >= 0.4
                                    else "weak"
                                ),
                            })
            correlations.sort(key=lambda x: abs(x["corr"]), reverse=True)
            correlations = correlations[:10]
        except Exception:
            pass


    # -- Column type breakdown --
    type_counts: dict[str, int] = {}
    for c in col_profiles:
        sem = c["semantic"]
        type_counts[sem] = type_counts.get(sem, 0) + 1

    return {
        "file_name":  name,
        "total_rows": total,
        "total_cols": len(df.columns),
        "memory_mb":  mem_mb,
        "duplicate_rows": dup_rows,
        "null_cols":  sum(1 for c in col_profiles if c["null_pct"] > 0),
        "complete_cols": sum(1 for c in col_profiles if c["null_pct"] == 0),
        "numeric_cols": len(numeric_cols),
        "key_candidates": key_candidates,
        "near_key_cols": near_key,
        "type_breakdown": type_counts,
        "correlations": correlations,
        "columns":  col_profiles,
    }


# -----------------------------------------------------------------------
# BFSI structural validators (check-digit / format correctness)
# -----------------------------------------------------------------------


def _luhn_checksum(digits: str) -> bool:
    # Validate a numeric string using the Luhn algorithm (used in ISIN check digit).
    total, odd = 0, True

    for d in reversed(digits):
        if not d.isdigit():
            return False

        n = int(d)

        if odd:
            total += n
        else:
            n *= 2


            total += n - 9 if n > 9 else n

        odd = not odd

    return total % 10 == 0


def _isin_valid(value: str) -> bool:
    # Validate ISIN: 2-char country + 9 alphanumeric + 1 Luhn check digit.
    v = value.strip().upper()

    if len(v) != 12 or not v[:2].isalpha() or not v[2:].isalnum():
        return False

    # Expand letters to digits (A=10 ... Z=35) then run Luhn
    expanded = "".join(str(ord(c) - 55) if c.isalpha() else c for c in v)

    return _luhn_checksum(expanded)


def _lei_valid(value: str) -> bool:
    # Validate LEI (ISO 17442): 18 alphanumeric chars + 2 numeric check digits (MOD 97).
    v = value.strip().upper()


    if len(v) != 20 or not v[:18].isalnum() or not v[18:].isdigit():
        return False

    # Move last 2 digits to front, expand letters, check MOD 97 == 1
    rearranged = v[18:] + v[:18]

    expanded = "".join(str(ord(c) - 55) if c.isalpha() else c for c in rearranged)

    try:
        return int(expanded) % 97 == 1
    except ValueError:
        return False


def _cusip_valid(value: str) -> bool:
    # Validate CUSIP: 8 base chars + 1 check digit.
    v = value.strip().upper()

    if len(v) != 9:
        return False

    total = 0

    for i, c in enumerate(v[:8]):
        if c.isdigit():


            n = int(c)
        elif c.isalpha():
            n = ord(c) - 55
        elif c == '*':
            n = 36
        elif c == '@':
            n = 37
        elif c == '#':
            n = 38
        else:
            return False

        if i % 2 == 1:
            n *= 2

        total += n // 10 + n % 10

    check = (10 - (total % 10)) % 10

    return str(check) == v[8]


def _uti_valid(value: str) -> bool:


    """Validate UTI: 1-52 chars, alphanumeric + hyphen/underscore/dot only."""
    v = value.strip()

    return bool(v) and len(v) <= 52 and bool(re.match(r'^[A-Za-z0-9._\-]{1,52}$', v))


def _bic_valid(value: str) -> bool:
    # Validate BIC/SWIFT: 8 or 11 chars, AAAA BB CC [DDD].
    v = value.strip().upper()

    return bool(re.match(r'^[A-Z]{6}[A-Z0-9]{2}([A-Z0-9]{3})?$', v))


# Map identifier type -> validator function
_BFSI_VALIDATORS: dict[str, callable] = {
    "isin":  _isin_valid,
    "lei":   _lei_valid,
    "cusip": _cusip_valid,
    "uti":   _uti_valid,
    "bic":   _bic_valid,
}


def validate_bfsi_identifiers(df: pd.DataFrame, sample_size: int = 500) -> dict[str, dict]:
    # Run structural validation (check-digit / format) on BFSI identifier columns.
    #
    # Returns a dict keyed by column name:
    # {
    #   "identifier_type": "isin",
    #   "total_checked":  450,
    #   "valid_count":    440,
    #   "invalid_count":  10,
    #   "invalid_pct":    2.2,
    #   "invalid_samples": ["US0231351067X", ...],  # up to 5 bad values
    #   "status":     "PASS" | "WARN" | "FAIL",
    # }
    #
    # Only columns whose names match _BFSI_ID_COL_HINTS are checked.
    results: dict[str, dict] = {}

    for col in df.columns:
        col_lc = col.lower().replace("-", "_").replace(" ", "_")
        matched_type = None
        for id_type, hints in _BFSI_ID_COL_HINTS.items():
            if id_type in _BFSI_VALIDATORS and any(h in col_lc for h in hints):
                matched_type = id_type
                break

        if matched_type is None:
            continue

        validator = _BFSI_VALIDATORS[matched_type]
        series = df[col].dropna().astype(str).str.strip()
        sample = series.head(sample_size)
        checked = len(sample)
        if checked == 0:
            continue

        invalid_mask    = ~sample.apply(validator)
        invalid_count   = int(invalid_mask.sum())
        invalid_samples = sample[invalid_mask].head(5).tolist()
        invalid_pct     = round(invalid_count / checked * 100, 1)

        results[col] = {
            "identifier_type": matched_type,
            "total_checked":  checked,
            "valid_count":    checked - invalid_count,
            "invalid_count":  invalid_count,
            "invalid_pct":    invalid_pct,
            "invalid_samples": invalid_samples,
            "status": (
                "PASS" if invalid_pct == 0
                else "WARN" if invalid_pct <= 5
                else "FAIL"
            ),
        }

    return results


# -----------------------------------------------------------------------
# Volume & freshness anomaly detection (Monte Carlo-style)
# -----------------------------------------------------------------------


# Stewardship routing table -- maps BFSI domain -> Jefferies owner group
_STEWARDSHIP_ROUTING: dict[str, str] = {
    "Transactional": "Trade Operations",
    "Analytics":    "Risk / Finance",
    "Reference":    "Reference Data",
    "Regulatory":   "Compliance / Legal",
    "Technical":    "Technology / Data Engineering",
}

# Regulatory domain for columns that trigger specific frameworks
_REGULATORY_DOMAIN_MAP: dict[str, str] = {
    "MiFID II": "Regulatory",
    "EMIR":    "Regulatory",
    "BCBS239": "Regulatory",
    "SOX":     "Regulatory",
    "CFTC":    "Regulatory",
    "GDPR":    "Compliance / Legal",
    "CCPA":    "Compliance / Legal",
    "HIPAA": "Compliance / Legal",
    "PCI":   "Compliance / Legal",
}


def check_volume_and_freshness(
    df: pd.DataFrame,
    fingerprint: str,
    file_name: str,
    volume_threshold_pct: float = 20.0,
    freshness_threshold_days: int = 2,
    username: str = "default",
) -> dict:

    # Monte Carlo-style volume and freshness anomaly detection.

    # Compares current row count against the last saved run stored in
    # feedback_rules.json under category 'run_stats'. Checks max date
    # in any date-like column against today.


    # Returns:
    # {
    # "row_count":       int,
    # "prev_row_count":  int | None,
    # "volume_delta_pct": float | None,
    # "volume_status":   "OK" | "VOLUME_ANOMALY" | "NO_HISTORY",
    # "freshness_checks": [{col, max_date, age_days, status}],
    # "anomalies":       ["..."],  # human-readable alerts
    # }

    from agent.feedback_store import get_rules, save_rule

    row_count = len(df)
    anomalies: list[str] = []

    # -- Volume check --------------------------------------------------------

    prev_row_count   = None
    volume_delta_pct = None


    volume_status  = "NO_HISTORY"

    saved = get_rules(username, fingerprint)

    run_stat = next(
        (r for r in reversed(saved) if r.get("category") == "run_stats"), None
    )

    if run_stat:
        try:
            prev_row_count   = int(run_stat["rule"].split("row_count=")[1].split(",")[0])

            volume_delta_pct = round(
                (row_count - prev_row_count) / max(prev_row_count, 1) * 100, 1
            )

            if abs(volume_delta_pct) > volume_threshold_pct:
                volume_status = "VOLUME_ANOMALY"

                anomalies.append(
                    f"VOLUME_ANOMALY: '{file_name}' has {row_count:,} rows vs "
                    f"{prev_row_count:,} in previous run "
                    f"({volume_delta_pct:+.1f}% -- threshold +-{volume_threshold_pct}%). "
                    f"Possible missing trade batch or duplicate load."


                )
            else:
                volume_status = "OK"

        except (IndexError, ValueError):
            pass

    # Persist current run stats for next comparison
    save_rule(
        username,
        fingerprint,
        f"row_count={row_count}, file={file_name}",
        category="run_stats",
    )


    # -- Freshness check ------------------------------------------------------

    freshness_checks: list[dict] = []

    _DATE_COL_HINTS = {"date", "dt", "time", "ts", "timestamp", "trade", "settle",
             "value_date", "maturity", "expiry", "payment"}

    today = pd.Timestamp.now().normalize()


    for col in df.columns:
        col_lc = col.lower()

        if not any(h in col_lc for h in _DATE_COL_HINTS):
            continue

        series = df[col]

        # Try to parse as datetime if not already
        if not pd.api.types.is_datetime64_any_dtype(series):
            try:
                series = pd.to_datetime(series, errors="coerce")
            except Exception:
                continue

        valid = series.dropna()

        if len(valid) == 0:
            continue

        max_date = valid.max()

        age_days = (today - max_date.normalize()).days

        status  = "STALE" if age_days > freshness_threshold_days else "OK"

        freshness_checks.append({
            "column":  col,
            "max_date": str(max_date.date()),
            "age_days": age_days,
            "status":  status,
        })

        if status == "STALE":
            anomalies.append(
                f"STALE_DATA: '{col}' max date is {max_date.date()} "
                f"({age_days} days old -- threshold {freshness_threshold_days} days). "
                f"Check ETL schedule or source system feed."
            )

    return {
        "row_count":      row_count,
        "prev_row_count": prev_row_count,
        "volume_delta_pct": volume_delta_pct,
        "volume_status":  volume_status,
        "freshness_checks": freshness_checks,
        "anomalies":      anomalies,


    }


# -----------------------------------------------------------------------
# Per-field z-score anomaly detection (Informatica CLAIRE-style)
# -----------------------------------------------------------------------

# BFSI analytics / trading field name hints -- module-level so the
# multi-variate and trend-anomaly detectors below can share the same
# "which columns are domain-critical" definition as the per-column detector.
_ANOMALY_COL_HINTS = {
    "notional", "nominal", "face_value", "principal",
    "price", "rate", "coupon", "yield", "strike", "premium",
    "pnl", "p_l", "mtm", "mark_to_market", "market_value", "npv",
    "delta", "gamma", "vega", "theta", "dv01", "cs01",
    "var", "cvar", "exposure", "stress", "fee", "commission",
    "amount", "amt", "quantity", "qty",
}


def _numeric_columns(df: pd.DataFrame, threshold: float = 0.9) -> dict[str, pd.Series]:
    # _load_file() loads every column as string, so a plain is_numeric_dtype
    # check never matches real uploaded data -- every anomaly/cluster/drift
    # detector in this module needs the same coerce-and-check-hit-rate
    # fallback already used by the column-profiling stats (see analyze_quality_full's
    # per-column stats loop): a column counts as numeric if it's genuinely
    # numeric-dtype OR at least `threshold` of its non-null values parse as
    # numbers. Returns {column: coerced_numeric_series} for qualifying columns
    # only, so callers can just iterate the dict instead of re-deriving this.
    out: dict[str, pd.Series] = {}
    for col in df.columns:
        s = df[col]
        if pd.api.types.is_numeric_dtype(s):
            out[col] = s
            continue
        non_null = s.notna().sum()
        if not non_null:
            continue
        coerced = pd.to_numeric(s, errors="coerce")
        if coerced.notna().sum() / non_null >= threshold:
            out[col] = coerced
    return out


def detect_numeric_anomalies(df: pd.DataFrame, z_threshold: float = 3.0) -> list[dict]:
    # Flag values that are statistical outliers (|z-score| > z_threshold) on
    # BFSI-critical numeric columns: notional, price, rate, P&L, risk metrics.
    #
    # Returns a list of per-column anomaly summaries:
    # {
    #   "column":  str,
    #   "mean":    float,
    #   "std":     float,
    #   "anomaly_count": int,
    #   "anomaly_pct":  float,
    #   "max_z":    float,     # worst offender
    #   "sample_values": [float, ...],
    #   "confidence":  "High" | "Medium" | "Low",  # inverse of anomaly rate
    #   "severity":   "CRITICAL" | "WARN" | "INFO",
    # }
    #
    # Checks BFSI-hinted columns with a z-score/MAD combo (sensitive,
    # severity-scored) and every other numeric column with a plain IQR
    # coverage pass (capped at INFO) -- see _score_column below.
    results: list[dict] = []
    _MOD_Z_THRESHOLD = 3.5  # standard modified-z-score cutoff (Iglewicz & Hoaglin)

    def _score_column(col: str, clean: pd.Series, method: str, default_severity_cap: str | None):
        # Shared scorer used for both the BFSI-hint pass (classic + robust
        # z-score combined) and the generic pass (IQR only, since columns with
        # no domain hint shouldn't get the same z-score sensitivity/severity).
        if len(clean) < 10:
            return None

        mean = float(clean.mean())
        std  = float(clean.std())
        median = float(clean.median())
        mad = float((clean - median).abs().median())

        if method == "zscore":
            if std == 0:
                return None
            z_scores = ((clean - mean) / std).abs()
            # Modified (MAD-based) z-score catches outliers that a mean/std
            # z-score misses when the outliers themselves have already
            # inflated std -- classic z-score is not robust to its own
            # anomalies, MAD is. Union of both flags anything either method
            # would catch.
            mod_z = ((clean - median) / mad).abs() * 0.6745 if mad else pd.Series(0.0, index=clean.index)
            anomaly_mask = (z_scores > z_threshold) | (mod_z > _MOD_Z_THRESHOLD)
            if not anomaly_mask.any():
                return None
            max_z = round(float(max(z_scores[anomaly_mask].max(), mod_z[anomaly_mask].max())), 2)
        else:  # IQR (Tukey) -- distribution-shape-agnostic, used for generic (non-hinted) columns
            q1, q3 = float(clean.quantile(0.25)), float(clean.quantile(0.75))
            iqr = q3 - q1
            if iqr == 0:
                return None
            lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
            anomaly_mask = (clean < lo) | (clean > hi)
            if not anomaly_mask.any():
                return None
            # express as an IQR-multiple "z-like" score for a consistent severity scale
            max_z = round(float(max(
                (clean[anomaly_mask] - hi).clip(lower=0).max(),
                (lo - clean[anomaly_mask]).clip(lower=0).max(),
            ) / (iqr or 1)), 2)

        anomaly_count = int(anomaly_mask.sum())
        anomaly_pct = round(anomaly_count / len(clean) * 100, 2)
        samples = clean[anomaly_mask].head(5).round(4).tolist()

        confidence = (
            "Low"    if anomaly_pct > 10
            else "Medium" if anomaly_pct > 2
            else "High"
        )

        if default_severity_cap == "INFO":
            severity = "INFO"  # generic (non-domain) columns never auto-escalate to CRITICAL/WARN
        else:
            severity = (
                "CRITICAL" if max_z > 6 or anomaly_pct > 5
                else "WARN" if max_z > 4
                else "INFO"
            )

        return {
            "column":  col,
            "mean":    round(mean, 4),
            "std":     round(std, 4),
            "anomaly_count": anomaly_count,
            "anomaly_pct":  anomaly_pct,
            "max_z":   max_z,
            "samples": samples,
            "confidence": confidence,
            "severity": severity,
            "method":  "robust z-score" if method == "zscore" else "IQR (Tukey)",
            "domain_hint": default_severity_cap != "INFO",
        }

    numeric_cols = _numeric_columns(df)

    hinted_cols = []
    for col, series in numeric_cols.items():
        col_lc = col.lower().replace("-", "_").replace(" ", "_")
        if not any(h in col_lc for h in _ANOMALY_COL_HINTS):
            continue
        hinted_cols.append(col)
        clean = series.dropna()
        r = _score_column(col, clean, "zscore", None)
        if r:
            results.append(r)

    # Generic pass -- BFSI-named columns get the sensitive z-score/severity
    # treatment above, but every other numeric column still gets a coverage
    # check (IQR, capped at INFO) rather than being silently skipped entirely
    # just because its name didn't match a domain hint.
    for col, series in numeric_cols.items():
        if col in hinted_cols:
            continue
        clean = series.dropna()
        r = _score_column(col, clean, "iqr", "INFO")
        if r:
            results.append(r)

    return results


# -----------------------------------------------------------------------
# Multi-variate anomaly detection -- robust Mahalanobis distance across the
# BFSI-critical numeric columns jointly. A row can be individually in-range
# on every single column (so the per-column z-score pass above finds
# nothing) and still be jointly inconsistent -- e.g. quantity and price are
# each normal, but their combination breaks the usual notional relationship.
# This is the kind of "ML-native" multi-column check that dedicated data
# observability platforms (Monte Carlo, Collibra, Ataccama) lead on; done
# here with plain numpy (covariance + pseudo-inverse), no sklearn dependency.
# -----------------------------------------------------------------------


def detect_multivariate_anomalies(df: pd.DataFrame, max_cols: int = 6, threshold: float = 3.5) -> dict:
    numeric_cols = list(_numeric_columns(df).keys())
    if len(numeric_cols) < 2:
        return {"checked_columns": [], "anomaly_count": 0, "rows": []}

    hinted = [c for c in numeric_cols
              if any(h in c.lower().replace("-", "_").replace(" ", "_") for h in _ANOMALY_COL_HINTS)]
    cols = (hinted if len(hinted) >= 2 else numeric_cols)[:max_cols]

    sub = df[cols].apply(pd.to_numeric, errors="coerce").dropna()
    if len(sub) < 30 or len(cols) < 2:
        return {"checked_columns": cols, "anomaly_count": 0, "rows": []}

    X = sub.to_numpy(dtype=float)
    # drop columns with zero variance -- singular covariance otherwise
    keep = X.std(axis=0) > 0
    if keep.sum() < 2:
        return {"checked_columns": cols, "anomaly_count": 0, "rows": []}
    X = X[:, keep]
    cols = [c for c, k in zip(cols, keep) if k]

    mean = X.mean(axis=0)
    cov = np.cov(X, rowvar=False)
    inv_cov = np.linalg.pinv(cov)
    diff = X - mean
    m_dist = np.sqrt(np.clip(np.einsum('ij,jk,ik->i', diff, inv_cov, diff), 0, None))

    # Robust (MAD-based) threshold on the distance itself, rather than a
    # fixed chi-square critical value -- keeps behaviour consistent with the
    # rest of this module's threshold-free philosophy.
    d_median = float(np.median(m_dist))
    d_mad = float(np.median(np.abs(m_dist - d_median))) or 1e-9
    mod_z_dist = 0.6745 * (m_dist - d_median) / d_mad
    mask = mod_z_dist > threshold

    anomaly_count = int(mask.sum())
    rows = []
    if anomaly_count:
        order = np.argsort(-mod_z_dist)
        for pos in order[:10]:
            if not mask[pos]:
                continue
            row_idx = sub.index[pos]
            row = {c: round(float(v), 4) for c, v in zip(cols, X[pos])}
            row["row_index"] = int(row_idx)
            row["distance_z"] = round(float(mod_z_dist[pos]), 2)
            rows.append(row)

    anomaly_pct = round(anomaly_count / len(sub) * 100, 2) if len(sub) else 0.0
    return {
        "checked_columns": cols,
        "row_count_checked": int(len(sub)),
        "anomaly_count": anomaly_count,
        "anomaly_pct": anomaly_pct,
        "rows": rows,
        "severity": "CRITICAL" if anomaly_pct > 5 else "WARN" if anomaly_count else "INFO",
        "detail": (
            f"{anomaly_count} row(s) are jointly anomalous across {', '.join(cols)} even though "
            f"no single column crossed its own threshold." if anomaly_count else ""
        ),
    }


# -----------------------------------------------------------------------
# Trend / time-aware anomaly detection -- buckets rows by day (or week, if
# too few distinct days) using the first date-like column found, then flags
# BFSI-critical numeric columns whose most recent period mean deviates from
# the trailing rolling baseline. A flat global z-score across the whole file
# would average a gradual drift or a sudden regime shift away; this catches
# it. Lightweight stand-in for the seasonality-aware checks dedicated data
# observability platforms (Monte Carlo, Collibra) run as their core product.
# -----------------------------------------------------------------------


def detect_trend_anomalies(df: pd.DataFrame, min_periods: int = 6) -> list[dict]:
    date_col = None
    dates = None
    for c in df.columns:
        cl = c.lower().replace("-", "_").replace(" ", "_")
        if any(h in cl for h in ("date", "as_of", "trade_date", "value_date", "timestamp", "settlement_date")):
            parsed = pd.to_datetime(df[c], errors="coerce")
            if parsed.notna().mean() > 0.8:
                date_col = c
                dates = parsed
                break
    if date_col is None:
        return []

    n_days = dates.dt.normalize().nunique()
    period = dates.dt.to_period("D" if n_days >= min_periods else "W")

    numeric_cols = _numeric_columns(df)
    results = []
    for col, series in numeric_cols.items():
        col_lc = col.lower().replace("-", "_").replace(" ", "_")
        if not any(h in col_lc for h in _ANOMALY_COL_HINTS):
            continue

        tmp = pd.DataFrame({"period": period, "val": series}).dropna()
        grp = tmp.groupby("period")["val"].mean().sort_index()
        if len(grp) < min_periods:
            continue

        rolling_mean = grp.shift(1).expanding().mean()
        rolling_std = grp.shift(1).expanding().std()
        base_mean = rolling_mean.iloc[-1]
        base_std = rolling_std.iloc[-1]
        if pd.isna(base_std) or base_std == 0:
            continue

        latest_val = float(grp.iloc[-1])
        z = abs((latest_val - float(base_mean)) / float(base_std))
        if z < 2.5:
            continue

        severity = "CRITICAL" if z > 4 else "WARN"
        results.append({
            "column": col,
            "date_column": date_col,
            "period": str(grp.index[-1]),
            "latest_value": round(latest_val, 4),
            "baseline_mean": round(float(base_mean), 4),
            "baseline_std": round(float(base_std), 4),
            "z_score": round(float(z), 2),
            "severity": severity,
            "detail": (
                f"Most recent period ({grp.index[-1]}) mean for '{col}' is {z:.1f} std "
                f"from the trailing baseline -- possible regime shift or data issue."
            ),
        })

    return results


# -----------------------------------------------------------------------
# Unsupervised numeric clustering -- 1D KMeans-style natural groupings
# -----------------------------------------------------------------------


def detect_numeric_clusters(df: pd.DataFrame, max_k: int = 5) -> list[dict]:
    # Detect natural value groupings in numeric columns without requiring
    # a predefined threshold (unlike z-score).
    #
    # Uses 1D gap-statistics heuristic:
    # 1. Sort values and compute gaps between consecutive unique values
    # 2. Identify large gaps (> mean + 2*std) as natural cluster boundaries
    # 3. Report clusters with their range, count, and centre
    #
    # Returns one entry per column that has 2+ distinct clusters.
    # Useful for detecting:
    # - Fee tiers (0-100, 100-500, 500+ basis points)
    # - Rating buckets (1-3, 4-6, 7-10)
    # - Transaction size bands (retail vs institutional)
    # - Erroneous data mixing two populations (e.g. USD and JPY amounts)
    results = []

    for col, series in _numeric_columns(df).items():
        clean = series.dropna()
        n = len(clean)
        if n < 20:
            continue

        unique_vals = clean.sort_values().unique()
        if len(unique_vals) < 4:
            continue

        # Compute gaps between consecutive unique values
        gaps = unique_vals[1:] - unique_vals[:-1]
        if gaps.std() == 0:
            continue

        # Natural cluster boundary = gap > mean + 1.5*std
        threshold = gaps.mean() + 1.5 * gaps.std()
        boundary_indices = [i for i, g in enumerate(gaps) if g > threshold]

        if not boundary_indices:
            continue

        # Build cluster ranges
        boundaries = [-1] + boundary_indices + [len(unique_vals) - 1]
        clusters = []
        for i in range(len(boundaries) - 1):
            lo_idx = boundaries[i] + 1
            hi_idx = boundaries[i + 1]
            cluster_vals = clean[(clean >= unique_vals[lo_idx]) & (clean <=
                                  unique_vals[hi_idx])]
            clusters.append({
                "min":    round(float(unique_vals[lo_idx]), 4),
                "max":    round(float(unique_vals[hi_idx]), 4),
                "centre": round(float(cluster_vals.mean()), 4),
                "count":  int(len(cluster_vals)),
                "pct":    round(len(cluster_vals) / n * 100, 1),
            })

        if len(clusters) < 2:
            continue

        # Flag if clusters look suspicious (very unequal sizes or extreme separation)
        sizes = [c["count"] for c in clusters]
        max_ratio = max(sizes) / (min(sizes) or 1)
        severity = "WARN" if max_ratio > 10 else "INFO"

        results.append({
            "column":  col,
            "clusters": clusters[:max_k],
            "n_clusters": len(clusters),
            "max_size_ratio": round(max_ratio, 1),
            "severity": severity,
            "detail":  f"{len(clusters)} natural value groups detected "
                       + (f"-- dominant cluster {max_ratio:.0f}x larger than smallest (possible "
                          f"mixed populations)" if max_ratio > 10 else ""),
        })

    return results


# -----------------------------------------------------------------------
# Categorical distribution clustering (Ataccama-style unsupervised drift)
# -----------------------------------------------------------------------


def detect_categorical_drift(
    df: pd.DataFrame,
    baseline_snapshot: dict | None = None,
    max_cardinality: int = 200,


) -> list[dict]:

    # Detect distribution shift in categorical (string/low-cardinality) columns
    # without requiring a predefined threshold.

    # Two modes:
    # 1. With baseline: compare value frequency distribution vs saved snapshot
    # using Jensen-Shannon divergence -- flags columns where distribution
    # changed significantly.
    # 2. Without baseline: flag columns with suspicious patterns --
    # very high singleton rate (each value appears once = possible ID leak),
    # sudden skew (one value dominates >95%), or high entropy collapse.

    # Returns list of findings per column.

    import math

    results = []


    def _entropy(counts: dict) -> float:
        total = sum(counts.values())
        if not total:
            return 0.0
        return -sum((c / total) * math.log2(c / total + 1e-12) for c in counts.values())


    def _js_divergence(p: dict, q: dict) -> float:
        """Jensen-Shannon divergence between two frequency dicts (0=identical, 1=max)."""
        keys = set(p) | set(q)

        tot_p = sum(p.values()) or 1

        tot_q = sum(q.values()) or 1

        pv = {k: p.get(k, 0) / tot_p for k in keys}

        qv = {k: q.get(k, 0) / tot_q for k in keys}

        mv = {k: (pv[k] + qv[k]) / 2 for k in keys}

        def kl(a, m):
            return sum(a[k] * math.log2(a[k] / m[k] + 1e-12) for k in keys if a[k] > 0)

        return (kl(pv, mv) + kl(qv, mv)) / 2


    baseline_cols = {}

    if baseline_snapshot:
        for bc in baseline_snapshot.get("columns", []):
            if bc.get("value_counts"):
                baseline_cols[bc["name"]] = bc["value_counts"]


    _numeric_cols_for_exclusion = _numeric_columns(df)
    for col in df.columns:
        s = df[col]

        # Skip numeric columns and very high cardinality (IDs). _load_file()
        # loads every column as string, so a plain is_numeric_dtype check
        # would never exclude anything -- use the same coerce-and-check-rate
        # fallback as the other detectors in this module.
        if col in _numeric_cols_for_exclusion:
            continue

        n_unique = int(s.nunique(dropna=True))

        n_total  = len(s.dropna())

        if n_total == 0 or n_unique > max_cardinality:
            continue


        vc = s.value_counts(dropna=True).to_dict()

        top_val, top_cnt = max(vc.items(), key=lambda x: x[1]) if vc else ("", 0)

        top_pct = round(top_cnt / n_total * 100, 1) if n_total else 0


        singleton_pct = round(sum(1 for c in vc.values() if c == 1) / n_total * 100, 1)

        entropy_val = round(_entropy(vc), 3)


        # Mode 1: compare vs baseline
        if col in baseline_cols:
            jsd = round(_js_divergence(vc, baseline_cols[col]), 4)

            if jsd < 0.05:
                continue  # negligible change

            severity = "CRITICAL" if jsd > 0.3 else "WARN" if jsd > 0.1 else "INFO"

            new_vals = [v for v in vc if v not in baseline_cols[col]][:5]

            gone_vals = [v for v in baseline_cols[col] if v not in vc][:5]

            results.append({
                "column":  col,
                "type":    "distribution_shift",
                "severity": severity,
                "js_divergence": jsd,
                "new_values":  new_vals,
                "removed_values": gone_vals,
                "top_value":  str(top_val),


                "top_pct":  top_pct,
                "entropy":  entropy_val,
                "detail": f"Distribution shifted (JSD={jsd:.3f}). "
                    + (f"New values: {new_vals}. " if new_vals else "")
                    + (f"Removed: {gone_vals}." if gone_vals else ""),
            })

        else:
            # Mode 2: flag suspicious patterns without baseline
            issues = []
            severity = "INFO"

            if top_pct > 95 and n_unique > 1:
                issues.append(f"'{top_val}' dominates at {top_pct}%")
                severity = "WARN"

            if singleton_pct > 80 and n_unique > 10:
                issues.append(f"{singleton_pct}% values are unique (possible ID column)")
                severity = "WARN"

            if not issues:
                continue

            results.append({


                "column":  col,
                "type":    "suspicious_distribution",
                "severity": severity,
                "js_divergence": None,
                "new_values":  [],
                "removed_values": [],
                "top_value":  str(top_val),
                "top_pct":  top_pct,
                "entropy":  entropy_val,
                "detail":  " | ".join(issues),
            })

    return results


# -----------------------------------------------------------------------
# Data Governance -- enriched with data dictionary
# -----------------------------------------------------------------------


_PII_REGEX = {
    "email":       re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b'),
    "phone":       re.compile(r'\b(\+\d{1,3}[\s.-]?)?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}\b'),
    "ssn":         re.compile(r'\b\d{3}-\d{2}-\d{4}\b'),
    "credit_card": re.compile(r'\b\d{4}[- ]?\d{4}[- ]?\d{4}[- ]?\d{4}\b'),
    "ip_address":  re.compile(r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b'),
    "date_dob":    re.compile(r'\b\d{4}[-/]\d{2}[-/]\d{2}\b'),
    "iban":        re.compile(r'\b[A-Z]{2}\d{2}[A-Z0-9]{4}\d{7,}\b'),
}

# BFSI-specific identifier patterns (separate from consumer PII)
_BFSI_ID_REGEX: dict[str, re.Pattern] = {
    "isin":  re.compile(r'\b[A-Z]{2}[A-Z0-9]{9}[0-9]\b'),
    "lei":   re.compile(r'\b[A-Z0-9]{18}[0-9]{2}\b'),
    "cusip": re.compile(r'\b[A-Z0-9]{9}\b'),
    "uti":   re.compile(r'\b[A-Z0-9]{4}[A-Z0-9]{1,48}\b'),
    "bic":   re.compile(r'\b[A-Z]{6}[A-Z0-9]{2}([A-Z0-9]{3})?\b'),
}


# BFSI column name hints for regulatory identifier detection
_BFSI_ID_COL_HINTS: dict[str, list[str]] = {
    "isin":       ["isin", "isin_code", "security_id", "sec_id"],
    "lei":        ["lei", "legal_entity_id", "lei_code", "reporting_entity"],
    "cusip":      ["cusip", "cusip_code"],
    "uti":        ["uti", "unique_trade_id", "usi", "usi_namespace"],
    "bic":        ["bic", "swift", "bic_code", "swift_code"],
    "mic_code":   ["mic", "mic_code", "venue", "exchange_code"],
    "cleared":    ["cleared", "is_cleared", "clearing_flag"],
    "ccp":        ["ccp", "clearing_member", "clearing_house"],
    "reporting_entity": ["reporting_entity", "reporting_obligation", "emir_flag",
"mifid_flag", "cftc_flag"],
    "pnl":        ["pnl", "p_l", "profit_loss", "profit_and_loss", "realised_pnl",
"unrealised_pnl"],
    "mtm":        ["mtm", "mark_to_market", "fair_value", "market_value", "mkt_val"],
    "var":        ["var", "cvar", "value_at_risk", "expected_shortfall"],
    "notional":   ["notional", "notional_amount", "face_value", "face_amount"],
    "risk_factor": ["risk_factor", "exposure", "stress", "sensitivity"],
}


# Columns that must not be null under specific regulatory frameworks
_BFSI_MANDATORY_FIELDS: dict[str, list[str]] = {
    "MiFID II": ["lei", "uti", "mic_code", "venue", "reporting_entity",
            "mifid_flag", "cleared", "instrument_type", "asset_class"],
    "EMIR":    ["uti", "cleared", "ccp", "reporting_entity", "emir_flag",
            "counterparty", "trade_date", "maturity_date"],
    "BCBS239": ["var", "cvar", "pnl", "mtm", "exposure", "risk_factor",
            "notional", "stress"],
    "SOX":     ["pnl", "revenue", "balance", "approval", "authorised_by",
            "audit_trail"],
    "CFTC":    ["usi", "usi_namespace", "cftc_flag", "swap_dealer",
            "reporting_entity"],
}

_PII_COL_HINTS: dict[str, list[str]] = {
    "name":   ["name", "firstname", "lastname", "fullname", "first_name", "last_name",
"surname"],
    "email":  ["email", "mail"],
    "phone":  ["phone", "tel", "mobile", "cell", "fax", "contact"],


    # "sin" removed -- too short, matches 'business', 'isin', 'processing'
    # "ssn" kept as exact word only (enforced by word-boundary match below)
    "ssn":    ["ssn", "social_security", "tax_id", "nin", "national_insurance"],
    "dob":    ["dob", "birth", "birthdate", "date_of_birth", "birthday"],
    "address": ["address", "addr", "street", "city", "zip", "postal", "postcode"],
    "gender": ["gender", "sex"],
    "salary": ["salary", "wage", "income", "pay", "compensation"],
    # "bic" removed -- it's a SWIFT/BIC banking identifier, not consumer PII
    "account": ["account", "acct", "iban", "routing"],
    "credit_card": ["card", "cc_num", "credit_card", "cvv", "pan"],
    # "ip" removed -- too short, matches 'cus_ip', 'equip' etc.
    "ip_address": ["ipaddr", "ip_address", "client_ip", "user_ip"],
    "national_id": ["passport", "national_id", "id_number", "driver", "license"],
    "race":   ["race", "ethnicity", "nationality"],
}


_REGULATORY_TAGS: dict[str, list[str]] = {
    # Consumer / personal data frameworks


    "GDPR": ["email", "phone", "name", "dob", "address", "ip_address", "national_id",
"gender", "race"],
    "CCPA": ["email", "phone", "name", "address", "ip_address"],
    "HIPAA": ["ssn", "dob", "name", "address", "phone", "national_id"],
    "PCI":  ["credit_card", "account"],
    # BFSI regulatory frameworks -- keyed on BFSI identifier / field types
    "MiFID II": ["lei", "uti", "mic_code", "cleared", "reporting_entity",
            "isin", "venue", "instrument_type", "asset_class"],
    "EMIR":    ["uti", "cleared", "ccp", "reporting_entity", "counterparty",
            "trade_date", "maturity_date"],
    "BCBS239": ["var", "pnl", "mtm", "exposure", "risk_factor", "notional", "stress"],
    "SOX":    ["pnl", "revenue", "balance", "approval", "audit_trail"],
    "CFTC":   ["usi", "usi_namespace", "cftc_flag", "reporting_entity"],
}


def _parse_sensitivity_from_text(text: str) -> str | None:

    # Extract intended sensitivity tier from a plain-English override rule,


    # handling negation correctly.

    # "CUSIP is Public" -> "Public"
    # "not Internal Use Only; should be Public" -> "Public"  (not the negated one)
    # "mark as Confidential" -> "Confidential"
    # "Highly Restricted" -> "Highly Restricted"

    import re as _re

    t = text.lower()

    # Build a map of tier -> positions where it appears NOT negated
    _TIERS = [
        ("Highly Restricted", ["highly restricted"]),
        ("Internal Use Only", ["jefferies use only", "jefferies"]),
        ("Confidential",     ["confidential"]),
        ("Public",           ["public", "reference/public", "reference data"]),
    ]

    # Negation words/phrases that precede a tier mention


    _NEG = r"(?:not|non|isn.t|is not|no longer|shouldn.t be|should not be|remove|clear|or|nor|proprietary,?\s+sensitive,?\s+or)\s+"  # OCR-UNCERTAIN: regex pattern approximate

    candidates = []

    for tier, keywords in _TIERS:
        for kw in keywords:
            # Check if this keyword appears WITHOUT a negation prefix (within 30 chars)
            for m in _re.finditer(_re.escape(kw), t):
                start = max(0, m.start() - 30)

                prefix = t[start:m.start()]

                if _re.search(_NEG, prefix):
                    continue  # this mention is negated -- skip

                candidates.append((m.start(), tier))

            break

    if not candidates:
        return None

    # Return the tier of the LAST non-negated mention
    # ("not Internal Use Only; classify as Public" -> Public wins)


    candidates.sort(key=lambda x: x[0])

    return candidates[-1][1]


def _parse_regulatory_override(text: str) -> list[str] | None:

    # Extract an explicit list of regulatory frameworks from a plain-English rule.

    # "MiFID II only; no other frameworks" -> ["MiFID II"]
    # "limited to MiFID II" -> ["MiFID II"]
    # "GDPR and CCPA" -> ["GDPR", "CCPA"]
    # "no regulatory frameworks" / "not subject to any regulation" -> []

    # Returns None if no regulatory instruction found (= don't override).

    t = text.lower()

    # Detect "no frameworks at all" instructions
    _NO_REG = ("no regulatory", "no regulation", "not subject to", "no framework",
        "not regulated", "exempt from", "no compliance", "none apply",


        "no other regulatory", "no other framework")
    if any(p in t for p in _NO_REG):
        # If "only" a specific framework is mentioned, return just that one
        pass  # fall through to detect which ones ARE mentioned

    _ALL_FRAMEWORKS = {
        "MiFID II": ["mifid ii", "mifid2", "mifid 2"],
        "EMIR":    ["emir"],
        "BCBS239":  ["bcbs239", "bcbs 239"],
        "SOX":    ["sox", "sarbanes"],
        "CFTC":    ["cftc"],
        "GDPR":    ["gdpr"],
        "CCPA":    ["ccpa"],
        "HIPAA":    ["hipaa"],
        "PCI":    ["pci", "pci-dss", "pci dss"],
    }

    found = []
    for framework, keywords in _ALL_FRAMEWORKS.items():


        if any(kw in t for kw in keywords):
            found.append(framework)

    # "only X; no other" -> return just X (already in found, no other present)
    # "no other regulatory frameworks apply" with specific one mentioned -> return found
    _limiting = ("only", "limited to", "solely", "exclusively", "just",
        "no other", "nothing else")
    has_limiting = any(p in t for p in _limiting)
    _none_apply  = any(p in t for p in _NO_REG)

    if found and (has_limiting or _none_apply):
        return found        # explicit subset
    if not found and _none_apply:
        return []        # explicitly no frameworks
    if found:
        return found        # frameworks explicitly mentioned without limiting
    return None        # no regulatory instruction -- don't override


def _sensitivity_level(pii: list[str], bfsi_types: set[str] | None = None) -> str:
    types = {p.split(" ")[0] for p in pii}
    if types & {"ssn", "credit_card", "iban", "national_id", "dob"}:
        return "Highly Restricted"
    if types & {"email", "phone", "address", "salary", "account", "name", "race"}:
        return "Confidential"
    # BFSI internal data: sensitive trading/risk/analytics columns that are not personal
    # data but must be restricted to authorised internal users
    if bfsi_types and bfsi_types & {"pnl", "mtm", "var", "notional", "risk_factor",
                    "cleared", "lei", "uti", "reporting_entity"}:
        return "Internal Use Only"
    if bfsi_types and bfsi_types & {"isin", "cusip", "mic_code", "bic", "ccp"}:
        return "Internal Use Only"
    return "Public"


def _regulatory_flags(pii_types: set[str], bfsi_types: set[str] | None = None) -> list[str]:
    combined = pii_types | (bfsi_types or set())
    return [reg for reg, types in _REGULATORY_TAGS.items() if combined & set(types)]


def _bfsi_access_recommendation(sensitivity: str, reg_flags: list[str],
                bfsi_types: set[str]) -> str:
    """Return a domain-specific access recommendation for a BFSI column."""
    if "MiFID II" in reg_flags or "EMIR" in reg_flags or "CFTC" in reg_flags:
        return (
            "Regulatory reporting access required -- restrict to compliance/legal team; "
            "immutable after submission; access log mandatory"
        )
    if "SOX" in reg_flags or bfsi_types & {"pnl", "mtm", "revenue", "balance"}:
        return (
            "Front-office read-only -- restrict write access; "
            "audit trail required (SOX); finance/risk team only"
        )
    if "BCBS239" in reg_flags or bfsi_types & {"var", "risk_factor", "exposure", "stress"}:
        return (
            "Risk team access only -- aggregation accuracy governed by BCBS239; "
            "changes require risk data owner sign-off"


        )
    if sensitivity == "Highly Restricted":
        return "Restrict – row-level security / masking required in all environments"
    if sensitivity == "Confidential":
        return "Limit – role-based access control recommended; mask in non-prod"
    if sensitivity == "Internal Use Only":
        return (
            "Internal use only -- restrict to authorised business users; "
            "do not expose in external APIs or reports"
        )
    return "Standard access controls"


def analyze_governance(df: pd.DataFrame, name: str,
            data_dict: dict | None = None,
            user_hints: dict | None = None) -> dict:
    hints = user_hints or {}
    data_dict = data_dict or {}
    sample = df.head(500)


    findings = []

    # Parse pii_context hint: "no PII", "internal only", or "no customer data"
    # signal that auto-detected PII findings should be downgraded to informational.
    _pii_ctx = hints.get("pii_context", "").lower()
    _hint_no_pii = any(kw in _pii_ctx for kw in ("no pii", "no customer", "internal only",
"non-pii"))

    # Dataset Controls column-level overrides (saved via dc_governance_override rules)
    # Format: {"CUSIP": {"sensitivity": "Public", "clear_pii": True}}
    # These take highest priority -- user explicitly saved these overrides.
    _dc_overrides: dict = hints.get("_dc_col_overrides", {})

    # Build a set of all column names (normalised) for cross-column conditional checks
    all_col_lc = {c.lower().replace("-", "_").replace(" ", "_") for c in df.columns}

    for col in df.columns:
        detected: list[str] = []
        bfsi_detected: list[str] = []


        col_lc = col.lower().replace("-", "_").replace(" ", "_")

        # 0. Dataset Controls override -- highest priority, applied before any detection
        _dc_col_override = _dc_overrides.get(col, {})

        # If excluded via Dataset Controls, skip this column entirely from governance
        if _dc_col_override.get("exclude"):
            continue

        # 1. Override from data dictionary
        dd = data_dict.get(col, {})
        if dd.get("sensitivity"):
            detected.append(f"classified:{dd['sensitivity']} (data dictionary)")

        # 2. Consumer PII -- name-based hints
        # Use word-boundary matching: hint must appear as a whole token
        # (delimited by _, space, or start/end of string) to avoid false positives
        # e.g. "ssn" should NOT match "businessdate", "isin", "processingstatus"
        import re as _re_pii


        _col_tokens = set(_re_pii.split(r'[_\s]+', col_lc))
        for pii_type, _col_hints in _PII_COL_HINTS.items():
            matched = False
            for h in _col_hints:
                # Short hints (<=3 chars) must match as exact token
                # Longer hints can match as substring for compound names
                if len(h) <= 3:
                    if h in _col_tokens or col_lc == h:
                        matched = True
                        break
                else:
                    if h in col_lc:
                        matched = True
                        break
            if matched:
                if not any(pii_type in p for p in detected):
                    detected.append(f"{pii_type} (column name)")
                break


        # 3. Consumer PII -- content regex scan
        # Skip columns whose names clearly indicate business/financial dates or IDs --
        # these are not personal data even if their values look date-like
        _BFSI_BUSINESS_COL_NAMES = {
            "businessdate", "positiondate", "tradedate", "settlementdate",
            "maturitydate", "expirydate", "valuationdate", "processingdate",
            "previousbusinessdate", "lastactivitydate", "snapshotdate",
            "snapshottime", "snapshotimestamp", "lastupdated", "previousdate",
            "marketvalue", "marketvaluebase", "bookvalue", "facevalue",
            "notional", "currentfaceqty", "origfaceqty", "actualqty",
        }
        _skip_content_scan = col_lc.replace("_","").replace(" ","") in _BFSI_BUSINESS_COL_NAMES
        if df[col].dtype == object and not _skip_content_scan:
            vals = sample[col].dropna().astype(str)
            n_vals = max(len(vals), 1)
            for pii_type, pat in _PII_REGEX.items():
                if not any(pii_type in p for p in detected):
                    hits = int(vals.str.contains(pat, regex=True, na=False).sum())


                    # Require at least 1% hit rate (min 2 hits) to reduce false positives
                    # from financial data that incidentally matches date/phone patterns
                    min_hits = max(2, int(n_vals * 0.01))
                    if hits >= min_hits:
                        detected.append(f"{pii_type} (content, {hits}/{len(vals)} sampled)")

        # 4. BFSI identifier -- name-based hints
        for bfsi_type, col_hints in _BFSI_ID_COL_HINTS.items():
            if any(h in col_lc for h in col_hints):
                if not any(bfsi_type in b for b in bfsi_detected):
                    bfsi_detected.append(f"{bfsi_type} (column name)")

        # 5. BFSI identifier -- content regex scan (object columns only)
        if df[col].dtype == object:
            vals = sample[col].dropna().astype(str)
            for bfsi_type, pat in _BFSI_ID_REGEX.items():
                if not any(bfsi_type in b for b in bfsi_detected):
                    hits = int(vals.str.fullmatch(pat, na=False).sum())
                    if hits >= max(1, len(vals) // 5):


                        bfsi_detected.append(
                            f"{bfsi_type} (content, {hits}/{len(vals)} sampled)"
                        )

        pii_types_set  = {p.split(" ")[0].replace("classified:", "") for p in detected}
        bfsi_types_set = {b.split(" ")[0] for b in bfsi_detected}

        # When user states this dataset has no customer PII, demote auto-detected
        # name/content findings (data-dict overrides are always kept).
        if _hint_no_pii:
            dict_detected = [p for p in detected if "data dictionary" in p]
            auto_detected  = [p for p in detected if "data dictionary" not in p]
            if auto_detected and not dict_detected:
                detected = [p + " [overridden: user context]" for p in auto_detected]
                pii_types_set = set()

        sensitivity = (
            dd.get("sensitivity")
            if dd.get("sensitivity") in (


                "Highly Restricted", "Confidential", "Internal Use Only", "Public"
            )
            else ("Public" if _hint_no_pii and not any("data dictionary" in p for p in detected)  # OCR-UNCERTAIN
                else _sensitivity_level(detected, bfsi_types_set))
        )

        reg_flags = _regulatory_flags(pii_types_set, bfsi_types_set)

        # 6. Mandatory field null check -- flag regulatory columns that contain nulls
        null_count = int(df[col].isna().sum())
        null_pct  = round(null_count / len(df) * 100, 1) if len(df) else 0
        mandatory_breach: list[str] = []
        for framework, mandatory_cols in _BFSI_MANDATORY_FIELDS.items():
            if framework in reg_flags and null_count > 0:
                if any(m in col_lc for m in mandatory_cols):
                    mandatory_breach.append(
                        f"{framework}: mandatory field has {null_pct}% nulls "
                        f"({null_count:,} rows) -- potential regulatory reporting breach"
                    )


        # 7. Cross-column conditional checks (EMIR: cleared=Y requires ccp)
        conditional_warnings: list[str] = []
        if any(h in col_lc for h in ["cleared", "is_cleared", "clearing_flag"]):
            if df[col].dtype == object:
                cleared_y = df[col].astype(str).str.upper().isin({"Y", "YES", "TRUE", "1"}).sum()
                if cleared_y > 0:
                    ccp_cols = [c for c in all_col_lc
                            if any(h in c for h in ["ccp", "clearing_member", "clearing_house"])]
                    if not ccp_cols:
                        conditional_warnings.append(
                            f"EMIR: {cleared_y:,} rows have cleared=Y but no CCP/clearing_member "
                            f"column found in dataset -- EMIR reporting incomplete"
                        )
                    else:
                        # CCP column exists -- check it's not null where cleared=Y
                        ccp_col_actual = next(
                            (c for c in df.columns
                                if c.lower().replace("-", "_").replace(" ", "_") in ccp_cols), None


                        )
                        if ccp_col_actual is not None:
                            cleared_mask = df[col].astype(str).str.upper().isin(
                                {"Y", "YES", "TRUE", "1"}
                            )
                            ccp_nulls = df.loc[cleared_mask, ccp_col_actual].isna().sum()
                            if ccp_nulls > 0:
                                conditional_warnings.append(
                                    f"EMIR: {ccp_nulls:,} rows have cleared=Y but "
                                    f"'{ccp_col_actual}' is null -- EMIR mandatory field breach"
                                )

        # -- Apply Dataset Controls overrides (highest priority, all types) ------
        # User saved these once via Dataset Controls -- apply on every run.
        if _dc_col_override.get("sensitivity"):
            sensitivity = _dc_col_override["sensitivity"]
        if _dc_col_override.get("regulatory") is not None:
            # Explicit list of frameworks (may be empty = no frameworks)
            reg_flags        = _dc_col_override["regulatory"]


            mandatory_breach    = []  # clear breaches -- user said these don't apply
            conditional_warnings = []
        if _dc_col_override.get("clear_pii"):
            detected     = [d for d in detected if "data dictionary" in d]
            bfsi_detected  = []
            pii_types_set  = set()
            bfsi_types_set = set()
            if _dc_col_override.get("regulatory") is None:
                # Only clear reg_flags if not already set by regulatory override
                reg_flags        = []
                mandatory_breach    = []
                conditional_warnings = []

        # Stewardship routing -- derive Jefferies owner group from domain + regulatory context
        op_usage = _infer_operational_usage(col)
        if reg_flags and any(r in _REGULATORY_DOMAIN_MAP for r in reg_flags):
            # Regulatory frameworks take precedence for routing
            escalate_to = next(


                (_REGULATORY_DOMAIN_MAP[r] for r in reg_flags if r in
_REGULATORY_DOMAIN_MAP),
                _STEWARDSHIP_ROUTING.get(op_usage, "Technology / Data Engineering")
            )
        else:
            escalate_to = _STEWARDSHIP_ROUTING.get(op_usage, "Technology / Data Engineering")

        findings.append({
            "column":       col,
            "dtype":       str(df[col].dtype),
            "null_count":     null_count,
            "null_pct":     null_pct,
            "pii_detected":    detected,
            "bfsi_identifiers": bfsi_detected,
            "sensitivity":    sensitivity,
            "regulatory":     reg_flags,
            "mandatory_breaches": mandatory_breach,
            "conditional_warnings": conditional_warnings,
            "owner":       dd.get("owner", ""),


            "description":    dd.get("description", ""),
            "business_term":   dd.get("business_term", ""),
            "from_dict":     col in data_dict,
            "escalate_to":    escalate_to,
            "access_rec":     _bfsi_access_recommendation(
                        sensitivity, reg_flags, bfsi_types_set
                    ),
        })


    pii_cols  = sum(1 for f in findings if f["pii_detected"])
    bfsi_cols = sum(1 for f in findings if f["bfsi_identifiers"])
    all_sens  = [f["sensitivity"] for f in findings]
    # Overall classification: highest sensitivity tier wins, with BFSI tier between
    # Confidential and Public
    _SENS_ORDER = ["Highly Restricted", "Confidential", "Internal Use Only", "Public"]
    overall = next(
        (s for s in _SENS_ORDER if s in all_sens),
        "Public"
    )


    all_regs = sorted({r for f in findings for r in f["regulatory"]})

    # Aggregate mandatory breaches and conditional warnings across all columns
    all_mandatory_breaches  = [b for f in findings for b in f["mandatory_breaches"]]
    all_conditional_warnings = [w for f in findings for w in f["conditional_warnings"]]

    # Stewardship routing summary -- group columns by owner for issue triage
    stewardship_summary: dict[str, list[str]] = {}
    for f in findings:
        if f["mandatory_breaches"] or f["conditional_warnings"] or f["pii_detected"]:
            owner = f["escalate_to"]
            stewardship_summary.setdefault(owner, []).append(f["column"])

    # Columns without data dictionary entry
    undocumented = [f["column"] for f in findings if not f["from_dict"]] if data_dict else []

    return {
        "file_name":    name,


        "overall_classification": overall,
        "regulatory_frameworks": all_regs,
        "pii_column_count":   pii_cols,
        "bfsi_identifier_col_count": bfsi_cols,
        "mandatory_breaches":   all_mandatory_breaches,
        "conditional_warnings":  all_conditional_warnings,
        "stewardship_routing":  stewardship_summary,
        "undocumented_columns":  undocumented,
        "has_data_dict":     bool(data_dict),
        "columns":        findings,
    }


# ----------------------------------------------------------------------
# Data Lineage – auto column alignment + transform + reconciliation
# ----------------------------------------------------------------------

# Common abbreviation expansions for semantic column name matching
_ABBREV_MAP = {


    "acct": "account", "amt": "amount", "ccy": "currency", "curr": "currency",
    "src": "source", "tgt": "target", "ref": "reference", "id": "identifier",
    "le": "legal entity", "prin": "principal", "dt": "date", "ts": "timestamp",
    "num": "number", "qty": "quantity", "pct": "percent", "cd": "code",
    "nm": "name", "desc": "description", "typ": "type", "stat": "status",
    "prod": "product", "rgn": "region", "sys": "system", "upd": "updated",
    "attr": "attribution", "use": "usage", "mgr": "manager", "txn": "transaction",
}

# Null sentinel values across systems
_LINEAGE_NULLS = {"null", "none", "nan", "n/a", "na", "#n/a", "-", "nil", "undefined", ""}

# Boolean/code translation domains
_BOOL_DOMAINS = [
    ({"y", "yes", "true", "1", "t"}, {"n", "no", "false", "0", "f"}),
]


def _expand_col_name(col: str) -> str:


    # Normalise a column name to lowercase words, expanding abbreviations.

    # Handles: CamelCase, PascalCase, snake_case, kebab-case, spaces, ALL-CAPS
    # acronyms.

    # Consecutive single-char tokens produced by CamelCase splitting are recombined
    # into
    # the original acronym before abbreviation lookup (e.g. 'ID', 'ISIN', 'FX').

    # Insert space before each uppercase letter to split CamelCase/PascalCase
    spaced = re.sub(r"([A-Z])", r" \1", col).replace("_", " ").replace("-", " ")
    raw_tokens = spaced.lower().split()

    # Recombine runs of single-char tokens back into acronyms (e.g. ['i','d'] -> 'id')
    merged: list[str] = []
    acron_buf: list[str] = []
    for t in raw_tokens:
        if len(t) == 1:
            acron_buf.append(t)
        else:
            if acron_buf:


                merged.append("".join(acron_buf))
                acron_buf = []
            merged.append(t)
    if acron_buf:
        merged.append("".join(acron_buf))

    expanded = [_ABBREV_MAP.get(t, t) for t in merged if t]
    return " ".join(expanded)


def _col_name_similarity(a: str, b: str) -> float:
    # Semantic similarity between two column names (0.0-1.0).
    ea, eb = _expand_col_name(a), _expand_col_name(b)
    # Exact after expansion
    if ea == eb:
        return 1.0
    # Token overlap (Jaccard)
    ta, tb = set(ea.split()), set(eb.split())
    if ta and tb:


        jaccard = len(ta & tb) / len(ta | tb)
    else:
        jaccard = 0.0
    # SequenceMatcher on expanded names
    seq = SequenceMatcher(None, ea, eb).ratio()
    return round(max(jaccard, seq), 3)


def _value_domain(series: pd.Series) -> set:
    # Return the set of distinct non-null normalised lowercase values.
    return {str(v).strip().lower() for v in series.dropna() if str(v).strip().lower() not in _LINEAGE_NULLS}  # OCR-UNCERTAIN


def _profile_similarity(s1: pd.Series, s2: pd.Series) -> float:
    # Compare two column value profiles to detect same-content columns.
    #
    # Uses the maximum of three overlap measures so that partial-domain matches
    # are detected correctly.  Example: InstrumentID (225 values) vs Security
    # (65 values) share 32 values -- Jaccard=0.12 but containment of target
    # in source = 32/65 = 0.49, which correctly signals a real match.
    #
    # jaccard      = |d1 n d2| / |d1 U d2|  (symmetric, penalises size diff)
    # containment12 = |d1 n d2| / |d2|      (fraction of target values in source)
    # containment21 = |d1 n d2| / |d1|      (fraction of source values in target)
    d1, d2 = _value_domain(s1), _value_domain(s2)
    if not d1 or not d2:
        return 0.0
    common = len(d1 & d2)
    if common == 0:
        return 0.0
    jaccard      = common / len(d1 | d2)
    containment12 = common / len(d2)  # fraction of target values found in source
    containment21 = common / len(d1)  # fraction of source values found in target
    return round(max(jaccard, containment12, containment21), 3)


def _detect_transform(s1: pd.Series, s2: pd.Series) -> list[dict]:
    # Auto-detect value-level transforms needed to align s1 (source) to s2 (target).
    # Returns a list of transform descriptors.
    transforms = []
    n = 200
    # Reset index on both slices so element-wise == comparisons align by position, not label.
    sample1 = s1.dropna().astype(str).str.strip().head(n).reset_index(drop=True)
    sample2 = s2.dropna().astype(str).str.strip().head(n).reset_index(drop=True)
    # Trim to the shorter length so comparisons are always same-length
    min_len = min(len(sample1), len(sample2))
    if min_len == 0:
        return transforms
    sample1 = sample1.iloc[:min_len]
    sample2 = sample2.iloc[:min_len]

    # Case normalisation
    if (sample1.str.upper() == sample2).mean() > 0.7:
        transforms.append({"type": "case", "rule": "uppercase", "description": "Source → UPPER CASE"})
    elif (sample1.str.lower() == sample2).mean() > 0.7:
        transforms.append({"type": "case", "rule": "lowercase", "description": "Source → lower case"})
    elif (sample1.str.title() == sample2).mean() > 0.7:
        transforms.append({"type": "case", "rule": "titlecase", "description": "Source → Title Case"})

    # Date format detection
    date_patterns = [
        (r"\d{2}-[A-Z]{3}-\d{4}", "%d-%b-%Y"),
        (r"\d{4}/\d{2}/\d{2}", "%Y/%m/%d"),
        (r"\d{2}/\d{2}/\d{4}", "%d/%m/%Y"),
        (r"\d{8}", "%Y%m%d"),
    ]
    for pat, fmt in date_patterns:
        if sample1.str.match(pat).mean() > 0.5:
            transforms.append({
                "type": "date",
                "rule": f"{fmt} → ISO 8601",
                "description": f"Date format {fmt} → YYYY-MM-DD",
            })
            break

    # Number format (thousands separators, currency symbols)
    if sample1.str.contains(r"[,$£€]", regex=True, na=False).mean() > 0.3:
        transforms.append({"type": "number", "rule": "strip_currency", "description": "Strip currency symbols / thousand separators"})

    # Null sentinel translation
    null_variants = {"NULL", "N/A", "-", "nil", "None"}
    src_nulls = set(sample1[sample1.isin(null_variants)].unique())
    if src_nulls:
        transforms.append({"type": "null", "rule": "normalise_nulls", "description": f"Null sentinels {src_nulls} → empty"})

    # Boolean/code translation
    d1 = {v.lower() for v in sample1.unique()}
    d2 = {v.lower() for v in sample2.unique()}
    for true_set, false_set in _BOOL_DOMAINS:
        if d1 <= (true_set | false_set) and d2 <= (true_set | false_set) and d1 != d2:
            transforms.append({"type": "boolean", "rule": "code_translate", "description":
                                f"Boolean domain translation {d1} → {d2}"})

    return transforms


def _apply_transform(series: pd.Series, transforms: list[dict]) -> pd.Series:
    # Apply detected transforms to align source values toward target format.
    s = series.copy().astype(object)
    for t in transforms:
        if t["type"] == "case":
            if t["rule"] == "uppercase":
                s = s.str.upper()
            elif t["rule"] == "lowercase":
                s = s.str.lower()


            elif t["rule"] == "titlecase":
                s = s.str.title()
        elif t["type"] == "number":
            s = s.str.replace(r"[,$£€\s]", "", regex=True)
        elif t["type"] == "null":
            s = s.replace(list(_LINEAGE_NULLS | {"NULL", "N/A", "-", "nil", "None", "NONE"}),
"")
        elif t["type"] == "boolean":
            bool_map = {}
            for v in s.dropna().unique():
                vl = str(v).lower()
                for true_set, false_set in _BOOL_DOMAINS:
                    if vl in true_set:
                        bool_map[v] = "true"
                    elif vl in false_set:
                        bool_map[v] = "false"
            if bool_map:
                s = s.replace(bool_map)
    return s


def _canonical(v) -> str:
    # Canonical comparable string -- null-safe, stripped, null-sentinel normalised.
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    s = str(v).strip()
    return "" if s.lower() in _LINEAGE_NULLS else s


def _compare_embedded_value(src_val: str, tgt_description: str) -> tuple[bool, str]:
    # Compare a source scalar value against a free-text target description field.
    #
    # Strategy (in priority order):
    # 1. Exact substring match (fastest, handles most cases).
    # 2. Numeric regex extraction -- build a pattern from the source number and extract
    # the nearest numeric token from the description, then compare as floats
    # (handles formatting differences like 1000 vs 1,000 vs 1.0e3).
    # 3. Case-insensitive substring match for non-numeric values.
    #
    # Returns (is_match: bool, extracted_from_target: str).
    sv = str(src_val).strip()
    tv = str(tgt_description).strip()
    if not sv:
        return True, ""    # empty source value -- nothing to validate

    # 1. Exact substring
    if sv in tv:
        return True, sv

    # 2. Numeric comparison
    _NUM_PAT = re.compile(r"[-+]?\d[\d,]*\.?\d*(?:[eE][-+]?\d+)?")  # OCR-UNCERTAIN
    try:
        src_num = float(sv.replace(",", ""))
        # Find all numbers in the description and pick the closest match
        nums_in_desc = _NUM_PAT.findall(tv)
        for raw in nums_in_desc:
            try:
                tgt_num = float(raw.replace(",", ""))
                # Accept if within 0.1% relative tolerance (handles rounding)
                if src_num == 0 and tgt_num == 0:
                    return True, raw
                if src_num != 0 and abs(tgt_num - src_num) / abs(src_num) <= 0.001:
                    return True, raw
            except ValueError:
                continue
        # If we got here, source is numeric but no matching number found in description
        all_extracted = ", ".join(nums_in_desc) or "(none)"
        return False, all_extracted
    except ValueError:
        pass  # source value is not numeric

    # 3. Case-insensitive substring for string values
    if sv.lower() in tv.lower():
        return True, sv

    return False, "(not found)"


def _llm_lineage_mapping(df_src: pd.DataFrame, df_tgt: pd.DataFrame,
            name_src: str, name_tgt: str,
            user_hints: dict | None = None) -> dict | None:
    # Ask Claude to analyse source and target column names + sample values and return:
    # - column_mappings  : list of {src_col, tgt_col, match_type, confidence, notes,
    #   regex_extract}
    # - key_cols      : {src: [col,...], tgt: [col,...]}
    # - not_in_target   : [src cols that have no counterpart in target]
    # - embedded_cols   : list of {src_cols:[...], tgt_col, regex_pattern, notes}
    # - reasoning      : free-text explanation of decisions made
    #
    # Returns None on any error so caller falls back to heuristics.
    _SAMPLE_ROWS = 5
    _MAX_COLS = 60

    def _col_sample(df: pd.DataFrame) -> str:
        cols = list(df.columns)[:_MAX_COLS]
        lines = []
        for c in cols:
            vals = df[c].dropna().astype(str).head(_SAMPLE_ROWS).tolist()
            lines.append(f"  {c}: {vals}")
        return "\n".join(lines)

    src_sample = _col_sample(df_src)
    tgt_sample = _col_sample(df_tgt)

    # Build optional user-supplied hints block
    hints_block = ""
    if user_hints:
        hints_lines = []

        if user_hints.get("domain_context"):
            hints_lines.append(f"Domain / business context:\n{user_hints['domain_context']}")
        if user_hints.get("key_hints"):
            hints_lines.append(f"Expected key columns (user-provided):\n{user_hints['key_hints']}")
        if user_hints.get("mapping_hints"):
            hints_lines.append(f"Known column mappings (user-provided):\n{user_hints['mapping_hints']}")
        if user_hints.get("exclude_hints"):
            hints_lines.append(f"Columns the user believes should be excluded / are metadata: {user_hints['exclude_hints']}")
        if user_hints.get("transform_hints"):
            hints_lines.append(f"Known transformations / derivation rules:\n{user_hints['transform_hints']}")
        if hints_lines:
            hints_block = (
                "\nADDITIONAL CONTEXT PROVIDED BY THE USER (treat as authoritative guidance):\n"
                + "\n".join(f"  - {l}" for l in hints_lines)
                + "\n"
            )

    # NOTE (reconstruction): this prompt string had lost its opening/closing
    # markers embedded mid-string during OCR assembly. Restored as an f-string
    # with the markers removed (they are pipeline artifacts, not original
    # content) -- verify against source in code review.
    prompt = f"""You are a data lineage expert. You are given two datasets:

SOURCE FILE: "{name_src}"
Columns and sample values:
{src_sample}

TARGET FILE: "{name_tgt}"
Columns and sample values:
{tgt_sample}
{hints_block}
Analyse the relationship between source and target and return a JSON object with exactly these keys:

{{
  "column_mappings": [
    {{
      "src_col": "exact source column name",
      "tgt_col": "exact target column name or null if not present",
      "match_type": "exact|renamed|embedded|merged|split|not_in_target",
      "confidence": "HIGH|MEDIUM|LOW",
      "notes": "brief explanation",
      "regex_extract": "regex pattern to extract src value from tgt cell, or null"
    }}
  ],
  "key_cols": {{
    "src": ["source key column(s)"],
    "tgt": ["corresponding target key column(s)"]
  }},
  "not_in_target": ["source columns with no counterpart in target"],
  "embedded_cols": [
    {{
      "src_cols": ["source columns whose values appear embedded in one target column"],
      "tgt_col": "target column name",
      "regex_pattern": "regex with named groups matching each src_col value, e.g. (?P<Price>[\\\\d.]+)",
      "notes": "description of the embedding pattern"
    }}
  ],
  "reasoning": "2-3 sentence summary of how you determined the mappings and key"
}}

Rules:
- match_type "embedded" = source value appears as a substring/token inside a larger target description field
- match_type "renamed"  = same data, different column name
- match_type "merged"   = multiple source columns concatenated into one target column
- match_type "split"    = one source column split into multiple target columns
- For embedded columns provide a named-group regex that can extract the value, e.g. Price[:\\s]*(\\d+[.,]?\\d*)
- Key columns must uniquely identify rows in BOTH source and target
- Only include columns in not_in_target if they genuinely have no equivalent in target
- Reply with ONLY the JSON object -- no markdown, no commentary outside the JSON"""

    try:
        raw = _ask_llm([{"role": "user", "content": [{"text": prompt}]}])

        # Strip any accidental markdown fences
        raw = re.sub(r"^```[a-z]*\n?", "", raw.strip(), flags=re.MULTILINE)
        raw = re.sub(r"```$", "", raw.strip())
        return json.loads(raw)
    except Exception:
        return None


def _classify_value_exception(src_val: str, tgt_val: str) -> str:
    # Classify what kind of value exception this is.
    if not src_val and tgt_val:
        return "NULL_IN_SOURCE"
    if src_val and not tgt_val:
        return "NULL_IN_TARGET"


    try:
        if abs(float(src_val) - float(tgt_val)) / max(abs(float(src_val)), 1) < 0.001:
            return "ROUNDING_DIFFERENCE"
    except Exception:
        pass
    if src_val.lower() == tgt_val.lower():
        return "CASE_DIFFERENCE"
    if src_val in tgt_val or tgt_val in src_val:
        return "TRUNCATION_OR_PREFIX"
    return "VALUE_MISMATCH"


def _exception_severity(src_val: str, tgt_val: str) -> str:
    # Assign severity level based on the classified exception type.
    exc_type = _classify_value_exception(src_val, tgt_val)
    if exc_type in ("CASE_DIFFERENCE", "ROUNDING_DIFFERENCE"):
        return "WARN"
    if exc_type in ("NULL_IN_TARGET", "VALUE_MISMATCH"):
        return "ERROR"


    if exc_type == "NULL_IN_SOURCE":
        return "INFO"
    return "ERROR"


# BFSI domain column sets used to classify exceptions into Business / Technical / Operational
_BFSI_BUSINESS_COLS: frozenset[str] = frozenset({
    # Trade economics
    "notional", "quantity", "qty", "amount", "price", "rate", "coupon", "yield",
    "market_value", "npv", "dirty_price", "clean_price", "accrued",
    "fx_rate", "spot_rate", "forward_rate", "strike", "premium", "fee", "commission",
    # Trade identification / classification
    "trade_id", "trade_ref", "order_id", "order_ref", "deal_id", "ticket_id",
    "isin", "cusip", "sedol", "ticker", "ric", "bbg_id", "bloomberg_id",
    "instrument_type", "asset_class", "product_type", "security_type",
    "ccy", "currency", "base_ccy", "quote_ccy", "settle_ccy",
    "side", "buy_sell", "direction", "action",
    "status", "trade_status", "order_status", "settlement_status", "lifecycle_status",


    "counterparty", "cpty", "counterparty_id", "client_id", "client_name",
    "book", "book_id", "portfolio", "fund", "account", "entity", "legal_entity",
    "trader", "desk", "strategy",
    # Settlement / lifecycle
    "settle_date", "settlement_date", "value_date", "trade_date", "maturity_date",
    "expiry_date", "delivery_date", "payment_date",
    "settlement_type", "settlement_method", "delivery_type",
    # Risk / P&L
    "pnl", "mtm", "delta", "gamma", "vega", "theta", "rho", "dv01", "cs01",
    "var", "cvar", "stress", "risk_factor",
    # Regulatory / reporting
    "lei", "uti", "usi", "usi_namespace", "reporting_entity", "reporting_obligation",
    "mifid_flag", "emir_flag", "cftc_flag", "cleared", "venue", "mic_code",
})

_BFSI_TECHNICAL_COLS: frozenset[str] = frozenset({
    "created_at", "updated_at", "modified_at", "timestamp", "etl_batch",
    "source_system", "load_dt", "run_id", "process_id", "record_id",
    "checksum", "hash", "version", "row_num", "seq", "sequence",


    "file_name", "source_file", "batch_id",
})


def _bfsi_exception_domain(col_name: str, exc_type: str) -> str:
    # Classify an exception into one of four generic domains:
    # Data       -- value-level mismatch on a business/content column
    # Technical  -- data format / quality issue (null, rounding, case, type)
    # Operational -- missing/unmatched row, completeness gap, extraction failure
    # Other      -- anything not fitting the above
    #
    # Generic -- does not assume BFSI column vocabulary so it works for any schema.

    # Operational -- row-level completeness breaks (column name irrelevant)
    if exc_type in ("MISSING_IN_TARGET", "MISSING_IN_SOURCE", "DUPLICATE_SOURCE_ROW",
                    "VALUE_NOT_EMBEDDED"):
        return "Operational"

    # Technical -- data format / quality problems regardless of column name
    if exc_type in ("NULL_IN_TARGET", "NULL_IN_SOURCE", "ROUNDING_DIFFERENCE",
                    "CASE_DIFFERENCE", "TRUNCATION_OR_PREFIX"):
        return "Technical"

    # Metadata / audit columns -- always Technical
    normalized = col_name.lower().strip()
    for pfx in ("src_", "tgt_", "source_", "target_", "old_", "new_"):
        if normalized.startswith(pfx):
            normalized = normalized[len(pfx):]
    for sfx in ("_src", "_tgt", "_old", "_new", "_1", "_2"):
        if normalized.endswith(sfx):
            normalized = normalized[:-len(sfx)]

    if normalized in _BFSI_TECHNICAL_COLS or any(
        tk in normalized for tk in ("etl", "batch", "load", "checksum", "hash", "seq")
    ):
        return "Technical"

    # VALUE_MISMATCH on any non-metadata column -> Data
    if exc_type == "VALUE_MISMATCH":
        return "Data"

    return "Other"


# -----------------------------------------------------------------------
# Field spec inference -- mirrors TransferPricing_element_mapping.xlsx layout
# -----------------------------------------------------------------------

_BFSI_ANALYTICS_COLS = frozenset({
    "pnl", "mtm", "delta", "gamma", "vega", "theta", "nav", "aum", "exposure", "charge",
    "alloc", "allocation", "fee", "rate", "spread", "yield", "return", "alpha", "beta",
    "sharpe", "vol", "volatility", "notional", "market_value", "mkt_value", "price",
    "cost", "accrual", "amortization", "unrealised", "realised", "fx", "swap",
})

_BFSI_REFERENCE_COLS = frozenset({
    "id", "code", "isin", "cusip", "sedol", "ric", "ticker", "lei", "uti", "ccy", "currency",
    "entity", "counterparty", "cpty", "account", "portfolio", "fund", "legal_entity",
    "issuer", "country", "region", "sector", "asset_class", "product", "strategy",
    "instrument", "security", "booking", "legal", "regulatory", "identifier",
})

_BFSI_TRANSACTIONAL_COLS = frozenset({
    "trade", "order", "execution", "settlement", "booking", "position", "transaction",
    "transfer", "payment", "delivery", "receipt", "confirm", "match", "break", "fail",
    "novation", "allocation", "affirmation", "clearing", "margin", "collateral",
})

# Date format sniffers (ordered most-specific first)
_DATE_PATTERNS = [
    (r"^\d{4}-\d{2}-\d{2}$",        "YYYY-MM-DD"),
    (r"^\d{2}/\d{2}/\d{4}$",        "DD/MM/YYYY"),
    (r"^\d{2}-\d{2}-\d{4}$",        "DD-MM-YYYY"),
    (r"^\d{8}$",              "YYYYMMDD"),
    (r"^\d{4}-\d{2}-\d{2}T\d{2}:",    "YYYY-MM-DDTHH:MM:SS"),
]


def _infer_operational_usage(col_name: str) -> str:
    n = col_name.lower()
    for pfx in ("src_", "tgt_", "source_", "target_", "old_", "new_"):
        if n.startswith(pfx): n = n[len(pfx):]
    for sfx in ("_src", "_tgt", "_old", "_new", "_1", "_2"):
        if n.endswith(sfx): n = n[:-len(sfx)]
    tokens = set(re.split(r"[_\s]+", n))
    if tokens & _BFSI_ANALYTICS_COLS or any(k in n for k in _BFSI_ANALYTICS_COLS):
        return "Analytics"
    if tokens & _BFSI_TRANSACTIONAL_COLS or any(k in n for k in
_BFSI_TRANSACTIONAL_COLS):
        return "Transactional"
    if tokens & _BFSI_REFERENCE_COLS or any(k in n for k in _BFSI_REFERENCE_COLS):
        return "Reference"
    return "Analytics"


def _infer_criticality(col_name: str, null_pct: float) -> str:
    n = col_name.lower()
    if any(k in n for k in ("id", "key", "isin", "cusip", "lei", "uti", "trade", "account", "entity")):
        return "High"
    if null_pct > 40:
        return "Low"
    if any(k in n for k in _BFSI_ANALYTICS_COLS):
        return "Medium"
    return "Medium"


def _infer_business_attribute(col_name: str) -> str:
    # Convert snake_case / camelCase / abbreviations to a readable business name.
    n = col_name
    # strip common prefixes
    for pfx in ("src_", "tgt_", "source_", "target_", "old_", "new_", "match_"):
        if n.lower().startswith(pfx):
            n = n[len(pfx):]


    # split on _ or camelCase boundaries
    parts = re.sub(r"([a-z])([A-Z])", r"\1 \2", n).replace("_", " ").split()
    # Expand known abbreviations
    _abbr = {
        "id": "ID", "ccy": "Currency", "qty": "Quantity", "mkt": "Market",
        "val": "Value", "pnl": "P&L", "mtm": "Mark-to-Market", "cpty": "Counterparty",
        "acct": "Account", "sec": "Security", "alloc": "Allocation",
        "ui": "UI", "dq": "Data Quality", "tp": "Transfer Pricing",
        "adj": "Adjusted", "calc": "Calculated", "ref": "Reference",
    }
    return " ".join(_abbr.get(p.lower(), p.title()) for p in parts)


def _infer_description(col_name: str, series: "pd.Series") -> str:
    ba = _infer_business_attribute(col_name)
    usage = _infer_operational_usage(col_name)
    dtype_label = _pandas_dtype_to_sql(str(series.dtype))
    return f"{ba} -- {dtype_label} field used for {usage.lower()} purposes."


def _pandas_dtype_to_sql(dtype_str: str) -> str:
    if "int" in dtype_str:  return "INTEGER"
    if "float" in dtype_str: return "DECIMAL"
    if "datetime" in dtype_str: return "TIMESTAMP"
    if "bool" in dtype_str:  return "BOOLEAN"
    return "VARCHAR"


def _infer_format(series: "pd.Series", dtype_str: str) -> str:
    import re as _re
    if "datetime" in dtype_str:
        return "YYYY-MM-DD HH:MM:SS"
    if "float" in dtype_str or "int" in dtype_str:
        return "Numeric"
    if "bool" in dtype_str:
        return "Y/N or True/False"
    # Sample non-null strings
    sample = series.dropna().astype(str).head(200)


    if sample.empty:
        return "--"
    # Date format detection
    for pattern, fmt in _DATE_PATTERNS:
        if sample.str.match(pattern).mean() > 0.7:
            return fmt
    # Case detection
    has_upper = sample.str.isupper().mean()
    has_lower = sample.str.islower().mean()
    if has_upper > 0.7:  return "UpperCase"
    if has_lower > 0.7:  return "LowerCase"
    return "Mixed"


def _infer_business_rules(col_name: str, series: "pd.Series", dtype_str: str) -> str:
    rules = []
    n = col_name.lower()
    if "int" in dtype_str or "float" in dtype_str:
        numeric = pd.to_numeric(series, errors="coerce").dropna()


        if not numeric.empty:
            mn, mx = numeric.min(), numeric.max()
            if any(k in n for k in ("qty","quantity","count","alloc")):
                rules.append(f"Must be >= 0 (range: {mn:.0f}-{mx:.0f})")
            elif any(k in n for k in ("pct","percent","rate","ratio")):
                rules.append(f"Typically 0-100 (range: {mn:.2f}-{mx:.2f})")
            else:
                rules.append(f"Numeric range: {mn:.4g}-{mx:.4g}")
    if any(k in n for k in ("date","dt","timestamp","time")):
        rules.append("Must be valid date")
    if any(k in n for k in ("ccy","currency","iso")):
        rules.append("3-char ISO 4217 currency code")
    if any(k in n for k in ("isin",)):
        rules.append("12-char alphanumeric ISIN format")
    if any(k in n for k in ("lei",)):
        rules.append("20-char alphanumeric LEI format")
    if any(k in n for k in ("email","mail")):
        rules.append("Valid email format")
    if not rules:


        return "--"  # OCR-UNCERTAIN (continuation from prior page context)
    return "; ".join(rules)


def _exc(exc_type: str, count: int, total: int, severity: str, sample: str, rule: str = "") -> dict:
    # Build a single exception entry.
    return {
        "type": exc_type,
        "count": count,
        "pct": round(count / total * 100, 1) if total else 0,
        "severity": severity,
        "sample": str(sample)[:80],
        "rule": rule,
    }


def _compute_field_exceptions(
    col_name: str,


    series_src: "pd.Series",
    series_tgt: "pd.Series | None",
    match_type: str,
) -> dict:
    """Validate field values against inferred business rules and cross-file exceptions."""
    exceptions: list[dict] = []
    n       = col_name.lower().strip()
    s       = series_src
    dtype_str = str(s.dtype)
    src_sql = _pandas_dtype_to_sql(dtype_str)
    total   = len(s)
    non_null = s.dropna()
    src_null = total - len(non_null)
    src_dup  = int(s.duplicated(keep=False).sum())

    # -- 1. MANDATORY / NULL check
    if src_null > 0:
        sev = "High" if src_null / total > 0.1 else "Medium"


        sample_idx = s[s.isna()].index[:3].tolist()
        exceptions.append(_exc(
            "NULL_IN_SOURCE", src_null, total, sev,
            f"rows {sample_idx}", "Mandatory field must not be null",
        ))

    # -- 2. DUPLICATE check
    if src_dup > 0:
        dup_vals = s[s.duplicated(keep=False)].dropna().astype(str).unique()[:3].tolist()
        exceptions.append(_exc(
            "DUPLICATE_IN_SOURCE", src_dup, total, "Medium",
            ", ".join(dup_vals), "Values should be unique",
        ))

    # -- 3. LIST OF VALUES (LOV) violation
    # Auto-detect LOV when distinct <= 15 on non-null sample, then validate rest
    if 1 < s.nunique(dropna=True) <= 15 and len(non_null) > 0:
        lov_set = set(non_null.astype(str).str.strip().str.upper().unique())


        invalid = non_null.astype(str).str.strip().str.upper().apply(
            lambda v: v not in lov_set
        )
        # This is always 0 since lov_set was built from the same series --
        # but if a target exists we check target values against the source LOV
        if series_tgt is not None:
            tgt_non_null = series_tgt.dropna().astype(str).str.strip().str.upper()
            lov_violations = tgt_non_null[~tgt_non_null.isin(lov_set)]
            if len(lov_violations) > 0:
                exceptions.append(_exc(
                    "LOV_VIOLATION", len(lov_violations), len(series_tgt),
                    "High",
                    ", ".join(lov_violations.unique()[:3].tolist()),
                    f"Allowed values: {{{', '.join(sorted(lov_set)[:8])}}}",
                ))

    # -- 4. BUSINESS FORMAT rules (column-name driven)
    # ISO 4217 currency code -- 3 uppercase letters
    if any(k in n for k in ("ccy", "currency", "iso_ccy", "curr")):


        _iso_re = re.compile(r"^[A-Z]{3}$")
        bad = non_null.astype(str).str.strip().apply(lambda v: not bool(_iso_re.match(v)))
        cnt = int(bad.sum())
        if cnt:
            samples = non_null.astype(str)[bad].unique()[:3].tolist()
            exceptions.append(_exc(
                "INVALID_CURRENCY_CODE", cnt, total, "High",
                ", ".join(samples), "Must be 3-char ISO 4217 uppercase code (e.g. USD, GBP)",
            ))

    # ISIN -- 12 alphanumeric
    if "isin" in n:
        _isin_re = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")
        bad = non_null.astype(str).str.strip().apply(lambda v: not bool(_isin_re.match(v)))
        cnt = int(bad.sum())
        if cnt:
            samples = non_null.astype(str)[bad].unique()[:3].tolist()
            exceptions.append(_exc(
                "INVALID_ISIN", cnt, total, "High",


                ", ".join(samples), "Must match ISIN format: 2 alpha + 9 alphanumeric + 1 digit",
            ))

    # CUSIP -- 9 alphanumeric
    if "cusip" in n:
        _cusip_re = re.compile(r"^[A-Z0-9]{9}$")
        bad = non_null.astype(str).str.strip().apply(lambda v: not bool(_cusip_re.match(v)))
        cnt = int(bad.sum())
        if cnt:
            samples = non_null.astype(str)[bad].unique()[:3].tolist()
            exceptions.append(_exc(
                "INVALID_CUSIP", cnt, total, "Medium",
                ", ".join(samples), "Must be 9-char alphanumeric CUSIP",
            ))

    # LEI -- 20 alphanumeric
    if "lei" in n:
        _lei_re = re.compile(r"^[A-Z0-9]{18}[0-9]{2}$")


        bad = non_null.astype(str).str.strip().apply(lambda v: not bool(_lei_re.match(v)))
        cnt = int(bad.sum())
        if cnt:
            samples = non_null.astype(str)[bad].unique()[:3].tolist()
            exceptions.append(_exc(
                "INVALID_LEI", cnt, total, "High",
                ", ".join(samples), "Must be 20-char alphanumeric LEI",
            ))

    # UTI -- 52 char max, alphanumeric + hyphen
    if "uti" in n:
        _uti_re = re.compile(r"^[A-Z0-9\-]{1,52}$")
        bad = non_null.astype(str).str.strip().apply(lambda v: not bool(_uti_re.match(v)))
        cnt = int(bad.sum())
        if cnt:
            samples = non_null.astype(str)[bad].unique()[:3].tolist()
            exceptions.append(_exc(
                "INVALID_UTI", cnt, total, "Medium",
                ", ".join(samples), "Must be <=52 alphanumeric/hyphen UTI",


            ))

    # Date format consistency
    if any(k in n for k in ("date", "_dt", "timestamp", "trade_dt", "settle_dt",
                "value_dt", "mat_dt", "expiry_dt", "start_dt")):
        if src_sql == "VARCHAR":
            # Check if values conform to any recognised date pattern
            _any_date = re.compile(
                r"^\d{4}-\d{2}-\d{2}$|^\d{2}/\d{2}/\d{4}$|^\d{8}$|^\d{2}-\d{2}-\d{4}$"
            )
            bad = non_null.astype(str).str.strip().apply(
                lambda v: not bool(_any_date.match(v))
            )
            cnt = int(bad.sum())
            if cnt:
                samples = non_null.astype(str)[bad].unique()[:3].tolist()
                exceptions.append(_exc(
                    "INVALID_DATE_FORMAT", cnt, total, "High",
                    ", ".join(samples), "Must be YYYY-MM-DD, DD/MM/YYYY, or YYYYMMDD",


                ))  # OCR-UNCERTAIN (closing paren carried from prior page context)

    # ==== GAP: the "numeric" variable's definition was not recoverable from the
    # scan (comment already flagged its origin as unclear) -- RECONSTRUCTED
    # (unverified) below as a numeric coercion of non_null, matching how the
    # subsequent negative-value / quantity checks use it; verify against source.
    numeric = pd.to_numeric(non_null, errors="coerce")
    valid_num = numeric.dropna()

    # Quantity / count fields must be >= 0
    if any(k in n for k in ("qty", "quantity", "count", "alloc", "units", "shares",
                "notional", "principal", "face_value")):
        neg = (valid_num < 0).sum()
        if neg:
            samples = valid_num[valid_num < 0].head(3).astype(str).tolist()
            exceptions.append(_exc(
                "NEGATIVE_VALUE", int(neg), total, "High",
                ", ".join(samples), f"'{col_name}' must be >= 0",
            ))

    # Rate / percentage fields must be in 0-100 (or 0-1 for decimal rates)
    if any(k in n for k in ("rate", "pct", "percent", "ratio", "spread",
                "yield", "coupon", "haircut")):
        # Detect scale: if max > 1 treat as 0-100, else 0-1
        mx = float(valid_num.max()) if len(valid_num) else 0
        if mx > 1:   # 0-100 scale


            out_of_range = ((valid_num < 0) | (valid_num > 100)).sum()
            rule_msg = "Rate must be 0-100"
        else:      # 0-1 scale
            out_of_range = ((valid_num < 0) | (valid_num > 1)).sum()
            rule_msg = "Rate must be 0.0-1.0"
        if out_of_range:
            samples = valid_num[
                (valid_num < 0) | (valid_num > (100 if mx > 1 else 1))
            ].head(3).astype(str).tolist()
            exceptions.append(_exc(
                "OUT_OF_RANGE", int(out_of_range), total, "High",
                ", ".join(samples), rule_msg,
            ))

    # Charge / fee / price must be non-negative
    if any(k in n for k in ("charge", "fee", "price", "cost", "premium",
                "strike", "barrier", "cap", "floor")):
        neg = (valid_num < 0).sum()
        if neg:


            samples = valid_num[valid_num < 0].head(3).astype(str).tolist()
            exceptions.append(_exc(
                "NEGATIVE_CHARGE", int(neg), total, "Medium",
                ", ".join(samples), f"'{col_name}' should be >= 0",
            ))

    # Non-parseable values in a supposedly numeric field
    coerce_fail = int(pd.to_numeric(s.dropna(), errors="coerce").isna().sum())
    if coerce_fail > 0:
        samples = s.dropna()[
            pd.to_numeric(s.dropna(), errors="coerce").isna()
        ].head(3).astype(str).tolist()
        exceptions.append(_exc(
            "NON_NUMERIC_VALUE", coerce_fail, total, "High",
            ", ".join(samples), "Non-numeric value in numeric field",
        ))

    # -- 6. CROSS-FILE EXCEPTIONS (when target series is available)
    if series_tgt is not None:


        tgt_sql  = _pandas_dtype_to_sql(str(series_tgt.dtype))
        tgt_null = int(series_tgt.isna().sum())
        tgt_total = len(series_tgt)

        # Null in target
        if tgt_null > 0:
            sev = "High" if tgt_null / tgt_total > 0.1 else "Medium"
            exceptions.append(_exc(
                "NULL_IN_TARGET", tgt_null, tgt_total, sev,
                "--", "Target column has null values",
            ))

        # Type mismatch
        if src_sql != tgt_sql:
            exceptions.append(_exc(
                "TYPE_MISMATCH", total, total, "High",
                f"{src_sql} -> {tgt_sql}", "Source and target data types differ",
            ))


        # Value mismatch (row-aligned) -- null-sentinel normalised
        try:
            min_len = min(total, tgt_total)
            s_a = s.iloc[:min_len].reset_index(drop=True)
            t_a = series_tgt.iloc[:min_len].reset_index(drop=True)
            # Normalise: NaN + all null sentinels -> "" so cross-format nulls don't flag
            def _norm(series):
                return (series.fillna("").astype(str).str.strip()
                        .apply(lambda v: "" if v.lower() in _NULL_SENTINELS else v))
            s_n = _norm(s_a)
            t_n = _norm(t_a)
            # Only flag when at least one side is non-blank after normalisation
            mismatch = (s_n != t_n) & ~((s_n == "") & (t_n == ""))
            cnt = int(mismatch.sum())
            if cnt:
                sev = "High" if cnt / min_len > 0.05 else "Medium"
                samples = s_a[mismatch].head(3).astype(str).tolist()
                exceptions.append(_exc(
                    "VALUE_MISMATCH", cnt, min_len, sev,


                    ", ".join(samples), "Row-level value differs between source and target",
                ))
        except Exception:
            pass

        # Case difference (values match case-insensitively)
        if src_sql == "VARCHAR" and tgt_sql == "VARCHAR":
            try:
                min_len = min(total, tgt_total)
                s_a = s.iloc[:min_len].dropna().astype(str).str.strip()
                t_a = series_tgt.iloc[:min_len].dropna().astype(str).str.strip()
                idx = s_a.index.intersection(t_a.index)
                if len(idx):
                    case_diff = int(
                        ((s_a[idx] != t_a[idx]) & (s_a[idx].str.lower() == t_a[idx].str.lower())).sum()
                    )
                    if case_diff:
                        exceptions.append(_exc(
                            "CASE_DIFFERENCE", case_diff, len(idx), "Low",


                            "--", "Values match case-insensitively but casing differs",
                        ))
            except Exception:
                pass

        # Rounding difference (numeric, diff < 0.01)
        if src_sql in ("DECIMAL", "INTEGER") and tgt_sql in ("DECIMAL", "INTEGER"):
            try:
                s_num = pd.to_numeric(s, errors="coerce")
                t_num = pd.to_numeric(series_tgt, errors="coerce")
                min_len = min(len(s_num), len(t_num))
                s_a = s_num.iloc[:min_len].reset_index(drop=True)
                t_a = t_num.iloc[:min_len].reset_index(drop=True)
                both = s_a.notna() & t_a.notna()
                diff = (s_a - t_a).abs()
                rounding = int((both & (diff > 0) & (diff < 0.01)).sum())
                if rounding:
                    exceptions.append(_exc(
                        "ROUNDING_DIFFERENCE", rounding,


                        int(both.sum()), "Low",
                        "diff < 0.01", "Numeric values differ by < 0.01 (rounding)",
                    ))
            except Exception:
                pass

        # Truncation (target string shorter than source by > 20%)
        if src_sql == "VARCHAR" and tgt_sql == "VARCHAR":
            try:
                min_len = min(total, tgt_total)
                s_len = s.iloc[:min_len].dropna().astype(str).str.len()
                t_len = series_tgt.iloc[:min_len].dropna().astype(str).str.len()
                idx = s_len.index.intersection(t_len.index)
                if len(idx):
                    trunc = int(
                        ((s_len[idx] > 0) & (t_len[idx] / s_len[idx] < 0.8)).sum()
                    )
                    if trunc:
                        exceptions.append(_exc(


                            "TRUNCATION", trunc, len(idx), "Medium",
                            "--", "Target value appears truncated (< 80% of source length)",
                        ))
            except Exception:
                pass

    # -- 7. DEPENDENCY / CONDITIONAL checks
    # If a "type" or "category" column is present in the same series context we
    # can't access sibling columns here -- those are handled at dataset level.
    # What we CAN check: conditional non-null (field is partially filled -- flag rows
    # where it's null when it appears it should be populated based on a pattern).
    if 0 < src_null < total:
        null_mask  = s.isna()
        filled_mask = s.notna()
        # Check if nulls are clustered (systematic gap) vs random
        if total > 10:
            null_run = null_mask.astype(int).diff().abs().sum()
            fill_run = filled_mask.astype(int).diff().abs().sum()


            # Low run count = values grouped in blocks -> conditional/dependent field
            if null_run < total * 0.2:
                exceptions.append(_exc(
                    "CONDITIONAL_NULL", src_null, total, "Low",
                    f"{src_null} nulls in systematic block",
                    "Nulls appear conditional/dependent -- verify population logic",
                ))

    # -- 8. UNMAPPED
    if match_type == "Unmapped-Source":
        exceptions.append(_exc(
            "MISSING_IN_TARGET", total, total, "High",
            f"No target column mapped",
            f"Column '{col_name}' exists in source but has no target mapping",
        ))
    elif match_type == "Unmapped-Target":
        exceptions.append(_exc(
            "MISSING_IN_SOURCE", total, total, "High",


            f"No source column mapped",
            f"Column '{col_name}' exists in target but has no source mapping",
        ))

    overall_status = "PASS"
    if any(e["severity"] == "High" for e in exceptions):
        overall_status = "FAIL"
    elif exceptions:
        overall_status = "WARN"

    return {
        "exceptions": exceptions,
        "exception_count": len(exceptions),
        "validation_status": overall_status,
    }


def _build_field_spec(
    col_name: str,


    series_src: "pd.Series",
    series_tgt: "pd.Series | None",
    source_name: str,
    target_name: str,
    match_type: str,  # 'Exact', 'Fuzzy', 'Unmapped-Source', 'Unmapped-Target'
    tgt_col_name: str | None = None,
    fuzzy_score: float | None = None,
) -> dict:
    """Infer all 21 spec fields for one column pair from raw data alone."""
    s = series_src
    dtype_str = str(s.dtype)
    total = len(s)
    null_count = int(s.isna().sum())
    null_pct  = round(null_count / total * 100, 1) if total else 0
    non_null  = s.dropna()
    distinct  = int(s.nunique(dropna=True))

    # Example value -- first non-null, truncated
    example_val = str(non_null.iloc[0])[:40] if len(non_null) else "--"


    # List of values -- only for low-cardinality cols
    lov = "--"
    if 1 < distinct <= 15:
        top_vals = s.value_counts(dropna=True).head(15).index.tolist()
        lov = ",".join(str(v)[:20] for v in top_vals)

    # Transformation Applied -- detect if src vs tgt differ in type or case
    transformation = "N"
    if series_tgt is not None:
        src_sql = _pandas_dtype_to_sql(dtype_str)
        tgt_sql = _pandas_dtype_to_sql(str(series_tgt.dtype))
        if src_sql != tgt_sql:
            transformation = f"Type cast {src_sql}->{tgt_sql}"
        else:
            # Case transformation check on string cols
            src_sample = s.dropna().astype(str).head(100)
            tgt_sample = series_tgt.dropna().astype(str).head(100)
            if len(src_sample) and len(tgt_sample):


                src_upper = src_sample.str.isupper().mean()
                tgt_upper = tgt_sample.str.isupper().mean()
                if abs(src_upper - tgt_upper) > 0.5:
                    transformation = "Case normalisation"

    # Mandatory / Conditional mandatory
    mandatory = "Y" if null_pct == 0 else "N"
    cond_mandatory = "Y" if 0 < null_pct < 30 else "N"

    # Duplicate check
    dup_count = int(s.duplicated(keep=False).sum())
    has_dup = "Y" if dup_count > 0 else "N"

    # DQ threshold Required
    dq_required = "Y" if null_pct > 5 or dup_count > 0 else "N"

    # Technical fields
    sql_type = _pandas_dtype_to_sql(dtype_str)
    # Length / precision


    if sql_type == "VARCHAR":
        max_len = int(s.dropna().astype(str).str.len().max()) if len(non_null) else 0
        length = f"{max_len} chars"
    elif sql_type in ("INTEGER","DECIMAL"):
        numeric = pd.to_numeric(s, errors="coerce").dropna()
        if not numeric.empty:
            int_digits = len(str(int(abs(numeric).max()))) if not numeric.empty else 1
            if sql_type == "DECIMAL":
                dec_digits = int(s.dropna().astype(str).str.extract(r"\.(\d+)")[0].str.len().max() or 0)
                length = f"({int_digits},{dec_digits})"
            else:
                length = f"{int_digits} digits"
        else:
            length = "--"
    else:
        length = "--"

    fmt = _infer_format(s, dtype_str)


    # Default value -- dominant value if > 40% frequency
    default_val = "--"
    if len(non_null):
        top_val = s.value_counts(dropna=True).iloc[0]
        top_freq = s.value_counts(dropna=True).iloc[0] / total
        if top_freq > 0.4:
            top_key = s.value_counts(dropna=True).index[0]
            default_val = str(top_key)[:30]

    # Error handling -- Y if data quality issues found
    error_handling = "Y" if (null_pct > 5 or dup_count > 0 or match_type in
        ("Fuzzy","Unmapped-Source","Unmapped-Target")) else "N"

    return {
        # Core
        "attribute_field": col_name,
        "target_field": tgt_col_name or "--",
        "source": source_name,


        "target": target_name,
        "match_type": match_type,
        "fuzzy_score": fuzzy_score,
        "example_value": example_val,
        "transformation_applied": transformation,
        "dq_threshold_required": dq_required,
        # Business Attribute Mapping
        "business_attribute": _infer_business_attribute(col_name),
        "description": _infer_description(col_name, s),
        "business_rules": _infer_business_rules(col_name, s, dtype_str),
        "criticality": _infer_criticality(col_name, null_pct),
        "mandatory": mandatory,
        "conditional_mandatory": cond_mandatory,
        "duplicate": has_dup,
        "list_of_values": lov,
        # Technical Attribute Mapping
        "data_type": sql_type,
        "length": length,
        "format": fmt,


        "default_value": default_val,
        # Others
        "error_handling": error_handling,
        "operational_usage": _infer_operational_usage(col_name),
        # Metadata
        "null_pct": null_pct,
        "null_count": null_count,
        "dup_count": dup_count,
        "distinct_count": distinct,
        "total_rows": total,
        # Validation exceptions (computed when both series are available)
        **_compute_field_exceptions(col_name, s, series_tgt, match_type),
    }


# ------------------------------------------------------------
# Data Mapping - enriched with mapping spec
# ------------------------------------------------------------


def analyze_mapping(df1, df2, name1, name2,
            mapping_spec: list[dict] | None = None,
            user_hints: dict | None = None) -> dict:
    hints = user_hints or {}
    mapping_spec = mapping_spec or []
    unmapped_f2  = list(df2.columns)
    exact, fuzzy, unmapped_f1 = [], [], []

    # Parse user-supplied mapping hints ("src=tgt, src2=tgt2") and inject as
    # pre-confirmed exact matches so they appear in results even when names differ.
    _hint_pairs: list[tuple[str, str]] = []
    for part in hints.get("mapping_hints", "").split(","):
        part = part.strip()
        if "=" in part:
            left, _, right = part.partition("=")
            left, right = left.strip(), right.strip()
            if left in df1.columns and right in df2.columns:
                _hint_pairs.append((left, right))


    # Columns the user wants excluded from mapping
    _hint_exclude = {c.strip() for c in hints.get("exclude_hints", "").split(",") if c.strip()}

    # Pre-populate hint-defined pairs as exact matches (highest priority)
    for src, tgt in _hint_pairs:
        exact.append({
            "f1_col": src, "f2_col": tgt,
            "score": 1.0,
            "type_ok": str(df1[src].dtype) == str(df2[tgt].dtype),
            "f1_type": str(df1[src].dtype),
            "f2_type": str(df2[tgt].dtype),
            "hint_matched": True,
            "domain": _bfsi_exception_domain(src, "VALUE_MISMATCH"),
        })
        if tgt in unmapped_f2:
            unmapped_f2.remove(tgt)

    # If a spec is provided, use it as the authoritative mapping
    spec_results = []


    if mapping_spec:
        for spec in mapping_spec:
            src, tgt = spec["source_column"], spec["target_column"]
            src_exists = src in df1.columns
            tgt_exists = tgt in df2.columns
            type_ok = (str(df1[src].dtype) == str(df2[tgt].dtype)
                if src_exists and tgt_exists else None)
            spec_results.append({
                **spec,
                "source_exists": src_exists,
                "target_exists": tgt_exists,
                "type_ok": type_ok,
                "f1_type": str(df1[src].dtype) if src_exists else "--",
                "f2_type": str(df2[tgt].dtype) if tgt_exists else "--",
                "status": (
                    "OK" if src_exists and tgt_exists
                    else "MISSING_SOURCE" if not src_exists
                    else "MISSING_TARGET"
                ),


            })

    # Track columns already resolved by hint pairs so auto-matching skips them
    _hint_src_matched = {src for src, _ in _hint_pairs}

    # Auto column matching (exact + fuzzy) - always computed
    for col1 in df1.columns:
        if col1 in _hint_exclude or col1 in _hint_src_matched:
            continue
        if col1 in df2.columns:
            exact.append({
                "f1_col": col1, "f2_col": col1,
                "score": 1.0,
                "type_ok": str(df1[col1].dtype) == str(df2[col1].dtype),
                "f1_type": str(df1[col1].dtype),
                "f2_type": str(df2[col1].dtype),
                "domain": _bfsi_exception_domain(col1, "VALUE_MISMATCH"),
            })
            if col1 in unmapped_f2:


                unmapped_f2.remove(col1)
        else:
            best_score, best_col = 0.0, None
            for col2 in unmapped_f2:
                s = SequenceMatcher(None, col1.lower(), col2.lower()).ratio()
                if s > best_score:
                    best_score, best_col = s, col2
            if best_col and best_score >= 0.55:
                fuzzy.append({
                    "f1_col": col1, "f2_col": best_col,
                    "score": round(best_score, 2),
                    "type_ok": str(df1[col1].dtype) == str(df2[best_col].dtype),
                    "f1_type": str(df1[col1].dtype),
                    "f2_type": str(df2[best_col].dtype),
                    "domain": _bfsi_exception_domain(col1, "VALUE_MISMATCH"),
                })
                unmapped_f2.remove(best_col)
            else:
                unmapped_f1.append({


                    "col": col1,
                    "dtype": str(df1[col1].dtype),
                    "domain": _bfsi_exception_domain(col1, "VALUE_MISMATCH"),
                })

    # Tag unmapped_f2 entries with domain
    unmapped_f2 = [
        {"col": c, "dtype": str(df2[c].dtype), "domain": _bfsi_exception_domain(c,
"VALUE_MISMATCH")}
        if isinstance(c, str) else c
        for c in unmapped_f2
    ]

    # Build field specs (one per column, mirrors TransferPricing mapping sheet
    # structure)
    field_specs = []
    for e in exact:
        try:
            field_specs.append(_build_field_spec(
                col_name=e["f1_col"], series_src=df1[e["f1_col"]],


                series_tgt=df2[e["f2_col"]] if e["f2_col"] in df2.columns else None,
                source_name=name1, target_name=name2,
                match_type="Exact", tgt_col_name=e["f2_col"],
            ))
        except Exception:
            pass
    for f in fuzzy:
        try:
            field_specs.append(_build_field_spec(
                col_name=f["f1_col"], series_src=df1[f["f1_col"]],
                series_tgt=df2[f["f2_col"]] if f["f2_col"] in df2.columns else None,
                source_name=name1, target_name=name2,
                match_type="Fuzzy", tgt_col_name=f["f2_col"],
                fuzzy_score=f["score"],
            ))
        except Exception:
            pass
    for u in unmapped_f1:
        try:


            field_specs.append(_build_field_spec(
                col_name=u["col"], series_src=df1[u["col"]],
                series_tgt=None, source_name=name1, target_name=name2,
                match_type="Unmapped-Source",
            ))
        except Exception:
            pass
    for u in unmapped_f2:
        try:
            field_specs.append(_build_field_spec(
                col_name=u["col"], series_src=df2[u["col"]],
                series_tgt=None, source_name=name2, target_name=name1,
                match_type="Unmapped-Target",
            ))
        except Exception:
            pass

    # Reconciliation: count and sum on numeric common columns
    _RECON_ROW_LIMIT = 50_000


    reconciliation = []
    # All null sentinel strings that should be treated as missing before numeric coercion
    _recon_null_map = {s: pd.NA for s in _NULL_SENTINELS}
    if len(df1) <= _RECON_ROW_LIMIT and len(df2) <= _RECON_ROW_LIMIT:
        for col in df1.columns:
            col1_numeric = pd.to_numeric(
                df1[col].astype(str).str.strip().str.lower().map(_recon_null_map).fillna(df1[col]),
                errors="coerce")
            col2_numeric = (pd.to_numeric(
                df2[col].astype(str).str.strip().str.lower().map(_recon_null_map).fillna(df2[col]),
                errors="coerce") if col in df2.columns else None)
            if col in df2.columns and col1_numeric.notna().any() and col2_numeric.notna().any():
                s1 = float(col1_numeric.sum())
                s2 = float(col2_numeric.sum())
                diff_pct = abs(s1 - s2) / abs(s1) * 100 if s1 != 0 else (0 if s2 == 0 else 100)
                reconciliation.append({
                    "column": col,


                    "f1_sum": round(s1, 4), "f2_sum": round(s2, 4),
                    "f1_count": int(df1[col].count()), "f2_count": int(df2[col].count()),
                    "sum_diff": round(s2 - s1, 4),
                    "diff_pct": round(diff_pct, 2),
                    "status": "OK" if diff_pct < 0.01 else ("WARN" if diff_pct < 1.0 else "FAIL"),
                    "domain": _bfsi_exception_domain(col, "VALUE_MISMATCH"),
                })

    total_mapped = len(exact) + len(fuzzy)
    completeness_pct = round(
        total_mapped / len(df1.columns) * 100 if df1.columns.any() else 0, 1
    )
    avg_fuzzy_score = round(
        sum(f["score"] for f in fuzzy) / len(fuzzy) if fuzzy else 0, 2
    )
    row_ratio = (
        min(len(df1), len(df2)) / max(len(df1), len(df2))
        if max(len(df1), len(df2)) > 0 else 1.0
    )


    # Detect files that are likely unrelated to the same business domain
    warning_reasons = []
    if completeness_pct < 30:
        warning_reasons.append(
            f"only {completeness_pct}% of columns match (threshold: 30%)"
        )
    if len(exact) == 0:
        warning_reasons.append("no exact column name matches found")
    if fuzzy and avg_fuzzy_score < 0.6:
        warning_reasons.append(
            f"fuzzy column matches are weak (avg score {avg_fuzzy_score}, threshold: 0.60)"
        )
    if row_ratio < 0.1:
        warning_reasons.append(
            f"row counts are very different ({len(df1):,} vs {len(df2):,})"
        )

    relatedness_warning = (


        {
            "level": "HIGH" if len(warning_reasons) >= 2 else "MEDIUM",
            "message": (
                "These files appear to be from DIFFERENT business domains "
                "and may not be meaningful to compare."
            ),
            "reasons": warning_reasons,
        }
        if warning_reasons else None
    )

    # -- Business rule validation
    # Run each spec rule against df1 (source file).  Only rules where the
    # source column actually exists in df1 are evaluated.
    rule_results = []
    if mapping_spec:
        for spec in mapping_spec:
            col = spec["source_column"]
            if not col or col not in df1.columns:


                continue

            series   = df1[col]
            violations: list[dict] = []
            severity = spec.get("severity", "ERROR")

            # -- not_null check
            if spec.get("not_null"):
                null_mask = series.isna() | (series.astype(str).str.strip() == "") | (series.astype(str).str.upper() == "NULL")
                for idx in series[null_mask].index[:100]:
                    violations.append({
                        "row_index": int(idx),
                        "value": None,
                        "message": f"{col}: value is null/blank (not_null=TRUE)",
                    })

            # -- unique check
            if spec.get("unique"):


                dupes = series[series.duplicated(keep=False) & series.notna()]
                for idx in dupes.index[:100]:
                    if len(violations) >= 100:
                        break
                    violations.append({
                        "row_index": int(idx),
                        "value": str(series.at[idx]),
                        "message": f"{col}: duplicate value '{series.at[idx]}' (unique=TRUE)",
                    })

            # -- value_in_list check
            allowed = spec.get("value_in_list") or []
            if allowed:
                non_null_mask = series.notna() & (series.astype(str).str.strip() != "")
                bad_mask = non_null_mask & ~series.astype(str).str.strip().isin(allowed)
                for idx in series[bad_mask].index[:100]:
                    if len(violations) >= 100:
                        break
                    violations.append({


                "row_index": int(idx),
                "value": str(series.at[idx]),
                "message": f"{col}: '{series.at[idx]}' not in allowed list {allowed}",
            })

        # --- min_value / max_value check ----------------------------------------
        min_val = spec.get("min_value")
        max_val = spec.get("max_value")
        if min_val is not None or max_val is not None:
            numeric = pd.to_numeric(series, errors="coerce")
            if min_val is not None:
                bad_min = numeric[numeric.notna() & (numeric < min_val)]
                for idx in bad_min.index[:100]:
                    if len(violations) >= 100:
                        break
                    violations.append({
                        "row_index": int(idx),
                        "value": str(series.at[idx]),
                        "message": f"{col}: {series.at[idx]} < min_value {min_val}",


                    })
            if max_val is not None:
                bad_max = numeric[numeric.notna() & (numeric > max_val)]
                for idx in bad_max.index[:100]:
                    if len(violations) >= 100:
                        break
                    violations.append({
                        "row_index": int(idx),
                        "value": str(series.at[idx]),
                        "message": f"{col}: {series.at[idx]} > max_value {max_val}",
                    })

        # --- regex_pattern check ----------------------------------------
        pattern = spec.get("regex_pattern", "")
        if pattern:
            non_null_mask = series.notna() & (series.astype(str).str.strip() != "")
            try:
                bad_re = series[non_null_mask & ~series.astype(str).str.match(pattern, na=False)]


                for idx in bad_re.index[:100]:
                    if len(violations) >= 100:
                        break
                    violations.append({
                        "row_index": int(idx),
                        "value": str(series.at[idx]),
                        "message": f"{col}: '{series.at[idx]}' does not match pattern '{pattern}'",
                    })
            except re.error:
                pass  # invalid regex -- skip silently

        # --- condition check ----------------------------------------
        # Supports only:  IF <col> IN (val1,val2,...) THEN <col2> IS NOT NULL
        condition = spec.get("condition", "")
        if condition:
            cond_match = re.match(
                r"IF\s+(\w+)\s+IN\s+\(([^)]+)\)\s+THEN\s+(\w+)\s+IS\s+NOT\s+NULL",
                condition.strip(),
                re.IGNORECASE,


            )
            if cond_match:
                trigger_col   = cond_match.group(1).strip()
                trigger_vals  = [v.strip() for v in cond_match.group(2).split(",")]
                required_col  = cond_match.group(3).strip()
                if trigger_col in df1.columns and required_col in df1.columns:
                    trigger_series  = df1[trigger_col].astype(str).str.strip().str.upper()
                    trigger_vals_up = [v.upper() for v in trigger_vals]
                    req_series      = df1[required_col]
                    trigger_active  = trigger_series.isin(trigger_vals_up)
                    req_missing     = req_series.isna() | (req_series.astype(str).str.strip() == "") | (req_series.astype(str).str.upper() == "NULL")
                    bad_cond        = trigger_active & req_missing
                    for idx in df1[bad_cond].index[:100]:
                        if len(violations) >= 100:
                            break
                        violations.append({
                            "row_index": int(idx),
                            "value": str(df1.at[idx, trigger_col]),


                            "message": (
                                f"Condition violated: {trigger_col}="
                                f"'{df1.at[idx, trigger_col]}' requires "
                                f"{required_col} to be non-null"
                            ),
                        })

        rule_results.append({
            "column":     col,
            "rule":       spec.get("business_rule") or spec.get("description") or col,
            "severity":   severity,
            "violations": violations,
            "violation_count": len(violations),
            "status":     (
                "PASS" if not violations
                else ("ERROR" if severity == "ERROR" else
                      ("WARN" if severity == "WARNING" else "INFO"))
            ),
            "sample_violations": violations[:5],


        })

    # Flat list of source-side column names that have a confirmed match (exact or fuzzy).
    # Used by compare_dataframes to restrict key inference to semantically equivalent
    # columns.
    matched_col_names = [m["f1_col"] for m in exact] + [m["f1_col"] for m in fuzzy]

    # Domain summary counts across all mapped + unmapped entries
    all_domain_items = (
        [e["domain"] for e in exact]
        + [f["domain"] for f in fuzzy]
        + [u["domain"] for u in unmapped_f1]
        + [u["domain"] for u in unmapped_f2]
    )
    domain_summary = {
        "Business":    all_domain_items.count("Business"),
        "Technical":   all_domain_items.count("Technical"),
        "Operational": all_domain_items.count("Operational"),
        "Other":       all_domain_items.count("Other"),


    }

    return {
        "file1_name": name1, "file2_name": name2,
        "has_spec": bool(mapping_spec),
        "spec_results": spec_results,
        "rule_results": rule_results,
        "exact": exact, "fuzzy": fuzzy,
        "unmapped_f1": unmapped_f1,
        "unmapped_f2": unmapped_f2,
        "total_mapped": total_mapped,
        "total_f1": len(df1.columns),
        "total_f2": len(df2.columns),
        "mapping_completeness_pct": completeness_pct,
        "avg_fuzzy_score": avg_fuzzy_score,
        "relatedness_warning": relatedness_warning,
        "reconciliation": reconciliation,
        "matched_col_names": matched_col_names,
        "domain_summary": domain_summary,


        "field_specs": field_specs,
    }


# ----------------------------------------------------------------------
# LLM helpers
# ----------------------------------------------------------------------

def _hints_block(hints: dict | None, fields: list[str]) -> str:
    # Build a formatted user-hints block to inject into any LLM prompt.

    # domain_context carries the raw free-text the user typed -- always rendered
    # verbatim.  Extracted sub-fields (key_hints, mapping_hints, etc.) are
    # appended as a compact structured reference so the LLM sees both.

    if not hints:
        return ""
    raw = hints.get("domain_context", "").strip()
    if not raw:
        return ""
    _labels = {
        "key_hints":       "Extracted key columns",
        "mapping_hints":   "Extracted column mappings",
        "exclude_hints":   "Extracted exclusions",
        "transform_hints": "Extracted transforms",
        "nullable_hints":  "Extracted nullable columns",
        "pii_context":     "Governance context",
    }
    structured = []
    for f in fields:
        if f != "domain_context" and hints.get(f):
            structured.append(f"  {_labels.get(f, f)}: {hints[f]}")
    block = (
        "\nADDITIONAL CONTEXT PROVIDED BY THE USER (treat as authoritative guidance):\n"
        f"{raw}\n"
    )
    if structured:
        block += "Parsed from above:\n" + "\n".join(structured) + "\n"
    return block


def _parse_prompt(filename: str, raw_text: str, hints: dict | None = None) -> str:
    # Build an LLM prompt to convert unstructured text into a structured JSON table.
    hb = _hints_block(hints, ["domain_context", "format_hint"])
    # Truncate to avoid hitting token limits (~6000 chars is safe)
    preview = raw_text[:6000]
    if len(raw_text) > 6000:
        preview += "\n... [truncated]"
    return (
        f"You are a data parsing expert. The file '{filename}' contains unstructured or "
        "semi-structured text. Extract ALL records into a structured tabular format.\n"
        + hb + "\n"
        f"FILE CONTENT:\n{preview}\n\n"


        "Return ONLY a valid JSON object with this exact structure (no markdown, no commentary):\n"
        '{\n'
        '  "columns": ["col1", "col2", ...],\n'
        '  "rows": [\n'
        '    {"col1": "value", "col2": "value", ...},\n'
        '    ...\n'
        '  ],\n'
        '  "format_detected": "brief description of detected format",\n'
        '  "notes": "any parsing caveats or assumptions"\n'
        '}\n\n'
        "If the content cannot be structured, return:\n"
        '{"columns": [], "rows": [], "format_detected": "unstructured", "notes": "reason"}'
    )


def parse_unstructured(raw_bytes: bytes, filename: str,
                        hints: dict | None = None) -> dict:


    # Use Claude LLM to parse an unstructured file into a structured table.
    # Accepts the raw bytes of the uploaded file so no temp file is needed.
    # Returns a dict with columns, rows, row_count, col_count, format_detected, notes,
    # error.

    try:
        encoding = chardet.detect(raw_bytes).get("encoding") or "utf-8"
        raw_text = raw_bytes.decode(encoding, errors="replace")
    except Exception as exc:
        return {
            "file_name": filename, "columns": [], "rows": [],
            "row_count": 0, "col_count": 0,
            "format_detected": "error", "notes": "", "error": str(exc),
        }

    prompt = _parse_prompt(filename, raw_text, hints=hints)
    try:
        response_text = _ask_llm(
            [{"role": "user", "content": [{"text": prompt}]}],


            system="You are a structured data extraction engine. Output ONLY valid JSON.",
        )
        # Strip any accidental markdown fences
        cleaned = response_text.strip()
        if cleaned.startswith("```"):
            cleaned = "\n".join(cleaned.splitlines()[1:])
        if cleaned.endswith("```"):
            cleaned = "\n".join(cleaned.splitlines()[:-1])
        parsed = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        return {
            "file_name": filename, "columns": [], "rows": [],
            "row_count": 0, "col_count": 0,
            "format_detected": "parse error",
            "notes": f"LLM returned non-JSON: {str(exc)[:120]}",
            "error": "JSON decode error",
        }
    except Exception as exc:
        return {


            "file_name": filename, "columns": [], "rows": [],
            "row_count": 0, "col_count": 0,
            "format_detected": "error", "notes": "", "error": str(exc),
        }

    columns = parsed.get("columns", [])
    rows = parsed.get("rows", [])
    format_detected = parsed.get("format_detected", "unknown")
    notes = parsed.get("notes", "")
    return {
        "file_name": filename,
        "columns": columns,
        "rows": rows,             # list of dicts
        "row_count": len(rows),
        "col_count": len(columns),
        "format_detected": format_detected,
        "notes": notes,
        "error": None,


    }


# ----------------------------------------------------------------------
# Excel report generation
# ----------------------------------------------------------------------

_HDR_FILL  = PatternFill("solid", fgColor="1a1a2e")
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
_PASS_FILL = PatternFill("solid", fgColor="D1FAE5")
_FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")
_WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
_ALT_FILL  = PatternFill("solid", fgColor="F8F9FF")


def _hdr(ws, row: int, cols: list[str]) -> None:
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = _HDR_FONT


        c.fill = _HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 18


def _autofit(ws, max_w: int = 50) -> None:
    for col in ws.columns:
        best = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 2, max_w)


def _status_fill(status: str) -> PatternFill:
    s = str(status).upper()
    return _PASS_FILL if s in ("PASS", "OK") else _FAIL_FILL if s == "FAIL" else _WARN_FILL


def generate_excel(data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()


    wb.remove(wb.active)  # remove default empty sheet
    action = data.get("action", "")
    files  = data.get("file_names", [])

    # --- Summary sheet
    ws = wb.create_sheet("Summary")

    # AI Executive Summary -- shown first if available
    _ai_summary = data.get("ai_summary", "")
    _start_row = 1
    if _ai_summary:
        ws.cell(1, 1, "AI Executive Summary").font = Font(bold=True, size=12,
            color="FFFFFF")
        ws.cell(1, 1).fill = PatternFill("solid", fgColor="1F3B6B")
        ws.merge_cells("A1:B1")
        ws.cell(2, 1, _ai_summary)
        ws.cell(2, 1).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = max(60, len(_ai_summary) // 3)


        ws.merge_cells("A2:B2")
        ws.cell(3, 1, "").fill = PatternFill("solid", fgColor="F0F4FF")
        ws.merge_cells("A3:B3")
        _start_row = 5

    _hdr(ws, _start_row, ["Item", "Value"])
    rows = [
        ("Action",      action),
        ("Data Files",  ", ".join(files)),
        ("Elapsed (s)", data.get("elapsed", "")),
    ]
    # Add DQ scores if present
    for qr in data.get("quality_reports", []):
        dqs = qr.get("dq_score", {})
        if dqs:
            rows += [
                ("DQ Score",      f"{dqs.get('score','')}/100 (Grade {dqs.get('grade','')})"),
                ("Completeness",  f"{dqs.get('completeness','')}%"),
                ("Uniqueness",    f"{dqs.get('uniqueness','')}%"),


                ("Validity",       f"{dqs.get('validity','')}%"),
                ("Rule Failures",  sum(1 for r in qr.get("rule_results",[]) if r.get("status")=="FAIL")),
                ("Duplicate Rows", qr.get("duplicate_rows",0)),
            ]
    for r in data.get("ref_log", []):
        rows.append((f"Ref Doc: {r['file']}", f"{r['type']} ({r.get('rows',0)} rows)"))
    for ri, (k, v) in enumerate(rows, _start_row + 1):
        ws.cell(ri, 1, k).font = Font(bold=True)
        ws.cell(ri, 2, str(v))
        if ri % 2 == 0:
            for ci in range(1, 3):
                ws.cell(ri, ci).fill = _ALT_FILL
    _autofit(ws)

    # --- Comparison sheets -- one tab per exception type per file pair
    def _safe_sn(name: str, suffix: str, used: set) -> str:
        """Build a unique sheet name ≤31 chars."""
        base = re.sub(r"[\\/*?:\[\]]", "", f"{suffix} {name}")[:28].strip()


        sn = base
        i = 2
        while sn in used:
            sn = f"{base[:25]}{i}"
            i += 1
        used.add(sn)
        return sn

    _used_sn: set = set()

    for p in data.get("pairs", []):
        d        = p["diff"]
        f1       = p["file1_name"]
        f2       = p["file2_name"]
        key_cols = d.get("key_columns", [])
        data_cols = d.get("data_columns", [])

        # --- Comparison Summary tab
        ws = wb.create_sheet(_safe_sn(f1, "Cmp Summary", _used_sn))


        _hdr(ws, 1, ["Metric", "Value"])
        meta = [
            ("File 1",                       f1),
            ("File 2",                       f2),
            ("Key method",                   d["key_method"]),
            ("Key columns",                  ", ".join(key_cols or [])),
            ("File 1 rows",                  d["file1_rows"]),
            ("File 2 rows",                  d["file2_rows"]),
            ("Extra rows from File 1",       d["removed_count"]),
            ("Extra rows from File 2",       d["added_count"]),
            ("Modified rows",                d["modified_count"]),
            ("Extra columns from File 2",    ", ".join(d["schema_added_columns"])),
            ("Extra columns from File 1",    ", ".join(d["schema_removed_columns"])),
        ]
        for ri, (k, v) in enumerate(meta, 2):
            ws.cell(ri, 1, k).font = Font(bold=True)
            ws.cell(ri, 2, str(v))
            if ri % 2 == 0:
                for ci in range(1, 3):


                    ws.cell(ri, ci).fill = _ALT_FILL
        _autofit(ws)

        # --- Rows only in File 1 (exception rows -- key columns only)
        if d.get("file1_only"):
            ws = wb.create_sheet(_safe_sn(f1, "Only", _used_sn))
            ws.cell(1, 1, f"Exception: rows in '{f1}' NOT found in '{f2}' ({d['file1_only_count']} rows)").font = Font(bold=True, color="991B1B")
            # Only write key + columns that actually have data (skip blanks)
            sample_rows = d["file1_only"][:10]
            used_data_cols = [c for c in data_cols
                if any(r["row_data"].get(c) not in (None, "", "nan") for r in sample_rows)]
            out_cols = key_cols + used_data_cols
            _hdr(ws, 2, out_cols)
            for ri, r_item in enumerate(d["file1_only"], 3):
                for ci, k in enumerate(key_cols, 1):
                    ws.cell(ri, ci, r_item["key_values"].get(k, ""))
                for ci, c in enumerate(used_data_cols, len(key_cols) + 1):
                    ws.cell(ri, ci, r_item["row_data"].get(c, ""))


                if ri % 2 == 0:
                    for ci in range(1, len(out_cols) + 1):
                        ws.cell(ri, ci).fill = _ALT_FILL
            _autofit(ws)

        # --- Rows only in File 2 (exception rows -- key columns only)
        if d.get("file2_only"):
            ws = wb.create_sheet(_safe_sn(f2, "Only", _used_sn))
            ws.cell(1, 1, f"Exception: rows in '{f2}' NOT found in '{f1}' ({d['file2_only_count']} rows)").font = Font(bold=True, color="166534")
            sample_rows = d["file2_only"][:10]
            used_data_cols = [c for c in data_cols
                if any(r["row_data"].get(c) not in (None, "", "nan") for r in sample_rows)]
            out_cols = key_cols + used_data_cols
            _hdr(ws, 2, out_cols)
            for ri, r_item in enumerate(d["file2_only"], 3):
                for ci, k in enumerate(key_cols, 1):
                    ws.cell(ri, ci, r_item["key_values"].get(k, ""))
                for ci, c in enumerate(used_data_cols, len(key_cols) + 1):


                    ws.cell(ri, ci, r_item["row_data"].get(c, ""))
                if ri % 2 == 0:
                    for ci in range(1, len(out_cols) + 1):
                        ws.cell(ri, ci).fill = _ALT_FILL
            _autofit(ws)

        # --- Modified rows (exceptions only -- key + changed column + old/new) --
        if d.get("modified_rows"):
            ws = wb.create_sheet(_safe_sn(f1, "Modified Rows", _used_sn))
            ws.cell(1, 1, f"Exception: rows changed between '{f1}' and '{f2}' ({d['modified_count']} rows, {sum(len(mr.get('changes',{})) for mr in d['modified_rows'])} field changes)").font = Font(bold=True, color="C0392B")
            _hdr(ws, 2, ["Key", "Changed Column", f"Old Value ({f1})", f"New Value ({f2})"])
            # Style header
            for ci, colour in [(3, "FF0000"), (4, "166534")]:
                ws.cell(2, ci).font = Font(bold=True, color=colour)
            ri = 3
            for mr in d["modified_rows"]:
                key_str = " | ".join(f"{k}={v}" for k, v in mr["key_values"].items())
                for col, chg in mr["changes"].items():


                    ws.cell(ri, 1, key_str)
                    ws.cell(ri, 2, col).font = Font(bold=True)
                    c3 = ws.cell(ri, 3, chg["file1"])
                    c4 = ws.cell(ri, 4, chg["file2"])
                    c3.fill = PatternFill("solid", fgColor="FFF0F0")
                    c4.fill = PatternFill("solid", fgColor="F0FFF4")
                    if ri % 2 == 0:
                        ws.cell(ri, 1).fill = _ALT_FILL
                        ws.cell(ri, 2).fill = _ALT_FILL
                    ri += 1
            _autofit(ws)

        # --- Null column exceptions
        null_col_exc = d.get("null_column_exceptions", [])
        if null_col_exc:
            ws = wb.create_sheet(_safe_sn(f1, "Null Exceptions", _used_sn))
            ws.cell(1, 1, f"Column null exceptions ({len(null_col_exc)})").font = Font(bold=True, color="7C3AED")
            _hdr(ws, 2, ["Column", "Has Data In", "All Null In", "Non-Null Count", "Sample Values"])


            for ri, exc in enumerate(null_col_exc, 3):
                ws.cell(ri, 1, exc["column"])
                ws.cell(ri, 2, exc["has_data_in"])
                ws.cell(ri, 3, exc["all_null_in"])
                ws.cell(ri, 4, exc["non_null_count"])
                ws.cell(ri, 5, ", ".join(exc["sample_values"]))
                if ri % 2 == 0:
                    for ci in range(1, 6):
                        ws.cell(ri, ci).fill = _ALT_FILL
            _autofit(ws)

        # --- Duplicate rows
        for label, dup_rows, dup_cols, fname in [
            (f"Duplicates in '{f1}'", d.get("file1_duplicate_rows", []), d.get("file1_dup_columns", []), f1),
            (f"Duplicates in '{f2}'", d.get("file2_duplicate_rows", []), d.get("file2_dup_columns", []), f2),
        ]:
            if dup_rows:
                ws = wb.create_sheet(_safe_sn(fname, "Duplicates", _used_sn))


                ws.cell(1, 1, label).font = Font(bold=True, color="3730A3")
                _hdr(ws, 2, dup_cols)
                for ri, dr in enumerate(dup_rows, 3):
                    for ci, c in enumerate(dup_cols, 1):
                        ws.cell(ri, ci, dr.get(c, ""))
                    if ri % 2 == 0:
                        for ci in range(1, len(dup_cols) + 1):
                            ws.cell(ri, ci).fill = _ALT_FILL
                _autofit(ws)

    # --- DQ Results sheet
    for q in data.get("quality_reports", []):
        sn = f"DQ_{q['file_name'][:24]}"[:31]
        ws = wb.create_sheet(sn)

        # --- Score summary (matches UI stat cards exactly)
        dq = q["dq_score"]
        summary_rows = [
            ("DQ Score",  f"{dq['score']}/100 Grade {dq['grade']}"),


            ("Base Score",         f"{dq.get('base_score', dq['score'])}/100"),
            ("Governance Penalty", f"-{dq.get('governance_penalty', 0)} pts"),
            ("Completeness",       f"{dq['completeness']}%"),
            ("Uniqueness",         f"{dq['uniqueness']}%"),
            ("Validity",           f"{dq['validity']}%"),
            ("Total Rows",         q["total_rows"]),
            ("Total Columns",      q.get("total_cols", len(q.get("columns", [])))),
            ("Duplicate Rows",     q["duplicate_rows"]),
            ("Rule FAILs",         sum(1 for r in q.get("rule_results", []) if r.get("status") == "FAIL")),
        ]
        for ri, (label, value) in enumerate(summary_rows, 1):
            c1 = ws.cell(ri, 1, label)
            c1.font = Font(bold=True)
            c2 = ws.cell(ri, 2, value)
            if ri == 1:
                c1.font = Font(bold=True, size=12)
                c2.font = Font(bold=True, size=12,
                    color={"A": "166534", "B": "0369A1",
                           "C": "92400E", "F": "991B1B"}.get(dq["grade"], "374151"))


        # Build governance lookup: col_name → gov_entry (from embedded governance)
        _gov_cols = {}
        gov = q.get("governance")
        if gov:
            for gc in gov.get("columns", []):
                _gov_cols[gc["column"]] = gc

        # --- Column details -- same column order as UI
        has_dict = q.get("has_data_dict", False) or bool(_gov_cols)
        col_hdrs = ["Column", "DQ Score", "Type", "Null %", "Unique %",
                    "Cardinality", "Min", "Max", "Mean", "Outlier Count", "Outlier %"]
        if has_dict:
            col_hdrs += ["Sensitivity", "Regulatory Frameworks", "Access Recommendation",
                         "Owner", "Description"]

        row = len(summary_rows) + 2
        ws.cell(row, 1, "Column Details").font = Font(bold=True, size=11)
        row += 1


        _hdr(ws, row, col_hdrs)
        row += 1

        for c in q.get("columns", []):
            col_score = c.get("dq_score", "")
            col_grade = c.get("dq_grade", "")
            score_str = f"{col_score} ({col_grade})" if col_score != "" else ""
            ws.cell(row, 1, c["name"]).font = Font(bold=True)
            c2 = ws.cell(row, 2, score_str)
            if col_grade:
                c2.font = Font(bold=True,
                    color={"A": "166534", "B": "0369A1",
                           "C": "92400E", "F": "991B1B"}.get(col_grade, "374151"))
            ws.cell(row, 3, c.get("dtype", ""))
            # Null % -- red if high
            c4 = ws.cell(row, 4, c.get("null_pct", 0))
            if (c.get("null_pct") or 0) > 20:
                c4.fill = _FAIL_FILL
            elif (c.get("null_pct") or 0) > 5:


                c4.fill = _WARN_FILL
            ws.cell(row, 5, c.get("uniqueness_pct", ""))
            ws.cell(row, 6, c.get("cardinality", ""))
            ws.cell(row, 7, c.get("min", ""))
            ws.cell(row, 8, c.get("max", ""))
            ws.cell(row, 9, c.get("mean", ""))
            ws.cell(row, 10, c.get("outlier_count", ""))
            ws.cell(row, 11, c.get("outlier_pct", ""))
            if has_dict:
                # Pull governance data from embedded governance or dd_ fields
                gv = _gov_cols.get(c["name"], {})
                sens  = gv.get("sensitivity") or c.get("dd_sensitivity", "")
                regs  = ", ".join(gv.get("regulatory", [])) or ""
                acc   = gv.get("access_rec") or ""
                owner = gv.get("owner") or c.get("dd_owner", "")
                desc  = gv.get("description") or c.get("dd_description", "")
                c12 = ws.cell(row, 12, sens)
                c12.fill = (_FAIL_FILL if sens == "Highly Restricted"
                            else _WARN_FILL if sens == "Confidential"


                            else _PASS_FILL if sens else PatternFill())
                ws.cell(row, 13, regs)
                ws.cell(row, 14, acc)
                ws.cell(row, 15, owner)
                ws.cell(row, 16, desc)
            if row % 2 == 0:
                for ci2 in range(1, len(col_hdrs) + 1):
                    if not ws.cell(row, ci2).fill.fgColor.rgb not in ("00000000", "FFFFFFFF"):
                        ws.cell(row, ci2).fill = _ALT_FILL
            row += 1

        # --- Business rules -- FAIL/WARN only (exceptions only, matches UI) --
        fail_warn = [r for r in q.get("rule_results", [])
                     if r.get("status") in ("FAIL", "WARN") and not r.get("skipped")]
        if fail_warn:
            row += 1
            ws.cell(row, 1, f"Rule Exceptions ({len(fail_warn)} FAIL/WARN)").font = Font(bold=True, size=11)
            row += 1


            _hdr(ws, row, ["Rule Name", "Column", "Rule Type", "Status",
                           "Pass %", "Failed Count", "Failing Examples"])
            row += 1
            for r in fail_warn:
                ws.cell(row, 1, r.get("rule_name", ""))
                ws.cell(row, 2, r.get("column_name", ""))
                ws.cell(row, 3, r.get("rule_type", ""))
                st = r.get("status", "")
                c_st = ws.cell(row, 4, st)
                c_st.fill = _status_fill(st)
                ws.cell(row, 5, r.get("pass_pct", ""))
                ws.cell(row, 6, r.get("failed", ""))
                ws.cell(row, 7, ", ".join(str(x) for x in r.get("failing_examples", [])[:5]))
                row += 1

        # --- BFSI structural validation failures
        bfsi_fails = [v for v in q.get("bfsi_validation", {}).values()
                      if v.get("status") != "PASS"]
        if bfsi_fails:


            row += 1
            ws.cell(row, 1, "BFSI Identifier Validation Failures").font = Font(bold=True, size=11)
            row += 1
            _hdr(ws, row, ["Column", "Identifier Type", "Invalid Count",
                           "Invalid %", "Status", "Sample Invalid Values"])
            row += 1
            for v in bfsi_fails:
                ws.cell(row, 1, v.get("column", ""))
                ws.cell(row, 2, v.get("identifier_type", ""))
                ws.cell(row, 3, v.get("invalid_count", ""))
                ws.cell(row, 4, v.get("invalid_pct", ""))
                c_st = ws.cell(row, 5, v.get("status", ""))
                c_st.fill = _FAIL_FILL if v.get("status") == "FAIL" else _WARN_FILL
                ws.cell(row, 6, ", ".join(str(x) for x in v.get("invalid_samples", [])[:5]))
                row += 1

        # --- Volume / freshness anomalies
        anomalies = q.get("volume_freshness", {}).get("anomalies", [])


        if anomalies:
            row += 1
            ws.cell(row, 1, "Volume & Freshness Anomalies").font = Font(bold=True, size=11)
            row += 1
            for a in anomalies:
                c_a = ws.cell(row, 1, a)
                c_a.fill = _WARN_FILL
                row += 1

        _autofit(ws)

    # --- Governance sheet
    for g in data.get("governance_reports", []):
        sn = f"Gov_{g['file_name'][:25]}"[:31]
        ws = wb.create_sheet(sn)
        ws.cell(1, 1, "Classification").font = Font(bold=True)
        ws.cell(1, 2, g["overall_classification"])
        ws.cell(2, 1, "Regulatory").font = Font(bold=True)


        ws.cell(2, 2, ", ".join(g["regulatory_frameworks"]))
        ws.cell(3, 1, "PII columns").font = Font(bold=True)
        ws.cell(3, 2, g["pii_column_count"])
        row = 5
        _hdr(ws, row, ["Column", "Type", "Sensitivity", "PII Findings",
                       "Regulatory", "Owner", "Access Recommendation", "Description"])
        row += 1
        for col in g.get("columns", []):
            ws.cell(row, 1, col["column"])
            ws.cell(row, 2, col["dtype"])
            sens = col["sensitivity"]
            c_s = ws.cell(row, 3, sens)
            c_s.fill = (_FAIL_FILL if sens == "Highly Restricted"
                        else _WARN_FILL if sens == "Confidential" else _PASS_FILL)
            ws.cell(row, 4, "; ".join(col.get("pii_detected", [])))
            ws.cell(row, 5, ", ".join(col.get("regulatory", [])))
            ws.cell(row, 6, col.get("owner", ""))
            ws.cell(row, 7, col.get("access_rec", ""))
            ws.cell(row, 8, col.get("description", ""))


            row += 1
        _autofit(ws)

    # --- Data Profile sheet
    for p in data.get("profile_reports", []):
        sn = f"Profile_{p['file_name'][:20]}"[:31]
        ws = wb.create_sheet(sn)

        # --- Summary header
        summary = [
            ("File",             p["file_name"]),
            ("Total Rows",       p.get("total_rows", "")),
            ("Total Columns",    p.get("total_cols", "")),
            ("Memory (MB)",      p.get("memory_mb", "")),
            ("Duplicate Rows",   p.get("duplicate_rows", "")),
            ("Key Candidates",   ", ".join(p.get("key_candidates", []))),
            ("Near-Key Columns", ", ".join(p.get("near_key_cols", []))),
        ]
        for ri, (label, value) in enumerate(summary, 1):


            ws.cell(ri, 1, label).font = Font(bold=True)
            ws.cell(ri, 2, str(value) if value is not None else "")

        # --- Semantic type breakdown
        row = len(summary) + 2
        ws.cell(row, 1, "Column Type Breakdown").font = Font(bold=True, size=11)
        row += 1
        for sem_type, cnt in p.get("type_breakdown", {}).items():
            ws.cell(row, 1, sem_type)
            ws.cell(row, 2, cnt)
            row += 1

        # --- Strong correlations (exceptions: strength = strong or moderate) --
        strong_corr = [c for c in p.get("correlations", [])
                       if c.get("strength") in ("strong", "moderate")]
        if strong_corr:
            row += 1
            ws.cell(row, 1, f"Notable Correlations ({len(strong_corr)})").font = Font(bold=True, size=11)


            row += 1
            _hdr(ws, row, ["Column A", "Column B", "Correlation", "Strength"])
            row += 1
            for c in strong_corr:
                ws.cell(row, 1, c.get("col1", ""))
                ws.cell(row, 2, c.get("col2", ""))
                corr_val = c.get("corr", 0)
                c3 = ws.cell(row, 3, corr_val)
                c3.fill = (_FAIL_FILL if abs(corr_val) >= 0.7 else _WARN_FILL)
                ws.cell(row, 4, c.get("strength", ""))
                row += 1

        # --- Column exceptions: nulls, outliers, low cardinality anomalies --
        # Exception = column has nulls OR has outliers OR identifier-like with dupes
        exc_cols = [
            c for c in p.get("columns", [])
            if (c.get("null_pct", 0) > 0)
            or (c.get("outlier_count", 0) or 0) > 0
            or (c.get("cardinality") == "low (enum)" and c.get("unique_count", 0) <= 1)


        ]
        row += 1
        ws.cell(row, 1, f"Column Exceptions ({len(exc_cols)} of {len(p.get('columns',[]))} columns have issues)").font = Font(bold=True, size=11)
        row += 1
        _hdr(ws, row, ["Column", "BFSI Domain", "Criticality", "Semantic Type",
                       "Dtype", "Null Count", "Null %",
                       "Unique Count", "Unique %", "Cardinality",
                       "Min", "Max", "Mean", "Std",
                       "Outlier Count", "Outlier %", "Top Values"])
        row += 1
        for c in exc_cols:
            top_vals = "; ".join(
                f"{tv.get('value','')}({tv.get('count','')})"
                for tv in (c.get("top_values") or [])[:5]
            )
            null_pct = c.get("null_pct", 0) or 0
            out_cnt  = c.get("outlier_count", 0) or 0


            ws.cell(row, 1, c.get("name", "")).font = Font(bold=True)
            ws.cell(row, 2, c.get("bfsi_domain", ""))
            crit_cell = ws.cell(row, 3, c.get("criticality", ""))
            if c.get("criticality") == "High":
                crit_cell.font = Font(bold=True, color="991B1B")
            ws.cell(row, 4, c.get("semantic", ""))
            ws.cell(row, 5, c.get("dtype", ""))
            c6 = ws.cell(row, 6, c.get("null_count", 0))
            c7 = ws.cell(row, 7, null_pct)
            if null_pct > 20:
                c6.fill = _FAIL_FILL; c7.fill = _FAIL_FILL
            elif null_pct > 5:
                c6.fill = _WARN_FILL; c7.fill = _WARN_FILL
            ws.cell(row, 8, c.get("unique_count", ""))
            ws.cell(row, 9, c.get("unique_pct", ""))
            ws.cell(row, 10, c.get("cardinality", ""))
            ws.cell(row, 11, c.get("min", ""))
            ws.cell(row, 12, c.get("max", ""))


            ws.cell(row, 13, c.get("mean", ""))
            ws.cell(row, 14, c.get("std", ""))
            c15 = ws.cell(row, 15, out_cnt)
            c16 = ws.cell(row, 16, c.get("outlier_pct", ""))
            if out_cnt > 0:
                c15.fill = _WARN_FILL; c16.fill = _WARN_FILL
            ws.cell(row, 17, top_vals)
            if row % 2 == 0:
                for ci2 in range(1, 18):
                    if ws.cell(row, ci2).fill.patternType is None:
                        ws.cell(row, ci2).fill = _ALT_FILL
            row += 1

        _autofit(ws)

    # --- Mapping sheet
    for m in data.get("mappings", []):
        sn = f"Map_{m['file1_name'][:12]}_{m['file2_name'][:12]}"[:31]


        ws = wb.create_sheet(sn)
        ws.cell(1, 1, "Mapping Completeness").font = Font(bold=True)
        ws.cell(1, 2, f"{m['mapping_completeness_pct']}%")
        # Spec results
        row = 3
        if m.get("spec_results"):
            _hdr(ws, row, ["Source Column", "Target Column", "Status", "Mandatory",
                           "F1 Type", "F2 Type", "Type OK", "Transformation"])
            row += 1
            for s in m["spec_results"]:
                ws.cell(row, 1, s["source_column"])
                ws.cell(row, 2, s["target_column"])
                c_s = ws.cell(row, 3, s["status"])
                c_s.fill = _status_fill(s["status"])
                ws.cell(row, 4, "Yes" if s["mandatory"] else "No")
                ws.cell(row, 5, s["f1_type"])
                ws.cell(row, 6, s["f2_type"])
                ws.cell(row, 7, "OK" if s.get("type_ok") else ("MISMATCH" if s.get("type_ok") is False else "--"))


                ws.cell(row, 8, s.get("transformation", ""))
                row += 1
            row += 1

        # Reconciliation
        if m.get("reconciliation"):
            _hdr(ws, row, ["Column", "F1 Sum", "F2 Sum", "Diff", "Diff %", "Status"])
            row += 1
            for r in m["reconciliation"]:
                ws.cell(row, 1, r["column"])
                ws.cell(row, 2, r["f1_sum"])
                ws.cell(row, 3, r["f2_sum"])
                ws.cell(row, 4, r["sum_diff"])
                ws.cell(row, 5, r["diff_pct"])
                c_s = ws.cell(row, 6, r["status"])
                c_s.fill = _status_fill(r["status"])
                row += 1
        _autofit(ws)


    # --- Cross Reference sheets
    xref = data.get("xref")
    if xref:
        xs = xref.get("summary", {})

        # XRef Summary
        ws = wb.create_sheet("XRef Summary")
        _hdr(ws, 1, ["Metric", "Value"])
        xr_meta = [
            ("Sources",        xs.get("source_count", "")),
            ("Total Keys",     xs.get("total_keys", "")),
            ("Matched in All", xs.get("matched_in_all", "")),
            ("Conflicts",      xs.get("conflict_count", "")),
            ("Coverage Gaps",  xs.get("coverage_gap_count", "")),
            ("Golden Source",  xs.get("golden_source", "auto")),
            ("Elapsed (s)",    xref.get("elapsed", "")),
        ]
        for ri, (k, v) in enumerate(xr_meta, 2):


            ws.cell(ri, 1, k).font = Font(bold=True)
            ws.cell(ri, 2, str(v))
            if ri % 2 == 0:
                for ci in range(1, 3):
                    ws.cell(ri, ci).fill = _ALT_FILL
        _autofit(ws)

        # XRef Coverage Matrix
        cov = xref.get("coverage_matrix")
        if cov and cov.get("rows"):
            ws = wb.create_sheet("XRef Coverage")
            src_names = cov.get("source_names", [])
            _hdr(ws, 1, ["Key"] + src_names)
            for ri, row in enumerate(cov["rows"], 2):
                ws.cell(ri, 1, row.get("key", ""))
                for ci, sn in enumerate(src_names, 2):
                    present = sn in row.get("present_in", [])
                    cell = ws.cell(ri, ci, "✓" if present else "--")
                    cell.alignment = Alignment(horizontal="center")


                    if present:
                        cell.font = Font(color="059669", bold=True)
                if ri % 2 == 0:
                    ws.cell(ri, 1).fill = _ALT_FILL
            _autofit(ws)

        # XRef Conflicts
        conflicts = xref.get("conflicts")
        if conflicts:
            ws = wb.create_sheet("XRef Conflicts")
            src_names = [s["name"] for s in xref.get("sources", [])]
            hdrs = ["Key", "Field", "Conflict Type"] + src_names + ["Golden Value"]
            _hdr(ws, 1, hdrs)
            for ri, c in enumerate(conflicts, 2):
                ws.cell(ri, 1, c.get("key", ""))
                ws.cell(ri, 2, c.get("field", ""))
                ws.cell(ri, 3, c.get("conflict_type", ""))
                for ci, sn in enumerate(src_names, 4):
                    ws.cell(ri, ci, str(c.get("source_values", {}).get(sn, "")))


                ws.cell(ri, 4 + len(src_names), str(c.get("golden_value", "")))
                if c.get("conflict_type") == "VALUE_CONFLICT":
                    for ci in range(1, len(hdrs) + 1):
                        ws.cell(ri, ci).fill = _FAIL_FILL
                elif ri % 2 == 0:
                    for ci in range(1, len(hdrs) + 1):
                        ws.cell(ri, ci).fill = _ALT_FILL
            _autofit(ws)

        # XRef Only-in-Source sheets
        for src in xref.get("sources", []):
            only = src.get("only_in_source")
            if only:
                ws = wb.create_sheet(_safe_sn(src["name"], "Only", _used_sn))
                ws.cell(1, 1, f"Rows only in '{src['name']}' -- not found in other sources "
                               f"({len(only)} rows)").font = Font(bold=True, color="991B1B")
                if only:
                    cols = list(only[0].keys())
                    _hdr(ws, 2, cols)


                    for ri, row in enumerate(only, 3):
                        for ci, col in enumerate(cols, 1):
                            ws.cell(ri, ci, str(row.get(col, "")))
                        if ri % 2 == 0:
                            for ci in range(1, len(cols) + 1):
                                ws.cell(ri, ci).fill = _ALT_FILL
                    _autofit(ws)

    # --- Logs sheet
    ws = wb.create_sheet("Logs")
    _hdr(ws, 1, ["Elapsed (s)", "Level", "Message"])
    for ri, lg in enumerate(data.get("proc_logs", []), 2):
        ws.cell(ri, 1, lg["elapsed"])
        ws.cell(ri, 2, lg["level"])
        ws.cell(ri, 3, lg["message"])
        if lg["level"] == "WARN":
            ws.cell(ri, 2).fill = _WARN_FILL
    _autofit(ws)


    return wb


def _build_email_html(data: dict) -> str:
    # Generate a clean HTML email body from stored results.
    _ACT_LABELS = {"compare": "Reconciliation", "lineage": "Complex Recon", "quality": "Data Quality",
                   "profile": "Data Profile", "parse": "Parse", "governance": "Governance"}
    action      = data.get("action", "")
    action_label = _ACT_LABELS.get(action, action.title())
    files = ", ".join(data.get("file_names", []))
    lines = [
        "<html><body style='font-family:Arial,sans-serif;max-width:800px;margin:auto'>",
        f"<h2 style='color:#1a1a2e'>AI Agent -- Data Validation -- {action_label} Report</h2>",
        f"<p><strong>Files:</strong> {files}</p>",
        f"<p><strong>Elapsed:</strong> {data.get('elapsed',0)}s</p>",
    ]


    for q in data.get("quality_reports", []):
        dq = q["dq_score"]
        lines.append(f"<h3>DQ: {q['file_name']}</h3>")
        lines.append(f"<p>Score: <strong>{dq['score']}/100</strong> (Grade {dq['grade']}) &nbsp;"
                      f"Completeness: {dq['completeness']}% | Uniqueness: {dq['uniqueness']}% | "
                      f"Validity: {dq['validity']}%</p>")
        fails = [r for r in q.get("rule_results", []) if r.get("status") == "FAIL"]
        if fails:
            lines.append("<table border='1' cellpadding='4' style='border-collapse:collapse;font-size:12px'>")
            lines.append("<tr style='background:#1a1a2e;color:#fff'><th>Rule</th><th>Column</th><th>Failed</th></tr>")
            for r in fails:
                lines.append(f"<tr style='background:#fee2e2'><td>{r['rule_name']}</td>"
                              f"<td>{r['column_name']}</td><td>{r['failed']}</td></tr>")
            lines.append("</table>")

    for g in data.get("governance_reports", []):
        lines.append(f"<h3>Governance: {g['file_name']}</h3>")


        lines.append(f"<p>Classification: <strong>{g['overall_classification']}</strong> | "
                      f"PII columns: {g['pii_column_count']} | Regulatory: {', '.join(g['regulatory_frameworks'])}</p>")

    for m in data.get("mappings", []):
        lines.append(f"<h3>Mapping: {m['file1_name']} → {m['file2_name']}</h3>")
        rw = m.get("relatedness_warning")
        if rw:
            colour = "#7f1d1d" if rw["level"] == "HIGH" else "#78350f"
            bg     = "#fef2f2" if rw["level"] == "HIGH" else "#fffbeb"
            border = "#fca5a5" if rw["level"] == "HIGH" else "#fcd34d"
            icon   = "&#9888;" if rw["level"] == "HIGH" else "&#9888;"
            reasons_html = "".join(f"<li>{r}</li>" for r in rw["reasons"])
            lines.append(
                f"<div style='border:2px solid {border};background:{bg};color:{colour};"
                f"border-radius:6px;padding:12px 16px;margin:8px 0'>"
                f"<strong>{icon} RELATEDNESS WARNING [{rw['level']}]:</strong> {rw['message']}"
                f"<ul style='margin:6px 0 0 16px'>{reasons_html}</ul></div>"
            )


        lines.append(f"<p>Completeness: "
                      f"<strong>{m['mapping_completeness_pct']}%</strong> | "
                      f"Exact: {len(m['exact'])} | Fuzzy: {len(m['fuzzy'])} | "
                      f"Unmapped F1: {len(m['unmapped_f1'])} | Unmapped F2: {len(m['unmapped_f2'])}</p>")

        lines.append("<hr/><p style='color:#9ca3af;font-size:11px'>Sent by Data Validation AGENT</p></body></html>")

    return "\n".join(lines)


# ----------------------------------------------------------------------
# Routes
# ----------------------------------------------------------------------


def _infer_column_config(df: pd.DataFrame) -> list[dict]:
    # Auto-infer per-column DQ configuration from the data itself.

    # For each column returns a config dict with:
    # name, dtype, mandatory (bool), exclude (bool),
    # null_threshold_pct, min_val, max_val, decimal_places,
    # timeliness_days, rule_type (suggested), infer_reason (explanation).

    # Mandatory inference rules (applied in priority order):
    # 1. 0% nulls in the data AND name matches ID/key/amount patterns -> mandatory
    # 2. 0% nulls AND high cardinality (identifier-like) -> mandatory
    # 3. >50% nulls -> optional (exclude from completeness)
    # 4. Name matches free-text/notes patterns -> optional
    # 5. Everything else -> mandatory by default (safe assumption)

    _MANDATORY_NAME_HINTS = {
        "id", "key", "pk", "ref", "num", "no", "number", "code",
        "isin", "cusip", "sedol", "lei", "bic", "iban",
        "amount", "amt", "price", "notional", "qty", "quantity",
        "value", "rate", "balance", "bal",
        "date", "dt", "time", "ts", "timestamp",
        "type", "side", "status", "currency", "ccy", "symbol", "ticker",
        "account", "acct", "counterparty", "broker", "trader",
        # Product / instrument classification
        "product", "instrument", "asset", "class", "category", "subtype",
        "sub", "leg", "strategy", "book", "portfolio", "fund", "entity",
        "market", "venue", "exchange", "desk", "region", "sector",
        # Trade / transaction fields
        "trade", "transaction", "order", "deal", "contract", "position",
        "direction", "action", "event", "lifecycle", "state",
        # Party / counterparty identifiers
        "party", "client", "customer", "firm", "issuer", "obligor",
        "name", "cpty",
    }
    _OPTIONAL_NAME_HINTS = {
        "remark", "remarks", "note", "notes", "comment", "comments",
        "description", "desc", "narrative", "memo", "text", "info",
        "optional", "extra", "misc", "other", "temp", "tmp",
        "internal", "flag",
    }
    _DATE_NAME_HINTS = {"date", "dt", "time", "ts", "timestamp", "created", "updated", "modified"}
    _AMOUNT_NAME_HINTS = {"amount", "amt", "price", "notional", "qty", "quantity", "value",
        "rate", "balance", "bal", "vol", "volume"}

    total = len(df)
    configs = []

    for col in df.columns:
        s = df[col]
        col_lower = col.lower().replace(" ", "_").replace("-", "_")
        null_pct = s.isna().mean() * 100

        # Split col name into tokens for hint matching
        tokens = set(re.split(r'[_\s\-\.]+', col_lower))

        name_is_mandatory = bool(tokens & _MANDATORY_NAME_HINTS)
        name_is_optional  = bool(tokens & _OPTIONAL_NAME_HINTS)
        name_is_date      = bool(tokens & _DATE_NAME_HINTS)
        name_is_amount    = bool(tokens & _AMOUNT_NAME_HINTS)

        uniq_ratio = s.nunique(dropna=True) / total if total else 0
        is_identifier_like = uniq_ratio > 0.95

        # --- Mandatory determination
        # Priority 1: name strongly signals optional -> always optional
        if name_is_optional and not name_is_mandatory:
            mandatory = False
            exclude   = null_pct > 80
            infer_reason = "Name suggests optional/free-text field"
        # Priority 2: name strongly signals mandatory -> mandatory regardless of null%
        #  (null% only affects the reason text, not the decision)
        elif name_is_mandatory or is_identifier_like:
            mandatory = True
            exclude   = False
            if null_pct == 0:
                infer_reason = "0% nulls + name matches mandatory pattern"
            elif null_pct <= 5:
                infer_reason = f"Name matches mandatory pattern ({null_pct:.1f}% nulls)"
            else:
                infer_reason = (
                    f"Name matches mandatory pattern -- review ({null_pct:.0f}% nulls"
                    )
        # Priority 3: high nulls with no name signal -> optional
        elif null_pct > 50:
            mandatory = False
            exclude   = null_pct > 80
            infer_reason = f"{null_pct:.0f}% nulls -- likely an optional column"
        # Priority 4: zero nulls, no name signal
        elif null_pct == 0:
            mandatory = True
            exclude   = False
            infer_reason = "0% nulls in current data -- treated as mandatory"
        # Priority 5: very low nulls
        elif null_pct <= 5:
            mandatory = True
            exclude   = False
            infer_reason = f"Only {null_pct:.1f}% nulls -- treated as mandatory"
        # Priority 6: moderate nulls, ambiguous
        else:
            mandatory = True
            exclude   = False
            infer_reason = f"{null_pct:.1f}% nulls -- defaulting to mandatory (review)"

        # --- Suggest unique rule for identifier-like columns
        # Only suggest when cardinality is near-perfect AND name hints at an ID.
        # We never *enforce* uniqueness without user confirmation -- the dropdown
        # shows the suggestion pre-selected so the user can change it.
        _ID_NAME_HINTS = {"id", "key", "pk", "ref", "num", "number", "code",
            "isin", "cusip", "sedol", "lei", "bic", "iban",
            "trade", "transaction", "order", "deal", "contract"}
        name_is_id = bool(tokens & _ID_NAME_HINTS)
        infer_rule = "auto"
        if name_is_id and is_identifier_like and null_pct == 0:
            infer_rule = "unique"
        elif is_identifier_like and null_pct == 0 and not name_is_amount and not name_is_date:
            infer_rule = "unique"

        cfg: dict = {
            "name":               col,
            "dtype":              str(s.dtype),
            "mandatory":          mandatory,
            "exclude":            exclude,
            "null_threshold_pct": 0 if mandatory else round(null_pct + 10, 0),
            "infer_reason":       infer_reason,
            "min_val":            "",
            "max_val":            "",
            "decimal_places":     "",
            "timeliness_days":    "",
            "rule_type":          infer_rule,
        }

        # --- Numeric defaults
        if pd.api.types.is_numeric_dtype(s):
            clean = s.dropna()
            if len(clean):
                cfg["min_val"] = round(float(clean.min()), 4)
                cfg["max_val"] = round(float(clean.max()), 4)

                # Infer decimal places from actual values
                def _dp(v):
                    try:
                        sv = str(float(v))
                        return len(sv.rstrip('0').split('.')[1]) if '.' in sv else 0
                    except Exception:
                        return 0
                max_dp = int(clean.apply(_dp).max())
                cfg["decimal_places"] = max_dp if max_dp > 0 else ""
            if name_is_amount:
                cfg["rule_type"] = "range"

        # --- Date defaults
        elif pd.api.types.is_datetime64_any_dtype(s) or name_is_date:
            cfg["rule_type"] = "freshness_days"
            cfg["timeliness_days"] = 1 if "trade" in col_lower else 7

        # --- String BFSI format auto-detect -- vectorised, capped at 50 rows --
        # Skip if values are numeric/financial quantities (>=80% parse as numbers).
        elif s.dtype == object:
            str_s = s.dropna().astype(str).str.strip().head(50)
            _num_rate = pd.to_numeric(str_s, errors="coerce").notna().mean()
            if _num_rate < 0.8:
                str_s_upper = str_s.str.upper()
                for fmt_name, pat in _DQ_FORMAT_PATTERNS.items():
                    if len(str_s_upper) and str_s_upper.str.fullmatch(pat.pattern).mean() >= 0.8:
                        cfg["rule_type"] = f"{fmt_name}_format"
                        break

        configs.append(cfg)

    return configs


@app.post("/api/dq/baseline/{session_id}/{file_name}")
async def save_dq_baseline(session_id: str, file_name: str):
    """Save the current DQ run's baseline_snapshot to the workspace for future drift comparison."""


    import json as _json_b
    try:
        cache_dir = Path("workspace") / "dq_baselines"
        cache_dir.mkdir(parents=True, exist_ok=True)
        _sid_report = _session_quality_cache.get(session_id, {}).get(file_name)
        if not _sid_report:
            return JSONResponse({"ok": False, "error": "No DQ report found for this session/file. Re-run Data Quality first."})
        baseline = _sid_report.get("baseline_snapshot")
        if not baseline:
            return JSONResponse({"ok": False, "error": "No baseline snapshot in report."})
        key = re.sub(r'[^\w\-.]', '_', file_name)
        (cache_dir / f"{key}.json").write_text(_json_b.dumps(baseline), encoding="utf-8")
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


@app.get("/api/dq/rules/{fingerprint}")


async def get_dq_rules(fingerprint: str):
    """Load saved DQ column config for a schema fingerprint."""
    path = Path("workspace") / "dq_rules" / f"{fingerprint}.json"
    if not path.exists():
        return JSONResponse({})
    try:
        return JSONResponse(json.loads(path.read_text(encoding="utf-8")))
    except Exception:
        return JSONResponse({})


@app.post("/api/dq/rules/{fingerprint}")
async def save_dq_rules(fingerprint: str, request: Request):
    """Save DQ column config for a schema fingerprint."""
    try:
        body = await request.json()
        path = Path("workspace") / "dq_rules"
        path.mkdir(parents=True, exist_ok=True)
        (path / f"{fingerprint}.json").write_text(json.dumps(body), encoding="utf-8")


        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/dq/history/{file_name}")
async def get_dq_history_endpoint(file_name: str, request: Request):
    """Get DQ score history for a file (for trend chart)."""
    # Must use _ws_resolve_username -- route is outside /api/ws/* so the
    # auth middleware never sets request.state.username here.
    try:
        username = _ws_resolve_username(request) or "default"
    except Exception:
        username = "default"
    try:
        rows = _ws_db.get_dq_history(file_name, username, days=30)
        return JSONResponse(rows)
    except Exception as e:
        return JSONResponse([], status_code=200)


@app.get("/api/dq/baseline/{file_name}")
async def get_dq_baseline_endpoint(file_name: str, request: Request):
    """Get the earliest recorded DQ score for a file (for the vs-baseline banner)."""
    try:
        username = _ws_resolve_username(request) or "default"
    except Exception:
        username = "default"
    try:
        baseline = _ws_db.get_dq_baseline(file_name, username)
        return JSONResponse(baseline)
    except Exception:
        return JSONResponse(None)


@app.get("/api/recon/history/{fingerprint}")
async def get_recon_history_endpoint(fingerprint: str, request: Request):
    """Break-rate trend for this schema's reconciliations (for the trend chart)."""
    try:
        username = _ws_resolve_username(request) or "default"
    except Exception:
        username = "default"
    try:
        rows = _ws_db.get_recon_history(fingerprint, username, days=30)
        return JSONResponse(rows)
    except Exception:
        return JSONResponse([], status_code=200)


@app.get("/api/recon/baseline/{fingerprint}")
async def get_recon_baseline_endpoint(fingerprint: str, request: Request):
    """Earliest recorded reconciliation run for this schema (for the vs-baseline banner)."""
    try:
        username = _ws_resolve_username(request) or "default"
    except Exception:
        username = "default"
    try:
        baseline = _ws_db.get_recon_baseline(fingerprint, username)
        return JSONResponse(baseline)
    except Exception:
        return JSONResponse(None)


@app.get("/api/recon/rule-templates")
async def list_rule_templates(request: Request, exclude_fingerprint: str = ""):
    """Every schema with saved reconciliation rules -- usable as a template
    library: pick one and apply its rules onto a different (new) schema,
    instead of rules only ever auto-loading for the exact schema they were
    written for. exclude_fingerprint hides the current session's own schema
    from the list, since copying a schema's rules onto itself is a no-op."""
    username = _ws_resolve_username(request) or "default"
    datasets = _fp_list_datasets(username)
    if exclude_fingerprint:
        datasets = [d for d in datasets if d["fingerprint"] != exclude_fingerprint]
    datasets.sort(key=lambda d: d.get("updated", ""), reverse=True)
    return JSONResponse(datasets)


@app.post("/api/recon/rule-templates/apply")
async def apply_rule_template(request: Request):
    """Copy every saved rule from one schema onto another. Duplicate rules
    (already present on the target) are skipped automatically."""
    username = _ws_resolve_username(request) or "default"
    body = await request.json()
    from_fp = str(body.get("from_fingerprint", "")).strip()
    to_fp = str(body.get("to_fingerprint", "")).strip()
    if not from_fp or not to_fp:
        raise HTTPException(400, "from_fingerprint and to_fingerprint are both required.")
    if from_fp == to_fp:
        raise HTTPException(400, "Source and target schema are the same -- nothing to apply.")
    copied = _fp_copy_rules(username, from_fp, to_fp)
    return JSONResponse({"copied": copied})


@app.post("/api/dq/mask/{session_id}")
async def mask_pii_data(session_id: str, request: Request):

    # Download a masked copy of the uploaded data with PII columns replaced.

    # Masking strategies per column type:
    # email       -> a***@***.com
    # phone       -> ***-***-XXXX (last 4 kept for debugging)
    # ssn/tax_id  -> ***-**-XXXX
    # credit_card -> ****-****-****-XXXX
    # name        -> [MASKED]
    # dob/date    -> YYYY-**-** (year kept for analytics)
    # address     -> [ADDRESS MASKED]
    # iban        -> GB**-****-****-XXXX (country+last4)
    # generic PII -> [REDACTED]

    # Returns a masked CSV file for download.


    import io, re as _re

    body = await request.json()
    columns_to_mask: list[str] = body.get("columns", [])  # list of column names to mask
    mask_strategy: dict[str, str] = body.get("strategies", {})  # column -> strategy override

    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        raise HTTPException(404, "Session not found or expired -- please re-upload.")

    dfs_raw = stored["dataframes"]
    if not dfs_raw:
        raise HTTPException(400, "No dataframes in session.")

    # Use first dataframe
    df_orig = dfs_raw[0]["df"].copy()
    fname   = dfs_raw[0]["name"]


    # Auto-detect PII columns from governance results if not explicitly provided
    if not columns_to_mask:
        for qr in stored.get("quality_reports", []):
            gov = qr.get("governance") or {}
            for col_finding in gov.get("columns", []):
                if col_finding.get("pii_detected"):
                    columns_to_mask.append(col_finding["column"])

    def _mask_value(val: str, col_name: str, strategy: str) -> str:
        if not val or val in ("nan", "None", ""):
            return val
        s = strategy or _infer_mask_strategy(col_name)
        if s == "email":
            m = _re.match(r'^([^@]+)(@[^@]+)$', val)
            return (m.group(1)[0] + "***" + m.group(2)) if m else "***@***.com"
        if s == "phone":
            digits = _re.sub(r'\D', '', val)
            return f"***-***-{digits[-4:]}" if len(digits) >= 4 else "***-***-****"
        if s == "ssn":


            parts = _re.sub(r'\D', '', val)
            return f"***-**-{parts[-4:]}" if len(parts) >= 4 else "***-**-****"
        if s == "credit_card":
            digits = _re.sub(r'\D', '', val)
            return f"****-****-****-{digits[-4:]}" if len(digits) >= 4 else "****-****-****-****"
        if s == "name":
            return "[MASKED]"
        if s == "dob":
            m = _re.match(r'^(\d{4})([-/]\d{2}[-/]\d{2})$', val)
            return (m.group(1) + "-**-**") if m else "[DOB MASKED]"
        if s == "address":
            return "[ADDRESS MASKED]"
        if s == "iban":
            clean = _re.sub(r'[\s-]', '', val.upper())
            return f"{clean[:4]}-****-****-{clean[-4:]}" if len(clean) >= 8 else "[IBAN MASKED]"
        return "[REDACTED]"

    def _infer_mask_strategy(col: str) -> str:


        cl = col.lower()
        if any(k in cl for k in ("email", "mail")):                    return "email"
        if any(k in cl for k in ("phone", "tel", "mobile")):            return "phone"
        if any(k in cl for k in ("ssn", "social", "tax_id", "nin")):    return "ssn"
        if any(k in cl for k in ("card", "cc_num", "credit")):          return "credit_card"
        if any(k in cl for k in ("name", "firstname", "lastname", "fullname")): return "name"
        if any(k in cl for k in ("dob", "birth", "birthdate")):         return "dob"
        if any(k in cl for k in ("address", "addr", "street", "zip", "postal")): return "address"
        if any(k in cl for k in ("iban", "account_no")):                return "iban"
        return "generic"

    masked_df = df_orig.copy()
    for col in columns_to_mask:
        if col not in masked_df.columns:
            continue
        strategy = mask_strategy.get(col, "")
        masked_df[col] = masked_df[col].astype(str).apply(
            lambda v, c=col, s=strategy: _mask_value(v, c, s)
        )


    # Add masking audit column
    masked_df["_pii_masked"] = f"Masked columns: {', '.join(columns_to_mask)} | Generated by Data Validation Agent"

    # Return as CSV
    buf = io.StringIO()
    masked_df.to_csv(buf, index=False)
    buf.seek(0)

    safe_name = fname.rsplit(".", 1)[0] if "." in fname else fname
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_masked.csv"'},
    )


@app.post("/dq-infer")


async def dq_infer(files: list[UploadFile] = File(...)):
    """Accept one or more files, return auto-inferred per-column DQ config as JSON."""
    try:
        dfs = []
        for upload in files:
            df = _load_file(upload)
            dfs.append(df)
        combined = pd.concat(dfs, ignore_index=True) if len(dfs) > 1 else dfs[0]
        configs = _infer_column_config(combined)
        return JSONResponse({"columns": configs, "total_rows": len(combined)})
    except Exception as exc:
        raise HTTPException(400, str(exc))


@app.post("/dq-suggest")
async def dq_suggest(files: list[UploadFile] = File(...)):
    # Send column profiles to the LLM and get back AI-suggested DQ rules.
    # Returns a list of suggestion dicts, one per column that the LLM thinks
    # needs a rule beyond what statistics already inferred:
    # {name, rule_type, value, reason, mandatory, allowed_values, pattern}
    try:
        dfs = []
        for upload in files:
            df = _load_file(upload)
            dfs.append(df)
        combined = pd.concat(dfs, ignore_index=True) if len(dfs) > 1 else dfs[0]

        # Build a compact column profile summary for the LLM -- keep it small
        total_rows = len(combined)
        col_summaries = []
        for col in combined.columns:
            s = combined[col]
            null_pct = round(s.isna().mean() * 100, 1)
            uniq_n   = int(s.nunique(dropna=True))
            dtype_str = str(s.dtype)
            sample   = s.dropna().astype(str).head(8).tolist()

            summary = {
                "name":     col,
                "dtype":    dtype_str,
                "null_pct": null_pct,
                "unique_n": uniq_n,
                "total":    total_rows,
                "sample":   sample,
            }
            if pd.api.types.is_numeric_dtype(s):
                clean = s.dropna()
                if len(clean):
                    summary["min"]  = round(float(clean.min()), 4)
                    summary["max"]  = round(float(clean.max()), 4)
                    summary["mean"] = round(float(clean.mean()), 4)
                    summary["std"]  = round(float(clean.std()), 4)
            col_summaries.append(summary)

        profile_json = json.dumps(col_summaries, indent=2)

        system_prompt = (
            "You are a senior data quality engineer specialising in financial services (BFSI) data. "
            "You are given a column-level profile of a dataset. Your job is to suggest data quality "
            "rules for each column based on its name, data type, sample values, and statistics. "
            "Return ONLY a valid JSON array. Each element is a suggestion for one column. "
            "Only suggest rules that add value beyond basic not-null or uniqueness checks. "
            "Focus on: business meaning from the column name, expected value ranges, "
            "enum constraints, format patterns, cross-column relationships. "
            "Use these rule_type values only: "
            "range, allowed_values, pattern, positive, non_negative, integer_only, "
            "not_future_date, not_past_date, isin_format, cusip_format, sedol_format, "
            "lei_format, bic_format, iban_format, currency_code_format, mic_format, "
            "email_format, date_format, freshness_days, decimal_places, uppercase, lowercase. "
            "Each suggestion must have: "

            "name (column name), rule_type, value (rule parameter if needed, else empty string), "
            "reason (1 sentence explaining why), mandatory (true/false), "
            "allowed_values (comma-separated if rule_type=allowed_values, else empty), "
            "pattern (regex if rule_type=pattern, else empty). "
            "Return only columns where you have a confident, specific suggestion. "
            "Do not suggest rules for columns you are unsure about. "
            "Return an empty array [] if no confident suggestions exist."
        )

        user_prompt = (
            f"Dataset has {total_rows} rows and {len(combined.columns)} columns.\n\n"
            f"Column profiles:\n{profile_json}\n\n"
            "Suggest data quality rules. Return a JSON array only, no explanation outside the JSON."
        )

        raw = await asyncio.to_thread(
            _ask_llm,
            [{"role": "user", "content": [{"text": user_prompt}]}],
            system=system_prompt,
        )

        # Extract JSON array from response (LLM may wrap in markdown code blocks)
        json_match = re.search(r'\[.*\]', raw, re.DOTALL)
        if not json_match:
            return JSONResponse({"suggestions": [], "raw": raw[:500]})

        suggestions = json.loads(json_match.group())
        # Validate structure -- keep only well-formed suggestions
        valid = []
        for s in suggestions:
            if isinstance(s, dict) and s.get("name") and s.get("rule_type"):
                valid.append({
                    "name":       str(s.get("name", "")),
                    "rule_type":  str(s.get("rule_type", "auto")),
                    "value":      str(s.get("value", "") or ""),
                    "reason":     str(s.get("reason", "")),
                    "mandatory":  bool(s.get("mandatory", True)),

                    "allowed_values": str(s.get("allowed_values", "") or ""),
                    "pattern":    str(s.get("pattern", "") or ""),
                })
        return JSONResponse({"suggestions": valid})

    except Exception as exc:
        # Non-fatal -- UI treats suggestions as optional enhancement
        return JSONResponse({"suggestions": [], "error": str(exc)})


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    # If local auth or LDAP-only login is enabled, redirect to login unless a
    # valid session exists.
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED, verify_session
        if LOCAL_AUTH_ENABLED:
            token = request.cookies.get("dv_local_session", "")
            if not token or not verify_session(token):
                from fastapi.responses import RedirectResponse


                return RedirectResponse(url="/login", status_code=302)
        else:
            from workspace.sso import LDAP_ENABLED as _LDAP_ENABLED, verify_sso_token as _verify_sso
            if _LDAP_ENABLED:
                token = request.cookies.get("dv_session", "")
                if not token or not _verify_sso(token):
                    from fastapi.responses import RedirectResponse
                    return RedirectResponse(url="/login", status_code=302)
    except Exception:
        pass
    resp = templates.TemplateResponse(request=request, name="index.html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, error: str = "", registered: str = "", reset: str = ""):
    """Serve the login page. Redirect to app if already logged in."""
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED, verify_session, has_any_users
        from workspace.sso import LDAP_ENABLED as _LDAP_ENABLED, verify_sso_token as _verify_sso
        if not (LOCAL_AUTH_ENABLED or _LDAP_ENABLED):
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url="/", status_code=302)
        show_register = False
        if LOCAL_AUTH_ENABLED:
            token = request.cookies.get("dv_local_session", "")


            if token and verify_session(token):
                from fastapi.responses import RedirectResponse
                return RedirectResponse(url="/", status_code=302)
            show_register = not has_any_users()
        else:
            token = request.cookies.get("dv_session", "")
            if token and _verify_sso(token):
                from fastapi.responses import RedirectResponse
                return RedirectResponse(url="/", status_code=302)
    except Exception:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/", status_code=302)

    return HTMLResponse(_render_login_page(
        error=error,
        show_register=show_register,
        registered=registered,
        reset=reset,
    ))


@app.post("/login")
async def login_submit(request: Request):
    from fastapi.responses import RedirectResponse
    try:


        from workspace.local_auth import LOCAL_AUTH_ENABLED, login as _local_login, AuthError, has_any_users

        form = await request.form()
        username = str(form.get("username", "")).strip()
        password = str(form.get("password", ""))

        if LOCAL_AUTH_ENABLED:
            token = _local_login(username, password)
            response = RedirectResponse(url="/", status_code=302)
            response.set_cookie(
                key="dv_local_session",
                value=token,
                httponly=True,
                samesite="lax",
                max_age=8 * 3600,
            )
            return response

        from workspace.sso import (
            LDAP_ENABLED as _LDAP_ENABLED, authenticate as _ldap_authenticate,
            create_sso_session_token as _sso_create_token, SSO_ROLE_MAP as _SSO_ROLE_MAP,
        )
        if not _LDAP_ENABLED:
            return RedirectResponse(url="/", status_code=302)

        result = _ldap_authenticate(username, password)
        if not result:
            raise ValueError("Invalid username or password.")
        from workspace.db import ensure_user as _ensure_user
        _ensure_user(result["username"], result["username"], "")
        if _SSO_ROLE_MAP:
            _ws_db.set_user_role(result["username"], result["role"])
        token = _sso_create_token(result["username"], result["role"], [])
        response = RedirectResponse(url="/", status_code=302)
        response.set_cookie(
            key="dv_session",
            value=token,
            httponly=True,
            samesite="lax",
            secure=request.url.scheme == "https",
            max_age=8 * 3600,
        )
        return response


    except Exception as exc:
        error_msg = str(exc) if "AuthError" in type(exc).__name__ or "Invalid" in str(exc) else "Login failed."
        return HTMLResponse(_render_login_page(error=error_msg, show_register=False))


@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED, has_any_users
        if not LOCAL_AUTH_ENABLED:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url="/", status_code=302)
    except Exception:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/", status_code=302)
    return HTMLResponse(_render_login_page(error="", show_register=True))


@app.post("/register")
async def register_submit(request: Request):
    from fastapi.responses import RedirectResponse
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED, register as _local_register, AuthError
        if not LOCAL_AUTH_ENABLED:
            return RedirectResponse(url="/", status_code=302)

        form = await request.form()
        username  = str(form.get("username", "")).strip()
        password  = str(form.get("password", ""))
        password2 = str(form.get("password2", ""))
        full_name = str(form.get("full_name", "")).strip()
        email     = str(form.get("email", "")).strip()

        if password != password2:
            return HTMLResponse(_render_login_page(
                error="Passwords do not match.", show_register=True,


                prefill={"username": username, "full_name": full_name, "email": email},
            ))

        _local_register(username, password, full_name=full_name, email=email)
        return RedirectResponse(url="/login?registered=1", status_code=302)

    except Exception as exc:
        error_msg = str(exc) if hasattr(exc, "args") else "Registration failed."
        form = await request.form() if not isinstance(request, str) else {}
        return HTMLResponse(_render_login_page(
            error=error_msg, show_register=True,
        ))


@app.get("/forgot-password", response_class=HTMLResponse)
async def forgot_password_page():
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED
        if not LOCAL_AUTH_ENABLED:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url="/login", status_code=302)
    except Exception:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/login", status_code=302)
    return HTMLResponse(_render_forgot_password_page())


@app.post("/forgot-password", response_class=HTMLResponse)
async def forgot_password_submit(request: Request):
    """Look up the account by email, and if found (and an email is on file),
    send the username plus a 30-minute password-reset link. Always shows the
    same generic confirmation regardless of whether a match was found, so
    this can't be used to enumerate registered emails/usernames."""
    generic_message = (
        "If that email matches an account, we've sent the username and a "
        "password reset link to it. Check your inbox (and spam folder)."
    )
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED, request_password_reset
        if not LOCAL_AUTH_ENABLED:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url="/login", status_code=302)
        form = await request.form()
        identifier = str(form.get("identifier", "")).strip()
        if identifier:
            result = request_password_reset(identifier)
            if result:
                username, email, token = result
                reset_link = f"{request.url.scheme}://{request.url.netloc}/reset-password?token={token}"
                html = (
                    f"<p>Hello,</p>"
                    f"<p>Your username is: <b>{username}</b></p>"
                    f"<p><a href=\"{reset_link}\">Click here to reset your password</a> "
                    f"(expires in 30 minutes).</p>"
                    f"<p>If you didn't request this, you can safely ignore this email.</p>"
                )
                try:
                    from workspace.scheduler import _deliver_email
                    _deliver_email(email, os.getenv("EMAIL_FROM", ""),
                                    "AI Agent -- Account Recovery", html)
                except Exception as exc:
                    import logging as _logging
                    _logging.getLogger(__name__).warning("Password reset email failed: %s", exc)
    except Exception:
        pass
    return HTMLResponse(_render_forgot_password_page(message=generic_message))


@app.get("/reset-password", response_class=HTMLResponse)
async def reset_password_page(token: str = ""):
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED, verify_reset_token
        if not LOCAL_AUTH_ENABLED:
            from fastapi.responses import RedirectResponse
            return RedirectResponse(url="/login", status_code=302)
        if not token or not verify_reset_token(token):
            return HTMLResponse(_render_forgot_password_page(
                error="This reset link is invalid or has expired. Please request a new one."))
    except Exception:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/login", status_code=302)
    return HTMLResponse(_render_reset_password_page(token))


@app.post("/reset-password", response_class=HTMLResponse)
async def reset_password_submit(request: Request, token: str = ""):
    from fastapi.responses import RedirectResponse
    try:
        from workspace.local_auth import LOCAL_AUTH_ENABLED, reset_password, AuthError
        if not LOCAL_AUTH_ENABLED:
            return RedirectResponse(url="/login", status_code=302)
        form = await request.form()
        password = str(form.get("password", ""))
        password2 = str(form.get("password2", ""))
        if password != password2:
            return HTMLResponse(_render_reset_password_page(token, error="Passwords do not match."))
        reset_password(token, password)
        return RedirectResponse(url="/login?reset=1", status_code=302)
    except Exception as exc:
        error_msg = str(exc) if hasattr(exc, "args") else "Password reset failed."
        return HTMLResponse(_render_reset_password_page(token, error=error_msg))


@app.get("/logout")
async def logout(request: Request):
    from fastapi.responses import RedirectResponse
    try:
        from workspace.local_auth import logout as _local_logout


        token = request.cookies.get("dv_local_session", "")
        if token:
            _local_logout(token)
    except Exception:
        pass
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("dv_local_session")
    # Also clear the SSO session cookie so this one link works regardless of
    # whether the user signed in via local auth or SAML/OIDC.
    response.delete_cookie("dv_session")
    return response


@app.get("/api/auth/change-password", response_class=HTMLResponse)
async def change_password_page(request: Request):
    return HTMLResponse(_render_change_password_page())


@app.post("/api/auth/change-password")
async def change_password_submit(request: Request):
    try:


        from workspace.local_auth import LOCAL_AUTH_ENABLED, verify_session, change_password, AuthError
        if not LOCAL_AUTH_ENABLED:
            return JSONResponse({"ok": False, "error": "Local auth not enabled."})
        token = request.cookies.get("dv_local_session", "")
        username = verify_session(token) if token else None
        if not username:
            return JSONResponse({"ok": False, "error": "Not authenticated."}, status_code=401)
        body = await request.json()
        change_password(username, body.get("old_password", ""), body.get("new_password", ""))
        return JSONResponse({"ok": True, "message": "Password changed. Please log in again."})
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)})


def _render_login_page(
    error: str = "",
    show_register: bool = False,
    registered: str = "",
    reset: str = "",
    prefill: dict = None,
) -> str:
    # ==== GAP: the source pages covering the rest of the registration form,
    # the entire login-form branch, and the title/subtitle assignments were
    # not recoverable from the scan (only the Full Name + Username fields of
    # the register form and the shared page wrapper survived). RECONSTRUCTED
    # (unverified) below using the exact field names read by login_submit()
    # (username, password) and register_submit() (username, password,
    # password2, full_name, email) so the forms POST correctly -- verify
    # wording/layout against the original if it becomes available.
    pf = prefill or {}
    err_html = f'<div class="auth-error">{error}</div>' if error else ""
    if reset:
        ok_html = '<div class="auth-ok">&#9989; Password reset -- please log in with your new password.</div>'
    elif registered:
        ok_html = '<div class="auth-ok">&#9989; Account created -- please log in.</div>'
    else:
        ok_html = ""

    if show_register:
        form_html = f"""<form method="post" action="/register" autocomplete="off">
{err_html}
<div class="auth-field">
<label>Full Name</label>
<input type="text" name="full_name" value="{pf.get('full_name','')}" placeholder="Your full name" autocomplete="name"/>
</div>
<div class="auth-field">
<label>Username</label>
<input type="text" name="username" value="{pf.get('username','')}" placeholder="Choose a username" autocomplete="username"/>
</div>
<div class="auth-field">
<label>Email</label>
<input type="email" name="email" value="{pf.get('email','')}" placeholder="you@example.com" autocomplete="email"/>
</div>
<div class="auth-field">
<label>Password</label>
<input type="password" name="password" placeholder="Min 8 characters" autocomplete="new-password"/>
</div>
<div class="auth-field">
<label>Confirm Password</label>
<input type="password" name="password2" placeholder="Re-enter password" autocomplete="new-password"/>
</div>
<button type="submit" class="auth-btn">Create Account</button>
</form>
<div class="auth-switch">Already have an account? <a href="/login">Sign in</a></div>"""
        title = "Create Account"
        subtitle = "AI Agent -- Data Validation"
    else:
        form_html = f"""<form method="post" action="/login" autocomplete="off">
{err_html}
{ok_html}
<div class="auth-field">
<label>Username</label>
<input type="text" name="username" placeholder="Your username" autocomplete="username"/>
</div>
<div class="auth-field">
<label>Password</label>
<input type="password" name="password" placeholder="********" autocomplete="current-password"/>
</div>
<button type="submit" class="auth-btn">Sign In</button>
</form>
<div class="auth-switch"><a href="/forgot-password">Forgot username or password?</a></div>
<div class="auth-switch">Need an account? <a href="/register">Create one</a></div>"""
        title = "Sign In"
        subtitle = "AI Agent -- Data Validation"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{title} -- AI Agent Data Validation</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Sora:wght@600;700;800&display=swap" rel="stylesheet"/>
<style>{_AUTH_PAGE_STYLE}</style>
</head>
<body>
<div class="auth-wrap">
  <div class="auth-card">
    <div class="auth-logo">
      <div class="auth-logo-mark"><svg viewBox="0 0 24 24" width="22" height="22" fill="none" xmlns="http://www.w3.org/2000/svg">
        <line x1="12" y1="2.6" x2="12" y2="5.2" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
        <circle cx="12" cy="2" r="1.15" fill="#2563eb"/>
        <rect x="4.5" y="5.6" width="15" height="12.4" rx="4" stroke="#2563eb" stroke-width="2"/>
        <circle cx="9.3" cy="11.6" r="1.35" fill="#2563eb"/>
        <circle cx="14.7" cy="11.6" r="1.35" fill="#2563eb"/>
        <path d="M9.2 15.1h5.6" stroke="#2563eb" stroke-width="1.6" stroke-linecap="round"/>
        <line x1="4.5" y1="10.5" x2="2.6" y2="10.5" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
        <line x1="19.5" y1="10.5" x2="21.4" y2="10.5" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
      </svg></div>
      <div class="auth-logo-text">
        <h1>AI Agent</h1>
        <p>Data Validation</p>
      </div>
    </div>
    <div class="auth-title">{title}</div>
    <div class="auth-subtitle">{subtitle}</div>
    {form_html}
  </div>
  <div class="auth-footer">&#128274; Your credentials are stored securely on this server.</div>
</div>
</body>
</html>"""


_AUTH_PAGE_STYLE = """
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:"Inter","Segoe UI",sans-serif;color:#0f172a;
 min-height:100vh;display:flex;align-items:center;justify-content:center;
 -webkit-font-smoothing:antialiased;
 background:
   radial-gradient(1100px 550px at 108% -8%, rgba(37,99,235,.10), transparent 60%),
   radial-gradient(900px 480px at -8% 105%, rgba(124,58,237,.08), transparent 55%),
   radial-gradient(700px 400px at 50% -15%, rgba(2,132,199,.06), transparent 60%),
   #eef1f6;
 background-attachment:fixed}
@keyframes gradShift{0%{background-position:0% 50%}50%{background-position:100% 50%}100%{background-position:0% 50%}}
@keyframes glowPulse{0%,100%{box-shadow:0 3px 14px -2px rgba(37,99,235,.55),0 0 0 3px rgba(37,99,235,.1),0 0 22px -4px rgba(124,58,237,.35)}
 50%{box-shadow:0 3px 18px -1px rgba(37,99,235,.7),0 0 0 5px rgba(37,99,235,.14),0 0 30px -2px rgba(124,58,237,.55)}}
.auth-wrap{width:100%;max-width:400px;padding:1.5rem}
.auth-card{background:linear-gradient(180deg,#ffffff,#f6f8fb);border:1.5px solid rgba(15,23,42,.13);
 border-radius:16px;padding:2rem 2rem 1.75rem;
 box-shadow:inset 0 1px 0 rgba(255,255,255,.7),0 20px 50px -20px rgba(37,99,235,.25),0 2px 8px rgba(15,23,42,.06)}
.auth-logo{display:flex;align-items:center;gap:.75rem;margin-bottom:1.75rem}
.auth-logo-mark{width:42px;height:42px;background:#ffffff;border:3px solid #2563eb;border-radius:12px;
 display:flex;align-items:center;justify-content:center;flex-shrink:0;overflow:hidden;
 animation:glowPulse 4.5s ease-in-out infinite}
.auth-logo-text h1{font-family:"Sora","Inter",sans-serif;font-size:1rem;font-weight:700;letter-spacing:.03em;
 background:linear-gradient(90deg,#2563eb,#7c3aed,#0284c7,#2563eb);background-size:300% auto;
 -webkit-background-clip:text;background-clip:text;color:transparent;animation:gradShift 7s ease infinite}
.auth-logo-text p{font-size:.68rem;color:#64748b;margin-top:.15rem;letter-spacing:.06em;text-transform:uppercase}
.auth-title{font-family:"Sora","Inter",sans-serif;font-size:1.2rem;font-weight:700;color:#0f172a;margin-bottom:.25rem}
.auth-subtitle{font-size:.78rem;color:#64748b;margin-bottom:1.5rem}
.auth-field{margin-bottom:.9rem}
.auth-field label{display:block;font-size:.72rem;font-weight:700;color:#64748b;
 text-transform:uppercase;letter-spacing:.04em;margin-bottom:.3rem}
.auth-field input{width:100%;padding:.5rem .75rem;border:1.5px solid rgba(15,23,42,.13);border-radius:9px;
 font-size:.86rem;background:#fff;color:#0f172a;outline:none;
 transition:border-color .15s,box-shadow .15s;font-family:inherit}
.auth-field input:focus{border-color:#2563eb;box-shadow:0 0 0 3.5px rgba(37,99,235,.12)}
.auth-field input::placeholder{color:#94a3b8}
.auth-btn{position:relative;overflow:hidden;width:100%;padding:.65rem;border:none;border-radius:10px;
 background:linear-gradient(90deg,#2563eb,#4f46e5);color:#fff;font-weight:700;font-size:.88rem;cursor:pointer;
 margin-top:.25rem;transition:filter .15s,transform .12s,box-shadow .2s;letter-spacing:.01em;
 box-shadow:0 2px 10px -3px rgba(37,99,235,.5)}
.auth-btn::after{content:'';position:absolute;inset:0;
 background:linear-gradient(115deg,transparent 20%,rgba(255,255,255,.35) 40%,transparent 60%);
 background-size:250% 100%;background-position:150% 0;transition:background-position .55s ease}
.auth-btn:hover::after{background-position:-50% 0}
.auth-btn:hover{filter:brightness(1.1);box-shadow:0 6px 18px -4px rgba(37,99,235,.65);transform:translateY(-1.5px)}
.auth-error{background:#fef2f2;border:1px solid #fecaca;
 color:#dc2626;border-radius:7px;padding:.5rem .75rem;font-size:.8rem;margin-bottom:.9rem}
.auth-ok{background:#f0fdf4;border:1px solid #bbf7d0;
 color:#16a34a;border-radius:7px;padding:.5rem .75rem;font-size:.8rem;margin-bottom:.9rem}
.auth-switch{margin-top:1.1rem;text-align:center;font-size:.76rem;color:#64748b}
.auth-switch a{color:#2563eb;text-decoration:none;font-weight:600}
.auth-switch a:hover{text-decoration:underline}
.auth-footer{margin-top:1.25rem;text-align:center;font-size:.7rem;color:#94a3b8}
"""


def _render_forgot_password_page(message: str = "", error: str = "") -> str:
    msg_html = f'<div class="auth-ok">{message}</div>' if message else ""
    err_html = f'<div class="auth-error">{error}</div>' if error else ""
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Forgot Password -- AI Agent Data Validation</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Sora:wght@600;700;800&display=swap" rel="stylesheet"/>
<style>{_AUTH_PAGE_STYLE}</style>
</head>
<body>
<div class="auth-wrap">
  <div class="auth-card">
    <div class="auth-logo">
      <div class="auth-logo-mark"><svg viewBox="0 0 24 24" width="22" height="22" fill="none" xmlns="http://www.w3.org/2000/svg">
        <line x1="12" y1="2.6" x2="12" y2="5.2" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
        <circle cx="12" cy="2" r="1.15" fill="#2563eb"/>
        <rect x="4.5" y="5.6" width="15" height="12.4" rx="4" stroke="#2563eb" stroke-width="2"/>
        <circle cx="9.3" cy="11.6" r="1.35" fill="#2563eb"/>
        <circle cx="14.7" cy="11.6" r="1.35" fill="#2563eb"/>
        <path d="M9.2 15.1h5.6" stroke="#2563eb" stroke-width="1.6" stroke-linecap="round"/>
        <line x1="4.5" y1="10.5" x2="2.6" y2="10.5" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
        <line x1="19.5" y1="10.5" x2="21.4" y2="10.5" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
      </svg></div>
      <div class="auth-logo-text"><h1>AI Agent</h1><p>Data Validation</p></div>
    </div>
    <div class="auth-title">Forgot username or password?</div>
    <div class="auth-subtitle">Enter the email you registered with. If we find a match, we'll email you your username and a link to reset your password.</div>
    <form method="post" action="/forgot-password" autocomplete="off">
{err_html}{msg_html}
<div class="auth-field">
<label>Email</label>
<input type="email" name="identifier" placeholder="you@example.com" autocomplete="email"/>
</div>
<button type="submit" class="auth-btn">Send Reset Instructions</button>
</form>
<div class="auth-switch"><a href="/login">&#8592; Back to sign in</a></div>
  </div>
  <div class="auth-footer">&#128274; Your credentials are stored securely on this server.</div>
</div>
</body>
</html>"""


def _render_reset_password_page(token: str, error: str = "") -> str:
    err_html = f'<div class="auth-error">{error}</div>' if error else ""
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Reset Password -- AI Agent Data Validation</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Sora:wght@600;700;800&display=swap" rel="stylesheet"/>
<style>{_AUTH_PAGE_STYLE}</style>
</head>
<body>
<div class="auth-wrap">
  <div class="auth-card">
    <div class="auth-logo">
      <div class="auth-logo-mark"><svg viewBox="0 0 24 24" width="22" height="22" fill="none" xmlns="http://www.w3.org/2000/svg">
        <line x1="12" y1="2.6" x2="12" y2="5.2" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
        <circle cx="12" cy="2" r="1.15" fill="#2563eb"/>
        <rect x="4.5" y="5.6" width="15" height="12.4" rx="4" stroke="#2563eb" stroke-width="2"/>
        <circle cx="9.3" cy="11.6" r="1.35" fill="#2563eb"/>
        <circle cx="14.7" cy="11.6" r="1.35" fill="#2563eb"/>
        <path d="M9.2 15.1h5.6" stroke="#2563eb" stroke-width="1.6" stroke-linecap="round"/>
        <line x1="4.5" y1="10.5" x2="2.6" y2="10.5" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
        <line x1="19.5" y1="10.5" x2="21.4" y2="10.5" stroke="#2563eb" stroke-width="1.8" stroke-linecap="round"/>
      </svg></div>
      <div class="auth-logo-text"><h1>AI Agent</h1><p>Data Validation</p></div>
    </div>
    <div class="auth-title">Set a new password</div>
    <div class="auth-subtitle">Choose a new password for your account.</div>
    <form method="post" action="/reset-password?token={token}" autocomplete="off">
{err_html}
<div class="auth-field">
<label>New Password</label>
<input type="password" name="password" placeholder="Min 8 characters" autocomplete="new-password"/>
</div>
<div class="auth-field">
<label>Confirm New Password</label>
<input type="password" name="password2" placeholder="Re-enter password" autocomplete="new-password"/>
</div>
<button type="submit" class="auth-btn">Reset Password</button>
</form>
<div class="auth-switch"><a href="/login">&#8592; Back to sign in</a></div>
  </div>
  <div class="auth-footer">&#128274; Your credentials are stored securely on this server.</div>
</div>
</body>
</html>"""


def _render_change_password_page() -> str:
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<title>Change Password</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Sora:wght@600;700;800&display=swap" rel="stylesheet"/>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:"Inter","Segoe UI",sans-serif;color:#0f172a;
min-height:100vh;display:flex;align-items:center;justify-content:center;
-webkit-font-smoothing:antialiased;
background:
  radial-gradient(1100px 550px at 108% -8%, rgba(37,99,235,.10), transparent 60%),
  radial-gradient(900px 480px at -8% 105%, rgba(124,58,237,.08), transparent 55%),
  #eef1f6;
background-attachment:fixed}}
.wrap{{width:100%;max-width:380px;padding:1.5rem}}
.card{{background:linear-gradient(180deg,#ffffff,#f6f8fb);border:1.5px solid rgba(15,23,42,.13);
border-radius:16px;padding:1.75rem;
box-shadow:inset 0 1px 0 rgba(255,255,255,.7),0 20px 50px -20px rgba(37,99,235,.25),0 2px 8px rgba(15,23,42,.06)}}
h2{{font-family:"Sora","Inter",sans-serif;font-size:1.05rem;font-weight:700;margin-bottom:1.25rem;color:#0f172a}}
.field{{margin-bottom:.85rem}}
.field label{{display:block;font-size:.7rem;font-weight:700;color:#64748b;
text-transform:uppercase;letter-spacing:.04em;margin-bottom:.28rem}}
.field input{{width:100%;padding:.48rem .7rem;border:1.5px solid rgba(15,23,42,.13);border-radius:8px;
font-size:.84rem;background:#fff;color:#0f172a;outline:none;
transition:border-color .15s,box-shadow .15s}}
.field input:focus{{border-color:#2563eb;box-shadow:0 0 0 3.5px rgba(37,99,235,.12)}}
.btn{{position:relative;overflow:hidden;width:100%;padding:.6rem;border:none;border-radius:9px;
background:linear-gradient(90deg,#2563eb,#4f46e5);
color:#fff;font-weight:700;font-size:.85rem;cursor:pointer;margin-top:.2rem;
transition:filter .15s,transform .12s,box-shadow .2s;box-shadow:0 2px 10px -3px rgba(37,99,235,.5)}}
.btn::after{{content:'';position:absolute;inset:0;
background:linear-gradient(115deg,transparent 20%,rgba(255,255,255,.35) 40%,transparent 60%);
background-size:250% 100%;background-position:150% 0;transition:background-position .55s ease}}
.btn:hover::after{{background-position:-50% 0}}
.btn:hover{{filter:brightness(1.1);box-shadow:0 6px 18px -4px rgba(37,99,235,.65);transform:translateY(-1.5px)}}
#msg{{margin-top:.75rem;font-size:.78rem;min-height:1.1rem;text-align:center}}
a{{color:#2563eb;font-size:.76rem;display:block;text-align:center;margin-top:1rem}}
</style>
</head>
<body>
<div class="wrap"><div class="card">
<h2>&#128274; Change Password</h2>
<div class="field"><label>Current Password</label>
<input type="password" id="old" placeholder="********"/></div>
<div class="field"><label>New Password</label>
<input type="password" id="new1" placeholder="Min 8 characters"/></div>
<div class="field"><label>Confirm New Password</label>
<input type="password" id="new2" placeholder="********"/></div>
<button class="btn" onclick="submit()">Update Password</button>
<div id="msg"></div>
<a href="/">&#8592; Back to app</a>
</div></div>
<script>
async function submit() {{
const msg = document.getElementById('msg');
const old = document.getElementById('old').value;
const n1  = document.getElementById('new1').value;
const n2  = document.getElementById('new2').value;
if (n1 !== n2) {{ msg.style.color='#dc2626'; msg.textContent='Passwords do not match.'; return; }}
msg.style.color='#64748b'; msg.textContent='Saving...';
const r = await fetch('/api/auth/change-password', {{
method:'POST', headers:{{'Content-Type':'application/json'}},
body: JSON.stringify({{old_password: old, new_password: n1}})
}});
const d = await r.json();
if (d.ok) {{ msg.style.color='#16a34a'; msg.textContent='✅ ' + d.message; setTimeout(()=>location.href='/login',1500); }}
else      {{ msg.style.color='#dc2626'; msg.textContent=d.error||'Failed.'; }}
}}
</script>
</body>
</html>"""


@app.post("/preview")
async def preview_file(request: Request):

    # Lightweight endpoint: parse one uploaded file and return a JSON preview
    # (first 10 rows, column names, detected format/delimiter, inferred key).
    # Called client-side after a file is dropped into the upload zone.

    try:
        form = await request.form()
    except Exception as exc:
        return JSONResponse({"error": f"Form parse error: {exc}"}, status_code=400)

    files = [v for v in form.getlist("files") if hasattr(v, "filename") and v.filename]
    delimiter = str(form.get("delimiter", "")).strip() or None


    if not files:
        return JSONResponse({"error": "No file provided"}, status_code=400)

    f = files[0]
    try:
        df = _load_file(f, delimiter=delimiter)
    except HTTPException as exc:
        return JSONResponse({"error": exc.detail}, status_code=exc.status_code)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)

    cols = list(df.columns)
    # Infer key on single file (compare df to itself -- perfect uniqueness wins)
    keys, key_method = infer_keys(df, df, cols)

    # First 10 rows, all values as strings
    preview = df.head(10).fillna("").astype(str).to_dict(orient="records")

    return JSONResponse({


        "filename":     f.filename,
        "format":       df.attrs.get("_format", "unknown"),
        "delimiter":    df.attrs.get("_delimiter", ""),
        "rows":         len(df),
        "columns":      len(cols),
        "column_names": cols,
        "inferred_keys": keys,
        "key_method":   key_method,
        "preview":      preview,
    })


@app.post("/xref")
async def xref_analyze(request: Request):

    # Cross Reference endpoint -- receives N source files/connections,
    # runs analyze_cross_reference(), returns rendered results HTML.

    _t0 = time.time()


    proc_logs: list[dict] = []

    def _log(msg: str, level: str = "INFO") -> None:
        proc_logs.append({"elapsed": round(time.time() - _t0, 3), "level": level, "message": msg})

    try:
        form = await request.form()
    except Exception as exc:
        raise HTTPException(400, f"Could not parse form: {exc}")

    session_id = str(uuid.uuid4())

    # Parse source count
    source_count = int(str(form.get("xref_source_count", "2")))
    source_count = max(2, min(5, source_count))

    # Load each source


    sources: list[tuple[str, pd.DataFrame]] = []
    _ws_username = _ws_resolve_username(request) if _ws_check_soft() else None

    for i in range(1, source_count + 1):
        name_raw = f"Source {i}"  # will be overridden by filename or connection name below
        uploads  = [v for v in form.getlist(f"xref_files_{i}")
                    if hasattr(v, "filename") and v.filename]
        conn_ids = [c.strip() for c in str(form.get(f"xref_conn_ids_{i}", "")).split(",") if c.strip()]

        df: pd.DataFrame | None = None

        # Try uploaded files first
        for upload in uploads:
            try:
                df = _load_file(upload)
                # Use filename without extension as the source label
                import os as _os
                _resolved_fname = df.attrs.get("_source_filename") or upload.filename
                name_raw = _os.path.splitext(_resolved_fname)[0] if _resolved_fname else name_raw


                _log(f"Loaded source {i} '{name_raw}' from upload: {len(df)} rows x {len(df.columns)} cols")
                break
            except Exception as exc:
                _log(f"Could not load upload for source {i}: {exc}", "WARN")

        # Try workspace connections
        if df is None:
            for cid in conn_ids:
                try:
                    if _ws_username:
                        rec = _ws_db.get_connection(cid, _ws_username)
                        if rec:
                            from workspace.connectors import BaseConnector as _BC
                            df = _BC.from_type(rec["source_type"], rec["config"]).fetch()
                            name_raw = rec.get("name", f"Source {i}")
                            _log(f"Fetched source {i} '{name_raw}' from connection: {len(df)} rows")
                            break
                except Exception as exc:
                    _log(f"Could not fetch connection {cid} for source {i}: {exc}", "WARN")

        if df is not None and len(df) > 0:
            sources.append((name_raw, df))
        else:
            _log(f"Source {i} '{name_raw}' has no data -- skipped", "WARN")
        # NOTE: dropped a duplicated/mis-indented re-photograph of this
        # "if df is not None..." + option-parsing block (source page 0701
        # overlapped 0700/0702) that would have re-run per source or
        # redefined key_col/compare_fields/etc. a second time.

    if len(sources) < 2:
        raise HTTPException(400, "Cross Reference requires at least 2 sources with data.")

    # Parse options
    key_col            = str(form.get("xref_key_col", "")).strip() or None
    compare_fields_raw = str(form.get("xref_compare_fields", "")).strip()
    compare_fields     = [c.strip() for c in compare_fields_raw.split(",") if c.strip()] or None
    golden_idx         = str(form.get("xref_golden_source", "")).strip()
    golden_source      = sources[int(golden_idx)-1][0] if golden_idx.isdigit() and 0 < int(golden_idx) <= len(sources) else None
    conflicts_only     = form.get("xref_conflicts_only") == "on"
    show_coverage      = form.get("xref_show_coverage", "on") == "on"


    _log(f"Cross Reference: {len(sources)} sources, key={key_col or 'auto'}, golden={golden_source or 'auto'}")

    # Run the engine
    try:
        xr = await asyncio.to_thread(
            analyze_cross_reference,
            sources,
            key_col=key_col,
            compare_fields=compare_fields,
            golden_source=golden_source,
            conflicts_only=conflicts_only,
            show_coverage=show_coverage,
        )
    except Exception as exc:
        raise HTTPException(500, f"Cross Reference failed: {exc}")

    total_elapsed = round(time.time() - _t0, 3)


    _log(f"Cross Reference complete in {total_elapsed}s -- {xr['summary']['conflict_count']} conflicts, {xr['summary']['matched_in_all']} matched keys")

    # ---- Reshape engine output for template compatibility ----
    # 1. Rename conflicts[].values -> conflicts[].source_values
    for c in xr.get("conflicts", []):
        if "values" in c and "source_values" not in c:
            c["source_values"] = c.pop("values")

    # 2. Add summary.coverage_gap_count
    xr["summary"].setdefault("coverage_gap_count",
        sum(len(v) for v in xr.get("only_in", {}).values()))
    xr["summary"].setdefault("golden_source",
        xr.get("golden_source", "auto"))

    # 3. Build sources list with only_in_source rows (for display + excel)
    keyed_dfs: dict[str, "pd.DataFrame"] = {}
    for name, df in sources:
        k_col = xr.get("key_col_per_source", {}).get(name)


        if k_col and k_col in df.columns:
            tmp = df.copy()
            tmp["__xref_key__"] = _xref_normalise_key(tmp[k_col])
            keyed_dfs[name] = tmp.set_index("__xref_key__")

    src_list = []
    for name, df in sources:
        stat = xr.get("source_stats", {}).get(name, {})
        only_keys = xr.get("only_in", {}).get(name, [])
        # Fetch actual rows for only-in-source keys
        only_rows: list[dict] = []
        if only_keys and name in keyed_dfs:
            kdf = keyed_dfs[name]
            for k in only_keys[:200]:
                if k in kdf.index:
                    only_rows.append({col: str(kdf.loc[k, col]) for col in kdf.columns[:20]})
        src_list.append({
            "name":      name,
            "key_col":   xr.get("key_col_per_source", {}).get(name, ""),


            "row_count":  stat.get("rows", len(df)),
            "matched":    stat.get("matched", 0),
            "exclusive":  stat.get("exclusive", 0),
            "conflicts":  stat.get("conflicts", 0),
            "coverage_pct": stat.get("coverage_pct", 0),
            "is_golden":  name == xr.get("golden_source", ""),
            "only_in_source": only_rows,
        })
    xr["sources"] = src_list

    # 4. Reshape coverage matrix for template ({col: {src: count}} -> {rows: [{key, present_in: [src]}]})
    raw_cov = xr.get("coverage_matrix", {})
    if isinstance(raw_cov, dict) and not ("rows" in raw_cov):
        src_names_for_cov = [s["name"] for s in src_list]
        # Coverage matrix per key: build from keyed_dfs
        all_keys_for_cov = set()
        for kdf in keyed_dfs.values():
            all_keys_for_cov.update(kdf.index.tolist())


        cov_rows = []
        for k in sorted(list(all_keys_for_cov))[:500]:
            present_in = [n for n in src_names_for_cov if n in keyed_dfs and k in keyed_dfs[n].index]
            cov_rows.append({"key": k, "present_in": present_in})
        xr["coverage_matrix"] = {
            "source_names": src_names_for_cov,
            "rows": cov_rows,
        }

    xr["proc_logs"] = proc_logs
    xr["elapsed"]   = total_elapsed
    xr["session_id"] = session_id
    _results_store[session_id] = {
        "action":     "xref",
        "file_names": [s[0] for s in sources],
        "xref":       xr,
        "proc_logs":  proc_logs,
        "elapsed":    total_elapsed,
        # Kept for /xref/run-llm/{session_id} (AI Copilot re-run with saved
        # rules) -- every other module already stores its dataframes for
        # rerun purposes, xref just never needed to until now.
        "dataframes": [{"name": n, "df": df} for n, df in sources],
    }

    # Build AI-summary context
    _xref_username = _ws_resolve_username(request) or "default"
    _dataset_fingerprint = _fp_resolve(
        _xref_username,
        _fp_compute([c for s in sources for c in s[1].columns], []),
    )
    # Persist so /dataset-controls/*, /rules/*, and /xref/run-llm can resolve
    # it later -- without this, every "remember: ..." rule saved from the
    # xref Copilot was silently written under an empty-string fingerprint
    # instead of this schema's real one (results_store.get(...,"") in every
    # rule endpoint), so saved rules never actually reloaded on a later run.
    _results_store[session_id]["dataset_fingerprint"] = _dataset_fingerprint
    _saved_rules_text = _fp_rules_text(_xref_username, _dataset_fingerprint)

    _summary = xr.get("summary", {})
    _cov = xr.get("coverage_matrix", {})
    _cov_sources = _cov.get("source_names", [])
    _cov_rows = [
        {"identifier": row["key"], **{s: (s in row["present_in"]) for s in _cov_sources}}
        for row in _cov.get("rows", [])
    ]
    _conflicts = [
        {
            "identifier": c["key"], "field": c["field"], "values": c.get("source_values", c.get("values", {})),
            "conflict_type": c.get("conflict_type", ""),
            "golden_value": c.get("golden_value", ""),
            "sources_agree": c.get("sources_agree", []),
            "sources_differ": c.get("sources_differ", []),
        }
        for c in xr.get("conflicts", [])
    ]
    return JSONResponse(_sanitize_json({
        "session_id": session_id,
        "counts": {
            "total_identifiers": _summary.get("total_keys", 0),
            "in_all_sources": _summary.get("matched_in_all", 0),
            "conflicts": _summary.get("conflict_count", 0),
        },
        "identifier_key": xr.get("key_col_used", ""),
        "golden_source": xr.get("golden_source", "auto (majority-wins)"),
        "sources": _cov_sources,
        "coverage_matrix": _cov_rows,
        "conflicts": _conflicts,
        "dataset_fingerprint": _dataset_fingerprint,
        "elapsed": total_elapsed,
        "logs": proc_logs,
    }))


def _parse_xref_rules_to_params(rules: list[dict], source_names: list[str], common_cols: list[str]) -> dict:
    """Ask the LLM to read saved Cross Reference rules and turn them into
    analyze_cross_reference() execution parameters. Mirrors
    _parse_recon_rules_to_params in spirit but with xref's smaller parameter
    set. Falls back to {} on no rules, {"_llm_error": ...} on LLM failure --
    same "surface the failure, don't run silently" convention as recon."""
    rule_text = "\n".join(
        f" [{r['category'].upper()}] {r['rule']}"
        for r in rules if r.get("category") not in ("recon_hints",)
    )
    if not rule_text.strip():
        return {}

    prompt = f"""You are a data cross-reference parameter extractor.

Given the user rules and the common columns shared across {len(source_names)} sources
({', '.join(source_names)}), produce ONLY a JSON object with execution parameters.

Common columns: {common_cols}

JSON schema (omit any key that is not needed):
{{
  "key_col": "col_name",           // the identifier column to match rows on across sources
  "compare_fields": ["col_a","col_b"],  // columns to compare for conflicts (all common columns if omitted)
  "exclude_cols": ["col_c"],       // columns to exclude entirely
  "golden_source": "Source Name"   // which source is authoritative on conflicts (majority-wins if omitted)
}}

Key rules:
- key_col and every entry in compare_fields/exclude_cols must be one of the common columns listed above.
- golden_source must exactly match one of: {source_names}.
- Only include what you are confident about. Reply with ONLY valid JSON, no commentary."""

    try:
        raw = _ask_llm([{"role": "user", "content": [{"text": prompt}]}])
        raw = re.sub(r"^```[a-z]*\n?", "", raw.strip(), flags=re.MULTILINE)
        raw = re.sub(r"```$", "", raw.strip())
        return json.loads(raw)
    except Exception as exc:
        return {"_llm_error": str(exc)}


@app.post("/xref/run-llm/{session_id}")
async def xref_run_llm(session_id: str, request: Request):
    """AI Copilot's "Run Cross Reference" -- reads saved rules for this
    schema, has the LLM turn them into key_col/compare_fields/exclude_cols/
    golden_source, reruns analyze_cross_reference() with them applied, and
    returns the same response shape /xref uses so the frontend can render it
    with the identical results UI."""
    _xrl_username = _ws_resolve_username(request) or "default"
    stored = _results_store.get(session_id, {})
    if not stored or "dataframes" not in stored:
        raise HTTPException(404, "Session not found or missing stored sources -- please re-run Cross Reference.")

    _t0 = time.time()
    proc_logs: list[dict] = []
    def _log(msg: str, level: str = "INFO") -> None:
        proc_logs.append({"elapsed": round(time.time() - _t0, 3), "level": level, "message": msg})

    sources: list[tuple[str, pd.DataFrame]] = [(item["name"], item["df"]) for item in stored["dataframes"]]
    common_cols = list(set.intersection(*(set(df.columns) for _, df in sources))) if sources else []

    fingerprint = stored.get("dataset_fingerprint") or _fp_resolve(
        _xrl_username,
        _fp_compute([c for s in sources for c in s[1].columns], []),
    )
    saved_rules = _fp_get_rules(_xrl_username, fingerprint, module="xref")
    params = await asyncio.to_thread(_parse_xref_rules_to_params, saved_rules, [s[0] for s in sources], common_cols)
    llm_error = params.pop("_llm_error", None)

    key_col = params.get("key_col")
    compare_fields = params.get("compare_fields") or None
    exclude_cols = set(params.get("exclude_cols") or [])
    if exclude_cols:
        sources = [(n, df.drop(columns=[c for c in exclude_cols if c in df.columns])) for n, df in sources]
    golden_source = params.get("golden_source")
    if golden_source and golden_source not in [s[0] for s in sources]:
        golden_source = None  # LLM hallucinated a source name -- fall back to auto

    _log(f"AI Copilot re-run: key={key_col or 'auto'}, golden={golden_source or 'auto'}, "
         f"{len(exclude_cols)} excluded col(s)")

    try:
        xr = await asyncio.to_thread(
            analyze_cross_reference, sources, key_col=key_col, compare_fields=compare_fields,
            golden_source=golden_source, conflicts_only=False, show_coverage=True,
        )
    except Exception as exc:
        raise HTTPException(500, f"Cross Reference failed: {exc}")

    total_elapsed = round(time.time() - _t0, 3)
    for c in xr.get("conflicts", []):
        if "values" in c and "source_values" not in c:
            c["source_values"] = c.pop("values")
    xr["summary"].setdefault("coverage_gap_count", sum(len(v) for v in xr.get("only_in", {}).values()))
    xr["summary"].setdefault("golden_source", xr.get("golden_source", "auto"))

    new_session_id = str(uuid.uuid4())
    xr["session_id"] = new_session_id
    _results_store[new_session_id] = {
        "action": "xref", "file_names": [s[0] for s in sources], "xref": xr,
        "proc_logs": proc_logs, "elapsed": total_elapsed,
        "dataframes": [{"name": n, "df": df} for n, df in sources],
        "dataset_fingerprint": fingerprint,
    }

    _summary = xr.get("summary", {})
    _cov = xr.get("coverage_matrix", {})
    _cov_sources = _cov.get("source_names", [s[0] for s in sources])
    _cov_rows = [
        {"identifier": row["key"], **{s: (s in row["present_in"]) for s in _cov_sources}}
        for row in _cov.get("rows", [])
    ]
    _conflicts = [
        {
            "identifier": c["key"], "field": c["field"], "values": c.get("source_values", c.get("values", {})),
            "conflict_type": c.get("conflict_type", ""),
            "golden_value": c.get("golden_value", ""),
            "sources_agree": c.get("sources_agree", []),
            "sources_differ": c.get("sources_differ", []),
        }
        for c in xr.get("conflicts", [])
    ]
    return JSONResponse(_sanitize_json({
        "session_id": new_session_id,
        "counts": {
            "total_identifiers": _summary.get("total_keys", 0),
            "in_all_sources": _summary.get("matched_in_all", 0),
            "conflicts": _summary.get("conflict_count", 0),
        },
        "identifier_key": xr.get("key_col_used", ""),
        "golden_source": xr.get("golden_source", "auto (majority-wins)"),
        "sources": _cov_sources,
        "coverage_matrix": _cov_rows,
        "conflicts": _conflicts,
        "dataset_fingerprint": fingerprint,
        "elapsed": total_elapsed,
        "logs": proc_logs,
        "llm_error": llm_error,
        "rules_applied": len(saved_rules),
    }))


def _ws_check_soft() -> bool:
    # Return True if workspace is available (non-raising).
    try:
        _ws_check()
        return True
    except Exception:
        return False


@app.post("/analyze")
async def analyze(request: Request):
    # --- Processing log ---
    _t0 = time.time()
    proc_logs: list[dict] = []

    def _log(msg: str, level: str = "INFO") -> None:
        elapsed = round(time.time() - _t0, 3)
        proc_logs.append({"elapsed": elapsed, "level": level, "message": msg})

    # Parse form data directly -- avoids FastAPI version-specific UploadFile quirks
    try:
        form = await request.form()
    except Exception as exc:
        raise HTTPException(400, f"Could not parse form data: {exc}")

    # -- scalar fields --
    # action is sent as a plain form field from a fresh FormData (no hidden-input collision)
    _action_vals = form.getlist("action")
    action     = str(_action_vals[0] if _action_vals else "compare")

    # -- License gate -- check feature is allowed for current subscription --
    _ai_to_feature = {"quality_ai": "quality", "profile_ai": "profile", "governance_ai": "governance"}

    _require_feature(_ai_to_feature.get(action, action))

    # -- Data Intelligence scope filter -- only return selected reports --
    # di_scope is a comma-separated list sent by the UI checkboxes e.g. "quality,profile"
    _di_scope_raw = str(form.get("di_scope", "")).strip()
    _di_scope = set(_di_scope_raw.split(",")) if _di_scope_raw else {"quality", "profile", "governance"}
    # BFSI Rule Pack selected in UI (e.g. "trade", "payments", "mifid2")
    _bfsi_pack    = str(form.get("bfsi_pack",     "")).strip().lower()
    _di_ai_enhanced = str(form.get("di_ai_enhanced", "")).strip() == "1"

    key_columns    = str(form.get("key_columns",    ""))
    exclude_columns = str(form.get("exclude_columns", ""))
    auto_keys     = str(form.get("auto_keys",     "on"))
    delimiter     = str(form.get("delimiter",     "")).strip() or None
    preprocess_a   = str(form.get("preprocess_a",   "")).strip()
    preprocess_b   = str(form.get("preprocess_b",   "")).strip()
    # Fuzzy/probabilistic matching -- opt-in, for when there's no clean shared key.
    fuzzy_fields   = [f.strip() for f in str(form.get("fuzzy_fields", "")).split(",") if f.strip()]
    fuzzy_threshold = str(form.get("fuzzy_threshold", "")).strip()
    try:
        fuzzy_threshold = float(fuzzy_threshold) if fuzzy_threshold else 0.75
    except ValueError:
        fuzzy_threshold = 0.75

    # -- Customize/Complex Compare criteria panel --

    # These fields are sent when the user runs a "lineage" (custom compare) action.
    # They feed the LLM column-mapping and transform hints so rules are schema-generic.
    _recon_prompt   = str(form.get("recon_prompt",   "")).strip()
    _mapping_hints  = str(form.get("mapping_hints",  "")).strip()
    _key_hints     = str(form.get("key_hints",     "")).strip()
    _transform_hints = str(form.get("transform_hints", "")).strip()
    _exclude_hints  = str(form.get("exclude_hints",  "")).strip()

    # Per-column DQ config table (JSON array from the interactive config table)
    _col_config_json = str(form.get("col_config_json", "")).strip()
    _col_config: list[dict] = []
    if _col_config_json:
        try:
            _col_config = json.loads(_col_config_json)
        except Exception:
            pass

    # DQ-specific hints from the quality panel hint fields

    _nullable_hints    = str(form.get("nullable_hints",     "")).strip()
    _range_hints      = str(form.get("range_hints",      "")).strip()
    _timeliness_hints   = str(form.get("timeliness_hints",   "")).strip()
    _precision_hints    = str(form.get("precision_hints",    "")).strip()
    _cross_col_rules    = str(form.get("cross_column_rules",   "")).strip()
    _ref_rules       = str(form.get("referential_rules",   "")).strip()
    _cross_file_rules   = str(form.get("cross_file_rules",    "")).strip()
    _cond_completeness   = str(form.get("conditional_completeness_rules","")).strip()

    user_hints: dict | None = None
    if any([_recon_prompt, _mapping_hints, _key_hints, _transform_hints, _exclude_hints,
         _col_config, _nullable_hints, _range_hints, _timeliness_hints,
         _precision_hints, _cross_col_rules, _ref_rules,
         _cross_file_rules, _cond_completeness]):
        user_hints = {
            "domain_context":    _recon_prompt,
            "mapping_hints":    _mapping_hints,
            "key_hints":      _key_hints,

            "transform_hints":    _transform_hints,
            "exclude_hints":     _exclude_hints,
            "col_config":       _col_config,
            "nullable_hints":    _nullable_hints,
            "range_hints":      _range_hints,
            "timeliness_hints":   _timeliness_hints,
            "precision_hints":    _precision_hints,
            "cross_column_rules":  _cross_col_rules,
            "referential_rules":   _ref_rules,
            "cross_file_rules":   _cross_file_rules,
            "conditional_completeness_rules": _cond_completeness,
        }

    _log(f"Action: {action} | Key columns: '{key_columns}' | "
       f"Exclude columns: '{exclude_columns}' | "
       f"Delimiter: {repr(delimiter) if delimiter else 'auto'}")

    # -- primary data files --
    # Compare / Lineage use two named groups (files_a, files_b) that are each

    # concatenated into one DataFrame before comparison.  All other actions
    # use the legacy "files" field.
    files_a_uploads: list[UploadFile] = [
        v for v in form.getlist("files_a")
        if hasattr(v, "filename") and v.filename
    ]
    files_b_uploads: list[UploadFile] = [
        v for v in form.getlist("files_b")
        if hasattr(v, "filename") and v.filename
    ]
    # Optional additional sources (files_c..files_f) for N-way reconciliation --
    # each is its own separate source, unlike files_a/files_b's row-concatenation.
    # Empty/absent for the ordinary 2-file case, so behaviour there is unchanged.
    _nway_extra_uploads: list[tuple[str, list[UploadFile]]] = [
        (label, [v for v in form.getlist(f"files_{label}") if hasattr(v, "filename") and v.filename])
        for label in ("c", "d", "e", "f")
    ]
    files: list[UploadFile] = [
        v for v in form.getlist("files")
        if hasattr(v, "filename") and v.filename
    ]

    # -- workspace connection IDs (comma-separated strings from hidden inputs) --
    # These supplement or replace file uploads with live data fetched from saved connections.
    # Use _ws_resolve_username (not get_current_user) because /analyze is outside

    # /api/ws/* so the middleware never sets request.state.username for this path.
    _ws_username = ""
    if _WS_ENABLED:
        try:
            _ws_username = _ws_resolve_username(request) or ""
        except Exception:
            pass

    def _parse_conn_ids(field: str) -> list[str]:
        return [c.strip() for c in str(form.get(field, "")).split(",") if c.strip()]

    _conn_a_ids = _parse_conn_ids("conn_a_ids")  # Dataset A connections (compare)
    _conn_b_ids = _parse_conn_ids("conn_b_ids")  # Dataset B connections (compare)
    _conn_ids  = _parse_conn_ids("conn_ids")   # Single-side connections (quality/parse/etc.)
    # Extra N-way source connections (files_c..files_f's "Saved Connector" pick) --
    # mirrors conn_a_ids/conn_b_ids so a source can be a live connection instead
    # of (or in addition to) an upload, same as Dataset A/B already support.
    _nway_extra_conn_ids: dict[str, list[str]] = {
        label: _parse_conn_ids(f"conn_{label}_ids") for label in ("c", "d", "e", "f")
    }

    # -- Knowledge Base: local uploads + workspace connections ----------------
    ref_uploads = [
        v for v in form.getlist("ref_docs")

        if hasattr(v, "filename") and v.filename
    ]
    # Fetch documents from saved workspace connections (SharePoint, OneDrive, S3, SFTP etc.)
    _kb_conn_ids = [c.strip() for c in str(form.get("kb_conn_ids", "")).split(",") if c.strip()]
    kb_raw: list[tuple[str, bytes]] = []
    for _kb_cid in _kb_conn_ids:
        try:
            _kb_rec = _ws_db.get_connection(_kb_cid, _ws_username) if _ws_username else None
            if _kb_rec:
                from workspace.connectors import BaseConnector as _BC
                import io as _kbio
                _kb_df = _BC.from_type(_kb_rec["source_type"], _kb_rec["config"]).fetch()
                # Serialise DataFrame back to bytes so _extract_kb_as_ref_doc can re-parse it
                _kb_buf = _kbio.BytesIO()
                if len(_kb_df.columns) > 1:
                    # Structured -- write as CSV bytes

                    _kb_df.to_csv(_kb_buf, index=False)
                    _kb_bytes = _kb_buf.getvalue().encode("utf-8")
                    _kb_fname = _kb_rec.get("name", _kb_cid) + ".csv"
                else:
                    # Single-column (text blob) -- write as plain text
                    _kb_bytes = "\n".join(_kb_df.iloc[:, 0].astype(str).tolist()).encode("utf-8")
                    _kb_fname = _kb_rec.get("name", _kb_cid) + ".txt"
                kb_raw.append((_kb_fname, _kb_bytes))
                _log(f"Knowledge Base: fetched '{_kb_rec.get('name', _kb_cid)}' ({len(_kb_df)} rows)")
        except Exception as _kb_err:
            _log(f"Knowledge Base: could not fetch connection '{_kb_cid}': {_kb_err}", level="WARN")

    # -- validation --
    # Allow connections to satisfy the "has data" requirement
    has_a = bool(files_a_uploads or _conn_a_ids)
    has_b = bool(files_b_uploads or _conn_b_ids)
    has_files = bool(files or _conn_ids)

    if action in ("compare", "lineage"):
        if not has_a or not has_b:
            raise HTTPException(
                400,
                f"{action.capitalize()} requires data in both Dataset A and Dataset B. "
                f"Got {len(files_a_uploads)} upload(s) + {len(_conn_a_ids)} connection(s) in A and "
                f"{len(files_b_uploads)} upload(s) + {len(_conn_b_ids)} connection(s) in B."
            )
    else:
        if not has_files:
            raise HTTPException(400, "No data files or connections provided. Please upload a file or select a saved connection.")
        if len(files) + len(_conn_ids) > 6:
            raise HTTPException(400, "Maximum 6 data sources supported.")

    def _load_upload_list(upload_list: list[UploadFile]) -> list[tuple[str, pd.DataFrame]]:
        # Load a list of UploadFile objects into (filename, DataFrame) tuples.
        # An archive (.zip/.tar/.tar.gz/.tgz/.gz) expands into one entry per
        # usable file inside it, so a single zip of several datasets becomes
        # several entries here -- not just its largest file.
        result = []
        for f in upload_list:

            _log(f"Loading data file: {f.filename}")
            f.file.seek(0)
            try:
                loaded = _load_file_multi_from_upload(f, delimiter=delimiter)
            except HTTPException as exc:
                _log(f"Skipped '{f.filename}': {exc.detail}")
                continue
            if len(loaded) > 1:
                _log(f"'{f.filename}' is an archive -> expanded into {len(loaded)} file(s)")
            for _resolved_fname, df in loaded:
                fmt  = df.attrs.get("_format",  "unknown")
                delim = df.attrs.get("_delimiter", "")
                delim_info = f" delim={repr(delim)}" if delim and delim != "auto" else ""
                _log(f"Loaded '{_resolved_fname}' [format: {fmt}{delim_info}] -> {len(df)} rows x {len(df.columns)} cols")
                result.append((_resolved_fname, df))
        return result

    def _concat_group(tuples: list[tuple[str, pd.DataFrame]], label: str) -> tuple[str, pd.DataFrame]:
        # Row-concatenate a list of same-schema DataFrames into one.
        if not tuples:
            raise HTTPException(400, f"Dataset {label} has no usable files -- every uploaded file "
                                      f"was unreadable, empty, or (if an archive) contained nothing parseable.")
        if len(tuples) == 1:
            return tuples[0]
        names = ", ".join(n for n, _ in tuples)
        dfs  = [df for _, df in tuples]
        combined = pd.concat(dfs, ignore_index=True)
        # Preserve format attribute from first file

        combined.attrs["_format"]  = tuples[0][1].attrs.get("_format", "unknown")
        combined.attrs["_delimiter"] = tuples[0][1].attrs.get("_delimiter", "")
        _log(f"Concatenated {len(tuples)} files into Dataset {label} ({names}) → {len(combined)} rows")
        return (names, combined)


    # Load data files
    dataframes: list[tuple[str, pd.DataFrame]] = []
    # Raw bytes keyed by filename -- used by the parse action
    raw_file_bytes: dict[str, bytes] = {}

    def _apply_preprocess(
        name: str, df: pd.DataFrame, cfg_json: str
    ) -> tuple[str, pd.DataFrame]:
        # Group-by + aggregate a DataFrame per user-supplied pre-process config.
        # cfg_json: JSON string {"group_by": ["col",...], "agg_col": "col", "agg_fn": "sum"}
        # Returns the (possibly transformed) (name, df) tuple unchanged if cfg is empty/invalid.

        if not cfg_json:
            return name, df
        try:
            cfg = json.loads(cfg_json)
            group_by: list[str] = cfg.get("group_by", [])
            agg_col: str    = cfg.get("agg_col", "")
            agg_fn: str    = cfg.get("agg_fn", "sum").lower()
            if not group_by or not agg_col:
                return name, df
            # Validate columns exist (case-insensitive lookup)
            col_map = {c.lower(): c for c in df.columns}
            resolved_keys = []
            for k in group_by:
                real = col_map.get(k.lower())
                if real is None:
                    _log(f"Pre-process: group-by column '{k}' not found in '{name}' -- skipping", level="WARN")
                    return name, df
                resolved_keys.append(real)

            real_agg = col_map.get(agg_col.lower())
            if real_agg is None:
                _log(f"Pre-process: aggregate column '{agg_col}' not found in '{name}' -- skipping", level="WARN")
                return name, df
            _fn_map = {"sum": "sum", "mean": "mean", "first": "first", "count": "count"}
            fn = _fn_map.get(agg_fn, "sum")
            before = len(df)
            agg_dict = {real_agg: fn}
            # Keep first value of all non-key, non-agg columns
            for c in df.columns:
                if c not in resolved_keys and c != real_agg:
                    agg_dict[c] = "first"
            df_out = df.groupby(resolved_keys, as_index=False, dropna=False).agg(agg_dict)
            df_out.attrs["_format"]  = df.attrs.get("_format", "unknown")
            df_out.attrs["_delimiter"] = df.attrs.get("_delimiter", "")
            _log(
                f"Pre-process '{name}': grouped by {resolved_keys}, "
                f"{fn}({real_agg}) → {before} rows → {len(df_out)} rows"

            )
            return name, df_out
        except Exception as pp_err:
            _log(f"Pre-process failed for '{name}': {pp_err}", level="WARN")
            return name, df

    # Helper: fetch a list of (name, df) from workspace connection IDs
    def _load_conn_list(conn_ids: list[str], label: str) -> list[tuple[str, pd.DataFrame]]:
        result = []
        for cid in conn_ids:
            _log(f"Fetching data from workspace connection '{cid}' (Dataset {label})")
            try:
                cname, df = _df_from_connection(cid, _ws_username)
                row_limit_hit = df.attrs.get("_row_limit_hit")
                if row_limit_hit:
                    # Silent truncation is worse than the row cap itself --
                    # the user could draw conclusions from a partial dataset
                    # without ever knowing rows were cut off. row_limit is
                    # per-connection config; 0 disables the cap entirely.
                    _log(
                        f"⚠ '{cname}' hit its row_limit ({row_limit_hit}) -- the source "
                        f"may have more rows than were fetched. Raise row_limit on this "
                        f"connection (or set it to 0 for unlimited) if you need the full data.",
                        level="WARN",
                    )
                _log(f"Fetched '{cname}' → {len(df)} rows × {len(df.columns)} cols")
                result.append((cname, df))
            except HTTPException:
                raise
            except Exception as exc:
                raise HTTPException(502, f"Connection fetch error: {exc}") from exc

        return result

    if action in ("compare", "lineage"):
      # Load both sides -- combine uploaded files + connections, then merge into one DF each
      a_tuples = _load_upload_list(files_a_uploads) + _load_conn_list(_conn_a_ids, "A")
      b_tuples = _load_upload_list(files_b_uploads) + _load_conn_list(_conn_b_ids, "B")
      a_name, a_df = _concat_group(a_tuples, "A")
      b_name, b_df = _concat_group(b_tuples, "B")
      a_name, a_df = _apply_preprocess(a_name, a_df, preprocess_a)
      b_name, b_df = _apply_preprocess(b_name, b_df, preprocess_b)
      dataframes = [(a_name, a_df), (b_name, b_df)]
      if action == "compare":
        # Each additional source (files_c..files_f) is its own separate dataset for
        # N-way reconciliation, not concatenated with anything else. A source can
        # be an upload, a saved connection, or both (combined like Dataset A/B).
        for _label, _uploads in _nway_extra_uploads:
          _extra_conn_ids = _nway_extra_conn_ids.get(_label, [])
          if not _uploads and not _extra_conn_ids:
            continue
          _extra_tuples = _load_upload_list(_uploads) + _load_conn_list(_extra_conn_ids, _label.upper())
          if _extra_tuples:
            _extra_name, _extra_df = _concat_group(_extra_tuples, _label.upper())
            dataframes.append((_extra_name, _extra_df))
    else:
      # Upload files first
      for f in files:
        _log(f"Loading data file: {f.filename}")
        f.file.seek(0)
        raw_file_bytes[f.filename] = f.file.read()
        f.file.seek(0)


        df = _load_file(f, delimiter=delimiter)
        fmt  = df.attrs.get("_format",  "unknown")
        delim = df.attrs.get("_delimiter", "")
        _resolved_fname = df.attrs.get("_source_filename") or f.filename
        dataframes.append((_resolved_fname, df))
        delim_info = f" delim={repr(delim)}" if delim and delim != "auto" else ""
        _log(f"Loaded '{_resolved_fname}' [format: {fmt}{delim_info}] -> {len(df)} rows x {len(df.columns)} cols")
      # Then connection sources
      for cname, df in _load_conn_list(_conn_ids, "data"):
        dataframes.append((cname, df))


    # Load & auto-classify all reference documents
    _log(f"Loading {len(ref_uploads)} reference document(s)")
    ref_result  = await asyncio.to_thread(_load_and_classify_ref_docs, ref_uploads, kb_raw=kb_raw)
    data_dict  = ref_result["data_dict"]
    rules    = ref_result["rules"]
    mapping_spec = ref_result["mapping_spec"]
    ref_log   = ref_result["classifications"]  # shown in results
    for r in ref_log:


      _log(f"Ref doc '{r['file']}' classified as: {r['type']}" +
         (f" ({r.get('rows',0)} rows)" if r.get("rows") else f" -- {r.get('detail','')}"),
         level="WARN" if r["type"] == "error" else "INFO")
    _log(f"Reference docs → DataDict: {len(data_dict)} columns | Rules: {len(rules)} | Mapping: {len(mapping_spec)} specs")


    manual_keys  = [k.strip() for k in key_columns.split(",")  if k.strip()] or None
    excluded_cols = [c.strip() for c in exclude_columns.split(",") if c.strip()]
    do_auto    = auto_keys == "on"


    # -- Merge saved Dataset Controls exclude rules into excluded_cols --------
    # Rules saved via Dataset Controls UI are stored under category
    # "dc_{action}_exclude" in feedback_rules.json, keyed by schema fingerprint.
    # They must be applied here so every fresh run respects them automatically.
    try:
      import json as _json_dc
      _fb_dc_path = _feedback_store_path
      if os.path.exists(_fb_dc_path):
        _dc_cat_prefix = f"dc_{action}_exclude"


        _all_dc: dict = {}
        with open(_fb_dc_path, encoding="utf-8") as _fb_dc:
          _all_dc = _json_dc.load(_fb_dc)
        # Build set of all columns present in the uploaded dataframes
        _all_df_cols = {c.lower(): c for df_tuple in dataframes for c in df_tuple[1].columns}
        for _fp_entry in _all_dc.values():
          for _r in _fp_entry.get("rules", []):
            if not _r.get("category", "").startswith(_dc_cat_prefix):
              continue
            _rule_text = _r.get("rule", "").lower()
            # Match any column name mentioned in the rule text
            for _col_lower, _col_actual in _all_df_cols.items():
              if len(_col_lower) >= 3 and _col_lower in _rule_text:
                if _col_actual not in excluded_cols:
                  excluded_cols.append(_col_actual)
      if excluded_cols:
        _log(f"Excluded columns (form + saved rules): {excluded_cols}")
    except Exception:


      pass

    # -- Large-file row cap --------------------------------
    # Analysis beyond this row count becomes very slow in a single-process sync
    # server; cap with a visible warning so the user knows results are sampled.
    _ANALYSIS_ROW_CAP = 500_000
    for _i, (_fn, _df) in enumerate(dataframes):
      if len(_df) > _ANALYSIS_ROW_CAP:
        dataframes[_i] = (_fn, _df.iloc[:_ANALYSIS_ROW_CAP].copy())
        _log(
          f"'{_fn}' has {len(_df):,} rows -- capped at {_ANALYSIS_ROW_CAP:,} for analysis. "
          "Full row count is preserved in export.",
          level="WARN",
        )


    pairs      = []
    nway_result = None
    fuzzy_result = None
    nway_fuzzy_result = None
    quality_reports  = []
    governance_reports = []


    mappings     = []
    parse_reports   = []
    lineage_reports  = []
    profile_reports  = []

    # ==== GAP: session_id was referenced by the per-file DQ-history save inside the
    # "quality" branch below before its original assignment point ("-- Session
    # storage --" further down); RECONSTRUCTED (unverified) by moving the
    # generation here so it is defined before first use in every branch; verify
    # against source if available.
    session_id = str(uuid.uuid4())

    # -- Per-action logic --
    if action == "compare" and fuzzy_fields and len(dataframes) > 2:
      # Opt-in probabilistic matching across 3+ sources -- picked up by the
      # response-building block below via nway_fuzzy_result.
      nway_fuzzy_result = await asyncio.to_thread(
        fuzzy_match_dataframes_nway, dataframes, fuzzy_fields, None, fuzzy_threshold, excluded_cols)

    elif action == "compare" and fuzzy_fields and len(dataframes) == 2:
      # Opt-in probabilistic matching when there's no clean shared key --
      # picked up by the response-building block below via fuzzy_result.
      fuzzy_result = await asyncio.to_thread(
        fuzzy_match_dataframes, dataframes[0][1], dataframes[1][1],
        fuzzy_fields, None, fuzzy_threshold, excluded_cols)

    elif action == "compare" and len(dataframes) > 2:
      # 3+ files uploaded -- N-way reconciliation instead of a pairwise diff.
      # nway_result is picked up by the response-building block below; pairs/
      # mappings stay empty for this action so the 2-file code path (and its
      # response shape) is completely untouched for the len==2 case.
      nway_result = await asyncio.to_thread(
        compare_dataframes_nway, dataframes, manual_keys, excluded_cols)

    elif action == "compare":
      # dataframes is always [(a_name, a_df), (b_name, b_df)] for compare/lineage
      (a_name, a_df), (b_name, b_df) = dataframes[0], dataframes[1]
      diff, mapping = await asyncio.gather(
        asyncio.to_thread(compare_dataframes, a_df, b_df, manual_keys, do_auto, excluded_cols,
                    user_hints=user_hints),
        asyncio.to_thread(analyze_mapping, a_df, b_df, a_name, b_name, mapping_spec,
                    user_hints=user_hints),
      )
      pairs.append({
        "file1_name":  a_name,
        "file2_name":  b_name,


        "file1_format": a_df.attrs.get("_format", ""),
        "file2_format": b_df.attrs.get("_format", ""),
        "file1_delimiter": a_df.attrs.get("_delimiter", ""),
        "file2_delimiter": b_df.attrs.get("_delimiter", ""),
        "diff":     diff,
        "mapping":    mapping,
      })


    elif action == "quality":
      # -- BFSI Rule Pack hint presets ----------------------------
      # Keyed by the bfsi_pack value sent from the UI dropdown.
      # Each entry supplies default hint values that are merged into user_hints
      # only when the user has not already supplied an explicit value for that field.
      _BFSI_PACK_HINTS = {
        "trade": {
          "key_hints": "trade_id",
          "nullable_hints": "remarks,comments,free_text",
          "timeliness_hints": "trade_date 1,settle_date 1,booking_date 1",
          "range_hints": "notional 0-1000000000000,price 0-1000000000",


          "bfsi_validators": ["isin_format:isin", "cusip_format:cusip", "lei_format:lei",
                    "positive:notional", "positive:price",
                    "allowed_values:side:BUY,SELL,B,S,BUY/SELL"],
        },
        "position": {
          "key_hints": "account_id,isin",
          "nullable_hints": "comments,benchmark,sector",
          "timeliness_hints": "valuation_date 1,as_of_date 1",
          "bfsi_validators": ["isin_format:isin", "positive:market_value",
                    "lei_format:lei", "currency_code_format:currency"],
        },
        "payments": {
          "key_hints": "transaction_id,uetr",
          "nullable_hints": "remittance_info,optional_ref,unstructured",
          "timeliness_hints": "value_date 1,settlement_date 1,created_date 1",
          "bfsi_validators": ["bic_format:bic", "bic_format:debtor_bic", "bic_format:creditor_bic",
                    "iban_format:iban", "iban_format:debtor_iban", "iban_format:creditor_iban",
                    "currency_code_format:currency", "positive:amount",


                    "positive:instructed_amount"],
        },
        "refdata": {
          "key_hints": "isin,cusip,sedol",
          "nullable_hints": "alias,local_code,notes,description",
          "bfsi_validators": ["isin_format:isin", "cusip_format:cusip", "sedol_format:sedol",
                    "lei_format:lei", "currency_code_format:currency",
                    "mic_format:exchange_mic", "mic_format:venue_mic"],
        },
        "mifid2": {
          "key_hints": "uti,trade_id",
          "nullable_hints": "waiver_indicator,post_trade_flag,commodity_derivative_indicator",
          "timeliness_hints": "trade_date_time 1,reporting_date 1,execution_timestamp 1",
          "bfsi_validators": ["lei_format:lei", "lei_format:counterparty_lei",
                    "isin_format:isin", "mic_format:venue_mic",
                    "positive:notional", "positive:price",
                    "currency_code_format:currency"],


        },
        "risk": {
          "key_hints": "book_id,risk_factor",
          "nullable_hints": "comments,override_reason,notes",
          "timeliness_hints": "as_of_date 1,run_date 1,value_date 1",
          "bfsi_validators": ["currency_code_format:currency",
                    "positive:exposure", "positive:var"],
        },
      }


      # Build cross-file ref dict -- every other file can be used as a reference
      _cross_file_map = {fname: df for fname, df in dataframes}
      # For merged analysis, pair up files: first file gets second as df2 (and vice versa)
      _df_map = {fname: df for fname, df in dataframes}
      _fnames = [fname for fname, _ in dataframes]
      for fname, df in dataframes:
        _log(f"Running Data Intelligence (DQ + Profile + Governance) on '{fname}'")
        _hints = dict(user_hints) if user_hints else {}


        _hints["cross_file_ref_data"] = {k: v for k, v in _cross_file_map.items() if k != fname}
        # Load saved baseline if exists
        _safe_fname = re.sub(r'[^\w\-.]', '_', fname)
        _baseline_path = Path("workspace") / "dq_baselines" / f"{_safe_fname}.json"
        if _baseline_path.exists():
          try:
            import json as _json_bl
            _hints["dq_baseline"] = _json_bl.loads(_baseline_path.read_text(encoding="utf-8"))
          except Exception:
            pass
        # Auto-load saved DQ rules for this schema if user didn't supply col_config
        if not _hints.get("col_config"):
          try:
            _fp_val = _dq_schema_fingerprint(df)
            _rules_path = Path("workspace") / "dq_rules" / f"{_fp_val}.json"
            if _rules_path.exists():
              _saved_cfg = json.loads(_rules_path.read_text(encoding="utf-8"))
              if _saved_cfg.get("col_config"):


                _hints["col_config"] = _saved_cfg["col_config"]
                _log(f"Auto-loaded DQ rules for schema {_fp_val}")
          except Exception as _e_fp:
            pass


        # -- Merge BFSI Rule Pack hints (only fills blanks, never overrides user input) --
        if _bfsi_pack and _bfsi_pack in _BFSI_PACK_HINTS:
          pack_cfg = _BFSI_PACK_HINTS[_bfsi_pack]
          if not _hints.get("timeliness_hints") and pack_cfg.get("timeliness_hints"):
            _hints["timeliness_hints"] = pack_cfg["timeliness_hints"]
          if not _hints.get("nullable_hints") and pack_cfg.get("nullable_hints"):
            _hints["nullable_hints"] = pack_cfg["nullable_hints"]
          if pack_cfg.get("bfsi_validators"):
            _hints["bfsi_validators"] = pack_cfg["bfsi_validators"]


        # -- Auto-inject domain accuracy rules from column names ----------------
        # Detects columns like "currency", "ccy", "country_code", "side" etc.
        # and injects domain_accuracy validators automatically.
        _domain_col_map = {


          "currency_code": ["currency","ccy","base_currency","quote_currency",
                    "settlement_currency","reporting_currency","traded_ccy"],
          "country_code": ["country","country_code","domicile","nationality",
                   "issuer_country","country_of_risk"],
          "asset_class":  ["asset_class","asset_type","instrument_class"],
          "side":     ["side","direction","buy_sell","buysell"],
          "trade_type":  ["trade_type","transaction_type","product_type"],
          "settlement_type": ["settlement_type","setl_type","delivery_type"],
        }
        _auto_domain_validators = list(_hints.get("bfsi_validators", []))
        _existing_v_cols = {v.split(":")[1].lower() for v in _auto_domain_validators if ":" in v}

        for domain, col_hints in _domain_col_map.items():
          for col in df.columns:
            if col.lower() in col_hints and col.lower() not in _existing_v_cols:
              _auto_domain_validators.append(f"domain_accuracy_{domain}:{col}")
              _existing_v_cols.add(col.lower())
        if _auto_domain_validators:
          _hints["bfsi_validators"] = _auto_domain_validators


        # -- Auto-inject BFSI temporal consistency rules ----------------
        # Detect trade_date/settle_date pairs and inject T+2 validation.
        # Also inject quantity>0 when side=BUY, notional>0 always.
        _auto_cross = list(_hints.get("cross_column_rules", "").split(";"))
        _auto_cross = [x for x in _auto_cross if x.strip()]
        _df_cols_lower = {c.lower(): c for c in df.columns}


        # trade_date < settle_date
        _td = next((v for k, v in _df_cols_lower.items() if "trade_date" in k or k == "trd_dt"), None)
        _sd = next((v for k, v in _df_cols_lower.items() if "settle_date" in k or "settlement_date" in k or k == "setl_dt"), None)
        if _td and _sd and not any("settle" in r and "trade" in r for r in _auto_cross):
          _hints["bfsi_validators"] = list(_hints.get("bfsi_validators", [])) + [
            f"not_future_date:{_td}"
          ]


        # notional > 0


        _notional = next((v for k, v in _df_cols_lower.items() if k in
          ("notional","notional_amount","face_value","face_amount")), None)
        if _notional:
          _existing_v = {v for v in _hints.get("bfsi_validators", [])}
          if f"positive:{_notional}" not in _existing_v:
            _hints.setdefault("bfsi_validators", [])
            _hints["bfsi_validators"] = list(_hints["bfsi_validators"]) + [f"positive:{_notional}"]


        # -- Auto-inject address completeness for known address columns ----
        _addr_cols = [v for k, v in _df_cols_lower.items()
              if any(h in k for h in ("address","street","addr","postal"))]
        if _addr_cols:
          _addr_v = list(_hints.get("bfsi_validators", []))
          for _ac in _addr_cols[:3]:
            if f"address_complete:{_ac}" not in _addr_v:
              _addr_v.append(f"address_complete:{_ac}")
          _hints["bfsi_validators"] = _addr_v


        # Pair with the other file for cross-file mapping recon (if two files uploaded)


        _other_fname = next((n for n in _fnames if n != fname), None)
        _df2 = _df_map.get(_other_fname) if _other_fname else None


      # -- Merged single call: DQ + Profile + Governance + Mapping recon --
        q = await asyncio.to_thread(
          analyze_quality_full,
          df, fname,
          data_dict, rules,
          _hints,
          _df2, _other_fname,
        )
        q["file_format"] = df.attrs.get("_format", "")


      # -- ML Anomaly detection -- z-score outliers across numeric columns --
        try:
          q["anomaly_results"] = await asyncio.to_thread(
            detect_numeric_anomalies, df, 3.0
          )
        except Exception:


          q["anomaly_results"] = []

        # -- Categorical distribution clustering (Ataccama-style) --------
        try:
          _cat_baseline = q.get("baseline_snapshot")
          q["categorical_drift"] = await asyncio.to_thread(
            detect_categorical_drift, df, _cat_baseline
          )
        except Exception:
          q["categorical_drift"] = []

        # -- Unsupervised numeric clustering --------------------
        try:
          q["numeric_clusters"] = await asyncio.to_thread(
            detect_numeric_clusters, df
          )
        except Exception:
          q["numeric_clusters"] = []

        # -- Multi-variate (Mahalanobis) + trend/time-aware anomaly detection --
        try:
          q["multivariate_anomalies"] = await asyncio.to_thread(
            detect_multivariate_anomalies, df
          )
        except Exception:
          q["multivariate_anomalies"] = {"checked_columns": [], "anomaly_count": 0, "rows": []}
        try:
          q["trend_anomalies"] = await asyncio.to_thread(
            detect_trend_anomalies, df
          )
        except Exception:
          q["trend_anomalies"] = []


        # -- AI-Enhanced: LLM suggests additional rules, re-run with them ------
        if _di_ai_enhanced and "quality" in _di_scope:
          try:
            _log(f"AI-Enhanced: calling LLM for additional DQ rules on '{fname}'")

            # Build compact column profile for LLM
            _col_summaries = []
            for _col in df.columns:
              _s = df[_col]
              _cs = {
                "name":   _col,
                "dtype":  str(_s.dtype),
                "null_pct": round(float(_s.isna().mean() * 100), 1),
                "unique_n": int(_s.nunique(dropna=True)),
                "total":  len(_s),
                "sample":  _s.dropna().astype(str).head(6).tolist(),
              }
              if pd.api.types.is_numeric_dtype(_s):
                _clean = _s.dropna()


                if len(_clean):
                  _cs["min"] = round(float(_clean.min()), 4)
                  _cs["max"] = round(float(_clean.max()), 4)
                  _cs["mean"] = round(float(_clean.mean()), 4)
                _col_summaries.append(_cs)

                # Include existing failures as context
                _existing_fails = [
                  {"rule": r.get("rule_name",""), "column": r.get("column_name",""), "failed": r.get("failed",0)}
                  for r in q["rule_results"] if r.get("status") == "FAIL"
                ][:10]

                _sys = (
                  "You are a senior BFSI data quality engineer. "
                  "You are given column profiles and existing rule failures for a dataset. "
                  "Suggest ADDITIONAL data quality rules that go beyond what was already checked. "
                  "Focus on business meaning, cross-column relationships, BFSI-specific patterns. "


                  "Use only these rule_type values: range, allowed_values, pattern, positive, "
                  "non_negative, integer_only, not_future_date, isin_format, cusip_format, "
                  "sedol_format, lei_format, bic_format, iban_format, currency_code_format, "
                  "mic_format, email_format, date_format, freshness_days, decimal_places, "
                  "domain_accuracy, address_complete, uppercase, lowercase. "
                  "Return ONLY a valid JSON array. Each element: "
                  "{name, rule_type, value, reason, severity} where severity is critical/major/minor. "
                  "Return [] if no confident additional suggestions."
                )
                _usr = (
                  f"Dataset: {fname} ({len(df)} rows, {len(df.columns)} cols)\n"
                  f"Existing failures: {json.dumps(_existing_fails)}\n"
                  f"Column profiles: {json.dumps(_col_summaries, indent=1)}\n"
                  "Suggest additional rules. JSON array only."
                )


                _ai_raw = await asyncio.to_thread(
                  _ask_llm,
                  [{"role": "user", "content": [{"text": _usr}]}],
                  _sys,
                )
                _ai_match = re.search(r'\[.*\]', _ai_raw, re.DOTALL)
                _ai_suggestions = json.loads(_ai_match.group()) if _ai_match else []

                # Validate and inject into a fresh hints copy
                _ai_validators = []
                for _sg in _ai_suggestions:
                  if not isinstance(_sg, dict) or not _sg.get("name") or not _sg.get("rule_type"):
                    continue
                  _ai_validators.append({
                    "name":   str(_sg["name"]),
                    "rule_type": str(_sg["rule_type"]),
                    "value":  str(_sg.get("value", "") or ""),
                    "severity": str(_sg.get("severity", "major")),


                    "reason":  str(_sg.get("reason", "")),
                  })


                if _ai_validators:
                  _log(f"AI-Enhanced: LLM suggested {len(_ai_validators)} additional rules -- re-running DQ")
                  _ai_hints = dict(_hints)
                  _ai_col_config = list(_ai_hints.get("col_config", []))
                  for _av in _ai_validators:
                    _col_match = next((c for c in df.columns if c.lower() == _av["name"].lower()), None)
                    if _col_match:
                      _ai_col_config.append({
                        "name":   _col_match,
                        "rule_type": _av["rule_type"],
                        "value":  _av.get("value", ""),
                        "severity": _av.get("severity", "major"),
                        "_ai_generated": True,
                        "_ai_reason": _av.get("reason", ""),
                      })


                  _ai_hints["col_config"] = _ai_col_config

                  # Re-run with AI rules
                  q_ai = await asyncio.to_thread(
                    analyze_quality_full, df, fname, data_dict, rules, _ai_hints, _df2,
                    _other_fname
                  )
                  # Extract only the AI-generated rule results
                  q["ai_rule_results"] = [
                    {**r, "_ai_reason": next(
                      (_av.get("reason","") for _av in _ai_validators
                       if _av["name"].lower() == (r.get("column_name","") or "").lower()
                       and _av["rule_type"] == r.get("rule_type","")), ""
                    )}
                    for r in q_ai.get("rule_results", [])
                    if r.get("_hint_injected") and not r.get("_auto")
                    and any(
                      _av["name"].lower() == (r.get("column_name","") or "").lower()
                      and _av["rule_type"] == r.get("rule_type","")


                      for _av in _ai_validators
                    )
                  ]
                  # Update score with AI findings included
                  q["ai_dq_score"] = q_ai.get("dq_score", {})
                  _log(f"AI-Enhanced re-run complete: {len(q.get('ai_rule_results',[]))} AI rule results")
                else:
                  q["ai_rule_results"] = []
                  q["ai_dq_score"] = {}
                  _log("AI-Enhanced: LLM returned no additional rules")

          except Exception as _e_ai:
            import traceback as _tb
            q["ai_rule_results"] = []
            q["ai_dq_score"] = {}
            _log(f"AI-Enhanced rule generation failed: {_e_ai} | {_tb.format_exc()[-300:]}", "WARN")

        # NOTE (reconstruction): the source pages here duplicated this block twice
        # (a page-repeat artifact seen elsewhere in this file) -- kept the more
        # complete, correctly-indented occurrence and dropped the earlier duplicate.
        if "quality" in _di_scope:
          quality_reports.append(q)

        # Extract embedded governance and profile -- only add if scope requested
        _gov = q.get("governance")
        if _gov and "governance" in _di_scope:
          _gov["file_format"] = q["file_format"]
          governance_reports.append(_gov)

        _prof = q.get("profile")
        if _prof and "profile" in _di_scope:
          # Normalise to the shape analyze_profile() returns so template works
          _prof_report = {
            "file_name":      fname,
            "file_format":    q["file_format"],
            "total_rows":     q["total_rows"],
            "total_cols":     q.get("total_cols", len(df.columns)),
            "memory_mb":      _prof.get("memory_mb", 0),
            "duplicate_rows": q["duplicate_rows"],
            "key_candidates": _prof.get("key_candidates", []),
            "near_key_cols": _prof.get("near_key_cols", []),
            "type_breakdown": _prof.get("type_breakdown", {}),
            "correlations":  _prof.get("correlations", []),
            "numeric_cols":  sum(
              1 for c in _prof.get("columns", []) if c.get("is_numeric") or c.get("mean") is not None
            ),
            "columns":    _prof.get("columns", []),
          }
          profile_reports.append(_prof_report)

        _log(f"Data Intelligence '{fname}' → "
           f"Score {q['dq_score']['score']} ({q['dq_score']['grade']}) | "
           f"Gov penalty: {q['dq_score'].get('governance_penalty', 0)} | "
           f"Sensitivity: {_gov.get('overall_classification') if _gov else 'N/A'} | "
           f"Reg frameworks: {_gov.get('regulatory_frameworks', []) if _gov else []}")

        # Persist DQ score to history for trend tracking
        try:


          _rule_fails_h = sum(1 for r in q["rule_results"] if r.get("status") == "FAIL")
          _crit_fails_h = q["dq_score"].get("severity_breakdown", {}).get("critical_fails", 0)
          _ws_db.save_dq_history(
            _ws_username or "default", fname, q.get("schema_fingerprint", ""),
            q["dq_score"], q["total_rows"], _rule_fails_h, _crit_fails_h,
            session_id=session_id,
            bfsi_pack=_bfsi_pack,
            di_scope=",".join(sorted(_di_scope)),
          )
        except Exception as _e_hist:
          _log(f"DQ history save failed: {_e_hist}")


    elif action == "profile":
      for fname, df in dataframes:
        _log(f"Running Data Profile on '{fname}'")
        p = await asyncio.to_thread(analyze_profile, df, fname)
        p["file_format"] = df.attrs.get("_format", "")
        profile_reports.append(p)


        _log(f"Profile '{fname}' → {p['total_cols']} cols | "
           f"{p['numeric_cols']} numeric | {len(p['key_candidates'])} key candidates")


    elif action == "governance":
      for fname, df in dataframes:
        _log(f"Running Governance analysis on '{fname}'")
        # Load saved dc_governance_override rules -- scan ALL fingerprints
        # so column-level overrides apply regardless of which dataset saved them
        _gov_hints_sa = dict(user_hints or {})
        try:
          import json as _json_gov2
          _fb_path2 = _feedback_store_path
          _all2: dict = {}
          if os.path.exists(_fb_path2):
            with open(_fb_path2, encoding="utf-8") as _fb2:
              _all2 = _json_gov2.load(_fb2)
          _col_ov_sa: dict = {}
          _cm_sa = {c.lower(): c for c in df.columns}


          for _fpe in _all2.values():
            for _r in _fpe.get("rules", []):
              _cat2 = _r.get("category", "")
              if not _cat2.startswith("dc_governance_"):
                continue
              _t = _r.get("rule", "").lower()
              _ns = _parse_sensitivity_from_text(_t)
              _nr = _parse_regulatory_override(_t)
              _np = any(k in _t for k in ("not pii","no pii","false positive","not a","not personal"))
              _ex2 = _cat2.startswith("dc_governance_exclude") or any(k in _t for k in ("exclude","skip","ignore","omit"))
              for _cl, _ca in _cm_sa.items():
                if len(_cl) >= 3 and _cl in _t:
                  _col_ov_sa.setdefault(_ca, {})
                  if _ex2: _col_ov_sa[_ca]["exclude"]  = True
                  if _ns: _col_ov_sa[_ca]["sensitivity"] = _ns
                  if _np: _col_ov_sa[_ca]["clear_pii"] = True
                  if _nr is not None: _col_ov_sa[_ca]["regulatory"] = _nr
          if _col_ov_sa:


            _gov_hints_sa["_dc_col_overrides"] = _col_ov_sa
        except Exception:
          pass

        g = await asyncio.to_thread(analyze_governance, df, fname, data_dict,
                       user_hints=_gov_hints_sa)
        g["file_format"] = df.attrs.get("_format", "")
        governance_reports.append(g)
        _log(f"Governance '{fname}' → Classification: {g['overall_classification']} | "
           f"PII columns: {g['pii_column_count']} | Reg: {g['regulatory_frameworks']}")


    elif action == "parse":
      # Convert each uploaded file to a structured table.
      # For well-structured formats (JSON, XML, SWIFT, FIX, CSV, Excel, Parquet)
      # _load_file() already produced a multi-column DataFrame -- use it directly.
      # Only fall back to LLM when the file is truly unstructured (single raw-text column).
      for fname, df in dataframes:
        fmt = df.attrs.get("_format", "auto-detected")
        _log(f"Parsing '{fname}' [format: {fmt}] → {len(df)} rows × {len(df.columns)} cols")


        if len(df.columns) > 1:
          # Well-structured: DataFrame already has meaningful columns
          columns = list(df.columns)
          # Cap at 200 rows for display; keeps browser responsive
          rows = (
            df.head(200)
            .astype(object)
            .fillna("")
            .to_dict("records")
          )
          result = {
            "file_name":  fname,
            "columns":    columns,
            "rows":     rows,
            "row_count":  len(df),
            "col_count":  len(columns),
            "format_detected": fmt,
            "notes":     (
              f"Parsed by built-in {fmt} loader."


              + (f" Showing first 1000 of {len(df)} rows." if len(df) > 1000 else "")
            ),
            "error":     None,
          }
          _log(f"Parse '{fname}' → {len(df)} rows × {len(columns)} cols (built-in loader)")
        else:
          # Truly unstructured (single column or loading produced a flat blob) -- use LLM
          _log(f"Single-column result for '{fname}' -- falling back to LLM parser")
          raw = raw_file_bytes.get(fname, b"")
          result = await asyncio.to_thread(parse_unstructured, raw, fname, hints=user_hints)
          if result["error"]:
            _log(f"LLM parse '{fname}' → error: {result['error']}", level="WARN")
          else:
            _log(
              f"LLM parse '{fname}' → {result['row_count']} rows × {result['col_count']} cols"
              f" | Format: {result['format_detected']}"
            )


        parse_reports.append(result)


    elif action in ("quality_ai", "profile_ai", "governance_ai"):
      # AI Copilot mode for DQ / Profile / Governance.
      # Same pattern as lineage: load schema, open Copilot, no analysis yet.
      # User chats, then says "run quality/profile/governance" to execute.
      fname, df = dataframes[0]


      def _ai_col_profile(df: pd.DataFrame) -> list[dict]:
        rows = []
        for col in df.columns:
          s = df[col]
          entry = {
            "column":  col,
            "dtype":  str(s.dtype),
            "null_pct": round(float(s.isna().mean() * 100), 1),
            "unique_n": int(s.nunique(dropna=True)),
            "sample":  s.dropna().astype(str).head(5).tolist(),
          }


          if pd.api.types.is_numeric_dtype(s):
            clean = s.dropna()
            if len(clean):
              entry["min"] = round(float(clean.min()), 4)
              entry["max"] = round(float(clean.max()), 4)
              entry["mean"] = round(float(clean.mean()), 4)
          rows.append(entry)
        return rows


      _module_label = {"quality_ai": "Data Quality", "profile_ai": "Data Profile",
               "governance_ai": "Governance"}[action]
      lineage_reports.append({
        "src_name":   fname,
        "tgt_name":   fname,
        "src_schema":  _ai_col_profile(df),
        "tgt_schema":  [],
        "src_rows":   len(df),
        "tgt_rows":   0,
        "_dq_ai_mode":  True,


        "_ai_module_label": _module_label,
      })
      _log(f"{_module_label} AI Mode: '{fname}' ({len(df)} rows × {len(df.columns)} cols)"
         " -- AI Copilot ready")


    elif action == "lineage":
      # Customize/Complex Compare -- files are loaded; AI Copilot drives the recon.
      # We skip the lineage engine entirely and hand the schema + saved rules to the
      # Copilot so the user can define mappings, transforms, and rules in chat.
      (src_name, src_df), (tgt_name, tgt_df) = dataframes[0], dataframes[1]


      def _schema_summary(df: pd.DataFrame) -> list[dict]:
        rows = []
        for col in df.columns:
          sample = df[col].dropna().astype(str).head(5).tolist()
          rows.append({"column": col, "dtype": str(df[col].dtype), "sample": sample})
        return rows


      lineage_reports.append({


        "src_name":  src_name,
        "tgt_name":  tgt_name,
        "src_schema": _schema_summary(src_df),
        "tgt_schema": _schema_summary(tgt_df),
        "src_rows":  len(src_df),
        "tgt_rows":  len(tgt_df),
      })
      _log(f"Customize/Complex Compare: '{src_name}' ({len(src_df)} rows) ↔ "
         f"'{tgt_name}' ({len(tgt_df)} rows) -- AI Copilot ready")


    else:
      raise HTTPException(400, f"Unknown action: {action}")


    total_elapsed = round(time.time() - _t0, 3)
    _log(f"Analysis complete in {total_elapsed}s")


    # -- Dataset Memory fingerprint --------------------------------
    # Lookup priority: file-name match → exact schema → fuzzy schema.


    # Rules are filtered by module (action) so recon rules never bleed into DQ.
    try:
      _dm_username = _ws_resolve_username(request) or "default"
    except Exception:
      _dm_username = "default"
    _all_cols = [list(df.columns) for _, df in dataframes]
    _cols1 = _all_cols[0] if len(_all_cols) > 0 else []
    _cols2 = _all_cols[1] if len(_all_cols) > 1 else []
    _file_names = [n for n, _ in dataframes]
    _dataset_fingerprint = _fp_resolve(
      _dm_username,
      _fp_compute(_cols1, _cols2),
      cols1=_cols1, cols2=_cols2,
      file_names=_file_names,
    )

    _saved_rules_text = _fp_rules_text(_dm_username, _dataset_fingerprint, module=action)


    # -- Persist recon hints for this schema so next run auto-reloads them ------
    # Saved as category "recon_hints" so they are distinct from user correction rules.
    if action == "lineage" and user_hints:
      _hints_to_save = {k: v for k, v in user_hints.items() if v}
      if _hints_to_save:
        _fp_save(
          _dm_username,
          _dataset_fingerprint,


          json.dumps(_hints_to_save),
          category="recon_hints",
          dataset_label=f"{dataframes[0][0]} vs {dataframes[1][0]}" if len(dataframes) >= 2 else "",
          cols1=_cols1,
          cols2=_cols2,
          file_names=_file_names,
        )


    # -- Session storage --
    _recon_ctx: dict = {}
    if action in ("lineage", "compare"):
      # Always set mode=recon so /recon/run/ can serve the session even when
      # lineage_reports is empty (e.g. first upload with no rules yet saved), and
      # so the reconciliation chat/"remember:" rule capture works whether the
      # session came from Standard mode's compare or AI Copilot's compare.
      _lr = lineage_reports[0] if lineage_reports else {}
      src_name_ctx, tgt_name_ctx = (dataframes[0][0], dataframes[1][0]) if len(dataframes) >= 2 else ("", "")

      def _chat_schema_summary(df: pd.DataFrame) -> list[dict]:
        return [{"column": col, "dtype": str(df[col].dtype),
                  "sample": df[col].dropna().astype(str).head(5).tolist()}
                for col in df.columns]

      _recon_ctx = {
        "mode": "recon",
        "src_name":  _lr.get("src_name", src_name_ctx),
        "tgt_name":  _lr.get("tgt_name", tgt_name_ctx),
        "src_rows":  _lr.get("src_rows",  len(dataframes[0][1]) if dataframes else 0),
        "tgt_rows":  _lr.get("tgt_rows",  len(dataframes[1][1]) if len(dataframes) > 1 else 0),
        "src_schema": _lr.get("src_schema") or (_chat_schema_summary(dataframes[0][1]) if dataframes else []),
        "tgt_schema": _lr.get("tgt_schema") or (_chat_schema_summary(dataframes[1][1]) if len(dataframes) > 1 else []),
        "saved_rules": _fp_get_rules(_dm_username, _dataset_fingerprint, cols1=_cols1, cols2=_cols2,
                      file_names=_file_names, module="recon"),
        "dataset_fingerprint": _dataset_fingerprint,
      }

    _chat_contexts[session_id] = {
      "action":    action,
      "files":    [n for n, _ in dataframes],
      "has_data_dict": bool(data_dict),
      "has_rules":  bool(rules),
      "has_mapping_spec": bool(mapping_spec),
      "comparisons": [
        {"pair": f"{p['file1_name']}→{p['file2_name']}",
         "added": p["diff"]["added_count"],


         "removed": p["diff"]["removed_count"],
         "modified": p["diff"]["modified_count"]}
        for p in pairs
      ],
      "quality": [
        {"file": q["file_name"], "score": q["dq_score"]["score"],
         "grade": q["dq_score"]["grade"],
         "rule_fails": sum(1 for r in q["rule_results"] if r.get("status") == "FAIL")}
        for q in quality_reports
      ],
      "governance": [
        {"file": g["file_name"], "classification": g["overall_classification"],
         "pii_cols": g["pii_column_count"], "regulatory": g["regulatory_frameworks"]}
        for g in governance_reports
      ],
      "mappings": [
        {"pair": f"{m['file1_name']}→{m['file2_name']}",
         "completeness": m["mapping_completeness_pct"]}
        for m in mappings


      ],
      **_recon_ctx,
    }

    # -- Session quality cache -- for baseline save endpoint ----------
    if quality_reports:
      if session_id not in _session_quality_cache:
        _session_quality_cache[session_id] = {}
      for _qr in quality_reports:
        _session_quality_cache[session_id][_qr["file_name"]] = _qr


    # Full results for download/email
    _results_store[session_id] = {
      "action":     action,
      "file_names":   [n for n, _ in dataframes],
      "ref_log":     ref_log,
      "pairs":      pairs,
      "quality_reports": quality_reports,
      "governance_reports": governance_reports,
      "mappings":    mappings,
      # Which saved connector IDs (if any) fed this run -- needed so "Save to
      # Workspace" can offer to turn this exact run into a recurring Job
      # without the user re-picking connections from scratch. Empty when the
      # session used file uploads instead, which POST /api/ws/jobs/from-session
      # checks for and rejects with a clear message (a schedule needs a
      # connector to re-fetch from, not a one-off uploaded file).
      "conn_a_ids": _conn_a_ids, "conn_b_ids": _conn_b_ids, "conn_ids": _conn_ids,


      "parse_reports":  parse_reports,
      "lineage_reports": lineage_reports,
      "profile_reports": profile_reports,
      "proc_logs":    proc_logs,
      "elapsed":     total_elapsed,
      # Store DataFrames so /rerun can re-compare without re-uploading
      "dataframes": [{"name": n, "df": df} for n, df in dataframes],
      "excluded_cols": excluded_cols,
      "key_columns":  key_columns,
      "dataset_fingerprint": _dataset_fingerprint,
    }

    # -- Register files with LangChain agent (enables /agent-chat memory + tools) --
    # Save DataFrames to temp files so agent/tools.py can reload them by path.
    try:
      import tempfile as _tempfile
      from agent.tools import register_files as _lc_register_files
      from agent.rag import invalidate_store as _lc_invalidate_store


      _sess_dir = Path(_tempfile.gettempdir()) / "dva_main_sessions" / session_id
      _sess_dir.mkdir(parents=True, exist_ok=True)

      _lc_f1_path = _lc_f1_name = _lc_f2_path = _lc_f2_name = None
      _lc_ref_paths: list[tuple[str, str]] = []

      for _idx, (_fname, _df) in enumerate(dataframes[:2]):
        # Use parquet for the session copy -- 10-50x faster than CSV for large frames
        _dest = _sess_dir / f"f{_idx + 1}_{Path(_fname).stem}.parquet"
        try:
          _df.to_parquet(_dest, index=False)
        except Exception:
          # Fall back to CSV if parquet serialisation fails (e.g. mixed types)
          _dest = _sess_dir / f"f{_idx + 1}_{_fname}"
          _df.to_csv(_dest, index=False)
        if _idx == 0:
          _lc_f1_path, _lc_f1_name = str(_dest), _fname
        else:
          _lc_f2_path, _lc_f2_name = str(_dest), _fname


      # Save reference docs too if they were uploaded
      for _ref in ref_uploads:
        try:
          _ref.file.seek(0)
          _ref_bytes = _ref.file.read()
          _ref_dest  = _sess_dir / f"ref_{_ref.filename}"
          _ref_dest.write_bytes(_ref_bytes)
          _lc_ref_paths.append((str(_ref_dest), _ref.filename))
        except Exception:
          pass

      _lc_register_files(
        session_id=session_id,
        file1_path=_lc_f1_path,
        file1_name=_lc_f1_name,
        file2_path=_lc_f2_path,
        file2_name=_lc_f2_name,
        ref_paths=_lc_ref_paths,


        dataset_fingerprint=_dataset_fingerprint,
      )
      _lc_invalidate_store(session_id)
    except Exception as _lc_err:
      pass  # Non-fatal -- structured report already rendered; agent just won't have file paths

    # Deduplicated union of all columns across loaded files (for rerun panel)
    all_file_columns = sorted(set(c for _, df in dataframes for c in df.columns))
    _resolved_fingerprint = _fp_resolve(_dm_username, _dataset_fingerprint, cols1=_cols1,
                                        cols2=_cols2, file_names=_file_names)

    # -- Build the JSON payload the frontend's fetch()-based UI expects for
    # each action. (The full session data above is still kept in
    # _results_store/_chat_contexts for downloads, chat, and reruns.)
    if action == "compare" and nway_fuzzy_result is not None:
        r = nway_fuzzy_result
        _resp = {
            "session_id": session_id,
            "nway": True,
            "fuzzy": True,
            "fuzzy_fields": r["fuzzy_fields"],
            "threshold": r["threshold"],
            "sources": r["sources"],
            "counts": r["counts"],
            "singleton_counts": r["singleton_counts"],
            "partial_rows": r["partial_rows"],
            "value_breaks": r["value_breaks"],
            "top_broken_columns": r["top_broken_columns"],
            "elapsed": total_elapsed,
            "logs": proc_logs,
        }
        if session_id in _results_store:
            _results_store[session_id]["_digest"] = _resp
        return JSONResponse(_sanitize_json(_resp))

    if action == "compare" and fuzzy_result is not None:
        (_ff1_name, _ff1_df), (_ff2_name, _ff2_df) = dataframes[0], dataframes[1]
        r = fuzzy_result
        _resp = {
            "session_id": session_id,
            "fuzzy": True,
            "fuzzy_fields": r["fuzzy_fields"],
            "threshold": r["threshold"],
            "counts": r["counts"],
            "files": {
                "file1": {"name": _ff1_name, "rows": len(_ff1_df), "columns": len(_ff1_df.columns)},
                "file2": {"name": _ff2_name, "rows": len(_ff2_df), "columns": len(_ff2_df.columns)},
            },
            "matched_rows": r["matched_rows"],
            "modified_rows": r["modified_rows"],
            "only_in_file1": r["file1_only"],
            "only_in_file2": r["file2_only"],
            "elapsed": total_elapsed,
            "logs": proc_logs,
        }
        if session_id in _results_store:
            _results_store[session_id]["_digest"] = _resp
        return JSONResponse(_sanitize_json(_resp))

    if action == "compare" and nway_result is not None:
        r = nway_result
        _resp = {
            "session_id": session_id,
            "nway": True,
            "sources": r["sources"],
            "counts": r["counts"],
            "keys": r["keys"],
            "method": r["method"],
            "fingerprint": _resolved_fingerprint,
            "common_columns_list": r["common_columns"],
            "excluded_meta_cols": r["excluded_meta_cols"],
            "duplicate_counts": r["duplicate_counts"],
            "singleton_counts": r["singleton_counts"],
            "partial_rows": [
                {"key": ", ".join(f"{k}={v}" for k, v in pr["key_values"].items()),
                 "present_in": pr["present_in"], "missing_from": pr["missing_from"]}
                for pr in r["partial_rows"]
            ],
            "value_breaks": [
                {"key": ", ".join(f"{k}={v}" for k, v in vb["key_values"].items()),
                 "changes": vb["changes"]}
                for vb in r["value_breaks"]
            ],
            "top_broken_columns": r["top_broken_columns"],
            "elapsed": total_elapsed,
            "logs": proc_logs,
        }
        if session_id in _results_store:
            _results_store[session_id]["_digest"] = _resp
        return JSONResponse(_sanitize_json(_resp))

    if action == "compare":
        diff = pairs[0]["diff"]
        modified = [
            {"key": ", ".join(f"{k}={v}" for k, v in mr["key_values"].items()),
             "changes": mr["changes"]}
            for mr in diff.get("modified_rows", [])
        ]
        _only1 = [
            {"key": ", ".join(f"{k}={v}" for k, v in r["key_values"].items()), "row": r.get("row_data", {})}
            for r in diff.get("file1_only", [])
        ]
        _only2 = [
            {"key": ", ".join(f"{k}={v}" for k, v in r["key_values"].items()), "row": r.get("row_data", {})}
            for r in diff.get("file2_only", [])
        ]
        (_f1_name, _f1_df), (_f2_name, _f2_df) = dataframes[0], dataframes[1]
        _resp = {
            "session_id": session_id,
            "counts": {
                "matched": diff.get("file1_rows", 0) - diff.get("removed_count", 0),
                "file1_only": diff.get("file1_only_count", 0),
                "file2_only": diff.get("file2_only_count", 0),
                "modified": diff.get("modified_count", 0),
            },
            "keys": diff.get("key_columns", []),
            "method": diff.get("key_method", ""),
            "fingerprint": _resolved_fingerprint,
            "type_mismatches": [{"column": k, **v} for k, v in diff.get("type_mismatches", {}).items()],
            "null_column_exceptions": diff.get("null_column_exceptions", []),
            "duplicates": {
                "file1": {"duplicate_rows": diff.get("file1_duplicate_count", 0), "rows": diff.get("file1_duplicate_rows", [])},
                "file2": {"duplicate_rows": diff.get("file2_duplicate_count", 0), "rows": diff.get("file2_duplicate_rows", [])},
            },
            "modified": modified,
            "only_in_file1": _only1,
            "only_in_file2": _only2,
            "schema_added_columns": diff.get("schema_added_columns", []),
            "schema_removed_columns": diff.get("schema_removed_columns", []),
            "excluded_meta_cols": diff.get("excluded_meta_cols", []),
            "waterfall": diff.get("waterfall", {}),
            "gross_break_total": diff.get("gross_break_total", 0),
            "net_break_total": diff.get("net_break_total", 0),
            "files": {
                "file1": {"name": _f1_name, "rows": len(_f1_df), "columns": len(_f1_df.columns),
                          "format": str(_f1_df.attrs.get("_format", "") or "").upper(),
                          "all_columns": list(_f1_df.columns)},
                "file2": {"name": _f2_name, "rows": len(_f2_df), "columns": len(_f2_df.columns),
                          "format": str(_f2_df.attrs.get("_format", "") or "").upper(),
                          "all_columns": list(_f2_df.columns)},
            },
            "common_columns": len(set(_f1_df.columns) & set(_f2_df.columns)),
            "common_columns_list": diff.get("common_columns", sorted(set(_f1_df.columns) & set(_f2_df.columns))),
            "elapsed": total_elapsed,
            "logs": proc_logs,
        }
        if session_id in _results_store:
            _results_store[session_id]["_digest"] = _resp
        if _WS_ENABLED:
            try:
                _ws_db.save_recon_history(
                    _ws_username or "default", f"{_f1_name} vs {_f2_name}", _resolved_fingerprint,
                    session_id, _resp["counts"]["matched"], _resp["counts"]["file1_only"],
                    _resp["counts"]["file2_only"], _resp["counts"]["modified"], method=_resp["method"],
                )
            except Exception as _e_hist:
                _log(f"Reconciliation history save failed: {_e_hist}")
        return JSONResponse(_sanitize_json(_resp))

    if action in ("quality", "profile"):
        q = quality_reports[0] if quality_reports else (profile_reports[0] if profile_reports else {})
        dq = q.get("dq_score", {})
        profile = q.get("profile", {}) or {}
        _profile_by_name = {c.get("name"): c for c in profile.get("columns", [])}

        dims = {
            "Completeness": dq.get("completeness"),
            "Uniqueness": dq.get("uniqueness"),
            "Validity": dq.get("validity"),
            "Consistency": dq.get("consistency"),
            "Conformity": dq.get("conformity"),
        }
        if dq.get("precision_active"):
            dims["Precision"] = dq.get("precision")
        if dq.get("timeliness_active"):
            dims["Timeliness"] = dq.get("timeliness")
        if dq.get("accuracy_active"):
            dims["Accuracy"] = dq.get("accuracy")

        columns_detail = []
        for c in q.get("columns", []):
            is_numeric = "min" in c and "max" in c
            prof_c = _profile_by_name.get(c.get("name"), {})
            columns_detail.append({
                "column": c.get("name"),
                "completeness_pct": round(100 - (c.get("null_pct") or 0), 1),
                "uniqueness_pct": c.get("uniqueness_pct"),
                "null_count": c.get("null_count"),
                "distinct": c.get("unique_count"),
                "detected_format": prof_c.get("semantic", c.get("cardinality", "")),
                "numeric": {"min": c["min"], "max": c["max"]} if is_numeric else False,
                "string": False,
            })

        correlations = [
            {"col_a": c.get("col1"), "col_b": c.get("col2"), "r": c.get("corr")}
            for c in profile.get("correlations", [])
        ]

        _resp = {
            "session_id": session_id,
            "name": q.get("file_name", ""),
            "dimensions": dims,
            "grade": dq.get("grade"),
            "score": dq.get("score"),
            "rows": q.get("total_rows"),
            "columns": q.get("total_cols"),
            "duplicate_rows": q.get("duplicate_rows"),
            "columns_detail": columns_detail,
            "near_key_columns": profile.get("near_key_cols", []),
            "correlations": correlations,
            "ai_rule_results": q.get("ai_rule_results", []),
            "ai_dq_score": q.get("ai_dq_score", {}),
            "rule_results": q.get("rule_results", []),
            "anomaly_results": q.get("anomaly_results", []),
            "categorical_drift": q.get("categorical_drift", []),
            "numeric_clusters": q.get("numeric_clusters", []),
            "multivariate_anomalies": q.get("multivariate_anomalies", {}),
            "trend_anomalies": q.get("trend_anomalies", []),
            "elapsed": total_elapsed,
            "logs": proc_logs,
        }
        if session_id in _results_store:
            _results_store[session_id]["_digest"] = _resp
        return JSONResponse(_sanitize_json(_resp))

    if action == "governance":
        g = governance_reports[0] if governance_reports else {}
        pii_col_count = g.get("pii_column_count", 0)
        breaches = g.get("mandatory_breaches", [])
        classification = str(g.get("overall_classification", "")).lower()
        if breaches:
            overall_risk = "critical"
        elif pii_col_count and classification in ("confidential", "restricted"):
            overall_risk = "high"
        elif pii_col_count:
            overall_risk = "medium"
        else:
            overall_risk = "low"

        columns = [
            {
                "name": c.get("column"),
                "pii_detected": bool(c.get("pii_detected")),
                "pii_types": c.get("pii_detected", []),
                "sensitivity": str(c.get("sensitivity", "")).lower(),
                "regulatory": c.get("regulatory", []),
                "steward": c.get("escalate_to", ""),
                "mandatory_breaches": c.get("mandatory_breaches", []),
                "conditional_warnings": c.get("conditional_warnings", []),
                "business_term": c.get("business_term", ""),
                "description": c.get("description", ""),
                "from_dict": bool(c.get("from_dict")),
                "access_rec": c.get("access_rec", ""),
            }
            for c in g.get("columns", [])
        ]
        recommendations = []
        if breaches:
            recommendations.append(f"{len(breaches)} mandatory field(s) have governance breaches -- review immediately.")
        if pii_col_count:
            recommendations.append(f"{pii_col_count} column(s) contain PII -- consider masking before sharing.")
        if g.get("regulatory_frameworks"):
            recommendations.append(f"Applicable regulatory frameworks: {', '.join(g['regulatory_frameworks'])}.")
        if g.get("undocumented_columns"):
            recommendations.append(f"{len(g['undocumented_columns'])} column(s) have no data-dictionary entry.")

        _resp = {
            "session_id": session_id,
            "overall_risk": overall_risk,
            "overall_classification": g.get("overall_classification", ""),
            "regulatory_frameworks": g.get("regulatory_frameworks", []),
            "pii_column_count": pii_col_count,
            "bfsi_identifier_col_count": g.get("bfsi_identifier_col_count", 0),
            "mandatory_breaches": breaches,
            "conditional_warnings": g.get("conditional_warnings", []),
            "stewardship_routing": g.get("stewardship_routing", {}),
            "undocumented_columns": g.get("undocumented_columns", []),
            "has_data_dict": bool(g.get("has_data_dict")),
            "columns": columns,
            "recommendations": recommendations,
            "elapsed": total_elapsed,
            "logs": proc_logs,
        }
        if session_id in _results_store:
            _results_store[session_id]["_digest"] = _resp
        return JSONResponse(_sanitize_json(_resp))

    if action == "parse":
        result = parse_reports[0] if parse_reports else {"columns": [], "rows": [], "error": "No file parsed."}
        result = dict(result)
        result["session_id"] = session_id
        result["elapsed"] = total_elapsed
        result["logs"] = proc_logs
        if session_id in _results_store:
            _results_store[session_id]["_digest"] = result
        return JSONResponse(_sanitize_json(result))

    # lineage / quality_ai / profile_ai / governance_ai -- these modes are
    # driven by the AI Copilot chat panel rather than a one-shot results
    # table; return the session/schema info the chat needs to continue.
    return JSONResponse(_sanitize_json({
        "session_id": session_id,
        "action": action,
        "file_names": [n for n, _ in dataframes],
        "all_file_columns": all_file_columns,
        "dataset_fingerprint": _resolved_fingerprint,
        "saved_rules": _fp_get_rules(_dm_username, _dataset_fingerprint, cols1=_cols1,
                                     cols2=_cols2, file_names=_file_names, module=action),
        "lineage_reports": lineage_reports,
        "message": "This mode is driven by the AI Copilot chat panel -- use the chat sidebar to continue.",
    }))


def _load_saved_recon_hints(fingerprint: str, username: str = "default") -> dict:
    # Return the most recently saved recon_hints dict for this schema fingerprint.
  for rule in reversed(_fp_get_rules(username, fingerprint)):
    if rule.get("category") == "recon_hints":
      try:
        return json.loads(rule["rule"])
      except Exception:
        pass
  return {}


@app.post("/rerun/{session_id}")


async def rerun(session_id: str, request: Request):
    # Re-run a compare using stored DataFrames from a previous session.
    # Allows the user to change key_columns / exclude_columns without
    # re-uploading files.  A fresh session_id is generated for the new result
    # so downloads, chat and further reruns all work correctly.

    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        raise HTTPException(404, "Session expired -- please re-upload files.")

    try:
        form = await request.form()
    except Exception as exc:
        raise HTTPException(400, f"Could not parse form data: {exc}")

    key_columns    = str(form.get("key_columns",    ""))
    exclude_columns = str(form.get("exclude_columns", ""))

    manual_keys  = [k.strip() for k in key_columns.split(",")  if k.strip()] or None
    excluded_cols = [c.strip() for c in exclude_columns.split(",") if c.strip()]

    # Reconstruct list-of-tuples from stored dicts
    dataframes: list[tuple[str, object]] = [
        (item["name"], item["df"]) for item in stored["dataframes"]
    ]

    if len(dataframes) < 2:
        raise HTTPException(400, "Rerun requires at least 2 stored files.")

    base_name, base_df = dataframes[0]
    pairs = []
    mappings = []
    for cmp_name, cmp_df in dataframes[1:]:
        diff  = compare_dataframes(base_df, cmp_df, manual_keys, True, excluded_cols)
        mapping = analyze_mapping(base_df, cmp_df, base_name, cmp_name, None)
        pairs.append({
            "file1_name":  base_name,

            "file2_name":  cmp_name,
            "file1_format": base_df.attrs.get("_format", ""),
            "file2_format": cmp_df.attrs.get("_format", ""),
            "file1_delimiter": base_df.attrs.get("_delimiter", ""),
            "file2_delimiter": cmp_df.attrs.get("_delimiter", ""),
            "diff":     diff,
            "mapping":    mapping,
        })

    new_session_id = str(uuid.uuid4())

    # Build chat context for the new session
    _chat_contexts[new_session_id] = {
        "action": "compare",
        "files": [n for n, _ in dataframes],
        "has_data_dict": False,
        "has_rules": False,
        "has_mapping_spec": False,
        "comparisons": [

            {"pair": f"{p['file1_name']}→{p['file2_name']}",
             "added": p["diff"]["added_count"],
             "removed": p["diff"]["removed_count"],
             "modified": p["diff"]["modified_count"]}
            for p in pairs
        ],
        "quality": [],
        "governance": [],
        "mappings": [],
    }

    # Carry DataFrames into new session so further reruns are possible
    _results_store[new_session_id] = {
        "action": "compare",
        "file_names": [n for n, _ in dataframes],
        "ref_log": [],
        "pairs": pairs,
        "quality_reports": [],
        "governance_reports": [],

        "mappings": mappings,
        "parse_reports": [],
        "proc_logs": [],
        "elapsed": 0,
        "dataframes": stored["dataframes"],  # preserve for chained reruns
        "excluded_cols": excluded_cols,
        "key_columns": key_columns,
    }

    diff = pairs[0]["diff"]
    modified = [
        {"key": ", ".join(f"{k}={v}" for k, v in mr["key_values"].items()),
         "changes": mr["changes"]}
        for mr in diff.get("modified_rows", [])
    ]
    _only1 = [
        {"key": ", ".join(f"{k}={v}" for k, v in r["key_values"].items()), "row": r.get("row_data", {})}
        for r in diff.get("file1_only", [])
    ]
    _only2 = [
        {"key": ", ".join(f"{k}={v}" for k, v in r["key_values"].items()), "row": r.get("row_data", {})}
        for r in diff.get("file2_only", [])
    ]
    (_f1_name, _f1_df), (_f2_name, _f2_df) = dataframes[0], dataframes[1]
    _resp = {
        "session_id": new_session_id,
        "counts": {
            "matched": diff.get("file1_rows", 0) - diff.get("removed_count", 0),
            "file1_only": diff.get("file1_only_count", 0),
            "file2_only": diff.get("file2_only_count", 0),
            "modified": diff.get("modified_count", 0),
        },
        "keys": diff.get("key_columns", []),
        "method": diff.get("key_method", ""),
        "fingerprint": "",
        "type_mismatches": [{"column": k, **v} for k, v in diff.get("type_mismatches", {}).items()],
        "null_column_exceptions": diff.get("null_column_exceptions", []),
        "duplicates": {
            "file1": {"duplicate_rows": diff.get("file1_duplicate_count", 0), "rows": diff.get("file1_duplicate_rows", [])},
            "file2": {"duplicate_rows": diff.get("file2_duplicate_count", 0), "rows": diff.get("file2_duplicate_rows", [])},
        },
        "modified": modified,
        "only_in_file1": _only1,
        "only_in_file2": _only2,
        "schema_added_columns": diff.get("schema_added_columns", []),
        "schema_removed_columns": diff.get("schema_removed_columns", []),
        "excluded_meta_cols": diff.get("excluded_meta_cols", []),
        "waterfall": diff.get("waterfall", {}),
        "gross_break_total": diff.get("gross_break_total", 0),
        "net_break_total": diff.get("net_break_total", 0),
        "files": {
            "file1": {"name": _f1_name, "rows": len(_f1_df), "columns": len(_f1_df.columns),
                      "format": str(_f1_df.attrs.get("_format", "") or "").upper(),
                      "all_columns": list(_f1_df.columns)},
            "file2": {"name": _f2_name, "rows": len(_f2_df), "columns": len(_f2_df.columns),
                      "format": str(_f2_df.attrs.get("_format", "") or "").upper(),
                      "all_columns": list(_f2_df.columns)},
        },
        "common_columns": len(set(_f1_df.columns) & set(_f2_df.columns)),
        "common_columns_list": diff.get("common_columns", sorted(set(_f1_df.columns) & set(_f2_df.columns))),
        "elapsed": 0,
        "logs": [],
    }
    _results_store[new_session_id]["_digest"] = _resp
    return JSONResponse(_sanitize_json(_resp))


@app.get("/dq-ai-results/{session_id}", response_class=HTMLResponse)
async def dq_ai_results_get(session_id: str, request: Request):


  # GET endpoint that runs DQ on stored session data and returns the full report.
  # Browser navigates here after AI Copilot finishes -- no POST needed.

  import asyncio as _asyncio
  stored = _results_store.get(session_id)
  if not stored or "dataframes" not in stored:
    raise HTTPException(404, "Session not found or expired -- please re-upload your file.")

  dataframes = [(item["name"], item["df"]) for item in stored["dataframes"]]
  if not dataframes:
    raise HTTPException(400, "No data in session.")

  t0 = time.time()
  proc_logs: list[dict] = []
  def _log(msg, level="INFO"):
    proc_logs.append({"elapsed": round(time.time()-t0,3), "level": level, "message": msg})


  quality_reports = []
  governance_reports = []
  profile_reports = []

  for fname, df in dataframes:
    _log(f"Running DQ on '{fname}' ({len(df)} rows)")
    try:
      q = await _asyncio.to_thread(analyze_quality_full, df, fname, {}, [], {}, None, None)
      q["file_format"] = df.attrs.get("_format", "")
      quality_reports.append(q)
      _gov = q.get("governance")
      if _gov:
        _gov["file_format"] = q["file_format"]
        governance_reports.append(_gov)
      _prof = q.get("profile")
      if _prof:
        profile_reports.append({
          "file_name": fname, "file_format": q["file_format"],


          "total_rows": q["total_rows"],
          "total_cols": q.get("total_cols", len(df.columns)),
          "memory_mb": _prof.get("memory_mb", 0),
          "duplicate_rows": q["duplicate_rows"],
          "key_candidates": _prof.get("key_candidates", []),
          "near_key_cols": _prof.get("near_key_cols", []),
          "type_breakdown": _prof.get("type_breakdown", {}),
          "correlations": _prof.get("correlations", []),
          "numeric_cols": sum(1 for c in _prof.get("columns",[]) if c.get("is_numeric") or c.get("mean") is not None),
          "columns": _prof.get("columns", []),
        })
      try: q["anomaly_results"] = await _asyncio.to_thread(detect_numeric_anomalies, df, 3.0)
      except: q["anomaly_results"] = []
      try: q["categorical_drift"] = await _asyncio.to_thread(detect_categorical_drift, df)
      except: q["categorical_drift"] = []
      try: q["multivariate_anomalies"] = await _asyncio.to_thread(detect_multivariate_anomalies, df)
      except: q["multivariate_anomalies"] = {"checked_columns": [], "anomaly_count": 0, "rows": []}
      try: q["trend_anomalies"] = await _asyncio.to_thread(detect_trend_anomalies, df)
      except: q["trend_anomalies"] = []
      try: q["numeric_clusters"] = await _asyncio.to_thread(detect_numeric_clusters, df)


      except: q["numeric_clusters"] = []
      _log(f"Score: {q['dq_score']['score']} ({q['dq_score']['grade']})")
    except Exception as e:
      _log(f"Error: {e}", "WARN")


  new_sid = str(uuid.uuid4())
  all_cols = sorted(set(c for _, df in dataframes for c in df.columns))
  _results_store[new_sid] = {
    "action": "quality", "file_names": [n for n,_ in dataframes],
    "quality_reports": quality_reports, "governance_reports": governance_reports,
    "profile_reports": profile_reports, "dataframes": stored["dataframes"],
    "proc_logs": proc_logs, "elapsed": round(time.time()-t0,3),
    "pairs":[], "mappings":[], "parse_reports":[], "lineage_reports":[],
    "excluded_cols":[],
  }
  return templates.TemplateResponse(request=request, name="index.html",
    context={
      "action": "quality", "di_scope": ["quality","profile","governance"],
      "di_ai_enhanced": False, "file_names": [n for n,_ in dataframes],
      "pairs":[], "quality_reports": quality_reports,


      "governance_reports": governance_reports, "mappings":[], "parse_reports":[],
      "lineage_reports":[], "profile_reports": profile_reports,
      "has_data_dict": False, "has_rules": False, "has_mapping_spec": False,
      "ref_log":[], "proc_logs": proc_logs, "elapsed": round(time.time()-t0,3),
      "session_id": new_sid, "all_file_columns": all_cols,
      "excluded_cols":[], "key_columns_val":"", "dataset_fingerprint":"",
      "saved_rules":[],
    })


@app.post("/rerun-quality-json/{session_id}")
async def rerun_quality_json(session_id: str, request: Request):
    # Run DQ and return JSON summary for AI Copilot chat card rendering.
    # Mirrors /recon/run/{session_id} pattern -- chat renders the result inline.

    import asyncio as _asyncio
    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        raise HTTPException(404, "Session not found or expired -- please re-upload your file.")

    _hints: dict = {}
    try:
        # ==== GAP: the LLM-hint-generation prompt (nullable_hints / key_hints /
        # timeliness_hints / bfsi_validators inference from column names) was not
        # recoverable from the scan -- only the tail of the prompt string and the
        # JSON-parsing that follows survived (source pages 0781/0782 don't connect).
        # RECONSTRUCTED (unverified) below, matching the JSON schema keys the
        # existing parsing code expects; verify against source if available.
        _cols_qj = list(stored["dataframes"][0][1].columns) if stored.get("dataframes") else []
        _raw = await _asyncio.to_thread(_ask_llm, [{"role": "user", "content": [{"text":
            "You are a data quality hint generator. Given these column names, suggest "
            "helpful hints as a single JSON object with these optional keys: "
            f"columns={_cols_qj}. "
            '"nullable_hints" (comma-separated column names that may legitimately be blank), '
            '"key_hints" (comma-separated column names likely to form a unique row key), '
            '"timeliness_hints" (comma-separated "column_name max_age_days" pairs), '
            '"bfsi_validators":["positive:price","allowed_values:side:BUY,SELL"]}'
            "\nReturn {} if nothing specific."
        }]}],
        )
        import re as _re
        _m = _re.search(r'\{.*\}', _raw, _re.DOTALL)
        if _m:
            _ex = json.loads(_m.group(0))
            for _k in ("nullable_hints","key_hints","timeliness_hints","bfsi_validators"):
                if _ex.get(_k): _hints[_k] = _ex[_k]
    except Exception:
        pass

    # ==== GAP: the "dataframes" list (reconstructed from stored["dataframes"]) was not
    # recoverable from the scan at this point -- RECONSTRUCTED (unverified) below,
    # matching the same pattern used by the sibling /dq-ai-results/{session_id} and
    # /rerun-quality/{session_id} endpoints; verify against source if available.
    dataframes: list[tuple[str, object]] = [
        (item["name"], item["df"]) for item in stored["dataframes"]
    ]
    fname, df = dataframes[0]

    # Inject regex_format rules saved via Rule Builder into col_config
    try:
        _fp_qj = stored.get("dataset_fingerprint", "")
        _file_names_qj = stored.get("file_names", [])
        _rqj_username = _ws_resolve_username(request) or "default"

        _saved_qj = _fp_get_rules(_rqj_username, _fp_qj, cols1=list(df.columns),
                    file_names=_file_names_qj, module="quality")
        _regex_entries: list[dict] = []
        for _r in _saved_qj:
            _rt = _r.get("rule", "")
            if _rt.startswith("regex_format:"):
                _parts = _rt.split(":", 3)
                if len(_parts) >= 3:
                    _col = _parts[1].strip()
                    _pat = _parts[2].strip().split(" -- ")[0].strip()
                    _sev = "minor" if "minor" in (_r.get("category") or "") else \
                        "critical" if "critical" in (_r.get("category") or "") else "major"
                    _regex_entries.append({"name": _col, "rule_type": "regex_format",
                                "value": _pat, "severity": _sev})
        if _regex_entries:
            _hints["col_config"] = list(_hints.get("col_config") or []) + _regex_entries
    except Exception:
        pass

    try:
        q = await _asyncio.to_thread(analyze_quality_full, df, fname, {}, [], _hints, None, None)
        dq = q["dq_score"]

        # Run anomaly detection, clustering and categorical drift (same as main pipeline)
        try:
            q["anomaly_results"] = await _asyncio.to_thread(detect_numeric_anomalies, df, 3.0)
        except Exception:
            q["anomaly_results"] = []

        try:
            q["categorical_drift"] = await _asyncio.to_thread(detect_categorical_drift, df,
                                    q.get("baseline_snapshot"))
        except Exception:
            q["categorical_drift"] = []

        try:
            q["numeric_clusters"] = await _asyncio.to_thread(detect_numeric_clusters, df)
        except Exception:
            q["numeric_clusters"] = []

        try:
            q["multivariate_anomalies"] = await _asyncio.to_thread(detect_multivariate_anomalies, df)
        except Exception:
            q["multivariate_anomalies"] = {"checked_columns": [], "anomaly_count": 0, "rows": []}

        try:
            q["trend_anomalies"] = await _asyncio.to_thread(detect_trend_anomalies, df)
        except Exception:
            q["trend_anomalies"] = []

        fails = [r for r in q.get("rule_results", []) if r.get("status") == "FAIL"]
        warnings = [r for r in q.get("rule_results", []) if r.get("status") == "WARN"]
        anomalies = q.get("anomaly_results", [])
        cat_issues = [c for c in q.get("categorical_drift", []) if c.get("severity") in ("CRITICAL","WARN")]

        # Store for download/follow-up questions
        new_sid = str(uuid.uuid4())
        _results_store[new_sid] = {
            "action": "quality", "file_names": [fname],
            "quality_reports": [q], "governance_reports": [],
            "profile_reports": [], "dataframes": stored["dataframes"],
            "proc_logs": [], "elapsed": 0,
            "pairs":[], "mappings":[], "parse_reports":[], "lineage_reports":[], "excluded_cols":[],
        }
        if new_sid not in _chat_contexts:
            _chat_contexts[new_sid] = {"action":"quality","files":[fname],"quality":[{"file":fname,"score":dq["score"],"grade":dq["grade"],"rule_fails":len(fails)}]}

            # Column completeness details
            col_completeness = [
                {"name": c["name"], "null_pct": c.get("null_pct", 0), "dq_score":
    c.get("dq_score"), "dq_grade": c.get("dq_grade")}
                for c in q.get("columns", [])
            ]

            # Build set of AI-injected column+ruletype combos for tagging
            _ai_validator_keys = set()
            for _bv in _hints.get("bfsi_validators", []):
                if ":" in _bv:
                    _bv_rule, _bv_col = _bv.split(":", 1)
                    if _bv_rule.startswith("domain_accuracy_"):
                        _bv_rule = "domain_accuracy"
                    _ai_validator_keys.add((_bv_col.strip().lower(), _bv_rule.strip()))

            # All rule results -- tag AI-defined ones


            all_rules = [
                {
                    "rule":   r.get("rule_name") or r.get("description", ""),
                    "col":    r.get("column_name", ""),
                    "type":   r.get("rule_type", ""),
                    "status": r.get("status", ""),
                    "pass_pct": r.get("pass_pct", 100),
                    "failed": r.get("failed", 0),
                    "severity": r.get("severity", "major"),
                    "examples": (r.get("failing_examples") or [])[:3],
                    "ai_defined": (
                        (r.get("column_name","").lower(), r.get("rule_type","")) in _ai_validator_keys
                    ),
                }
                for r in q.get("rule_results", []) if not r.get("skipped")
            ]

            # Separate AI-defined rules for dedicated display
            ai_rules = [r for r in all_rules if r.get("ai_defined")]


            # Anomalies full
            anomaly_data = [
                {"col": a["column"], "pct": a.get("anomaly_pct",0), "max_z": a.get("max_z",0),
                 "severity": a.get("severity",""), "samples": (a.get("samples") or a.get("sample_values") or [])[:3]}
                for a in anomalies
            ]

            # Categorical drift full
            cat_data = [
                {"col": c["column"], "detail": c.get("detail",""), "severity": c.get("severity",""),
                 "top_val": c.get("top_value",""), "top_pct": c.get("top_pct",0)}
                for c in q.get("categorical_drift", [])
            ]

            # Consistency issues
            consistency_data = [
                {"check": c.get("check",""), "col": c.get("column",""), "failed": c.get("failed",0),
                 "status": c.get("status","")}


                for c in q.get("consistency_issues", []) if c.get("status") in ("FAIL","WARN")
            ][:10]

            # Drift alerts vs baseline
            drift_data = [
                {"col": a["column"], "metric": a["metric"], "detail": a["detail"], "severity": a["severity"]}
                for a in q.get("drift_alerts", [])
            ]

            # Numeric clusters
            cluster_data = [
                {"col": c["column"], "n": c["n_clusters"], "detail": c["detail"], "severity": c["severity"]}
                for c in q.get("numeric_clusters", []) if c.get("severity") == "WARN"
            ][:5]

            # -- LLM Executive Summary
            _exec_summary = ""
            _exec_summary_error = ""


            try:
                _summary_context = (
                    f"Dataset: {fname} ({q['total_rows']} rows, {q.get('total_cols', len(df.columns))} columns)\n"
                    f"DQ Score: {dq['score']}/100 (Grade {dq['grade']})\n"
                    f"Completeness: {dq.get('completeness',100):.1f}% | Uniqueness: {dq.get('uniqueness',100):.1f}% | "
                    f"Validity: {dq.get('validity',100):.1f}% | Conformity: {dq.get('conformity',100):.1f}%\n"
                    f"Rule failures: {len(fails)} FAIL, {len(warnings)} WARN\n"
                )
                if fails:
                    _summary_context += "Key failures: " + "; ".join(
                        f"{r.get('column_name','?')} ({r.get('rule_name','?')})" for r in fails[:5]
                    ) + "\n"
                if anomalies:
                    _summary_context += f"Anomalies: {len(anomalies)} columns with statistical outliers\n"
                if cat_issues:
                    _summary_context += f"Categorical issues: {len(cat_issues)} columns with distribution problems\n"


                if ai_rules:
                    _summary_context += f"AI-defined rules checked: {len(ai_rules)} ({sum(1 for r in ai_rules if r.get('status')=='FAIL')} failed)\n"

                _sum_prompt = (
                    f"You are a BFSI data quality analyst. Write a concise 3-4 sentence executive summary "
                    f"of this DQ report for a business user. Be direct -- state the overall quality level, "
                    f"the most important issues, and one clear recommendation. No bullet points, plain sentences.\n\n"
                    f"{_summary_context}"
                )
                _dq_username = getattr(request.state, "username", None) or ""
                _exec_summary = await asyncio.to_thread(
                    _ask_llm,
                    [{"role": "user", "content": [{"text": _sum_prompt}]}],
                    "", "quality", "ai_summary", _dq_username,
                )
                _exec_summary = _exec_summary.strip()


            except Exception as _e_sum:
                _exec_summary = (
                    f"Dataset '{fname}' scored {dq['score']}/100 (Grade {dq['grade']}). "
                    f"{len(fails)} rule failure(s) detected. "
                    + ("Immediate attention required." if dq['grade'] in ('D','F') else
                       "Minor issues to review." if dq['grade'] == 'C' else "Data quality is acceptable.")
                )
                _exec_summary_error = _llm_error_message(_e_sum) + " Showing a rule-based summary instead."

            # Store summary with session for Excel download
            _results_store[new_sid]["ai_summary"] = _exec_summary

            # Save to 30-day history so trend chart reflects this run
            try:
                _ws_username = getattr(request.state, "username", None) or "default"
                _rule_fails_h = len(fails)
                _crit_fails_h = dq.get("severity_breakdown", {}).get("critical_fails", 0)
                _ws_db.save_dq_history(
                    _ws_username, fname, q.get("schema_fingerprint", ""),
                    dq, q["total_rows"], _rule_fails_h, _crit_fails_h,


                    session_id=new_sid, bfsi_pack=";".join(_hints.keys()), di_scope="quality",
                )
            except Exception as _eh:
                pass  # non-fatal

            return JSONResponse(_sanitize_json({
                "session_id":  new_sid,
                "file_name":   fname,
                "file_format": df.attrs.get("_format", ""),
                "total_rows":  q["total_rows"],
                "total_cols":  q.get("total_cols", len(df.columns)),
                "duplicate_rows": q.get("duplicate_rows", 0),
                "score":     dq["score"],
                "grade":     dq["grade"],
                "completeness": dq.get("completeness"),
                "uniqueness":   dq.get("uniqueness"),
                "validity":     dq.get("validity"),
                "consistency":  dq.get("consistency"),
                "conformity":   dq.get("conformity"),


                "precision":    dq.get("precision"),
                "timeliness":   dq.get("timeliness"),
                "rule_fails":   len(fails),
                "rule_warns":   len(warnings),
                "rule_total":   len(q.get("rule_results", [])),
                "all_rules":    all_rules,
                "col_completeness": col_completeness,
                "anomalies":    anomaly_data,
                "cat_drift":    cat_data,
                "consistency_issues": consistency_data,
                "drift_alerts": drift_data,
                "numeric_clusters": cluster_data,
                "multivariate_anomalies": q.get("multivariate_anomalies", {}),
                "trend_anomalies": q.get("trend_anomalies", []),
                "ai_hints_used": list(_hints.keys()),
                "ai_rules":     ai_rules,
                "ai_summary":   _exec_summary,
                "ai_summary_error": _exec_summary_error,
            }))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/rerun-quality/{session_id}", response_class=HTMLResponse)
async def rerun_quality(session_id: str, request: Request):

    # Re-run Data Quality on stored DataFrames from a quality_ai session.
    # Extracts any rules discussed in the AI Copilot conversation and injects
    # them as hints -- making this genuinely agent-driven like Reconciliation.

    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        raise HTTPException(404, "Session expired -- please re-upload your file.")

    dataframes: list[tuple[str, object]] = [
        (item["name"], item["df"]) for item in stored["dataframes"]
    ]
    if not dataframes:
        raise HTTPException(400, "No dataframes found in session.")

    import asyncio as _asyncio


    t0 = time.time()
    proc_logs: list[dict] = []
    def _log(msg: str, level: str = "INFO") -> None:
        proc_logs.append({"elapsed": round(time.time() - t0, 3), "level": level, "message": msg})

    # -- Extract rules from AI Copilot conversation (passed in POST body) ----
    _hints: dict = {}
    try:
        _body = await request.json() if request.headers.get("content-type","").startswith("application/json") else {}
        _conversation = _body.get("conversation", "")
        if _conversation:
            _extract_prompt = (
                f"From this data quality conversation:\n{_conversation}\n\n"
                "Extract any specific validation rules mentioned. Return JSON only:\n"
                '{"nullable_hints":"col1,col2","key_hints":"col","timeliness_hints":"date_col",'
                '"bfsi_validators":["positive:price","allowed_values:side:BUY,SELL"]}\n'
                "Return {} if nothing specific was mentioned."


            )
            _raw = await asyncio.to_thread(
                _ask_llm,
                [{"role": "user", "content": [{"text": _extract_prompt}]}],
            )
            import re as _re
            _m = _re.search(r'\{.*\}', _raw, _re.DOTALL)
            if _m:
                _extracted = json.loads(_m.group(0))
                for _k in ("nullable_hints","key_hints","timeliness_hints","bfsi_validators"):
                    if _extracted.get(_k):
                        _hints[_k] = _extracted[_k]
            if _hints:
                _log(f"AI Copilot injected hints from conversation: {list(_hints.keys())}")
    except Exception as _he:
        _log(f"Hint extraction skipped: {_he}", "INFO")

    quality_reports = []
    governance_reports = []


    profile_reports = []

    # quality_ai shows only DQ tab -- governance and profile are in their own modules now
    _di_scope = {"quality"}
    _hints: dict = {}
    # OCR-UNCERTAIN: this second "_hints: dict = {}" re-declaration appears clearly in the photo
    # immediately after the extraction block above, which would discard the AI-extracted hints --
    # transcribed exactly as shown, not corrected.

    for fname, df in dataframes:
        _log(f"Re-running DQ on '{fname}' ({len(df)} rows × {len(df.columns)} cols)")
        try:
            q = await _asyncio.to_thread(
                analyze_quality_full, df, fname, {}, [], _hints, None, None
            )
            q["file_format"] = df.attrs.get("_format", "")
            quality_reports.append(q)

            _gov = q.get("governance")
            if _gov:
                _gov["file_format"] = q["file_format"]


                governance_reports.append(_gov)

            _prof = q.get("profile")
            if _prof:
                profile_reports.append({
                    "file_name":   fname,
                    "file_format": q["file_format"],
                    "total_rows":  q["total_rows"],
                    "total_cols":  q.get("total_cols", len(df.columns)),
                    "memory_mb":   _prof.get("memory_mb", 0),
                    "duplicate_rows": q["duplicate_rows"],
                    "key_candidates": _prof.get("key_candidates", []),
                    "near_key_cols":  _prof.get("near_key_cols", []),
                    "type_breakdown": _prof.get("type_breakdown", {}),
                    "correlations":   _prof.get("correlations", []),
                    "numeric_cols":   sum(1 for c in _prof.get("columns", []) if c.get("is_numeric") or c.get("mean") is not None),
                    "columns":        _prof.get("columns", []),
                })


            # Anomaly + clustering
            # OCR-UNCERTAIN: nesting level of this inner try/except block (inside the outer
            # for/try) is inferred from context across pages 798-801; exact indentation depth
            # could not be fully confirmed from the photo.
            try:
                q["anomaly_results"] = await _asyncio.to_thread(detect_numeric_anomalies, df, 3.0)
            except Exception:
                q["anomaly_results"] = []
            try:
                q["categorical_drift"] = await _asyncio.to_thread(detect_categorical_drift, df)
            except Exception:
                q["categorical_drift"] = []
            try:
                q["numeric_clusters"] = await _asyncio.to_thread(detect_numeric_clusters, df)
            except Exception:
                q["numeric_clusters"] = []
            try:
                q["multivariate_anomalies"] = await _asyncio.to_thread(detect_multivariate_anomalies, df)
            except Exception:
                q["multivariate_anomalies"] = {"checked_columns": [], "anomaly_count": 0, "rows": []}
            try:
                q["trend_anomalies"] = await _asyncio.to_thread(detect_trend_anomalies, df)
            except Exception:
                q["trend_anomalies"] = []

            _log(f"DQ complete: Score {q['dq_score']['score']} ({q['dq_score']['grade']})")

        except Exception as exc:


            _log(f"DQ failed for '{fname}': {exc}", "WARN")

    new_session_id = str(uuid.uuid4())
    total_elapsed = round(time.time() - t0, 3)
    all_file_columns = sorted(set(c for _, df in dataframes for c in df.columns))

    _results_store[new_session_id] = {
        "action": "quality",
        "file_names": [n for n, _ in dataframes],
        "quality_reports": quality_reports,
        "governance_reports": governance_reports,
        "profile_reports": profile_reports,
        "dataframes": stored["dataframes"],
        "proc_logs": proc_logs,
        "elapsed": total_elapsed,
        "pairs": [], "mappings": [], "parse_reports": [],
        "lineage_reports": [], "excluded_cols": [],
    }


    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "action":        "quality",
            "di_scope":      list(_di_scope),
            "di_ai_enhanced": False,
            "file_names":    [n for n, _ in dataframes],
            "pairs":         [],
            "quality_reports": quality_reports,
            "governance_reports": governance_reports,
            "mappings":      [],
            "parse_reports": [],
            "lineage_reports": [],
            "profile_reports": profile_reports,
            "has_data_dict": False,
            "has_rules":     False,
            "has_mapping_spec": False,
            "ref_log":       [],


            "proc_logs":     proc_logs,
            "elapsed":       total_elapsed,
            "session_id":    new_session_id,
            "all_file_columns": all_file_columns,
            "excluded_cols": [],
            "key_columns_val": "",
            "dataset_fingerprint": "",
            "saved_rules":   [],
        },
    )


@app.post("/rerun-profile-json/{session_id}")
async def rerun_profile_json(session_id: str, request: Request):
    """Run Data Profile and return JSON for AI Copilot chat card."""
    import asyncio as _asyncio
    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        return JSONResponse({"error": "Session expired."}, status_code=404)


    dataframes = [(item["name"], item["df"]) for item in stored["dataframes"]]
    if not dataframes:
        return JSONResponse({"error": "No dataframes."}, status_code=400)

    _hints: dict = {}
    try:
        _body = await request.json() if request.headers.get("content-type","").startswith("application/json") else {}
        _conv = _body.get("conversation","")
        if _conv:
            _raw = await _asyncio.to_thread(_ask_llm,[{"role":"user","content":[{"text":
                f"From this profiling conversation:\n{_conv}\n\nExtract columns of interest. "
                'Return JSON: {"nullable_hints":"col","key_hints":"col","timeliness_hints":"col"} or {}'
            }]}])
            import re as _re4
            _m4 = _re4.search(r'\{.*\}',_raw,_re4.DOTALL)
            if _m4:
                _ex4=json.loads(_m4.group(0))
                for _k in ("nullable_hints","key_hints","timeliness_hints"):


                    if _ex4.get(_k): _hints[_k]=_ex4[_k]
    except Exception: pass

    fname, df = dataframes[0]
    try:
        q = await _asyncio.to_thread(analyze_quality_full, df, fname, {}, [], _hints, None, None)
        _prof = q.get("profile", {})
        cols = _prof.get("columns", [])

        new_sid = str(uuid.uuid4())
        _results_store[new_sid] = {"action":"profile","file_names":[fname],
            "quality_reports":[],"governance_reports":[],"profile_reports":[{
                "file_name":fname,"file_format":df.attrs.get("_format",""),
                "total_rows":q["total_rows"],"total_cols":q.get("total_cols",len(df.columns)),
"memory_mb":_prof.get("memory_mb",0),"duplicate_rows":q["duplicate_rows"],
"key_candidates":_prof.get("key_candidates",[]),"near_key_cols":_prof.get("near_key_cols",[]),


"type_breakdown":_prof.get("type_breakdown",{}),"correlations":_prof.get("correlations",[]),
        "numeric_cols":sum(1 for c in cols if c.get("is_numeric") or c.get("mean") is not None),
        "columns":cols,
        }],"dataframes":stored["dataframes"],"proc_logs":[],"elapsed":0,
"pairs":[],"mappings":[],"parse_reports":[],"lineage_reports":[],"excluded_cols":[]}

        # Build per-column summary
        col_summary = [{
            "name":   c["name"],
            "dtype":  c.get("dtype",""),
            "null_pct": c.get("null_pct",0),
            "unique_n": c.get("unique_count",0),
            "cardinality": c.get("cardinality",""),
            "mean":   c.get("mean"),
            "min":    c.get("min"),
            "max":    c.get("max"),


            "outlier_pct": c.get("outlier_pct",0),
            "is_numeric": c.get("is_numeric", c.get("mean") is not None),
        } for c in cols]

        # The executive summary is a nice-to-have on top of the profile that was
        # already computed above -- an LLM hiccup (rate limit, no provider
        # configured, etc.) must not take down the whole report, same as the
        # quality endpoint's ai_summary is already protected.
        _profile_ai_summary = ""
        _profile_ai_summary_error = ""
        try:
            _profile_ai_summary = (await _asyncio.to_thread(_ask_llm,
                [{"role":"user","content":[{"text":
                    f"Write a 2-3 sentence executive summary of this data profile for {fname} "
                    f"({q['total_rows']} rows, {q.get('total_cols',len(df.columns))} cols). "
                    f"Key candidates: {_prof.get('key_candidates',[])}. "
                    f"Numeric cols: {sum(1 for c in cols if c.get('mean') is not None)}. "
                    f"Duplicate rows: {q.get('duplicate_rows',0)}. Be concise and business-focused."
                }]}], "", "profile", "ai_summary",
                    getattr(request.state, "username", None) or "")).strip()
        except Exception as _e_sum:
            _profile_ai_summary_error = _llm_error_message(_e_sum)

        return JSONResponse(_sanitize_json({
            "session_id":  new_sid,
            "file_name":   fname,
            "file_format": df.attrs.get("_format",""),
            "total_rows":  q["total_rows"],
            "total_cols":  q.get("total_cols", len(df.columns)),
            "duplicate_rows": q.get("duplicate_rows",0),
            "memory_mb":   round(_prof.get("memory_mb",0),2),
            "numeric_cols": sum(1 for c in cols if c.get("is_numeric") or c.get("mean") is not None),
            "key_candidates": _prof.get("key_candidates",[])[:5],
            "near_key_cols":  _prof.get("near_key_cols",[])[:5],
            "type_breakdown": _prof.get("type_breakdown",{}),
            "correlations":   _prof.get("correlations",[])[:8],
            "columns":        col_summary,
            "ai_hints_used": list(_hints.keys()),
            "ai_summary":    _profile_ai_summary,
            "ai_summary_error": _profile_ai_summary_error,
        }))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/rerun-governance-json/{session_id}")
async def rerun_governance_json(session_id: str, request: Request):
    """Run Governance and return JSON for AI Copilot chat card."""
    import asyncio as _asyncio


    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        return JSONResponse({"error": "Session expired."}, status_code=404)
    dataframes = [(item["name"], item["df"]) for item in stored["dataframes"]]
    if not dataframes:
        return JSONResponse({"error": "No dataframes."}, status_code=400)

    _hints: dict = {}
    try:
        _body = await request.json() if request.headers.get("content-type","").startswith("application/json") else {}
        _conv = _body.get("conversation","")
        if _conv:
            _raw = await _asyncio.to_thread(_ask_llm,[{"role":"user","content":[{"text":
                f"From this governance conversation:\n{_conv}\n\nExtract PII overrides. "
                'Return JSON: {"not_pii":"col1,col2","sensitivity":"Confidential"} or {}'
            }]}])
            import re as _re5
            _m5=_re5.search(r'\{.*\}',_raw,_re5.DOTALL)


        if _m5:
            _ex5=json.loads(_m5.group(0))
            if _ex5.get("not_pii"): _hints["not_pii_columns"]=_ex5["not_pii"]
            if _ex5.get("sensitivity"): _hints["sensitivity_override"]=_ex5["sensitivity"]
    except Exception: pass

    fname, df = dataframes[0]
    try:
        g = await _asyncio.to_thread(analyze_governance, df, fname, {}, user_hints=_hints)
        g["file_format"] = df.attrs.get("_format","")

        new_sid = str(uuid.uuid4())
        _results_store[new_sid] = {"action":"governance","file_names":[fname],
            "quality_reports":[],"governance_reports":[g],"profile_reports":[],
            "dataframes":stored["dataframes"],"proc_logs":[],"elapsed":0,

"pairs":[],"mappings":[],"parse_reports":[],"lineage_reports":[],"excluded_cols":[]}

        # PII column details


        pii_cols = [c for c in g.get("columns", []) if c.get("pii_detected")]
        clean_cols = [c for c in g.get("columns",[]) if not c.get("pii_detected")]

        # Same reasoning as the profile endpoint: the executive summary is a
        # nice-to-have on top of the already-computed governance audit, so an
        # LLM hiccup must not take down the whole report.
        _gov_ai_summary = ""
        _gov_ai_summary_error = ""
        try:
            _gov_ai_summary = (await _asyncio.to_thread(_ask_llm,[{"role":"user","content":[{"text":
                f"Write a 2-3 sentence executive summary of this governance audit for {fname}. "
                f"Classification: {g.get('overall_classification','')}. "
                f"PII columns: {g.get('pii_column_count',0)}. "
                f"Regulatory frameworks: {g.get('regulatory_frameworks',[])}. "
                f"Be direct about data sensitivity and compliance risk."
            }]}], "", "governance", "ai_summary",
                getattr(request.state, "username", None) or "")).strip()
        except Exception as _e_sum:
            _gov_ai_summary_error = _llm_error_message(_e_sum)

        return JSONResponse(_sanitize_json({
            "session_id":   new_sid,
            "file_name":    fname,
            "file_format":  df.attrs.get("_format",""),
            "total_rows":   len(df),
            "total_cols":   len(df.columns),
            "classification": g.get("overall_classification",""),
            "pii_col_count":  g.get("pii_column_count",0),
            "bfsi_id_col_count": g.get("bfsi_identifier_col_count",0),
            "regulatory_frameworks": g.get("regulatory_frameworks",[]),
            "mandatory_breaches": g.get("mandatory_breaches",[])[:5],
            "pii_columns": [{
                "name":   c["column"],
                "sensitivity": c.get("sensitivity",""),
                "pii_types": c.get("pii_detected",[])[:3],
                "regulatory": c.get("regulatory_flags",[])[:3],
                "access":   c.get("access_recommendation",""),
            } for c in pii_cols[:15]],
            "undocumented": g.get("undocumented_columns",[])[:10],
            "ai_hints_used": list(_hints.keys()),
            "ai_summary": _gov_ai_summary,
            "ai_summary_error": _gov_ai_summary_error,
        }))
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/rerun-profile/{session_id}", response_class=HTMLResponse)


async def rerun_profile(session_id: str, request: Request):
    """Re-run Data Profile using AI Copilot conversation context."""
    import asyncio as _asyncio
    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        raise HTTPException(404, "Session expired -- please re-upload your file.")
    dataframes = [(item["name"], item["df"]) for item in stored["dataframes"]]
    if not dataframes:
        raise HTTPException(400, "No dataframes in session.")
    t0 = time.time()
    proc_logs: list[dict] = []
    def _log(msg, level="INFO"):
        proc_logs.append({"elapsed": round(time.time()-t0,3), "level": level, "message": msg})

    # Extract profile focus hints from AI conversation
    _profile_hints: dict = {}
    try:
        _body = await request.json() if request.headers.get("content-type","").startswith("application/json") else {}


        _conv = _body.get("conversation", "")
        if _conv:
            _raw = await _asyncio.to_thread(
                _ask_llm,
                [{"role": "user", "content": [{"text":
                    f"From this data profiling conversation:\n{_conv}\n\n"
                    "Extract any columns of interest mentioned. Return JSON only:\n"
                    '{"nullable_hints":"col1,col2","key_hints":"col","timeliness_hints":"date_col"}\n'
                    "Return {} if nothing specific was mentioned."
                }]}],
            )
            import re as _re2
            _m2 = _re2.search(r'\{.*\}', _raw, _re2.DOTALL)
            if _m2:
                _extracted2 = json.loads(_m2.group(0))
                for _k in ("nullable_hints","key_hints","timeliness_hints"):
                    if _extracted2.get(_k):
                        _profile_hints[_k] = _extracted2[_k]


        if _profile_hints:
            _log(f"Profile AI hints from conversation: {list(_profile_hints.keys())}")
    except Exception as _he2:
        _log(f"Profile hint extraction skipped: {_he2}", "INFO")

    profile_reports = []
    for fname, df in dataframes:
        _log(f"Profiling '{fname}'")
        try:
            q = await _asyncio.to_thread(analyze_quality_full, df, fname, {}, [], _profile_hints, None, None)
            _prof = q.get("profile")
            if _prof:
                profile_reports.append({
                    "file_name": fname, "file_format": df.attrs.get("_format",""),
                    "total_rows": q["total_rows"], "total_cols": q.get("total_cols", len(df.columns)),
                    "memory_mb": _prof.get("memory_mb",0), "duplicate_rows": q["duplicate_rows"],


                    "key_candidates": _prof.get("key_candidates",[]), "near_key_cols": _prof.get("near_key_cols",[]),
                    "type_breakdown": _prof.get("type_breakdown",{}), "correlations": _prof.get("correlations",[]),
                    "numeric_cols": sum(1 for c in _prof.get("columns",[]) if c.get("is_numeric") or c.get("mean") is not None),
                    "columns": _prof.get("columns",[]),
                })
            _log(f"Profile complete for '{fname}'")
        except Exception as e:
            _log(f"Profile failed: {e}", "WARN")

    new_sid = str(uuid.uuid4())
    elapsed = round(time.time()-t0, 3)
    all_cols = sorted(set(c for _, df in dataframes for c in df.columns))
    _results_store[new_sid] = {"action":"profile","file_names":[n for n,_ in dataframes],
        "quality_reports":[],"governance_reports":[],"profile_reports":profile_reports,
        "dataframes":stored["dataframes"],"proc_logs":proc_logs,"elapsed":elapsed,
        "pairs":[],"mappings":[],"parse_reports":[],"lineage_reports":[],"excluded_cols":[]}
    return templates.TemplateResponse(request=request, name="index.html", context={


        "action":"profile","di_scope":["profile"],"di_ai_enhanced":False,
        "file_names":[n for n,_ in dataframes],"pairs":[],"quality_reports":[],
        "governance_reports":[],"mappings":[],"parse_reports":[],"lineage_reports":[],
        "profile_reports":profile_reports,"has_data_dict":False,"has_rules":False,
        "has_mapping_spec":False,"ref_log":[],"proc_logs":proc_logs,"elapsed":elapsed,
        "session_id":new_sid,"all_file_columns":all_cols,
        "excluded_cols":[],"key_columns_val":"","dataset_fingerprint":"","saved_rules":[],
    })


@app.post("/rerun-governance/{session_id}", response_class=HTMLResponse)
async def rerun_governance(session_id: str, request: Request):
    """Re-run Governance using AI Copilot conversation context."""
    import asyncio as _asyncio
    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored:
        raise HTTPException(404, "Session expired -- please re-upload your file.")
    dataframes = [(item["name"], item["df"]) for item in stored["dataframes"]]
    if not dataframes:


        raise HTTPException(400, "No dataframes in session.")
    t0 = time.time()
    proc_logs: list[dict] = []
    def _log(msg, level="INFO"):
        proc_logs.append({"elapsed": round(time.time()-t0,3), "level": level, "message": msg})

    # Extract governance hints from AI conversation
    _gov_hints: dict = {}
    try:
        _body = await request.json() if request.headers.get("content-type","").startswith("application/json") else {}
        _conv = _body.get("conversation", "")
        if _conv:
            _raw = await _asyncio.to_thread(
                _ask_llm,
                [{"role": "user", "content": [{"text":
                    f"From this data governance conversation:\n{_conv}\n\n"
                    "Extract any columns identified as PII/sensitive or any overrides mentioned. "


                    "Return JSON only:\n"
                    '{"pii_columns":"col1,col2","not_pii":"col3","sensitivity":"Confidential"}\n'
                    "Return {} if nothing specific was mentioned."
                }]}],
            )
            import re as _re3
            _m3 = _re3.search(r'\{.*\}', _raw, _re3.DOTALL)
            if _m3:
                _extracted3 = json.loads(_m3.group(0))
                # Map to governance user_hints format
                if _extracted3.get("not_pii"):
                    _gov_hints["not_pii_columns"] = _extracted3["not_pii"]
                if _extracted3.get("sensitivity"):
                    _gov_hints["sensitivity_override"] = _extracted3["sensitivity"]
                if _gov_hints:
                    _log(f"Governance AI hints from conversation: {list(_gov_hints.keys())}")
    except Exception as _he3:
        _log(f"Governance hint extraction skipped: {_he3}", "INFO")


    governance_reports = []
    for fname, df in dataframes:
        _log(f"Governance analysis on '{fname}'")
        try:
            g = await _asyncio.to_thread(analyze_governance, df, fname, {}, user_hints=_gov_hints)
            g["file_format"] = df.attrs.get("_format","")
            governance_reports.append(g)
            _log(f"Governance: {g['overall_classification']} | PII cols: {g['pii_column_count']}")
        except Exception as e:
            _log(f"Governance failed: {e}", "WARN")

    new_sid = str(uuid.uuid4())
    elapsed = round(time.time()-t0, 3)
    all_cols = sorted(set(c for _, df in dataframes for c in df.columns))
    _results_store[new_sid] = {"action":"governance","file_names":[n for n,_ in dataframes],
        "quality_reports":[],"governance_reports":governance_reports,"profile_reports":[],
        "dataframes":stored["dataframes"],"proc_logs":proc_logs,"elapsed":elapsed,
        "pairs":[],"mappings":[],"parse_reports":[],"lineage_reports":[],"excluded_cols":[]}


    return templates.TemplateResponse(request=request, name="index.html",
        context={
            "action":"governance","di_scope":["governance"],"di_ai_enhanced":False,
            "file_names":[n for n,_ in dataframes],"pairs":[],"quality_reports":[],
            "governance_reports":governance_reports,"mappings":[],"parse_reports":[],"lineage_reports":[],
            "profile_reports":[],"has_data_dict":False,"has_rules":False,
            "has_mapping_spec":False,"ref_log":[],"proc_logs":proc_logs,"elapsed":elapsed,
            "session_id":new_sid,"all_file_columns":all_cols,
            "excluded_cols":[],"key_columns_val":"","dataset_fingerprint":"","saved_rules":[],
        })


@app.post("/help-chat")
async def help_chat(request: Request):
    """Stateless AI assistant for the User Guide modal -- no session required."""
    body     = await request.json()
    question = body.get("question", "").strip()
    history  = body.get("history", [])


    system   = body.get("system", "You are a helpful assistant for the AI Agent -- Data Validation.")
    if not question:
        return JSONResponse({"error": "Empty question"}, status_code=400)
    messages = [{"role": h["role"], "content": [{"text": h["text"]}]} for h in history[-8:]]
    messages.append({"role": "user", "content": [{"text": question}]})
    try:
        reply = await asyncio.to_thread(_ask_llm, messages, system=system)
        return JSONResponse({"reply": reply})
    except Exception as exc:
        return JSONResponse({"reply": f"Error: {exc}"}, status_code=500)


@app.post("/chat")
async def chat(request: Request):
    body       = await request.json()
    session_id = body.get("session_id", "")
    question   = body.get("question", "").strip()
    history    = body.get("history", [])
    _chat_username = _ws_resolve_username(request) or "default"


    if not question:
        return JSONResponse({"error": "Empty question"}, status_code=400)

    context = _chat_contexts.get(session_id, {})

    if context.get("mode") == "recon" and re.match(r"^(please\s+)?run\s*recon(ciliation)?[.!]?$", question, re.IGNORECASE):
        result = await asyncio.to_thread(_run_llm_recon_full, session_id, _chat_username)
        if result is None:
            return JSONResponse({"reply": "I couldn't find both files for this session anymore -- please re-upload and try again."})
        c = result["counts"]
        n_rules = result.get("rules_applied", 0)
        if result.get("llm_error"):
            rule_note = (f"⚠️ You have {n_rules} saved rule(s) for this schema, but the LLM call to convert "
                         f"them into parameters failed ({result['llm_error']}) -- ran with plain auto-detected "
                         f"key and NO rules applied. Check Settings → LLM Provider for a valid API key. ")
        elif n_rules:
            rule_note = f"Applied {n_rules} saved rule(s). "
        else:
            rule_note = "No saved rules for this schema yet -- used auto-detected key and no transforms. "
        reply = (
            f"✅ Ran reconciliation. {rule_note}"
            f"Key: {result.get('method','n/a')}. "
            f"Matched {c['matched']}, {c['file1_only']} only in {result['files']['file1']['name']}, "
            f"{c['file2_only']} only in {result['files']['file2']['name']}, {c['modified']} modified. "
            f"Full detail is in the Summary/Exceptions tabs above."
        )
        if result.get("key_warning"):
            reply += f"\n⚠️ {result['key_warning']}"
        return JSONResponse({"reply": reply, "recon_result": result})

    saved_rule_text = ""
    new_rule = None

    if context.get("mode") == "recon":
        # -- Recon / Customize mode -- rich schema-aware system prompt ----
        src_schema = context.get("src_schema", [])
        tgt_schema = context.get("tgt_schema", [])

        def _fmt_schema(schema: list[dict]) -> str:
            return "\n".join(
                f"  {s['column']} ({s['dtype']}) -- e.g. {', '.join(s['sample'][:3])}"
                for s in schema
            )

        existing_rules = context.get("saved_rules", [])


        if existing_rules:
            saved_rule_text = "\n\nSaved rules for this schema (always apply these):\n" + "\n".join(
                f"  [{r['category'].upper()}] {r['rule']}" for r in existing_rules
                if r.get("category") != "recon_hints"
            )

        comparisons = context.get("comparisons") or []
        results_text = ""
        if comparisons:
            lines = []
            for cmp_r in comparisons:
                lines.append(
                    f"  {cmp_r.get('pair','')}: {cmp_r.get('added',0)} rows only in target, "
                    f"{cmp_r.get('removed',0)} rows only in source, {cmp_r.get('modified',0)} rows with changed values")
            results_text = "\n\nCOMPARISON RESULTS (already computed by the deterministic recon engine -- these are the real, final numbers, not estimates):\n" + "\n".join(lines)

        system = (
            "You are an AI Copilot for data reconciliation. "
            "The user has uploaded two datasets and wants to compare them. "
            "Your job is to help them:\n"
            "1. Map columns between source and target (they may have different names)\n"
            "2. Define value transformation rules (e.g. date formats, number formats, casing)\n"
            "3. Identify the correct key columns for row matching\n"
            "4. Understand and explain any differences found\n"
            "5. Save rules so the same schema is handled automatically next time\n\n"
            f"SOURCE FILE: {context.get('src_name')} ({context.get('src_rows')} rows)\n"
            f"{_fmt_schema(src_schema)}\n\n"
            f"TARGET FILE: {context.get('tgt_name')} ({context.get('tgt_rows')} rows)\n"
            f"{_fmt_schema(tgt_schema)}\n"
            f"{results_text}\n"
            f"{saved_rule_text}\n\n"
            "When the user defines a rule (mapping, transform, key, exclusion), "
            "confirm it and tell them to type: remember: <the rule> -- to save it for future runs.\n"
            "If comparison results are given above, ground your answer in those exact numbers -- "
            "do not guess or invent counts.\n"
            "Be concise. Use bullet points. Always refer to actual column names from the schemas above."
        )

        # Auto-detect and save rules the user types as "remember: ..."
        _remember_match = re.match(r"^remember[:\s]+(.+)$", question, re.IGNORECASE)
        if _remember_match:
            rule_text = _remember_match.group(1).strip()
            fp = context.get("dataset_fingerprint", "")
            if fp and rule_text:
                label = f"{context.get('src_name','')} vs {context.get('tgt_name','')}"
                _fp_save(_chat_username, fp, rule_text, category="recon_rule", dataset_label=label)
                # Refresh saved rules in context
                context["saved_rules"] = _fp_get_rules(_chat_username, fp)


            _chat_contexts[session_id] = context
            new_rule = rule_text
    else:
        # -- Standard mode ----
        system = (
            "You are a data quality and governance assistant. "
            "The user is analysing data files using Data Validation AGENT.\n"
            f"Context:\n{json.dumps(context, indent=2)[:4000]}\n\n"
            "Answer precisely. Use bullet points. If asked for recommendations, be specific and actionable."
        )

    messages = [{"role": h["role"], "content": [{"text": h["text"]}]} for h in history[-8:]]
    messages.append({"role": "user", "content": [{"text": question}]})

    try:
        reply = await asyncio.to_thread(_ask_llm, messages, system=system)
        if new_rule:
            reply = f"✅ Rule saved: *{new_rule}*\n\n" + reply

        return JSONResponse({"reply": reply, "new_rule": new_rule})

    except Exception as exc:
        return JSONResponse({"reply": f"Error: {exc}"}, status_code=500)


def _parse_recon_rules_to_params(rules: list[dict], src_cols: list[str], tgt_cols: list[str]) -> dict:
    # Ask the LLM to read the saved recon rules and return structured execution parameters.
    #
    # Supports composite keys, regex-based column extraction from free-text fields,
    # multi-column aggregation, and value transforms.
    # Falls back to empty params on any failure so the caller can run a plain compare.
    rule_text = "\n".join(
        f" [{r['category'].upper()}] {r['rule']}"
        for r in rules if r.get("category") not in ("recon_hints",)
    )

    if not rule_text.strip():
        return {}

    prompt = f"""You are a data reconciliation parameter extractor.

Given the user rules and two file schemas, produce ONLY a JSON object with execution parameters.

SOURCE columns: {src_cols}
TARGET columns: {tgt_cols}

User rules:
{rule_text}

JSON schema (omit any key that is not needed):
{{
  "key_cols": ["col_a", "col_b"],     // composite key -- list of column names present
                                       in BOTH files after all transforms/parses/renames. Use exact final column names.
  "key_col": "col",                   // single key shorthand -- ignored if key_cols is set
  "col_map": {{"src_col": "tgt_col"}},  // rename columns so both sides share the same
                                       name BEFORE key resolution
  "parse_cols": [                     // extract new columns from free-text fields via regex
    {{
      "side": "src|tgt|both",
      "source_col": "Description",     // column to parse from
      "new_col": "Side",               // name of new column to create
      "pattern": "(?i)(buy|sell)",     // Python regex with one capture group
      "transform": "upper|lower|title|"  // optional post-transform on the captured value
    }}
  ],
  "transforms": [                     // value normalisation on existing columns
    {{
      "side": "src|tgt|both",
      "col": "col_name_or_*",
      "op": "upper|lower|strip|trim_quotes|strip_commas|strip_suffix|strip_prefix|replace_text|pad_left|truncate|regex_replace|to_numeric|round_numeric|negate|scale|fx_convert|abs_numeric|fillna_text|fillna_numeric|floor_numeric|parse_date|date_format|extract_date|ticker_strip|isin_strip|side_normalize|sign_to_side|map_values|coalesce",
      "arg": "meaning depends on op -- suffix string / regex pattern / fill value / width /
              decimals / scale factor / date format / mapping dict / fallback column name /
              FX rate table (JSON object, for fx_convert)",
      "replacement": "replacement string for replace_text and regex_replace (default empty string)",
      "ccy_col": "column holding each row's currency code -- REQUIRED for op fx_convert only"
    }}
  ],
  "src_agg": {{"group_by": ["col_a","col_b"], "agg_col": "vol_col", "agg_fn": "sum"}},
  "tgt_agg": {{"group_by": ["col_a","col_b"], "agg_col": "vol_col", "agg_fn": "sum"}},
  "src_value_col": "FirmNeedsQuantity",  // the single value column from SOURCE
                                          to compare (when it has a different name from the target value col)
  "tgt_value_col": "borrowReqNet",       // the single value column from TARGET to
                                          compare (set this when target has aggregation or a renamed value col)
  "exclude": ["col1", "col2"],
  "fuzzy_fields": ["customer_name", "address"],  // set ONLY when the user describes an approximate/
                                          // probabilistic match (typos, formatting drift, no clean
                                          // shared key) -- column name(s) to match by similarity
                                          // instead of exact key equality. Omit entirely for a normal
                                          // exact-key reconciliation.
  "fuzzy_threshold": 0.75                // 0-1 minimum similarity score to count as a match (default
                                          // 0.75) -- only meaningful together with fuzzy_fields.
}}

Key rules:
- If a composite key is needed (e.g. InstrumentID + Side), list ALL parts in key_cols.
- Set fuzzy_fields (not key_cols) when the user asks for "fuzzy match", "approximate match", "similar
  values", "no exact key available", or names a field known to have typos/formatting differences
  across the two sources (e.g. "match on customer name even if spelled slightly differently").
  fuzzy_fields and key_cols are mutually exclusive -- fuzzy matching does not use a join key.
- parse_cols runs BEFORE key resolution and aggregation -- use it to create derived columns from free text.
- col_map renames columns so both sides share identical names; apply after parse_cols.
- key_cols / key_col must reference column names that will exist AFTER parse_cols and col_map are applied.
- For "floor numeric" / "ignore decimals" use op "floor_numeric".
- For "strip .US suffix" / "remove .US" use op "strip_suffix" with arg ".US".
- For any exchange suffix (.US .LN .HK /UN etc.) use op "ticker_strip" -- no arg needed.
- For regex-based removal use op "regex_replace" with arg as the pattern and replacement as "".
- For NaN text fill use op "fillna_text" with arg "".
- For NaN numeric fill use op "fillna_numeric" with arg 0.
- For absolute value use op "abs_numeric".
- For rounding use op "round_numeric" with arg as decimal places (e.g. 2).
- For sign flip use op "negate".
- For unit scaling (e.g. units->thousands) use op "scale" with arg as the factor (e.g. 0.001).
- For cross-currency amount comparison (e.g. "convert amount to USD using EUR=1.087, GBP=1.27")
  use op "fx_convert" with arg as a JSON rate table (units of the common/base currency per 1 unit
  of that currency -- the base currency itself gets rate 1.0) and ccy_col naming the column that
  holds each row's currency code. Apply to the amount column on whichever side(s) need converting;
  rows whose currency isn't in the rate table are left unconverted rather than dropped.
- For buy/sell normalisation use op "side_normalize" -- maps B/S/BOT/SLD/1/-1 -> BUY/SELL.
- For positive=BUY negative=SELL use op "sign_to_side".
- For fixed value remapping use op "map_values" with arg as a JSON dict e.g. {{"B":"BUY","S":"SELL"}}.
- For zero-padding (CUSIP/ISIN) use op "pad_left" with arg as target width (e.g. 9).
- For date reformatting use op "date_format" with arg as strftime format (e.g. "%Y-%m-%d").
- For stripping time from datetime use op "extract_date".
- For currency/comma removal before numeric cast use op "to_numeric".
- For coalescing nulls from another column use op "coalesce" with arg as the fallback column name.
- When rules say "use only X and Y columns" or "compare X vs Y", set src_value_col and tgt_value_col accordingly.
- Only include what you are confident about. Reply with ONLY valid JSON, no commentary."""

    try:
        raw = _ask_llm([{"role": "user", "content": [{"text": prompt}]}])
        raw = re.sub(r"^```[a-z]*\n?", "", raw.strip(), flags=re.MULTILINE)
        raw = re.sub(r"```$", "", raw.strip())
        result = json.loads(raw)
        # Normalise: if key_cols not set but key_col is, promote it
        if not result.get("key_cols") and result.get("key_col"):
            result["key_cols"] = [result["key_col"]]

        return result
    except Exception as exc:
        # Surface the failure instead of silently running a plain compare --
        # otherwise saved rules look like they're being ignored with no explanation.
        return {"_llm_error": str(exc)}


def _apply_recon_params(
    df: pd.DataFrame,
    agg: dict | None,
    transforms: list[dict],
    side: str,
    col_map: dict | None = None,
    parse_cols: list[dict] | None = None,
) -> pd.DataFrame:

    # Apply parse_cols (regex extraction), col_map rename, value transforms,
    # and pre-aggregation to a DataFrame before comparison.

    # Order: parse_cols -> col_map -> transforms -> agg
    # This order ensures derived columns exist before renaming/aggregation.


    df = df.copy()

    # 1. Regex-based column extraction from free-text fields
    for pc in (parse_cols or []):
        if pc.get("side") not in (side, "both"):
            continue
        src_col = pc.get("source_col", "")
        new_col = pc.get("new_col", "")
        pattern = pc.get("pattern", "")
        xform   = pc.get("transform", "")
        if not src_col or not new_col or not pattern:
            continue
        # Case-insensitive column name lookup
        actual = next((c for c in df.columns if c.lower() == src_col.lower()), None)
        if actual is None:
            continue
        try:
            extracted = df[actual].astype(str).str.extract(pattern, expand=False)


            if xform == "upper":
                extracted = extracted.str.upper()
            elif xform == "lower":
                extracted = extracted.str.lower()
            elif xform == "title":
                extracted = extracted.str.title()
            df[new_col] = extracted.fillna("")
        except Exception:
            pass

    # 2. Column rename (col_map applied to both sides so keys align)
    if col_map:
        df = df.rename(columns=col_map)

    # 3. Value transforms
    for t in transforms:
        if t.get("side") not in (side, "both"):
            continue
        col_spec = t.get("col", "*")


        cols = [col_spec] if col_spec != "*" else list(df.columns)
        for c in cols:
            if c not in df.columns:
                continue
            op = t.get("op", "")
            if op == "upper":
                df[c] = df[c].astype(str).str.strip().str.upper()
            elif op == "lower":
                df[c] = df[c].astype(str).str.strip().str.lower()
            elif op == "strip":
                df[c] = df[c].astype(str).str.strip()
            elif op == "strip_commas":
                df[c] = df[c].astype(str).str.replace(",", "", regex=False)
            elif op == "strip_suffix":
                # Remove a fixed suffix (value in "arg") -- e.g. strip ".US" from "AAPL.US"
                suffix = str(t.get("arg", ""))
                if suffix:
                    df[c] = df[c].astype(str).str.removesuffix(suffix).str.strip()
            elif op == "strip_prefix":


                prefix = str(t.get("arg", ""))
                if prefix:
                    df[c] = df[c].astype(str).str.removeprefix(prefix).str.strip()
            elif op == "regex_replace":
                # Replace regex pattern (t["arg"]) with replacement (t["replacement"], default "")
                pattern     = t.get("arg", "")
                replacement = t.get("replacement", "")
                if pattern:
                    df[c] = df[c].astype(str).str.replace(pattern, replacement, regex=True).str.strip()
            elif op == "fillna_text":
                # Replace NaN / "nan" / empty string with t["arg"] (default "")
                fill_val = t.get("arg", "")
                df[c] = df[c].where(df[c].notna(), fill_val)
                df[c] = df[c].astype(str).replace({"nan": fill_val, "NaN": fill_val, "None": fill_val})
                df[c] = df[c].str.strip().replace({"": fill_val})
            elif op == "fillna_numeric":
                # Replace NaN with numeric value in t["arg"] (default 0)
                fill_val = t.get("arg", 0)


                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(fill_val)
            elif op == "abs_numeric":
                df[c] = pd.to_numeric(df[c], errors="coerce").abs()
            elif op == "parse_date":
                df[c] = pd.to_datetime(df[c], errors="coerce").dt.strftime("%Y-%m-%d")
            elif op == "floor_numeric":
                cleaned = df[c].astype(str).str.replace(",", "", regex=False).str.strip()
                df[c] = pd.to_numeric(cleaned, errors="coerce").apply(
                    lambda x: int(x) if pd.notna(x) else x
                )

            # -- String cleanup ----------------------------------------
            elif op == "trim_quotes":
                df[c] = df[c].astype(str).str.strip().str.strip("\"'")
            elif op == "replace_text":
                # Fixed string find/replace -- arg=find, replacement=replace_with
                find = str(t.get("arg", ""))
                repl = str(t.get("replacement", ""))
                if find:


                    df[c] = df[c].astype(str).str.replace(find, repl, regex=False)
            elif op == "pad_left":
                # Zero-pad to length in arg (default 9 for CUSIP)
                width = int(t.get("arg", 9))
                fill  = str(t.get("replacement", "0"))
                df[c] = df[c].astype(str).str.strip().str.zfill(width) if fill == "0" \
                    else df[c].astype(str).str.strip().str.rjust(width, fill)
            elif op == "truncate":
                # Limit string length to arg characters
                n = int(t.get("arg", 50))
                df[c] = df[c].astype(str).str[:n]

            # -- Numeric ----------------------------------------
            elif op == "round_numeric":
                # Round to N decimal places (arg, default 2)
                decimals = int(t.get("arg", 2))
                df[c] = pd.to_numeric(df[c], errors="coerce").round(decimals)
            elif op == "negate":
                df[c] = pd.to_numeric(df[c], errors="coerce") * -1


            elif op == "to_numeric":
                # Strip commas/currency symbols then cast to float
                df[c] = pd.to_numeric(
                    df[c].astype(str).str.replace(r"[,$£€\s]", "", regex=True),
                    errors="coerce"
                )
            elif op == "scale":
                # Multiply by factor in arg (e.g. 0.001 to convert units -> thousands)
                factor = float(t.get("arg", 1))
                df[c] = pd.to_numeric(df[c], errors="coerce") * factor
            elif op == "fx_convert":
                # Convert a per-row currency-denominated amount into a common base
                # currency before comparison, so e.g. a USD amount and a EUR amount
                # for "the same" trade can actually be tolerance-compared instead of
                # always looking like a value break. arg = JSON rate table (units of
                # base currency per 1 unit of that currency), e.g.
                # {"USD": 1.0, "EUR": 1.087, "GBP": 1.27}. ccy_col names the column
                # holding each row's currency code; rows with an unrecognised or
                # missing currency are left unconverted (rate 1.0) rather than
                # dropped, since a wrong guess is worse than a visible discrepancy.
                ccy_col = t.get("ccy_col", "")
                rates_raw = t.get("arg", "{}")
                try:
                    rates = json.loads(rates_raw) if isinstance(rates_raw, str) else (rates_raw or {})
                except Exception:
                    rates = {}
                if ccy_col and ccy_col in df.columns and rates:
                    rate_lookup = {str(k).strip().upper(): float(v) for k, v in rates.items()}
                    ccy_series = df[ccy_col].astype(str).str.strip().str.upper()
                    rate_series = ccy_series.map(rate_lookup).fillna(1.0)
                    df[c] = pd.to_numeric(df[c], errors="coerce") * rate_series

            # -- Date / Time ----------------------------------------
            elif op == "date_format":
                # Reformat date to arg format string (default YYYY-MM-DD)
                fmt = str(t.get("arg", "%Y-%m-%d"))
                df[c] = pd.to_datetime(df[c], errors="coerce").dt.strftime(fmt)
            elif op == "extract_date":
                # Strip time component -- keep date only
                df[c] = pd.to_datetime(df[c], errors="coerce").dt.strftime("%Y-%m-%d")


            # -- Financial / domain-specific ----------------------------------------
            elif op == "ticker_strip":
                # Remove exchange suffix: .US .LN .HK /UN .A etc.
                df[c] = df[c].astype(str).str.strip() \
                    .str.replace(r"\.[A-Z]{1,4}$", "", regex=True) \
                    .str.replace(r"/[A-Z]{1,3}$", "", regex=True) \
                    .str.strip()
            elif op == "isin_strip":
                # Keep only alphanumeric (strips country prefix/suffix from ISIN/CUSIP)
                df[c] = df[c].astype(str).str.strip().str.replace(r"[^A-Z0-9]", "", regex=True)
            elif op == "side_normalize":
                # Map common buy/sell variants -> BUY / SELL
                _side_map = {
                    "B": "BUY", "BUY": "BUY", "BOT": "BUY", "BOUGHT": "BUY",
                    "1": "BUY", "L": "BUY", "LONG": "BUY",
                    "S": "SELL", "SELL": "SELL", "SLD": "SELL", "SOLD": "SELL",
                    "-1": "SELL", "SS": "SELL", "SHORT": "SELL",
                }


                df[c] = df[c].astype(str).str.strip().str.upper().map(
                    lambda v: _side_map.get(v, v)
                )
            elif op == "sign_to_side":
                # Positive numeric qty -> BUY, negative -> SELL
                numeric = pd.to_numeric(df[c], errors="coerce")
                df[c] = numeric.apply(
                    lambda v: "BUY" if pd.notna(v) and v > 0
                    else ("SELL" if pd.notna(v) and v < 0 else "")
                )

            # -- Conditional ----------------------------------------
            elif op == "map_values":
                # Remap specific values -- arg is a JSON dict string or dict object
                mapping = t.get("arg", {})
                if isinstance(mapping, str):
                    try:
                        mapping = json.loads(mapping)
                    except Exception:


                        mapping = {}
                if mapping:
                    df[c] = df[c].astype(str).str.strip().replace(mapping)
            elif op == "coalesce":
                # Fill nulls in col c from fallback column named in arg
                fallback_col = str(t.get("arg", ""))
                if fallback_col in df.columns:
                    df[c] = df[c].where(df[c].notna() & (df[c].astype(str).str.strip() != ""),
                                         df[fallback_col])

    # 4. Pre-aggregation
    # Column names are looked up case-insensitively so col_map renames don't break agg specs.
    if agg and agg.get("group_by") and agg.get("agg_col"):
        col_upper = {c.upper(): c for c in df.columns}
        group_by = [col_upper[g.upper()] for g in agg["group_by"] if g.upper() in col_upper]
        agg_col  = col_upper.get(agg["agg_col"].upper())
        agg_fn   = agg.get("agg_fn", "sum").lower()
        if group_by and agg_col:


            # Keep only key + agg column after groupby -- drop extra columns so
            # auto-align in _prepare_recon can match the single value column.
            df = df.groupby(group_by, as_index=False).agg({agg_col: agg_fn})

    return df


def _resolve_keys(key_cols_raw: list[str], src_df: pd.DataFrame, tgt_df: pd.DataFrame):
    # Resolve a list of desired key column names to names that actually exist in
    # both DataFrames (case-insensitive).  Returns (manual_keys, tgt_df, missing).
    # - manual_keys: list of resolved column names (may be empty if resolution fails)
    # - tgt_df: possibly updated DataFrame with columns renamed to match src names
    # - missing: list of key parts that could not be found in one or both files
    manual_keys: list[str] = []
    missing: list[str] = []

    for desired in key_cols_raw:
        src_match = next((c for c in src_df.columns if c.upper() == desired.upper()), None)
        tgt_match = next((c for c in tgt_df.columns if c.upper() == desired.upper()), None)

        if src_match and tgt_match:
            # Rename tgt column to src name so both sides align
            if tgt_match != src_match:
                tgt_df = tgt_df.rename(columns={tgt_match: src_match})
            manual_keys.append(src_match)
        elif src_match:
            missing.append(f"'{desired}' found in source but not in target")
        elif tgt_match:
            missing.append(f"'{desired}' found in target but not in source")
        else:
            missing.append(f"'{desired}' not found in either file")

    return manual_keys, tgt_df, missing


def _prepare_recon(src_df: pd.DataFrame, tgt_df: pd.DataFrame, params: dict):
    # Apply the full pipeline: parse_cols -> col_map -> transforms -> agg -> key resolution.
    # Returns (src_df, tgt_df, manual_keys, exclude, key_warning).
    transforms   = params.get("transforms", [])
    col_map      = params.get("col_map", {}) or {}
    parse_cols   = params.get("parse_cols", []) or []
    key_cols_raw = params.get("key_cols") or (
        [params["key_col"]] if params.get("key_col") else []
    )
    exclude = params.get("exclude", [])

    # Apply parse_cols + col_map + transforms + agg to each side
    src_df = _apply_recon_params(src_df, params.get("src_agg"), transforms, "src",
                                  col_map=col_map, parse_cols=parse_cols)
    tgt_df = _apply_recon_params(tgt_df, params.get("tgt_agg"), transforms, "tgt",
                                  col_map=col_map, parse_cols=parse_cols)

    # Align value columns so both sides share the same column name for comparison.
    # Priority: explicit src_value_col / tgt_value_col from params (LLM-extracted),
    # then fall back to auto-detecting a single unmatched column.
    key_cols_upper = {k.upper() for k in key_cols_raw}
    tgt_agg_col = (params.get("tgt_agg") or {}).get("agg_col", "") or params.get("tgt_value_col", "")
    src_agg_col = (params.get("src_agg") or {}).get("agg_col", "") or params.get("src_value_col", "")

    # Prefer explicit src_value_col -> rename it to match tgt_value_col / tgt_agg_col
    explicit_src_val = params.get("src_value_col", "")
    explicit_tgt_val = params.get("tgt_value_col", "") or tgt_agg_col
    if explicit_src_val and explicit_tgt_val and explicit_src_val != explicit_tgt_val:
        col_upper_src = {c.upper(): c for c in src_df.columns}
        actual_src = col_upper_src.get(explicit_src_val.upper())
        if actual_src and explicit_tgt_val not in src_df.columns:
            src_df = src_df.rename(columns={actual_src: explicit_tgt_val})

    if tgt_agg_col and tgt_agg_col in tgt_df.columns and tgt_agg_col not in src_df.columns:
        # Fall back: find the single non-key, non-excluded src column not in tgt -- rename it
        src_unmatched = [c for c in src_df.columns
                          if c.upper() not in key_cols_upper
                          and c not in exclude
                          and c not in tgt_df.columns]
        if len(src_unmatched) == 1:
            src_df = src_df.rename(columns={src_unmatched[0]: tgt_agg_col})

    if src_agg_col and src_agg_col in src_df.columns and src_agg_col not in tgt_df.columns:
        tgt_unmatched = [c for c in tgt_df.columns
                          if c.upper() not in key_cols_upper
                          and c not in exclude
                          and c not in src_df.columns]
        if len(tgt_unmatched) == 1:
            tgt_df = tgt_df.rename(columns={tgt_unmatched[0]: src_agg_col})

    # Resolve composite key
    manual_keys, tgt_df, missing = _resolve_keys(key_cols_raw, src_df, tgt_df)

    key_warning = None
    if missing:
        key_warning = (
            f"Could not resolve key part(s): {', '.join(missing)}. "
            f"Source columns after transform: {list(src_df.columns)}. "
            f"Target columns after transform: {list(tgt_df.columns)}."
        )
    elif not manual_keys and key_cols_raw:
        key_warning = (
            f"Key resolution failed for {key_cols_raw}. "
            f"Fell back to content-based comparison."
        )

    # Collect explicit value columns -- from agg specs and from src_value_col/tgt_value_col.
    # Passed to compare_dataframes as force_data_cols so _key_based_diff skips
    # _split_meta_cols for them (aggregated numeric columns can look like surrogate keys).
    explicit_value_cols: set[str] = set()

    col_upper_final = {c.upper(): c for c in list(src_df.columns) + list(tgt_df.columns)}
    for agg in [params.get("tgt_agg"), params.get("src_agg")]:
        if agg and agg.get("agg_col"):
            resolved = col_upper_final.get(agg["agg_col"].upper(), agg["agg_col"])
            explicit_value_cols.add(resolved)
    for vkey in ("src_value_col", "tgt_value_col"):
        v = params.get(vkey, "")
        if v:
            resolved = col_upper_final.get(v.upper(), v)
            explicit_value_cols.add(resolved)

    force_data_cols = list(explicit_value_cols) if explicit_value_cols else None
    return src_df, tgt_df, manual_keys or None, exclude, key_warning, force_data_cols


def _run_llm_recon_full(session_id: str, username: str = "default") -> dict | None:
    """"run recon" -- the AI Copilot's fully automated reconciliation: reads the
    saved rules for this schema, has the LLM turn them into structured execution
    parameters (key_cols, col_map, parse_cols, transforms, aggregation, exclude),
    applies them to both DataFrames, runs the same deterministic compare_dataframes()
    engine as Standard mode, and returns the same rich response shape /analyze and
    /rerun use so the frontend can render it with the identical report UI.
    Returns None if the session has no stored DataFrames to rerun against."""
    stored = _results_store.get(session_id)
    if not stored or "dataframes" not in stored or len(stored["dataframes"]) < 2:
        return None

    context = _chat_contexts.get(session_id, {})
    dataframes: list[tuple[str, object]] = [
        (item["name"], item["df"]) for item in stored["dataframes"]
    ]
    base_name, base_df = dataframes[0]
    cmp_name, cmp_df = dataframes[1]

    fingerprint = stored.get("dataset_fingerprint") or context.get("dataset_fingerprint", "")
    saved_rules = _fp_get_rules(username, fingerprint, cols1=list(base_df.columns), cols2=list(cmp_df.columns),
                                 file_names=[base_name, cmp_name])
    recon_rules = [r for r in saved_rules if r.get("category") not in ("recon_hints",)]
    params = _parse_recon_rules_to_params(recon_rules, list(base_df.columns), list(cmp_df.columns))
    llm_error = params.pop("_llm_error", None)

    src_df, tgt_df, manual_keys, exclude, key_warning, force_data_cols = _prepare_recon(
        base_df.copy(), cmp_df.copy(), params
    )

    # A saved rule like "fuzzy match on customer_name" routes to the
    # probabilistic engine instead of the exact-key one -- same transforms
    # applied above (parse_cols/col_map/transforms via _prepare_recon), just
    # matched by similarity score instead of exact key equality.
    fuzzy_fields = [f for f in (params.get("fuzzy_fields") or [])
                     if f in src_df.columns and f in tgt_df.columns]
    if fuzzy_fields:
        threshold = float(params.get("fuzzy_threshold") or 0.75)
        try:
            fr = fuzzy_match_dataframes(src_df, tgt_df, fuzzy_fields, threshold=threshold, exclude_cols=exclude)
        except HTTPException as exc:
            fr = None
            llm_error = (llm_error + " " if llm_error else "") + f"Fuzzy match failed: {exc.detail}"
        if fr is not None:
            new_session_id = str(uuid.uuid4())
            _chat_contexts[new_session_id] = {
                **context,
                "action": "compare",
                "mode": "recon",
                "dataset_fingerprint": fingerprint,
                "comparisons": [{"pair": f"{base_name}→{cmp_name}",
                                  "matched": fr["counts"]["matched"], "modified": fr["counts"]["modified"]}],
            }
            _results_store[new_session_id] = {
                **stored,
                "dataframes": stored["dataframes"],
                "dataset_fingerprint": fingerprint,
                "recon_params_applied": params,
                "recon_key_warning": key_warning,
            }
            return {
                "session_id": new_session_id,
                "fuzzy": True,
                "fuzzy_fields": fr["fuzzy_fields"],
                "threshold": fr["threshold"],
                "counts": fr["counts"],
                "files": {
                    "file1": {"name": base_name, "rows": len(src_df), "columns": len(src_df.columns)},
                    "file2": {"name": cmp_name, "rows": len(tgt_df), "columns": len(tgt_df.columns)},
                },
                "matched_rows": fr["matched_rows"],
                "modified_rows": fr["modified_rows"],
                "only_in_file1": fr["file1_only"],
                "only_in_file2": fr["file2_only"],
                "fingerprint": fingerprint,
                "rules_applied": len(recon_rules),
                "llm_error": llm_error,
            }

    diff = compare_dataframes(src_df, tgt_df, manual_keys, True, exclude, force_data_cols=force_data_cols)
    mapping = analyze_mapping(src_df, tgt_df, base_name, cmp_name, None)

    new_session_id = str(uuid.uuid4())
    _chat_contexts[new_session_id] = {
        **context,
        "action": "compare",
        "mode": "recon",
        "dataset_fingerprint": fingerprint,
        "comparisons": [{"pair": f"{base_name}→{cmp_name}",
                          "added": diff.get("added_count", 0),
                          "removed": diff.get("removed_count", 0),
                          "modified": diff.get("modified_count", 0)}],
    }
    _results_store[new_session_id] = {
        **stored,
        "dataframes": stored["dataframes"],  # preserve original (pre-transform) files for chained runs
        "pairs": [{
            "file1_name": base_name, "file2_name": cmp_name,
            "file1_format": base_df.attrs.get("_format", ""), "file2_format": cmp_df.attrs.get("_format", ""),
            "file1_delimiter": base_df.attrs.get("_delimiter", ""), "file2_delimiter": cmp_df.attrs.get("_delimiter", ""),
            "diff": diff, "mapping": mapping,
        }],
        "dataset_fingerprint": fingerprint,
        "recon_params_applied": params,
        "recon_key_warning": key_warning,
    }

    modified = [
        {"key": ", ".join(f"{k}={v}" for k, v in mr["key_values"].items()), "changes": mr["changes"]}
        for mr in diff.get("modified_rows", [])
    ]
    only1 = [
        {"key": ", ".join(f"{k}={v}" for k, v in r["key_values"].items()), "row": r.get("row_data", {})}
        for r in diff.get("file1_only", [])
    ]
    only2 = [
        {"key": ", ".join(f"{k}={v}" for k, v in r["key_values"].items()), "row": r.get("row_data", {})}
        for r in diff.get("file2_only", [])
    ]
    resp = {
        "session_id": new_session_id,
        "counts": {
            "matched": diff.get("file1_rows", 0) - diff.get("removed_count", 0),
            "file1_only": diff.get("file1_only_count", 0),
            "file2_only": diff.get("file2_only_count", 0),
            "modified": diff.get("modified_count", 0),
        },
        "keys": diff.get("key_columns", []),
        "method": diff.get("key_method", ""),
        "fingerprint": fingerprint,
        "type_mismatches": [{"column": k, **v} for k, v in diff.get("type_mismatches", {}).items()],
        "null_column_exceptions": diff.get("null_column_exceptions", []),
        "duplicates": {
            "file1": {"duplicate_rows": diff.get("file1_duplicate_count", 0), "rows": diff.get("file1_duplicate_rows", [])},
            "file2": {"duplicate_rows": diff.get("file2_duplicate_count", 0), "rows": diff.get("file2_duplicate_rows", [])},
        },
        "modified": modified,
        "only_in_file1": only1,
        "only_in_file2": only2,
        "schema_added_columns": diff.get("schema_added_columns", []),
        "schema_removed_columns": diff.get("schema_removed_columns", []),
        "excluded_meta_cols": diff.get("excluded_meta_cols", []),
        "waterfall": diff.get("waterfall", {}),
        "gross_break_total": diff.get("gross_break_total", 0),
        "net_break_total": diff.get("net_break_total", 0),
        "files": {
            "file1": {"name": base_name, "rows": len(src_df), "columns": len(src_df.columns),
                      "format": str(base_df.attrs.get("_format", "") or "").upper(),
                      "all_columns": list(src_df.columns)},
            "file2": {"name": cmp_name, "rows": len(tgt_df), "columns": len(tgt_df.columns),
                      "format": str(cmp_df.attrs.get("_format", "") or "").upper(),
                      "all_columns": list(tgt_df.columns)},
        },
        "common_columns": len(set(src_df.columns) & set(tgt_df.columns)),
        "common_columns_list": diff.get("common_columns", sorted(set(src_df.columns) & set(tgt_df.columns))),
        "elapsed": 0,
        "logs": [],
        "rules_applied": len(recon_rules),
        "params_applied": params,
        "key_warning": key_warning,
        "llm_error": llm_error,
    }
    _results_store[new_session_id]["_digest"] = resp
    return resp


@app.post("/recon/run-llm/{session_id}")
async def recon_run_llm(session_id: str, request: Request):
    """Direct entry point for AI Copilot mode's "Run Reconciliation" button --
    same LLM-driven rule pipeline as typing "run recon" in chat, but triggered
    immediately on run instead of requiring a follow-up chat message."""
    _rrl_username = _ws_resolve_username(request) or "default"
    result = await asyncio.to_thread(_run_llm_recon_full, session_id, _rrl_username)
    if result is None:
        raise HTTPException(404, "Session not found or missing stored files -- please re-upload.")
    return JSONResponse(_sanitize_json(result))


@app.post("/recon/run/{session_id}")
async def recon_run(session_id: str, request: Request):

    # Execute a reconciliation for a lineage session using the saved rules.


    # Reads the DataFrames from the session store, applies parse_cols + col_map +
    # transforms + pre-aggregation from saved rules, runs compare_dataframes,
    # and returns a structured result the Copilot can render.

    stored = _results_store.get(session_id, {})
    context = _chat_contexts.get(session_id, {})

    if not stored or context.get("mode") != "recon":
        return JSONResponse({"error": "Session not found or not a recon session."},
                             status_code=404)

    dfs = stored.get("dataframes", [])
    if len(dfs) < 2:
        return JSONResponse({"error": "Two files are required for reconciliation."},
                             status_code=400)

    src_name, src_df = dfs[0]["name"], dfs[0]["df"].copy()
    tgt_name, tgt_df = dfs[1]["name"], dfs[1]["df"].copy()

    try:


        saved_rules = _fp_get_rules(
            _ws_resolve_username(request) or "default",
            stored.get("dataset_fingerprint", ""),
            cols1=list(src_df.columns),
            cols2=list(tgt_df.columns),
        )

        recon_rules = [r for r in saved_rules if r.get("category") not in ("recon_hints",)]

        params = await asyncio.to_thread(
            _parse_recon_rules_to_params, recon_rules, list(src_df.columns), list(tgt_df.columns)
        )

        src_df, tgt_df, manual_keys, exclude, key_warning, force_data_cols = _prepare_recon(
            src_df, tgt_df, params
        )

        # Debug snapshot -- columns + 3 sample rows after all transforms/parses
        debug_info = {
            "params_parsed": params,


            "src_cols_after": list(src_df.columns),
            "tgt_cols_after": list(tgt_df.columns),
            "manual_keys":    manual_keys,
            "force_data_cols": force_data_cols,
            "src_sample":    src_df.head(3).astype(str).to_dict(orient="records"),
            "tgt_sample":    tgt_df.head(3).astype(str).to_dict(orient="records"),
        }

        diff = compare_dataframes(
            src_df, tgt_df, manual_keys, True, exclude,
            force_data_cols=force_data_cols,
        )

        summary = {
            "src_name":    src_name,
            "tgt_name":    tgt_name,
            "src_rows":    diff.get("file1_rows", len(src_df)),
            "tgt_rows":    diff.get("file2_rows", len(tgt_df)),
            "added":       diff.get("added_count", 0),


            "removed":    diff.get("removed_count", 0),
            "modified":   diff.get("modified_count", 0),
            "key_used":   diff.get("key_columns", manual_keys or []),
            "key_method": diff.get("key_method", ""),
            "params_applied": params,
            "col_stats":  diff.get("col_stats", [])[:20],
            "modified_rows": diff.get("modified_rows", [])[:50],
            "added_rows":  diff.get("file2_only", [])[:50],
            "removed_rows": diff.get("file1_only", [])[:50],
            "key_warning": key_warning,
            "debug":      debug_info,
            "waterfall":  diff.get("waterfall", {}),
            "gross_break_total": diff.get("gross_break_total", 0),
            "net_break_total":  diff.get("net_break_total", 0),
        }
        return JSONResponse(_sanitize_json(summary))

    except Exception as exc:
        import traceback


        return JSONResponse(
            {"error": f"{type(exc).__name__}: {exc}", "traceback": traceback.format_exc()},
            status_code=500,
        )


@app.get("/recon/download/{session_id}")
async def recon_download(session_id: str, request: Request):
    _require_not_readonly(request)

    # Download the last recon run for a lineage session as a formatted Excel workbook.
    # Sheets: Summary, Value_Breaks, Source_Only, Target_Only.

    stored = _results_store.get(session_id, {})
    context = _chat_contexts.get(session_id, {})

    if not stored or context.get("mode") != "recon":
        raise HTTPException(404, "Session not found or not a recon session.")

    dfs = stored.get("dataframes", [])


    if len(dfs) < 2:
        raise HTTPException(400, "Two files are required for reconciliation.")

    src_name = dfs[0]["name"]
    tgt_name = dfs[1]["name"]
    src_df  = dfs[0]["df"].copy()
    tgt_df  = dfs[1]["df"].copy()

    saved_rules = _fp_get_rules(_ws_resolve_username(request) or "default",
                                 stored.get("dataset_fingerprint", ""))
    recon_rules = [r for r in saved_rules if r.get("category") not in ("recon_hints",)]
    params = await asyncio.to_thread(_parse_recon_rules_to_params, recon_rules, list(src_df.columns),
                                      list(tgt_df.columns))

    src_df, tgt_df, manual_keys, exclude, _, force_data_cols = _prepare_recon(
        src_df, tgt_df, params)

    try:
        diff = compare_dataframes(
            src_df, tgt_df, manual_keys, True, exclude,
            force_data_cols=force_data_cols,


        )
    except Exception as exc:
        raise HTTPException(500, f"Comparison failed: {exc}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    key_cols = diff.get("key_columns", manual_keys or [])
    key_str_list = key_cols  # alias

    # -- Summary sheet ----------------------------------------

    ws = wb.create_sheet("Summary")
    _hdr(ws, 1, ["Metric", "Value"])
    total = max(diff.get("file1_rows", 0), diff.get("file2_rows", 0))
    breaks = (diff.get("added_count", 0) + diff.get("removed_count", 0) +
              diff.get("modified_count", 0))
    summary_rows = [
        (src_name,      src_name),
        (tgt_name,      tgt_name),


        ("Key column(s)",   ", ".join(key_cols) or "content-based"),
        ("Key method",      diff.get("key_method", "")),
        (f"{src_name} rows", diff.get("file1_rows", len(src_df))),
        (f"{tgt_name} rows", diff.get("file2_rows", len(tgt_df))),
        ("Matched rows",    diff.get("file1_rows", 0) - diff.get("removed_count", 0) -
                            diff.get("modified_count", 0)),
        ("Value breaks",    diff.get("modified_count", 0)),
        (f"{src_name}-only rows", diff.get("removed_count", 0)),
        (f"{tgt_name}-only rows", diff.get("added_count", 0)),
        ("Total breaks",    breaks),
        ("Break rate",      f"{(breaks/total*100):.1f}%" if total else "--"),
        ("Status",          "PASS" if breaks == 0 else ("WARN" if breaks < total * 0.05 else "FAIL")),
    ]
    col_map = params.get("col_map") or {}
    if params.get("tgt_agg"):
        agg = params["tgt_agg"]
        summary_rows.append((f"{tgt_name} aggregation",
            f"{agg.get('agg_fn','sum')}({agg.get('agg_col','?')}) grouped by {', '.join(agg.get('group_by', []))}"))


    if params.get("src_agg"):
        agg = params["src_agg"]
        summary_rows.append((f"{src_name} aggregation",
            f"{agg.get('agg_fn','sum')}({agg.get('agg_col','?')}) grouped by {', '.join(agg.get('group_by', []))}"))
    if col_map:
        summary_rows.append(("Column mapping", ", ".join(f"{k}->{v}" for k, v in col_map.items())))
    for pc in (params.get("parse_cols") or []):
        summary_rows.append((f"Parsed column ({pc.get('side','?')})",
            f"{pc.get('source_col','?')} -> {pc.get('new_col','?')} via /{pc.get('pattern','?')}/"))
    for ri, (k, v) in enumerate(summary_rows, 2):
        ws.cell(ri, 1, k).font = Font(bold=True)
        c = ws.cell(ri, 2, str(v))
        if k == "Status":
            c.fill = _status_fill(str(v))
        elif ri % 2 == 0:
            ws.cell(ri, 1).fill = _ALT_FILL
            c.fill = _ALT_FILL
    _autofit(ws)


    # -- Value Breaks sheet
    mod_rows = diff.get("modified_rows", [])
    if mod_rows:
        ws = wb.create_sheet("Value Breaks")
        _hdr(ws, 1, ["Key", "Column", src_name, tgt_name, "Difference"])
        ri = 2
        for mr in mod_rows:
            key_str = ", ".join(f"{k}={v}" for k, v in mr.get("key_values", {}).items())
            for col, chg in mr.get("changes", {}).items():
                ws.cell(ri, 1, key_str)
                ws.cell(ri, 2, col)
                src_val = chg.get("file1", "")
                tgt_val = chg.get("file2", "")
                ws.cell(ri, 3, src_val)
                ws.cell(ri, 4, tgt_val)
                try:
                    diff_val = float(tgt_val) - float(src_val)
                    c = ws.cell(ri, 5, round(diff_val, 6))


                    c.font = Font(color="006400" if diff_val >= 0 else "8B0000")
                except (TypeError, ValueError):
                    ws.cell(ri, 5, "")
                if ri % 2 == 0:
                    for ci in range(1, 6):
                        ws.cell(ri, ci).fill = _ALT_FILL
                ri += 1
        _autofit(ws)

    # -- Source-Only sheet
    src_only = diff.get("file1_only", [])
    if src_only:
        ws = wb.create_sheet(f"{src_name[:28]} Only"[:31])
        # Derive column list from actual row_data to avoid _MAX_DATA_COLS truncation
        seen: dict = {}
        for row in src_only:
            for col in row.get("row_data", {}).keys():
                seen[col] = True
        data_cols = list(seen.keys())


        all_cols = key_cols + [c for c in data_cols if c not in key_cols]
        _hdr(ws, 1, all_cols)
        for ri, row in enumerate(src_only, 2):
            kv = row.get("key_values", {})
            rd = row.get("row_data", {})
            for ci, col in enumerate(all_cols, 1):
                ws.cell(ri, ci, kv.get(col, rd.get(col, "")))
            if ri % 2 == 0:
                for ci in range(1, len(all_cols) + 1):
                    ws.cell(ri, ci).fill = _ALT_FILL
        _autofit(ws)

    # -- Target-Only sheet
    tgt_only = diff.get("file2_only", [])
    if tgt_only:
        ws = wb.create_sheet(f"{tgt_name[:28]} Only"[:31])
        seen = {}
        for row in tgt_only:
            for col in row.get("row_data", {}).keys():


                seen[col] = True
        data_cols = list(seen.keys())
        all_cols = key_cols + [c for c in data_cols if c not in key_cols]
        _hdr(ws, 1, all_cols)
        for ri, row in enumerate(tgt_only, 2):
            kv = row.get("key_values", {})
            rd = row.get("row_data", {})
            for ci, col in enumerate(all_cols, 1):
                ws.cell(ri, ci, kv.get(col, rd.get(col, "")))
            if ri % 2 == 0:
                for ci in range(1, len(all_cols) + 1):
                    ws.cell(ri, ci).fill = _ALT_FILL
        _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Use actual file names (stripped of extension and unsafe chars) for the download filename


    def _safe_stem(name: str) -> str:
        stem = re.sub(r"\.[^.]+$", "", name)     # drop extension
        return re.sub(r"[^a-zA-Z0-9_-]", "_", stem)[:30]

    fname = f"recon_{_safe_stem(src_name)}_vs_{_safe_stem(tgt_name)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.post("/agent-chat")
async def agent_chat(request: Request):

    # AI Copilot endpoint -- backed by the LangChain agent (agent/executor.py).

    # Uses ConversationBufferWindowMemory for multi-turn context and the full
    # LangChain tool suite (compare_files, check_data_quality, map_columns,


    # run_governance_check, save_rule, list_rules, delete_rule).

    # Files are registered with the agent during /analyze so tools can reload
    # them without re-upload.

    # Request JSON:
    # { "session_id": "<id>", "question": "<user message>" }

    # Response JSON:
    # { "reply": "<agent answer>", "tools_called": ["compare_files", ...] }

    from agent.executor import run_agent
    from agent.tools import make_tool_dispatch

    body    = await request.json()
    session_id = body.get("session_id", "")
    question   = body.get("question", "").strip()

    if not question:
        return JSONResponse({"error": "Empty question"}, status_code=400)


    if not session_id:
        return JSONResponse({"error": "session_id is required."}, status_code=400)

    try:
        _raw_results = _results_store.get(session_id, {})
        # _summarize_result() (agent/tools.py) expects the same digested shape
        # /analyze already returns to the frontend (dimensions/score/counts/etc),
        # not the raw internal storage (quality_reports/pairs/dataframes/...) --
        # merge in the digest snapshot captured when that response was built.
        session_results = {**_raw_results, **_raw_results.get("_digest", {})}
        dispatch = make_tool_dispatch(session_results, {})
        tools_called: list[str] = []

        def _tracked_dispatch(name, arguments):
            tools_called.append(name)
            return dispatch(name, arguments)

        history = _agent_chat_history.get(session_id, [])
        fp = session_results.get("dataset_fingerprint", "")
        extra_rules = _fp_rules_text(_ws_resolve_username(request) or "default", fp, module="agent") if fp else ""

        answer = run_agent(question, history, _tracked_dispatch, extra_rules)

        history = history + [
            {"role": "user", "content": question},
            {"role": "assistant", "content": answer},
        ]
        _agent_chat_history[session_id] = history[-12:]

        return JSONResponse({"reply": answer, "tools_called": tools_called})
    except ValueError as exc:
        # Raised by agent tools when no files are registered yet
        return JSONResponse({"reply": str(exc), "tools_called": []})
    except Exception as exc:
        return JSONResponse({"reply": f"Agent error: {exc}", "tools_called": []},
                             status_code=500)


# -- Dataset Memory REST endpoints


@app.get("/rules/{session_id}")
async def get_rules_endpoint(session_id: str, request: Request):
    """Return saved rules for the dataset loaded in this session."""
    username = _ws_resolve_username(request) or "default"
    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")
    if not fp:
        return JSONResponse({"fingerprint": "", "rules": [], "label": ""})
    dfs = _results_store.get(session_id, {}).get("dataframes", [])
    cols1 = list(dfs[0]["df"].columns) if len(dfs) > 0 else []
    cols2 = list(dfs[1]["df"].columns) if len(dfs) > 1 else []
    # Resolve to the best matching fingerprint (exact or fuzzy) so the UI
    # always operates on the fingerprint that actually holds the rules.
    file_names = _results_store.get(session_id, {}).get("file_names", [])
    action    = _results_store.get(session_id, {}).get("action", None)
    resolved_fp = _fp_resolve(username, fp, cols1=cols1, cols2=cols2, file_names=file_names)
    if resolved_fp != fp:
        _results_store[session_id]["dataset_fingerprint"] = resolved_fp
    return JSONResponse({
        "fingerprint": resolved_fp,


        "rules": _fp_get_rules(username, resolved_fp, cols1=cols1, cols2=cols2,
                               file_names=file_names, module=action),
        "label": _fp_get_label(username, resolved_fp),
    })


@app.get("/dataset-controls/{session_id}/rules")
async def dataset_controls_get_rules(session_id: str, request: Request, context: str = "quality"):
    """Return saved Dataset Controls rules for this session + context only."""
    username = _ws_resolve_username(request) or "default"
    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")
    if not fp:
        return JSONResponse({"ctx_rules": [], "count": 0})
    dfs     = _results_store.get(session_id, {}).get("dataframes", [])
    file_names = _results_store.get(session_id, {}).get("file_names", [])
    cols1 = list(dfs[0]["df"].columns) if len(dfs) > 0 else None
    cols2 = list(dfs[1]["df"].columns) if len(dfs) > 1 else None
    all_rules = _fp_get_rules(username, fp, cols1=cols1, cols2=cols2, file_names=file_names)
    ctx_prefix = f"dc_{context}_"
    ctx_rules = [


        {"rule": r["rule"], "category": r["category"],
         "base_cat": r["category"].replace(ctx_prefix, ""),
         "index": i + 1}  # i+1 = 1-based global index into all_rules -- matches /rules/delete
        for i, r in enumerate(all_rules)
        if r.get("category", "").startswith(ctx_prefix)
    ]
    return JSONResponse({"ctx_rules": ctx_rules, "count": len(ctx_rules)})


@app.post("/dataset-controls/{session_id}/apply")
async def dataset_controls_apply(session_id: str, request: Request):

    # Interpret a plain-English dataset control instruction, apply it to the
    # current results and save it to Dataset Memory.

    # The LLM returns a structured action:
    # - category:  filter | exclude | nullable | rule | override | pin | general
    # - rule_text: canonical rule to save (plain English)
    # - filter:    optional JS filter expression to apply to the results table


    # - feedback: user-facing confirmation message

    body = await request.json()
    instruction = body.get("instruction", "").strip()
    context    = body.get("context", "quality")  # quality | governance | profile
    columns    = body.get("columns", [])          # column names currently visible

    if not instruction:
        return JSONResponse({"error": "No instruction provided."}, status_code=400)

    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")
    file_names = _results_store.get(session_id, {}).get("file_names", [])
    label = " / ".join(file_names[:2]) if file_names else "this dataset"

    # Build a context-aware prompt for the LLM
    col_sample = ", ".join(columns[:30]) if columns else "unknown"
    prompt = f"""You are a data controls assistant for a BFSI data platform.
The user is viewing {context} results for dataset: "{label}".
Available columns (sample): {col_sample}

The user typed this instruction:
"{instruction}"

Interpret it and return ONLY a JSON object with these fields:
{{
"category": "<one of: filter, exclude, nullable, rule, override, pin, general>",
"rule_text": "<plain English rule to save to Dataset Memory, max 120 chars>",
"filter_type": "<one of: show_issues_only, show_pii_only, show_excluded, show_all, column_filter, none>",
"filter_columns": ["<col1>", "<col2>"],
"feedback": "<short user-facing confirmation, max 80 chars>"
}}

Category guide:
- filter: user wants to show/hide rows or columns in the current view
- exclude: user wants to exclude columns from analysis/scoring
- nullable: user wants to mark columns as optional/nullable
- rule: user is setting a validation rule (not-null, range, format etc.)

- override: user is overriding a classification (e.g. "not PII", change sensitivity)
- pin: user wants to pin/highlight specific columns
- general: anything else to remember

Examples:
"show only columns with issues" -> filter, show_issues_only
"exclude ETL_BATCH and LOAD_DT" -> exclude
"BusinessDate is not a date of birth" -> override, filter_columns: [BusinessDate]
"mark CUSIP as not PII" -> override, filter_columns: [CUSIP]
"set remarks as nullable" -> nullable, filter_columns: [remarks]
"show only PII columns" -> filter, show_pii_only
"notional must be positive" -> rule

Return ONLY valid JSON, no explanation."""
    try:
        raw = await asyncio.to_thread(_ask_llm, [{"role": "user", "content": [{"text": prompt}]}])
        import re as _re
        m = _re.search(r'\{.*\}', raw, _re.DOTALL)


        parsed = json.loads(m.group(0)) if m else {}
    except Exception:
        parsed = {
            "category":  "general",
            "rule_text": instruction,
            "filter_type": "none",
            "filter_columns": [],
            "feedback":  f"Saved: {instruction[:60]}",
        }

    # Save to Dataset Memory using a namespaced category:
    # "dc_{context}_{type}" -- completely isolated from recon rules
    # e.g. dc_quality_exclude, dc_governance_override, dc_profile_filter
    rule_text = parsed.get("rule_text", instruction)
    base_cat  = parsed.get("category", "general")
    # Namespace: prefix with context so rules don't bleed between recon/DQ/governance
    dc_category = f"dc_{context}_{base_cat}"


    if fp and rule_text:
        _dca_username = _ws_resolve_username(request) or "default"
        dfs  = _results_store.get(session_id, {}).get("dataframes", [])
        cols1 = list(dfs[0]["df"].columns) if len(dfs) > 0 else None
        cols2 = list(dfs[1]["df"].columns) if len(dfs) > 1 else None
        idx, _ = _fp_save(_dca_username, fp, rule_text, dc_category, dataset_label=label, cols1=cols1,
                           cols2=cols2, file_names=file_names)
        all_rules = _fp_get_rules(_dca_username, fp, cols1=cols1, cols2=cols2, file_names=file_names)
    else:
        idx, all_rules = 0, []

    # Return ONLY rules for this specific context with GLOBAL indices for deletion
    ctx_prefix = f"dc_{context}_"
    ctx_rules = [
        {"rule": r["rule"], "category": r["category"],
         "base_cat": r["category"].replace(ctx_prefix, ""),
         "index": i + 1}  # global 1-based index -- used by /rules/delete
        for i, r in enumerate(all_rules)
        if r.get("category", "").startswith(ctx_prefix)
    ]


    return JSONResponse({
        "ok":        True,
        "category":  dc_category,
        "rule_text": rule_text,
        "filter_type": parsed.get("filter_type", "none"),
        "filter_columns": parsed.get("filter_columns", []),
        "feedback":  parsed.get("feedback", f"Saved: {rule_text[:60]}"),
        "rule_index": idx,
        "ctx_rules": ctx_rules,  # only this context's rules
    })


@app.post("/rules/{session_id}/save")
async def save_rule_endpoint(session_id: str, request: Request):
    """Save a rule for the dataset loaded in this session."""
    body   = await request.json()
    rule   = body.get("rule", "").strip()
    category = body.get("category", "general")


    username = _ws_resolve_username(request) or "default"
    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")
    if not fp:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    if not rule:
        return JSONResponse({"error": "Empty rule"}, status_code=400)
    label = " / ".join(_results_store[session_id].get("file_names", [])[:2])
    dfs = _results_store.get(session_id, {}).get("dataframes", [])
    cols1 = list(dfs[0]["df"].columns) if len(dfs) > 0 else None
    cols2 = list(dfs[1]["df"].columns) if len(dfs) > 1 else None
    idx, queued = _fp_save(username, fp, rule, category, dataset_label=label, cols1=cols1,
                           cols2=cols2)
    return JSONResponse({"index": idx, "rules": _fp_get_rules(username, fp, cols1=cols1,
                         cols2=cols2), "queued": queued})


@app.post("/rules/{session_id}/delete")
async def delete_rule_endpoint(session_id: str, request: Request):
    """Delete a saved rule by 1-based index."""
    username = _ws_resolve_username(request) or "default"
    body  = await request.json()
    idx  = int(body.get("rule_index", 0))


    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")

    if not fp:

        return JSONResponse({"error": "Session not found"}, status_code=404)

    _, queued = _fp_delete(username, fp, idx)

    return JSONResponse({"rules": _fp_get_rules(username, fp), "queued": queued})


@app.post("/rules/{session_id}/update")

async def update_rule_endpoint(session_id: str, request: Request):


    # Edit a rule in-place (rule_index, rule, category) or reorder (rule_index, direction: up|down).

    # Returns the full updated rules list.


    body      = await request.json()

    idx       = int(body.get("rule_index", 0))

    direction = body.get("direction", "").strip()

    new_rule  = body.get("rule", "").strip()

    # No default here -- absent/empty means "leave the category as-is". A
    # default of "general" would silently strip a rule's dc_{module}_ module
    # namespace on every text-only edit (any caller that omits category,
    # which is every caller that's just fixing a typo) and make it vanish
    # from that module's context-filtered rule list.
    new_cat   = body.get("category")


    username = _ws_resolve_username(request) or "default"
    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")

    if not fp:

        return JSONResponse({"error": "Session not found"}, status_code=404)

    _, queued = _fp_update(username, fp, idx, rule_text=new_rule or None, category=new_cat or None,

                direction=direction or None)

    return JSONResponse({"rules": _fp_get_rules(username, fp), "queued": queued})


@app.get("/rules/{session_id}/export")

async def export_rules_endpoint(session_id: str, request: Request):

    """Export all saved rules for this session's dataset as a downloadable JSON file."""

    username = _ws_resolve_username(request) or "default"
    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")

    if not fp:

        raise HTTPException(404, "Session not found.")

    dfs       = _results_store.get(session_id, {}).get("dataframes", [])

    cols1     = list(dfs[0]["df"].columns) if dfs else []

    file_names = _results_store.get(session_id, {}).get("file_names", [])


    rules    = _fp_get_rules(username, fp, cols1=cols1, file_names=file_names)

    label    = _fp_get_label(username, fp) or " / ".join(file_names[:2])

    payload  = json.dumps({

      "version":     "1.0",

      "fingerprint": fp,

      "label":       label,

      "exported_at": datetime.utcnow().isoformat(),

      "rules":       rules,

    }, indent=2)

    safe = re.sub(r"[^a-zA-Z0-9_-]", "_", label)[:40] or "rules"

    return Response(

      content=payload,

      media_type="application/json",

      headers={"Content-Disposition": f'attachment; filename="{safe}_rules.json"'},

    )


@app.post("/rules/{session_id}/import")

async def import_rules_endpoint(session_id: str, request: Request):


    # Import rules from a JSON file (produced by /rules/export).

    # Merges rules into the current session's dataset fingerprint -- skips duplicates.


    username = _ws_resolve_username(request) or "default"
    fp = _results_store.get(session_id, {}).get("dataset_fingerprint", "")

    if not fp:

        raise HTTPException(404, "Session not found.")

    try:

        form = await request.form()

        f    = form.get("file")

        if not f or not hasattr(f, "read"):

            raise HTTPException(400, "No file uploaded.")

        content = await f.read()

        data    = json.loads(content)

    except json.JSONDecodeError as e:

        raise HTTPException(400, f"Invalid JSON: {e}")


    incoming = data.get("rules", [])

    if not isinstance(incoming, list):


        raise HTTPException(400, "rules must be a JSON array.")


    dfs       = _results_store.get(session_id, {}).get("dataframes", [])

    cols1     = list(dfs[0]["df"].columns) if dfs else None

    cols2     = list(dfs[1]["df"].columns) if len(dfs) > 1 else None

    file_names = _results_store.get(session_id, {}).get("file_names", [])

    label     = "/".join(file_names[:2])


    existing_texts = {r["rule"].strip().lower() for r in _fp_get_rules(username, fp)}

    imported = 0

    for r in incoming:

        rule_text = str(r.get("rule", "")).strip()

        category  = str(r.get("category", "general"))

        if not rule_text or rule_text.lower() in existing_texts:

            continue

        _fp_save(username, fp, rule_text, category, dataset_label=label, cols1=cols1, cols2=cols2)

        existing_texts.add(rule_text.lower())

        imported += 1


    return JSONResponse({

      "imported": imported,

      "skipped":  len(incoming) - imported,

      "rules":    _fp_get_rules(username, fp, cols1=cols1, cols2=cols2, file_names=file_names),

    })


@app.post("/api/dq/autofix/{session_id}")

async def autofix_endpoint(session_id: str, request: Request):


    # Generate auto-fix suggestions and a corrected CSV download for common

    # fixable DQ issues: leading/trailing whitespace, mixed date formats,

    # case standardisation for allowed-value columns, null sentinel normalisation.

    # Returns a CSV of the corrected DataFrame.


    stored = _results_store.get(session_id, {})

    if not stored:

        raise HTTPException(404, "Session not found.")


    body        = await request.json()

    target_file = (body.get("file_name") or "").strip()

    fix_types   = body.get("fix_types") or ["trim", "date_format", "case", "null_sentinel"]

    quality_reports = stored.get("quality_reports", [])

    dataframes      = stored.get("dataframes", [])


    # Find matching dataframe

    df_entry = None

    for entry in dataframes:

        if not target_file or entry["name"] == target_file:

            df_entry = entry

            break

    if df_entry is None:

        raise HTTPException(404, f"File '{target_file}' not found in session.")


    df = df_entry["df"].copy()


    # Find the quality report for this file to get rule failures

    qr = next((q for q in quality_reports if q.get("file_name") == df_entry["name"]), {})


    rule_results = qr.get("rule_results", [])


    fixes_applied: list[dict] = []


    # 1. Trim whitespace on all string columns

    if "trim" in fix_types:

        for col in df.select_dtypes(include="object").columns:

            before = df[col].astype(str).str.strip()

            changed = int((df[col].astype(str) != before).sum())

            if changed:

                df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)

                fixes_applied.append({"col": col, "fix": "trim_whitespace", "rows_fixed": changed})


    # 2. Normalise null sentinels ("N/A", "null", "NULL", "none", "-") -> empty string

    if "null_sentinel" in fix_types:

        _SENTINELS = {"n/a", "null", "none", "-", "na", "nan", "#n/a", "unknown", "?", ""}

        for col in df.select_dtypes(include="object").columns:

            mask = df[col].astype(str).str.strip().str.lower().isin(_SENTINELS - {""})


            changed = int(mask.sum())

            if changed:

                df.loc[mask, col] = None

                fixes_applied.append({"col": col, "fix": "null_sentinel", "rows_fixed": changed})


    # 3. Standardise case for columns with allowed_value failures (UPPER for short codes)

    if "case" in fix_types:

        failed_av_cols = [

          r.get("column") for r in rule_results

          if r.get("rule_type") == "allowed_values" and r.get("status") == "FAIL"

        ]

        for col in set(failed_av_cols):

            if col and col in df.columns and df[col].dtype == object:

                # Attempt upper-case standardisation (BUY/SELL, USD etc.)

                sample = df[col].dropna().astype(str).head(10)

                if sample.str.len().mean() <= 10:

                    before_vals = df[col].copy()

                    df[col] = df[col].apply(lambda v: v.strip().upper() if isinstance(v, str) else v)


                    changed = int((df[col] != before_vals).sum())

                    if changed:

                        fixes_applied.append({"col": col, "fix": "case_upper", "rows_fixed": changed})


    # 4. Standardise date formats -- normalise to YYYY-MM-DD

    if "date_format" in fix_types:

        for col in df.columns:

            if df[col].dtype == object:

                sample = df[col].dropna().astype(str).head(50)

                if sample.str.match(r'\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}').mean() > 0.5:

                    try:

                        converted = pd.to_datetime(df[col], dayfirst=False, errors="coerce")

                        changed = int(converted.notna().sum())

                        if changed:

                            df[col] = converted.dt.strftime("%Y-%m-%d").where(converted.notna(), df[col])

                            fixes_applied.append({"col": col, "fix": "date_format_iso", "rows_fixed": changed})

                    except Exception:


                        pass


    # Return corrected CSV + summary

    import io as _io

    buf = _io.StringIO()

    df.to_csv(buf, index=False)

    csv_bytes = buf.getvalue().encode("utf-8")


    safe = re.sub(r"[^a-zA-Z0-9_-]", "_", Path(df_entry["name"]).stem)[:30]

    return Response(

      content=csv_bytes,

      media_type="text/csv",

      headers={

        "Content-Disposition": f'attachment; filename="{safe}_fixed.csv"',

        "X-Fixes-Applied":     json.dumps(fixes_applied),

        "X-Fix-Count":         str(len(fixes_applied)),

      },

    )


@app.get("/download/{session_id}")

async def download(session_id: str, request: Request, fmt: str = "excel"):

    """Download analysis results as Excel or JSON."""
    _require_not_readonly(request)

    data = _results_store.get(session_id)

    if not data:

        raise HTTPException(404, "Session not found. Please re-run the analysis.")


    # Build a descriptive filename from the actual file names used in the run

    file_names = data.get("file_names", [])

    action     = data.get("action", "report")

    if file_names:

        parts = [re.sub(r"[^a-zA-Z0-9_-]", "_", Path(n).stem)[:20] for n in file_names[:2]]

        safe_name = f"{action}_{'_vs_'.join(parts)}"

    else:

        safe_name = f"{action}_{re.sub(r'[^a-zA-Z0-9_-]', '', session_id[:8])}"


    fmt = (fmt or "excel").strip().lower()

    if fmt == "json":
        buf = io.BytesIO(json.dumps(_sanitize_json(data), indent=2, default=str).encode("utf-8"))
        return StreamingResponse(
            buf, media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.json"'},
        )

    wb = generate_excel(data)

    if fmt == "csv":
        # Pick the largest non-Summary sheet and export it as CSV.
        best_ws = None
        for ws in wb.worksheets:
            if ws.title == "Summary":
                continue
            if best_ws is None or ws.max_row > best_ws.max_row:
                best_ws = ws
        best_ws = best_ws or wb.worksheets[0]
        csv_buf = io.StringIO()
        writer = _csv.writer(csv_buf)
        for row in best_ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
        buf = io.BytesIO(csv_buf.getvalue().encode("utf-8"))
        return StreamingResponse(
            buf, media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
        )

    # Default: Excel


    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    return StreamingResponse(

        buf,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},

    )


@app.post("/send-email")
async def send_email(request: Request):
    # Send analysis report via Outlook (Windows) or SMTP fallback.
    # For parse action, parse_idx selects which parsed file's CSV to attach.
    # For all other actions, the Excel report is attached.
    _require_not_readonly(request)

    body = await request.json()

    session_id = body.get("session_id", "")
    recipients = [e.strip() for e in body.get("emails", "").split(",") if e.strip()]

    if not recipients:
        return JSONResponse({"error": "No email address provided."}, status_code=400)

    data = _results_store.get(session_id)
    if not data:
        return JSONResponse({"error": "Session not found. Please re-run the analysis."}, status_code=404)

    from_email = (
        body.get("from_email", "").strip()
        or os.getenv("EMAIL_FROM", "").strip()
    )

    action = data.get("action", "analysis")
    _ACTION_LABELS = {
        "compare": "Reconciliation",

        "lineage": "Complex Recon",
        "quality": "Data Quality",
        "profile": "Data Profile",
        "parse":   "Parse",
        "governance": "Governance",
    }
    action_label = _ACTION_LABELS.get(action, action.title())
    subject = f"AI Agent -- Data Validation -- {action_label} Report"
    html_body = _build_email_html(data)

    # Build attachment -- CSV for parse, Excel for everything else
    attach_bytes: Optional[bytes] = None
    attach_name: str  = "report.xlsx"
    attach_mime: str  = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if action == "parse":
        parse_idx = body.get("parse_idx")
        reports   = data.get("parse_reports", [])

        pr     = reports[int(parse_idx)] if parse_idx is not None and int(parse_idx) < len(reports) else (reports[0] if reports else None)
        if pr and pr.get("rows"):
            import csv as _csv
            cols = pr.get("columns", [])
            buf = io.StringIO()
            w   = _csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(pr["rows"])
            base_name   = (pr.get("file_name") or "parsed").rsplit(".", 1)[0]
            attach_bytes = buf.getvalue().encode("utf-8")
            attach_name  = f"{base_name}_parsed.csv"
            attach_mime  = "text/csv"

    elif action == "lineage":
        # Complex Reconciliation -- use the same recon Excel as Download button
        # Rebuild Excel from saved rules + current session dataframes
        try:
            context = _chat_contexts.get(session_id, {})

            dfs = data.get("dataframes", [])
            if len(dfs) >= 2 and context.get("mode") == "recon":
                src_name = dfs[0]["name"]
                tgt_name = dfs[1]["name"]
                src_df  = dfs[0]["df"].copy()
                tgt_df  = dfs[1]["df"].copy()
                saved_rules = _fp_get_rules(_ws_resolve_username(request) or "default",
                                             data.get("dataset_fingerprint", ""))
                recon_rules = [r for r in saved_rules if r.get("category") not in ("recon_hints",)]
                params = await asyncio.to_thread(_parse_recon_rules_to_params, recon_rules, list(src_df.columns), list(tgt_df.columns))
                src_df, tgt_df, manual_keys, exclude, _, force_data_cols = _prepare_recon(src_df, tgt_df, params)
                diff = compare_dataframes(src_df, tgt_df, manual_keys, True, exclude, force_data_cols=force_data_cols)
                import openpyxl as _opxl
                wb = _opxl.Workbook()
                wb.remove(wb.active)
                key_cols = diff.get("key_columns", manual_keys or [])
                # Summary sheet
                ws = wb.create_sheet("Summary")

                _hdr(ws, 1, ["Metric", "Value"])
                for ri, (k, v) in enumerate([
                    (src_name, src_name), (tgt_name, tgt_name),
                    ("Key column(s)", ", ".join(key_cols)),
                    (f"{src_name} rows", diff.get("file1_rows", len(src_df))),
                    (f"{tgt_name} rows", diff.get("file2_rows", len(tgt_df))),
                    ("Value breaks", diff.get("modified_count", 0)),
                    (f"{src_name}-only rows", diff.get("file1_only_count", 0)),
                    (f"{tgt_name}-only rows", diff.get("file2_only_count", 0)),
                ], 2):
                    ws.cell(ri, 1, k).font = Font(bold=True)
                    ws.cell(ri, 2, str(v))
                _autofit(ws)
                # Value Breaks sheet
                if diff.get("modified_rows"):
                    ws2 = wb.create_sheet("Value Breaks")
                    _hdr(ws2, 1, ["Key", "Column", src_name, tgt_name, "Difference"])
                    ri = 2
                    for mr in diff["modified_rows"]:

                        key_str = " | ".join(f"{k}={v}" for k, v in mr.get("key_values", {}).items())
                        for col, chg in mr.get("changes", {}).items():
                            ws2.cell(ri, 1, key_str)
                            ws2.cell(ri, 2, col)
                            ws2.cell(ri, 3, str(chg.get("file1", "")))
                            ws2.cell(ri, 4, str(chg.get("file2", "")))
                            try:
                                diff_val = float(str(chg.get("file1","")).replace(",","")) - float(str(chg.get("file2","")).replace(",",""))
                                ws2.cell(ri, 5, round(diff_val, 6))
                            except Exception:
                                ws2.cell(ri, 5, "")
                            ri += 1
                    _autofit(ws2)
                buf = io.BytesIO()
                wb.save(buf)
                attach_bytes = buf.getvalue()
                def _safe_stem(n):
                    import re as _re

                    return _re.sub(r"[^a-zA-Z0-9_-]", "_", _re.sub(r"\.[^.]+$", "", n))[:30]
                attach_name = f"recon_{_safe_stem(src_name)}_vs_{_safe_stem(tgt_name)}.xlsx"
            else:
                # Fallback: lineage session context not available -- use generate_excel
                wb  = generate_excel(data)
                buf = io.BytesIO(); wb.save(buf)
                attach_bytes = buf.getvalue()
        except Exception:
            pass

    else:
        try:
            wb  = generate_excel(data)
            buf = io.BytesIO()
            wb.save(buf)
            attach_bytes = buf.getvalue()
        except Exception:
            pass

    if not from_email:
        return JSONResponse(
            {"error": "No From address provided. Set EMAIL_FROM in .env or supply from_email in the request."},
            status_code=400,
        )

    # Reuse the scheduler's send path instead of a second, independently
    # maintained SMTP implementation -- this one had drifted from it and
    # picked up three real bugs: defaulted to port 25 (blocked outbound on
    # Railway and most cloud platforms), used STARTTLS even for port 465
    # (which needs implicit TLS via SMTP_SSL, not a plaintext-then-upgrade
    # handshake), and never forced IPv4 DNS resolution (the same
    # "[Errno 101] Network is unreachable" issue already fixed for scheduled
    # jobs, since Outlook/Gmail's SMTP hosts resolve to IPv6 first and
    # Railway has no outbound IPv6 route).
    from workspace.scheduler import _deliver_email
    attachment = (attach_bytes, attach_name, attach_mime) if attach_bytes else None
    try:
        await asyncio.to_thread(
            _deliver_email, ", ".join(recipients), from_email, subject, html_body, attachment
        )
        return JSONResponse({"ok": True, "sent_to": recipients})
    except Exception as exc:
        return JSONResponse({"error": f"SMTP error: {exc}"}, status_code=500)

    # ------------------------------------------------------------------
    # Parse File -- dedicated standalone module
    # ------------------------------------------------------------------


@app.get("/parse-file", response_class=HTMLResponse)

async def parse_file_page(request: Request):

    """Render the dedicated Parse File page."""

    return templates.TemplateResponse(request=request, name="parse.html")


@app.post("/parse-file")

async def parse_file_api(request: Request):


    # Accept one or more uploaded files, parse each into a structured DataFrame,

    # and return JSON results suitable for the standalone Parse File UI.


    try:

        form = await request.form()

    except Exception as exc:

        return JSONResponse({"error": f"Form parse error: {exc}"}, status_code=400)

    files = [v for v in form.getlist("files") if hasattr(v, "filename") and v.filename]


# NOTE: page image very blurry/low-resolution; best-effort reconstruction below.  # OCR-UNCERTAIN
    if not files:

        return JSONResponse({"error": "No files provided"}, status_code=400)  # OCR-UNCERTAIN


    MAX_PREVIEW_ROWS = 1000  # OCR-UNCERTAIN

    results = []


    for upload in files:

        entry: dict = {"file_name": upload.filename}

        try:

            df = _load_file(upload)  # OCR-UNCERTAIN

            fmt = df.attrs.get("_format", "unknown")  # OCR-UNCERTAIN

            total_rows = len(df)

            truncated = total_rows > MAX_PREVIEW_ROWS

            df_preview = df.head(MAX_PREVIEW_ROWS) if truncated else df

            # Serialise all values as strings for safe JSON transport

            rows = df_preview.fillna("").astype(str).to_dict(orient="records")

            entry.update({


              "format_detected": fmt,

              "row_count":     total_rows,

              "col_count":     len(df.columns),

              "columns":       list(df.columns),

              "rows":          rows,

              "notes":         f"Showing first {MAX_PREVIEW_ROWS:,} rows of {total_rows:,}"

                                if truncated else "",

              "error":         None,

            })

        except HTTPException as exc:

            entry.update({"error": exc.detail, "notes": "", "rows": [], "columns": [],

                          "row_count": 0, "col_count": 0, "format_detected": "unknown"})

        except Exception as exc:

            entry.update({"error": str(exc), "notes": "", "rows": [], "columns": [], "row_count":

0, "col_count": 0, "format_detected": "unknown"})


        results.append(entry)


    return JSONResponse({"results": results})


#
# ================================================================
# ⚡ PERSONAL AUTOMATION WORKSPACE -- /api/ws/*
#
# ================================================================
# All routes require authentication (enforced by WorkspaceAuthMiddleware).

# Username is read from request.state.username (set by the middleware).

#
# ================================================================

def _ws_check() -> None:


    # Raise 503 if the workspace feature failed to initialise.

    if not _WS_ENABLED:

        raise HTTPException(status_code=503, detail="Workspace feature is not available.")


# -- connections
# --------------------------------------------------------------------


@app.get("/api/ws/connections")

async def ws_list_connections(request: Request):

    _ws_check()

    username = _ws_get_user(request)

    return JSONResponse(_ws_db.list_connections(username))


@app.post("/api/ws/connections")

async def ws_save_connection(request: Request):

    _ws_check()

    username = _ws_get_user(request)

    body = await request.json()

    name        = str(body.get("name", "")).strip()

    source_type = str(body.get("source_type", "")).strip().lower()

    config      = body.get("config", {})

    conn_id     = body.get("id")


    owner        = (body.get("owner") or "").strip() or None

    business_domain = (body.get("business_domain") or "").strip() or None

    sensitivity  = (body.get("sensitivity") or "").strip() or None

    description  = (body.get("description") or "").strip() or None


    if not name:

        raise HTTPException(400, "Connection name is required.")

    _VALID_SOURCE_TYPES = {

      "sftp","ftp","s3","mssql","postgres","mysql","oracle","snowflake","api",

      "azure_blob","azure_sql","db2","sharepoint","databricks","gcs","bloomberg",

      "refinitiv","murex","calypso","kafka","salesforce",

      "redshift","teradata","bigquery","mongodb",

    }

    if source_type not in _VALID_SOURCE_TYPES:

        raise HTTPException(400, f"Invalid source_type: {source_type!r}")

    if not isinstance(config, dict):

        raise HTTPException(400, "config must be a JSON object.")

    if conn_id:
        # Editing an existing connection: GET /api/ws/connections/{id} never
        # returns secret fields to the browser (by design), so an edit form
        # that doesn't touch the password/key field submits it blank. Without
        # this, a plain overwrite would silently wipe the working credential.
        # Preserve the previously stored value for any secret-bearing field
        # the incoming payload didn't include; everything else is a full
        # replace, so intentionally clearing a non-secret field still works.
        existing = _ws_db.get_connection(conn_id, username)
        if existing:
            for _secret_key in (
                "password", "private_key_path", "aws_secret_access_key",
                "client_secret", "security_token", "access_token",
                "api_key", "sas_token", "credentials_json",
            ):
                if not config.get(_secret_key) and existing["config"].get(_secret_key):
                    config[_secret_key] = existing["config"][_secret_key]

    saved_id = _ws_db.save_connection(

        username, name, source_type, config, conn_id,


        owner=owner, business_domain=business_domain,

        sensitivity=sensitivity, description=description,

    )

    return JSONResponse({"id": saved_id, "status": "ok"})


@app.get("/api/ws/connections/{conn_id}")

async def ws_get_connection(conn_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    rec = _ws_db.get_connection(conn_id, username)

    if not rec:

        raise HTTPException(404, f"Connection '{conn_id}' not found.")

    # Never return passwords / private keys -- return config with sensitive fields masked

    safe_config = {k: v for k, v in rec["config"].items() if k not in ("password",

"private_key_path", "aws_secret_access_key")}

    return JSONResponse({"id": rec["id"], "name": rec["name"], "source_type":

rec["source_type"], "config": safe_config})


@app.delete("/api/ws/connections/{conn_id}")

async def ws_delete_connection(conn_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    _ws_db.delete_connection(conn_id, username)

    return JSONResponse({"status": "deleted"})


@app.post("/api/ws/connections/{conn_id}/test")

async def ws_test_connection(conn_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    rec = _ws_db.get_connection(conn_id, username)

    if not rec:

        raise HTTPException(404, "Connection not found.")

    try:

        connector = _ws_connectors.BaseConnector.from_type(rec["source_type"], rec["config"])

        ok = connector.test_connection()


        return JSONResponse({"ok": ok, "message": "Connected successfully." if ok else "Connection failed."})

    except Exception as exc:

        return JSONResponse({"ok": False, "message": str(exc)})


@app.get("/api/ws/connections/{conn_id}/tables")
async def ws_list_tables(conn_id: str, request: Request):
    """List tables visible to a saved database connection, so the user can
    pick a real table name instead of guessing one before running a job."""
    _ws_check()
    username = _ws_get_user(request)
    rec = _ws_db.get_connection(conn_id, username)
    if not rec:
        raise HTTPException(404, "Connection not found.")
    try:
        connector = _ws_connectors.BaseConnector.from_type(rec["source_type"], rec["config"])
        tables = connector.list_tables()
        return JSONResponse({"tables": tables})
    except NotImplementedError as exc:
        raise HTTPException(400, str(exc))
    except Exception as exc:
        raise HTTPException(400, f"Could not list tables: {exc}")


@app.get("/api/ws/connections/{conn_id}/preview")
async def ws_preview_connection(conn_id: str, request: Request):
    """Fetch a small sample of data from a saved connection so the user can
    sanity-check the config before using it in a job."""
    _ws_check()
    username = _ws_get_user(request)
    rec = _ws_db.get_connection(conn_id, username)
    if not rec:
        raise HTTPException(404, "Connection not found.")
    try:
        connector = _ws_connectors.BaseConnector.from_type(rec["source_type"], rec["config"])
        df = connector.fetch()
        preview = df.head(20)
        return JSONResponse({
            "columns": list(preview.columns),
            "rows": preview.fillna("").astype(str).to_dict(orient="records"),
            "total_rows": len(df),
        })
    except Exception as exc:
        raise HTTPException(400, f"Preview failed: {exc}")


# -- rule sets
# --------------------------------------------------------------------


@app.get("/api/ws/rulesets")

async def ws_list_rulesets(request: Request):

    _ws_check()

    username = _ws_get_user(request)

    return JSONResponse(_ws_db.list_rulesets(username))


@app.post("/api/ws/rulesets")

async def ws_save_ruleset(request: Request):

    _ws_check()

    username = _ws_get_user(request)


    body = await request.json()

    name        = str(body.get("name", "")).strip()

    description = str(body.get("description", "")).strip()

    rules       = body.get("rules", [])

    rs_id       = body.get("id")


    if not name:

        raise HTTPException(400, "Rule set name is required.")

    if not isinstance(rules, list):

        raise HTTPException(400, "rules must be a JSON array.")


    saved_id = _ws_db.save_ruleset(username, name, description, rules, rs_id)

    return JSONResponse({"id": saved_id, "status": "ok"})


@app.get("/api/ws/rulesets/{rs_id}")

async def ws_get_ruleset(rs_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)


    rec = _ws_db.get_ruleset(rs_id, username)

    if not rec:

        raise HTTPException(404, "Rule set not found.")

    return JSONResponse(rec)


@app.delete("/api/ws/rulesets/{rs_id}")

async def ws_delete_ruleset(rs_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    _ws_db.delete_ruleset(rs_id, username)

    return JSONResponse({"status": "deleted"})


# -- jobs
# --------------------------------------------------------------------


@app.get("/api/ws/jobs")

async def ws_list_jobs(request: Request):


    _ws_check()

    username = _ws_get_user(request)

    return JSONResponse(_ws_db.list_jobs(username))


@app.post("/api/ws/jobs")

async def ws_save_job(request: Request):

    _ws_check()

    username = _ws_get_user(request)

    body = await request.json()

    name        = str(body.get("name", "")).strip()

    action      = (body.get("action") or "compare").strip().lower()

    source_conn_id = body.get("source_conn_id") or None

    conn_a_id   = body.get("conn_a_id") or None

    conn_b_id   = body.get("conn_b_id") or None

    key_columns = (body.get("key_columns") or "").strip() or None

    exclude_columns = (body.get("exclude_columns") or "").strip() or None

    ruleset_id  = body.get("ruleset_id") or None

    schedule_cron  = (body.get("schedule_cron") or "").strip() or None


    from_email  = (body.get("from_email")   or "").strip() or None

    notify_email = (body.get("notify_email") or "").strip() or None

    job_id      = body.get("id")


    if not name:

        raise HTTPException(400, "Job name is required.")


    valid_actions = {"compare", "quality", "profile", "lineage", "governance", "parse", "xref", "xref_ai"}

    if action not in valid_actions:

        raise HTTPException(400, f"action must be one of: {sorted(valid_actions)}")


    # Validate cron expression if provided (must be 5 fields)

    if schedule_cron and len(schedule_cron.split()) != 5:

        raise HTTPException(400, "schedule_cron must be a 5-field cron expression (min hour day month dow).")


    # SLA config

    sla_raw = body.get("sla") or {}

    sla_json = json.dumps(sla_raw) if sla_raw else None


    ai_hints = body.get("ai_hints") or {}

    ai_hints_json = json.dumps(ai_hints) if ai_hints else None


    fan_out_pairs = body.get("fan_out_pairs") or None

    saved_id = _ws_db.save_job(

        username=username, name=name, action=action, source_conn_id=source_conn_id,

        conn_a_id=conn_a_id,

        conn_b_id=conn_b_id,

        key_columns=key_columns,

        exclude_columns=exclude_columns,

        ruleset_id=ruleset_id,

        schedule_cron=schedule_cron,

        from_email=from_email,

        notify_email=notify_email,

        job_id=job_id,

        fan_out_pairs=fan_out_pairs,

        sla_json=sla_json,

        ai_hints_json=ai_hints_json,

    )


    # Register / update the APScheduler entry

    if schedule_cron:

        _ws_schedule_job(saved_id, username, schedule_cron)


    elif job_id:

        _ws_unregister_job(job_id)


    return JSONResponse({"id": saved_id, "status": "ok"})


@app.get("/api/ws/jobs/{job_id}")
async def ws_get_job(job_id: str, request: Request):
    """Single-job fetch for the Edit flow -- mirrors GET /api/ws/connections/{id}.
    get_job() already decodes sla_json/fan_out_pairs/ai_hints_json into plain
    dicts/lists, unlike the raw list endpoint, so this is the version the
    edit form should actually consume."""
    _ws_check()
    username = _ws_get_user(request)
    job = _ws_db.get_job(job_id, username)
    if not job:
        raise HTTPException(404, f"Job '{job_id}' not found.")
    return JSONResponse(_sanitize_json(job))


@app.delete("/api/ws/jobs/{job_id}")

async def ws_delete_job(job_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    _ws_unregister_job(job_id)

    _ws_db.delete_job(job_id, username)

    return JSONResponse({"status": "deleted"})


@app.post("/api/ws/jobs/{job_id}/run")

async def ws_run_job(job_id: str, request: Request):

    """Manually trigger a job run."""

    _ws_check()


    username = _ws_get_user(request)

    job = _ws_db.get_job(job_id, username)

    if not job:

        raise HTTPException(404, "Job not found.")

    try:
        # Fire-and-poll instead of blocking the request on the full fetch +
        # analyze + email pipeline: a slow or unreachable connector can take
        # minutes, and a platform proxy (e.g. Railway) will drop the
        # connection well before that -- the button would look like it did
        # nothing even though the job kept running server-side. The run row
        # is created synchronously (fast) and returned immediately; the
        # frontend polls GET /api/ws/runs/{run_id} for the real outcome.
        run_id = await asyncio.to_thread(_ws_trigger_now_bg, job_id, username)
        return JSONResponse({"run_id": run_id, "status": "running"})

    except Exception as exc:

        raise HTTPException(500, str(exc)) from exc


@app.get("/api/ws/runs/{run_id}")
async def ws_get_run(run_id: str, request: Request):
    """Poll a single run's status -- used by the "Run Now" button to show
    the real pass/fail/summary/email outcome once a background run finishes,
    without blocking the original request on it."""
    _ws_check()
    username = _ws_get_user(request)
    run = _ws_db.get_run(run_id, username)
    if not run:
        raise HTTPException(404, "Run not found.")
    summary = json.loads(run["summary_json"]) if run.get("summary_json") else None
    return JSONResponse({
        "run_id": run["id"],
        "status": run.get("status"),
        "summary": summary,
        "error": run.get("error_msg"),
        "email_sent": bool(run.get("email_sent")),
        "email_skipped_reason": run.get("email_skipped_reason"),
        "email_error": run.get("email_error"),
    })


@app.get("/api/ws/jobs/{job_id}/runs")
async def ws_job_runs(job_id: str, request: Request, limit: int = 50):
    """Run history for a single job."""
    _ws_check()
    username = _ws_get_user(request)
    return JSONResponse(_ws_db.list_runs(username, job_id=job_id, limit=limit))


# -- Dashboard
# --------------------------------------------------------------------


@app.get("/api/ws/dashboard")
async def ws_dashboard(request: Request):
    """Return workspace dashboard summary: job health, recent runs, DQ trend, saved analyses."""
    _ws_check()
    username = _ws_get_user(request)

    # ==== GAP: the body of this dashboard-summary endpoint was not recoverable from
    # the scan (source pages jump straight from 0916 to 0917/0918, which begin the next
    # route, /api/ws/jobs/{job_id}/sla). RECONSTRUCTED (unverified) below using the
    # _ws_db methods available elsewhere in this file (list_jobs, list_saved_runs) to
    # produce a summary matching the docstring's stated scope; verify against source
    # if available.
    jobs = _ws_db.list_jobs(username)
    saved_runs = _ws_db.list_saved_runs(username, limit=10)

    return JSONResponse({
        "job_count":   len(jobs),
        "jobs":        jobs,
        "recent_runs": saved_runs,
    })


@app.get("/api/ws/jobs/{job_id}/sla")

async def ws_get_sla(job_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    return JSONResponse(_ws_db.get_job_sla(job_id, username))


@app.post("/api/ws/jobs/from-session")
async def ws_job_from_session(request: Request):
    """Turn an already-completed interactive run into a recurring scheduled
    job, reusing the same saved connector(s) it used -- the "Save to
    Workspace, then also keep it running automatically" flow, so the user
    never has to re-pick connections from scratch in a separate form.
    Requires the session to have used at least one Saved Connector (not a
    raw file upload) for its source(s), since a schedule needs something to
    re-fetch fresh data from on every run."""
    _ws_check()
    username = _ws_get_user(request)
    body = await request.json()
    session_id = body.get("session_id")
    stored = _results_store.get(session_id)
    if not stored:
        raise HTTPException(404, "Session not found or expired -- please re-run the analysis.")

    action = stored.get("action")
    if action not in ("compare", "quality", "profile"):
        raise HTTPException(400, f"Scheduling isn't supported for '{action}' yet (only compare/quality/profile).")

    name = (body.get("name") or "").strip() or f"{action} job"
    schedule_cron = (body.get("schedule_cron") or "").strip() or None
    notify_email = (body.get("notify_email") or "").strip() or None
    if schedule_cron and len(schedule_cron.split()) != 5:
        raise HTTPException(400, "schedule_cron must be a 5-field cron expression (min hour day month dow).")

    conn_a_id = conn_b_id = source_conn_id = None
    if action == "compare":
        conn_a_ids = stored.get("conn_a_ids") or []
        conn_b_ids = stored.get("conn_b_ids") or []
        if not conn_a_ids or not conn_b_ids:
            raise HTTPException(
                400,
                "This run didn't use a Saved Connector on both Dataset A and B -- scheduling "
                "needs a connector to re-fetch from, not a one-off file upload. Re-run using "
                "Saved Connectors on both sides, then try again.",
            )
        conn_a_id, conn_b_id = int(conn_a_ids[0]), int(conn_b_ids[0])
    else:
        conn_ids = stored.get("conn_ids") or []
        if not conn_ids:
            raise HTTPException(
                400,
                "This run didn't use a Saved Connector -- scheduling needs a connector to "
                "re-fetch from, not a one-off file upload. Re-run using a Saved Connector, "
                "then try again.",
            )
        source_conn_id = int(conn_ids[0])

    sla_raw = body.get("sla") or {}
    sla_json = json.dumps(sla_raw) if sla_raw else None

    job_id = _ws_db.save_job(
        username=username, name=name, action=action,
        source_conn_id=source_conn_id, conn_a_id=conn_a_id, conn_b_id=conn_b_id,
        schedule_cron=schedule_cron, notify_email=notify_email, sla_json=sla_json,
    )
    if schedule_cron:
        _ws_schedule_job(job_id, username, schedule_cron)
    return JSONResponse({"job_id": job_id, "status": "ok"})


# -- AI job assistant -- suggest key columns, schedule, thresholds
# --------------------------------------------------------------------


@app.post("/api/ws/jobs/ai-suggest")

async def ws_ai_suggest(request: Request):


    # Ask the LLM to suggest key columns, schedule, and SLA thresholds

    # for a new job based on the source schema.


    _ws_check()

    username = _ws_get_user(request)

    body = await request.json()


    conn_a_id = body.get("conn_a_id") or body.get("source_conn_id")

    action    = body.get("action", "compare")


    conn = _ws_db.get_connection(conn_a_id, username) if conn_a_id else None

    conn_name = conn.get("name", "unknown source") if conn else "unknown source"

    source_type = conn.get("source_type", "") if conn else ""


    prompt = (

      f"You are a BFSI data engineer. A user wants to create a scheduled "

      f"{'comparison' if action=='compare' else 'data quality'} job for the data source "

      f"'{conn_name}' (type: {source_type}).\n\n"

      f"Suggest:\n"

      f"1. key_columns: likely unique identifier column names for this source "

      f"(e.g. trade_id, isin, account_id). Return as comma-separated string.\n"

      f"2. schedule: best cron schedule for BFSI ops "

      f"(e.g. '0 8 * * 1-5' = weekdays at 8am). Return the cron string.\n"

      f"3. max_breaks: SLA threshold for acceptable break count (integer).\n"

      f"4. min_dq_score: minimum acceptable DQ score 0-100 (integer).\n"

      f"5. description: one sentence explaining what this job monitors.\n\n"


      f"Return ONLY valid JSON: "

      f'{{\"key_columns\":\"\",\"schedule\":\"\",\"max_breaks\":null,'

      f'\"min_dq_score\":null,\"description\":\"\"}}'

    )

    try:

        raw = await asyncio.to_thread(_ask_llm, [{"role": "user", "content": [{"text": prompt}]}])

        import re as _re

        m = _re.search(r'\{.*\}', raw, _re.DOTALL)

        suggestion = json.loads(m.group(0)) if m else {}

    except Exception:

        suggestion = {}

    return JSONResponse({"suggestion": suggestion})


# -- saved runs (manual "Save to Workspace" from results page)
# --------------------------------------------------------------------


@app.post("/api/ws/saved-runs")

async def ws_save_run(request: Request):

    _ws_check()


    username = _ws_get_user(request)

    body = await request.json()

    name        = str(body.get("name", "")).strip()

    action      = str(body.get("action", "")).strip()

    sources     = body.get("sources", [])

    session_id  = body.get("session_id") or None

    summary     = body.get("summary") or {}

    conn_a_id   = body.get("conn_a_id") or None

    conn_b_id   = body.get("conn_b_id") or None

    source_conn_id = body.get("source_conn_id") or conn_a_id or None

    key_columns = (body.get("key_columns") or "").strip() or None


    if not name:

        raise HTTPException(400, "Run name is required.")

    if not isinstance(sources, list):

        sources = [str(sources)]


    run_id = _ws_db.save_manual_run(

        username, name, action, sources, session_id, summary,


        conn_a_id=conn_a_id, conn_b_id=conn_b_id, key_columns=key_columns,

        source_conn_id=source_conn_id,

    )

    return JSONResponse({"id": run_id, "status": "saved"})


@app.get("/api/ws/saved-runs")

async def ws_list_saved_runs(request: Request, limit: int = 100, job_id: str = ""):

    _ws_check()

    username = _ws_get_user(request)

    source_conn_id = conn_a_id = conn_b_id = None
    if job_id:
        job = _ws_db.get_job(job_id, username)
        if job:
            source_conn_id = job.get("source_conn_id")
            conn_a_id = job.get("conn_a_id")
            conn_b_id = job.get("conn_b_id")

    runs = _ws_db.list_saved_runs(username, limit=min(limit, 500),
                                   source_conn_id=source_conn_id,
                                   conn_a_id=conn_a_id, conn_b_id=conn_b_id)

    return JSONResponse(runs)


@app.patch("/api/ws/saved-runs/{run_id}")

async def ws_update_saved_run(run_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    body = await request.json()


    new_name = str(body.get("name", "")).strip()

    if not new_name:

        raise HTTPException(400, "Name is required.")

    _ws_db.rename_saved_run(run_id, username, new_name)

    return JSONResponse({"status": "updated"})


@app.delete("/api/ws/saved-runs/{run_id}")

async def ws_delete_saved_run(run_id: str, request: Request):

    _ws_check()

    username = _ws_get_user(request)

    _ws_db.delete_saved_run(run_id, username)

    return JSONResponse({"status": "deleted"})


# -- current user info
# --------------------------------------------------------------------


@app.get("/api/ws/me")


async def ws_me(request: Request):

    _ws_check()

    username = _ws_get_user(request)

    return JSONResponse({
        "username": username,
        "display_name": getattr(request.state, "display_name", username),
        "role": getattr(request.state, "role", None) or _ws_db.get_user_role(username),
    })


@app.get("/api/ws/users")
async def ws_list_users(request: Request):
    """List all workspace users with role, last-active, token usage this
    month, and Dataset Memory (saved rules) usage. Admin only."""
    _ws_check()
    username = _ws_get_user(request)
    role = getattr(request.state, "role", None) or _ws_db.get_user_role(username)
    if role != "admin":
        raise HTTPException(403, "Admin access required.")
    from datetime import datetime as _dt
    month_key = _dt.utcnow().strftime("%Y-%m")
    users = _ws_db.list_users()
    for u in users:
        try:
            u["tokens_this_month"] = _ws_db.get_token_usage_month_total(u["username"], month_key)["total_tokens"]
        except Exception:
            u["tokens_this_month"] = 0
        try:
            u["dataset_memory_count"] = len(_fp_list_datasets(u["username"]))
        except Exception:
            u["dataset_memory_count"] = 0
    return JSONResponse(users)


@app.put("/api/ws/users/{username}/role")
async def ws_set_user_role(username: str, request: Request):
    """Change a user's role. Admin only."""
    _ws_check()
    caller = _ws_get_user(request)
    caller_role = getattr(request.state, "role", None) or _ws_db.get_user_role(caller)
    if caller_role != "admin":
        raise HTTPException(403, "Admin access required.")
    body = await request.json()
    role = str(body.get("role", "")).strip()
    if role not in ("admin", "analyst", "readonly"):
        raise HTTPException(400, "role must be one of: admin, analyst, readonly.")
    if username == caller and role != "admin":
        raise HTTPException(400, "You cannot remove your own admin access.")
    _ws_db.set_user_role(username, role)
    return JSONResponse({"ok": True})


@app.put("/api/ws/users/{username}/block")
async def ws_set_user_blocked(username: str, request: Request):
    """Block or unblock a user -- blocked users cannot log in and every
    Workspace API call for them is rejected immediately. Admin only."""
    _ws_check()
    caller = _ws_get_user(request)
    caller_role = getattr(request.state, "role", None) or _ws_db.get_user_role(caller)
    if caller_role != "admin":
        raise HTTPException(403, "Admin access required.")
    if username == caller:
        raise HTTPException(400, "You cannot block your own account.")
    body = await request.json()
    blocked = bool(body.get("blocked", True))
    _ws_db.set_user_blocked(username, blocked)
    return JSONResponse({"ok": True, "blocked": blocked})


@app.delete("/api/ws/users/{username}")
async def ws_delete_user(username: str, request: Request):
    """Permanently delete a user and everything scoped to them (connections,
    rulesets, jobs, run history, saved runs, audit log, token usage, and
    Dataset Memory rules). Admin only."""
    _ws_check()
    caller = _ws_get_user(request)
    caller_role = getattr(request.state, "role", None) or _ws_db.get_user_role(caller)
    if caller_role != "admin":
        raise HTTPException(403, "Admin access required.")
    if username == caller:
        raise HTTPException(400, "You cannot delete your own account.")
    _ws_db.delete_user(username)
    try:
        _fp_delete_user(username)
    except Exception:
        pass
    return JSONResponse({"ok": True})


# ------------------------------------------------------------------

# Audit Log  (persist + SSE stream)

# ------------------------------------------------------------------


try:

    from workspace.audit_log import (

        log_action  as _audit_log,

        subscribe   as _audit_subscribe,

        unsubscribe as _audit_unsubscribe,

        list_audit_log as _audit_list,

    )

    _AUDIT_ENABLED = True


except Exception as _audit_import_err:

    import logging as _lg_audit

    _lg_audit.getLogger(__name__).warning("Audit log feature disabled: %s", _audit_import_err)

    _AUDIT_ENABLED = False


@app.get("/api/logs/stream")

async def audit_log_stream(request: Request):


    # SSE endpoint -- streams new audit log events as they happen.

    # Each event is a JSON object serialised as a 'data:' SSE line.


    if not _AUDIT_ENABLED:

        raise HTTPException(status_code=503, detail="Audit log not available.")


    import asyncio


    q = _audit_subscribe()


    async def _event_generator():

        try:

            while True:

                # Race: next log event vs client disconnect

                get_task  = asyncio.ensure_future(q.get())

                disc_task = asyncio.ensure_future(request.is_disconnected())

                done, pending = await asyncio.wait(

                  [get_task, disc_task],

                  timeout=15.0,

                  return_when=asyncio.FIRST_COMPLETED,

                )

                # Cancel and await losing tasks to avoid "Task destroyed but pending" warnings

                for t in pending:

                    t.cancel()

                    try:

                        await t

                    except (asyncio.CancelledError, Exception):


                        pass

                if not done:

                    # timeout -- send keep-alive

                    yield ": keep-alive\n\n"

                    continue

                for fut in done:

                    try:

                        result = fut.result()

                    except Exception:

                        continue

                    if result is True:

                        # is_disconnected() returned True -- client gone

                        return

                    if isinstance(result, str):

                        yield f"data: {result}\n\n"

        finally:

            _audit_unsubscribe(q)


    return StreamingResponse(

        _event_generator(),

        media_type="text/event-stream",

        headers={

          "Cache-Control": "no-cache",

          "X-Accel-Buffering": "no",

        },

    )


@app.get("/api/logs")

async def audit_log_list(

    request: Request,

    username: Optional[str] = None,

    action: Optional[str] = None,

    limit: int = 200,

):

    """Return recent audit log entries as JSON."""

    if not _AUDIT_ENABLED:


        raise HTTPException(status_code=503, detail="Audit log not available.")

    entries = _audit_list(username=username, action=action, limit=min(limit, 1000))

    return JSONResponse(entries)


# -- SSO Routes -- SAML 2.0 and OIDC / Azure AD
# --------------------------------------------------------------------


import secrets as _secrets

_sso_state_store: dict[str, str] = {}  # state -> "pending" (in-memory CSRF store)


@app.get("/sso/login")

async def sso_login(request: Request):


    # Redirect the browser to the configured IdP login page.

    # Supports both SAML 2.0 (AD FS) and OIDC (Azure AD).

    # Returns 404 if SSO is not configured.


    try:

        if not _sso_enabled():

            raise HTTPException(status_code=404, detail="SSO not configured.")

        mode = _sso_mode()

        if mode == "saml":

            url = await _saml_login_url(request)

            return RedirectResponse(url=url, status_code=302)

        if mode == "oidc":

            state = _secrets.token_urlsafe(16)

            _sso_state_store[state] = "pending"

            url = _oidc_login_url(state)

            return RedirectResponse(url=url, status_code=302)

        raise HTTPException(status_code=400, detail=f"Unknown SSO mode: {mode!r}")

    except HTTPException:

        raise

    except Exception as exc:

        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/sso/saml/acs")

async def sso_saml_acs(request: Request):


    # SAML Assertion Consumer Service endpoint.

    # AD FS POSTs the SAML response here after successful login.

    # Creates a session cookie and redirects to the home page.


    try:

        if not _sso_enabled() or _sso_mode() != "saml":

            raise HTTPException(status_code=404, detail="SAML SSO not configured.")

        user_info = await _saml_process(request)

        token = _sso_create_token(

          user_info["username"],

          user_info["role"],

          user_info["groups"],

        )

        # Ensure user is in workspace DB


        from workspace.db import ensure_user as _ensure_user

        _ensure_user(

          user_info["username"],

          user_info.get("display_name", user_info["username"]),

          user_info.get("email", ""),

        )

        # Sync the group-resolved role on every login -- but only when
        # SSO_ROLE_MAP is actually configured, so a manually-promoted role
        # (set via the workspace user admin UI) is never silently
        # overwritten by SSO's own "analyst" default when no mapping exists.
        from workspace.sso import SSO_ROLE_MAP as _SSO_ROLE_MAP
        if _SSO_ROLE_MAP:
            _ws_db.set_user_role(user_info["username"], user_info["role"])

        response = RedirectResponse(url="/", status_code=302)

        response.set_cookie(

          key      = "dv_session",

          value    = token,

          httponly = True,

          samesite = "lax",

          secure   = request.url.scheme == "https",

        )

        return response

    except HTTPException:

        raise

    except Exception as exc:


        raise HTTPException(status_code=401, detail=f"SAML authentication failed: {exc}") from exc


@app.get("/sso/oidc/callback")

async def sso_oidc_callback(

    request: Request,

    code:  Optional[str] = None,

    state: Optional[str] = None,

    error: Optional[str] = None,

):


    # OIDC authorisation code callback.

    # Azure AD redirects here after successful login with ?code=...&state=...


    try:

        if not _sso_enabled() or _sso_mode() != "oidc":

            raise HTTPException(status_code=404, detail="OIDC SSO not configured.")

        if error:


            raise HTTPException(status_code=401, detail=f"OIDC error: {error}")

        if not code:

            raise HTTPException(status_code=400, detail="Missing authorisation code.")

        # CSRF state check

        if state not in _sso_state_store:

            raise HTTPException(status_code=400, detail="Invalid or expired CSRF state.")

        del _sso_state_store[state]


        user_info = _oidc_exchange(code)

        token   = _sso_create_token(

          user_info["username"],

          user_info["role"],

          user_info["groups"],

        )

        from workspace.db import ensure_user as _ensure_user

        _ensure_user(

          user_info["username"],

          user_info.get("display_name", user_info["username"]),

          user_info.get("email", ""),


        )

        from workspace.sso import SSO_ROLE_MAP as _SSO_ROLE_MAP
        if _SSO_ROLE_MAP:
            _ws_db.set_user_role(user_info["username"], user_info["role"])

        response = RedirectResponse(url="/", status_code=302)

        response.set_cookie(

          key      = "dv_session",

          value    = token,

          httponly = True,

          samesite = "lax",

          secure   = request.url.scheme == "https",

        )

        return response

    except HTTPException:

        raise

    except Exception as exc:

        raise HTTPException(status_code=401, detail=f"OIDC authentication failed: {exc}") from exc


@app.get("/sso/logout")

async def sso_logout():


    """Clear the SSO session cookie and redirect to login."""

    response = RedirectResponse(url="/sso/login", status_code=302)

    response.delete_cookie("dv_session")

    return response


@app.get("/api/auth/me")

async def auth_me(request: Request):

    """Return the current user's identity and role. Useful for UI role-gating."""

    username = _ws_resolve_username(request)

    if not username:

        raise HTTPException(status_code=401, detail="Not authenticated.")

    from workspace.auth import resolve_role_for_user as _resolve_role

    role = getattr(request.state, "role", _resolve_role(username))

    # ==== GAP: this is the last page of the scanned document (0937) -- the photo
    # cuts off mid dict-literal here, right after "sso_mode". Only the closing
    # `})` below is RECONSTRUCTED (unverified); every key up to this point is
    # verbatim from the scan. Verify against source if a later page becomes available.
    return JSONResponse({
      "username": username,
      "role":     role,
      "sso_enabled": _sso_enabled(),
      "sso_mode":    _sso_mode(),
    })

